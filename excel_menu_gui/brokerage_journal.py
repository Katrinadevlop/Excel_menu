import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path
from datetime import datetime
import re
from typing import List, Dict, Optional, Tuple

from dish_extractor import extract_categorized_dishes_from_menu, extract_date_from_menu

class BrokerageJournalGenerator:
    """Генератор бракеражного журнала на основе меню"""
    
    def __init__(self):
        pass
    
    def extract_date_from_menu(self, menu_path: str) -> Optional[datetime]:
        """Извлекает дату из файла меню"""
        return extract_date_from_menu(menu_path)
    

    def extract_categorized_dishes(self, menu_path: str) -> Dict[str, List[str]]:
        """Извлекает блюда из меню, распределяя по категориям:
        завтраки, салаты и холодные закуски, первые блюда, блюда из мяса, блюда из птицы, блюда из рыбы, гарниры
        """
        return extract_categorized_dishes_from_menu(menu_path)


    def _is_section_header(self, text: str) -> bool:
        """Определяет, является ли строка заголовком раздела (не блюдом)."""
        if not text:
            return False
        txt = str(text).strip()
        lower = txt.lower()
        
        # Точные заголовки разделов, которые точно нужно исключить
        exact_headers = [
            'салаты и холодные закуски',
            'сэндвичи',
            'пельмени', 
            'вареники',
            'соусы',
            'первые блюда',
            'блюда из мяса',
            'блюда из птицы',
            'блюда из рыбы',
            'гарниры',
            'завтраки'
        ]
        
        # Проверяем точное совпадение с заголовками
        if lower in exact_headers:
            return True
            
        # Строки полностью в верхнем регистре и очень короткие (вероятно заголовки)
        if txt.isupper() and len(txt) <= 15 and any(h in lower for h in ['блюд', 'салат', 'сэндвич']):
            return True
            
        return False
    
    def create_brokerage_journal(self, menu_path: str, template_path: str, output_path: str) -> Tuple[bool, str]:
        """Создает бракеражный журнал на основе меню с листа касс, заполняя столбец A завтраками, а столбец G - остальными блюдами. Время не изменяем."""
        try:
            # Проверяем существование шаблона
            if not Path(template_path).exists():
                return False, f"Шаблон бракеражного журнала не найден: {template_path}"
            
            # Извлекаем дату из меню
            menu_date = self.extract_date_from_menu(menu_path)
            if not menu_date:
                menu_date = datetime.now()
            
            # Извлекаем блюда по категориям
            categories = self.extract_categorized_dishes(menu_path)
            
            # Отладочный вывод результата
            print(f"\nРезультат извлечения категорий:")
            for category, dishes in categories.items():
                print(f"{category}: {len(dishes)} блюд - {dishes}")
            
            # Собираем завтраки и салаты для левого столбца (A)
            left_list: List[str] = []
            left_list.extend(categories.get('завтрак', []))
            left_list.extend(categories.get('салат', []))
            # Удаляем заголовки разделов
            left_list = [d for d in left_list if not self._is_section_header(d)]
            print(f"\nКоличество блюд для левого столбца (A): {len(left_list)}")
            
            # Собираем остальные блюда для правого столбца (G)
            right_list: List[str] = []
            right_list.extend(categories.get('первое', []))
            right_list.extend(categories.get('мясо', []))
            right_list.extend(categories.get('курица', []))
            right_list.extend(categories.get('птица', []))
            right_list.extend(categories.get('рыба', []))
            right_list.extend(categories.get('гарнир', []))
            # Удаляем заголовки разделов
            right_list = [d for d in right_list if not self._is_section_header(d)]
            print(f"Количество блюд для правого столбца (G): {len(right_list)}")
            
            # Открываем шаблон
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # Обновляем дату в шаблоне (строка 3, колонка 1)
            russian_months = {
                1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля', 5: 'мая', 6: 'июня',
                7: 'июля', 8: 'августа', 9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'
            }
            date_str_display = f"{menu_date.day} {russian_months.get(menu_date.month, 'unknown')}"
            ws.cell(row=3, column=1, value=date_str_display)
            
            # Форматируем дату для названия листа
            date_str = menu_date.strftime('%d.%m.%y')
            ws.title = date_str
            
            # Находим строку заголовков таблицы ("НАИМЕНОВАНИЕ БЛЮДА" в колонке A)
            header_row = None
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=1).value
                if v and 'наименование' in str(v).lower() and 'блюд' in str(v).lower():
                    header_row = r
                    break
            if header_row is None:
                return False, 'Не удалось определить заголовок таблицы в шаблоне'
            
            start_row = header_row + 1  # первая строка для блюд
            
            # Определяем первую полностью пустую строку: дальше НИЧЕГО не трогаем
            stop_row = start_row
            while stop_row <= ws.max_row:
                row_empty = True
                for c in range(1, 10):  # A..I
                    if ws.cell(row=stop_row, column=c).value not in (None, ''):
                        row_empty = False
                        break
                if row_empty:
                    break
                stop_row += 1
            
            # Заполняем ТОЛЬКО пустые ячейки первого столбца (A) блюдами из left_list
            inserted_left = 0
            dish_idx = 0
            for r in range(start_row, stop_row):
                if dish_idx >= len(left_list):
                    break
                current_val = ws.cell(row=r, column=1).value
                if current_val in (None, ''):
                    ws.cell(row=r, column=1, value=left_list[dish_idx])
                    # НЕ МЕНЯЕМ время - оставляем как в шаблоне
                    dish_idx += 1
                    inserted_left += 1
                else:
                    # Ячейка занята — не трогаем её и идем дальше
                    continue
            
            # Заполняем ТОЛЬКО пустые ячейки седьмого столбца (G) блюдами из right_list
            inserted_right = 0
            dish_idx = 0
            for r in range(start_row, stop_row):
                if dish_idx >= len(right_list):
                    break
                current_val = ws.cell(row=r, column=7).value  # Столбец G = 7
                if current_val in (None, ''):
                    ws.cell(row=r, column=7, value=right_list[dish_idx])
                    # НЕ МЕНЯЕМ время - оставляем как в шаблоне
                    dish_idx += 1
                    inserted_right += 1
                else:
                    # Ячейка занята — не трогаем её и идем дальше
                    continue
            
            wb.save(output_path)
            return True, f"Бракеражный журнал создан успешно для даты {date_str} (вставлено {inserted_left} блюд в колонку A, {inserted_right} блюд в колонку G)"
        except Exception as e:
            return False, f"Ошибка при создании бракеражного журнала: {str(e)}"
    
    def _create_header(self, ws, date: datetime):
        """Создает заголовок бракеражного журнала"""
        # Дата
        date_str = date.strftime("%d %B").replace('September', 'сентября').replace('August', 'августа')
        ws.cell(row=3, column=1, value=date_str)
        
        # Заголовок
        ws.cell(row=1, column=1, value="БРАКЕРАЖНЫЙ ЖУРНАЛ")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        
        # Объединяем ячейки для заголовка
        ws.merge_cells('A1:G1')


def create_brokerage_journal_from_menu(menu_path: str, template_path: str, output_path: str) -> Tuple[bool, str]:
    """Удобная функция для создания бракеражного журнала"""
    generator = BrokerageJournalGenerator()
    return generator.create_brokerage_journal(menu_path, template_path, output_path)
