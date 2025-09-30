
import os
import re
import math
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import List, Tuple, Optional, Set

import openpyxl

try:
    import xlrd  # for .xls reading
except ImportError:  # optional
    xlrd = None

try:
    import xlwings as xw  # optional, for .xls -> .xlsx conversion when needed
except ImportError:
    xw = None


class ColumnParseError(ValueError):
    pass


def col_to_index0(col: str) -> int:
    s = col.strip()
    if not s:
        raise ColumnParseError("Колонка не указана")
    if s.isdigit():
        n = int(s)
        if n <= 0:
            raise ColumnParseError("Номер колонки должен быть > 0")
        return n - 1
    acc = 0
    for ch in s.upper():
        if not ('A' <= ch <= 'Z'):
            raise ColumnParseError("Некорректная колонка. Пример: A, B, 1, 2, AA")
        acc = acc * 26 + (ord(ch) - ord('A') + 1)
    return acc - 1


def index0_to_col(idx: int) -> str:
    n = idx + 1
    out = []
    while n > 0:
        rem = (n - 1) % 26
        out.append(chr(ord('A') + rem))
        n = (n - 1) // 26
    return ''.join(reversed(out))


def normalize(s: Optional[str], ignore_case: bool) -> str:
    if not s:
        return ''
    s = str(s).strip()
    s = re.sub(r"\s+", " ", s)
    if ignore_case:
        s = s.lower()
    return s


def normalize_dish(s: Optional[str], ignore_case: bool) -> str:
    s = normalize(s, ignore_case)
    if not s:
        return ''
    # Убираем числовые данные в скобках (калории, граммы, цены)
    s = re.sub(r"\((?:[^)]*?\d\s*(?:к?кал|ккал|г|гр|грамм(?:а|ов)?|мл|л|кг|руб\.?|р\.?|₽)[^)]*?)\)", "", s, flags=re.IGNORECASE)
    # Убираем числовые данные без скобок
    s = re.sub(r"\b\d+[\.,]?\d*\s*(?:к?кал|ккал|г|гр|грамм(?:а|ов)?|мл|л|кг)\b\.?", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\b\d+[\.,]?\d*\s*(?:руб\.?|р\.?|₽)\b\.?", "", s, flags=re.IGNORECASE)
    s = re.sub(r"(?:₽|руб\.?|р\.?)\s*\d+[\.,]?\d*", "", s, flags=re.IGNORECASE)
    # Нормализуем варианты приготовления: убираем /вариант и оставляем основное блюдо
    # Например: "яйцо отварное/жареное" -> "яйцо отварное", "яйцо жареное" 
    s = re.sub(r"[\s,;:.-]+$", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_dish_with_variants(s: Optional[str], ignore_case: bool) -> List[str]:
    """
    Нормализует название блюда и возвращает все возможные варианты.
    Например: "Яйцо отварное/жареное" -> ["яйцо отварное", "яйцо жареное"]
    "Сосиска с сыром/с беконом" -> ["сосиска с сыром", "сосиска с беконом"]
    """
    base = normalize_dish(s, ignore_case)
    if not base:
        return []
    
    # Если нет слэша, возвращаем исходное блюдо
    if '/' not in base:
        return [base]
    
    # Разбиваем по слэшам для обработки множественных вариантов
    parts = [part.strip() for part in base.split('/')]
    if len(parts) < 2:
        return [base]
    
    variants = []
    
    # Определяем общий префикс и суффикс
    # Ищем самое длинное общее начало и конец всех частей
    def find_common_prefix(strings):
        if not strings:
            return ""
        prefix = strings[0]
        for s in strings[1:]:
            while prefix and not s.startswith(prefix):
                prefix = prefix[:-1]
        return prefix
    
    def find_common_suffix(strings):
        if not strings:
            return ""
        suffix = strings[0]
        for s in strings[1:]:
            while suffix and not s.endswith(suffix):
                suffix = suffix[1:]
        return suffix
    
    # Пытаемся найти структуру "префикс вариант1/вариант2 суффикс"
    # Сначала попробуем найти общие слова в начале и конце
    words = [part.split() for part in parts]
    
    # Простой подход: ищем паттерны
    found_pattern = False
    
    # Специальные случаи для повторяющихся предлогов
    
    # Паттерн: "с сыром/с беконом" или "по цене/по вкусу"
    pattern_with_prep = re.search(r'^(\S+\s+.+)/(\S+\s+.+)$', base.strip())
    if pattern_with_prep:
        part1 = pattern_with_prep.group(1).strip()
        part2 = pattern_with_prep.group(2).strip()
        
        # Проверяем, есть ли общий предлог (первое слово)
        part1_words = part1.split()
        part2_words = part2.split()
        
        if (len(part1_words) >= 2 and len(part2_words) >= 2 and 
            part1_words[0] == part2_words[0] and  # одинаковые предлоги
            part1_words[0] in ['с', 'по', 'для', 'от', 'на']):  # известные предлоги
            # Нашли повторяющиеся предлоги, пока просто возвращаем как есть
            variants = [part1, part2]
            found_pattern = True
        elif len(parts) == 2:
            # Проверяем общую основу для случаев типа "яйцо отварное/жареное"
            # Находим общие слова в начале
            words1 = part1.split()
            words2 = part2.split()
            
            # Находим общие префиксы
            common_prefix = []
            min_len = min(len(words1), len(words2))
            for i in range(min_len):
                if words1[i] == words2[i]:
                    common_prefix.append(words1[i])
                else:
                    break
            
            # Находим общие суффиксы
            common_suffix = []
            for i in range(1, min_len + 1):
                if len(words1) >= i and len(words2) >= i and words1[-i] == words2[-i]:
                    common_suffix.insert(0, words1[-i])
                else:
                    break
            
            if common_prefix or common_suffix:
                prefix_str = ' '.join(common_prefix) if common_prefix else ''
                suffix_str = ' '.join(common_suffix) if common_suffix else ''
                
                # Создаем варианты
                for part in [part1, part2]:
                    words = part.split()
                    start_idx = len(common_prefix)
                    end_idx = len(words) - len(common_suffix) if common_suffix else len(words)
                    middle_words = words[start_idx:end_idx] if start_idx < end_idx else []
                    
                    variant_parts = []
                    if prefix_str:
                        variant_parts.append(prefix_str)
                    if middle_words:
                        variant_parts.append(' '.join(middle_words))
                    if suffix_str:
                        variant_parts.append(suffix_str)
                    
                    variant = ' '.join(variant_parts).strip()
                    if variant:
                        variants.append(variant)
                
                if variants:
                    found_pattern = True
    
    # Паттерн 1: "слово1 слово2 вариант1/вариант2" 
    if not found_pattern:
        for i in range(len(parts)):
            current_words = parts[i].split()
            if len(current_words) > 1:
                # Проверяем, можно ли объединить с другими частями
                base_part = ' '.join(current_words[:-1])  # все кроме последнего слова
                variant_word = current_words[-1]  # последнее слово
                
                # Строим варианты, заменяя части после слэша
                temp_variants = []
                for j, part in enumerate(parts):
                    if j == i:
                        temp_variants.append(f"{base_part} {variant_word}")
                    else:
                        # Пытаемся найти соответствующую часть
                        other_words = part.split()
                        if len(other_words) >= 1:
                            temp_variants.append(f"{base_part} {' '.join(other_words)}")
                
                if len(temp_variants) == len(parts) and all(v.strip() for v in temp_variants):
                    variants.extend(temp_variants)
                    found_pattern = True
                    break
    
    # Паттерн 2: "префикс вариант1/вариант2 суффикс"
    if not found_pattern:
        # Ищем общие слова в начале всех частей
        all_words = [part.split() for part in parts if part.strip()]
        if all_words and all(len(words) > 0 for words in all_words):
            
            # Находим общий префикс по словам
            common_prefix_words = []
            min_len = min(len(words) for words in all_words)
            
            for i in range(min_len):
                word = all_words[0][i]
                if all(len(w) > i and w[i] == word for w in all_words):
                    common_prefix_words.append(word)
                else:
                    break
            
            # Находим общий суффикс по словам
            common_suffix_words = []
            for i in range(1, min_len + 1):
                word = all_words[0][-i]
                if all(len(w) >= i and w[-i] == word for w in all_words):
                    common_suffix_words.insert(0, word)
                else:
                    break
            
            prefix_str = ' '.join(common_prefix_words) if common_prefix_words else ''
            suffix_str = ' '.join(common_suffix_words) if common_suffix_words else ''
            
            # Создаем варианты
            for part in parts:
                words_in_part = part.split()
                
                # Убираем общий prefix и suffix, оставляем middle
                start_idx = len(common_prefix_words)
                end_idx = len(words_in_part) - len(common_suffix_words) if common_suffix_words else len(words_in_part)
                
                middle_words = words_in_part[start_idx:end_idx]
                
                # Собираем полный вариант
                variant_parts = []
                if prefix_str:
                    variant_parts.append(prefix_str)
                if middle_words:
                    variant_parts.append(' '.join(middle_words))
                if suffix_str:
                    variant_parts.append(suffix_str)
                
                variant = ' '.join(variant_parts).strip()
                if variant:
                    variants.append(variant)
            
            if variants:
                found_pattern = True
    
    # Если паттерн не найден, просто используем исходные части как варианты
    if not found_pattern or not variants:
        variants = [part.strip() for part in parts if part.strip()]
    
    # Убираем дубликаты и пустые строки, нормализуем пробелы
    final_variants = []
    for v in variants:
        normalized = re.sub(r'\s+', ' ', v).strip()
        if normalized and normalized not in final_variants:
            final_variants.append(normalized)
    
    return final_variants


def levenshtein(a: str, b: str) -> int:
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    dp = [list(range(len(b) + 1))]
    dp += [[i] + [0] * len(b) for i in range(1, len(a) + 1)]
    for i in range(1, len(a) + 1):
        for j in range(1, len(b) + 1):
            cost = 0 if a[i - 1] == b[j - 1] else 1
            dp[i][j] = min(dp[i - 1][j] + 1, dp[i][j - 1] + 1, dp[i - 1][j - 1] + cost)
    return dp[-1][-1]


def sim_percent(a: str, b: str) -> int:
    if not a and not b:
        return 100
    dist = levenshtein(a, b)
    m = max(len(a), len(b))
    if m == 0:
        return 100
    return int(round((1.0 - dist / m) * 100.0))


def get_sheet_names(path: str) -> List[str]:
    ext = Path(path).suffix.lower()
    if ext in ('.xlsx', '.xlsm'):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            return wb.sheetnames
        finally:
            wb.close()
    elif ext == '.xls':
        if xlrd is None:
            raise RuntimeError("Для .xls установите xlrd==1.2.0")
        book = xlrd.open_workbook(path, on_demand=True)
        try:
            return book.sheet_names()
        finally:
            book.release_resources()
    else:
        raise RuntimeError("Неподдерживаемый формат файла. Используйте .xls/.xlsx/.xlsm")


def read_cell_values(path: str, sheet_name: str) -> List[List[Optional[str]]]:
    """Читает лист как сетку значений (строки -> ячейки -> строка)."""
    ext = Path(path).suffix.lower()
    values: List[List[Optional[str]]] = []
    if ext in ('.xlsx', '.xlsm'):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sh = wb[sheet_name]
            for row in sh.iter_rows(values_only=True):
                values.append([None if v is None else str(v) for v in row])
        finally:
            wb.close()
    elif ext == '.xls':
        if xlrd is None:
            raise RuntimeError("Для .xls установите xlrd==1.2.0")
        book = xlrd.open_workbook(path)
        try:
            sh = book.sheet_by_name(sheet_name)
            for r in range(sh.nrows):
                row = []
                for c in range(sh.ncols):
                    v = sh.cell_value(r, c)
                    row.append(None if v == '' else str(v))
                values.append(row)
        finally:
            book.release_resources()
    else:
        raise RuntimeError("Неподдерживаемый формат")
    return values


def auto_detect_dish_column(path: str, sheet_name: str) -> Tuple[str, int]:
    vals = read_cell_values(path, sheet_name)
    keys = ["блюд", "наимен", "назван", "позици", "меню"]
    max_r = min(10, len(vals))
    for r in range(max_r):
        row = vals[r] if r < len(vals) else []
        for c, v in enumerate(row):
            s = normalize(v or '', True)
            if not s:
                continue
            for k in keys:
                if k in s:
                    return (index0_to_col(c), r + 1)  # column letter, header row 1-based
    # default
    return ("A", 1)


def ensure_xlsx(path: str) -> str:
    ext = Path(path).suffix.lower()
    if ext in ('.xlsx', '.xlsm'):
        return path
    if ext == '.xls':
        if xw is None:
            raise RuntimeError("Для .xls нужен xlwings (и установленный Excel), либо преобразуйте файл в .xlsx вручную.")
        # convert via Excel automation
        src = Path(path)
        dst = src.with_suffix('.xlsx')
        app = xw.App(visible=False, add_book=False)
        try:
            wb = app.books.open(str(src))
            wb.save(str(dst))
            wb.close()
        finally:
            app.quit()
        return str(dst)
    raise RuntimeError("Неподдерживаемый формат")


def _extract_dates_from_text(s: str) -> List[date]:
    """Ищет в строке даты в популярных форматах и возвращает список дат.
    Поддерживает: dd.mm.yyyy, dd-mm-yyyy, yyyy-mm-dd, dd.mm, dd-mm (год подставляется текущий),
    а также русские месяцы вида "5 сентября 2025" или "05 сен 2025".
    """
    if not s:
        return []
    s_norm = str(s).strip()
    out: List[date] = []
    today = date.today()

    # dd.mm.yyyy, dd-mm-yyyy, dd/mm/yyyy
    for m in re.finditer(r"\b(\d{1,2})[./\-](\d{1,2})[./\-](\d{2,4})\b", s_norm):
        d, mth, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            out.append(date(y, mth, d))
        except ValueError:
            pass

    # yyyy-mm-dd, yyyy.mm.dd
    for m in re.finditer(r"\b(\d{4})[./\-](\d{1,2})[./\-](\d{1,2})\b", s_norm):
        y, mth, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            out.append(date(y, mth, d))
        except ValueError:
            pass

    # dd.mm or dd-mm (assume nearest year to today if year is missing)
    for m in re.finditer(r"\b(\d{1,2})[./\-](\d{1,2})(?![./\-]\d)\b", s_norm):
        d, mth = int(m.group(1)), int(m.group(2))
        try:
            cand = date(today.year, mth, d)
            delta = (cand - today).days
            if delta > 183:
                # слишком далеко в будущем — считаем, что это прошедший год
                cand = date(today.year - 1, mth, d)
            elif delta < -183:
                # слишком далеко в прошлом — считаем, что это следующий год
                cand = date(today.year + 1, mth, d)
            out.append(cand)
        except ValueError:
            pass

    # Russian month names (both nominative and genitive, short forms too)
    months = {
        'янв': 1, 'январь': 1, 'января': 1,
        'фев': 2, 'февраль': 2, 'февраля': 2,
        'мар': 3, 'март': 3, 'марта': 3,
        'апр': 4, 'апрель': 4, 'апреля': 4,
        'май': 5, 'мая': 5,
        'июн': 6, 'июнь': 6, 'июня': 6,
        'июл': 7, 'июль': 7, 'июля': 7,
        'авг': 8, 'август': 8, 'августа': 8,
        'сен': 9, 'сентябрь': 9, 'сентября': 9,
        'oct': 10, 'окт': 10, 'октябрь': 10, 'октября': 10,
        'ноя': 11, 'ноябрь': 11, 'ноября': 11,
        'дек': 12, 'декабрь': 12, 'декабря': 12,
    }
    # 5 сентября 2025 / 5 сен 2025 / 5 сентября
    for m in re.finditer(r"\b(\d{1,2})\s+([A-Za-zА-Яа-яёЁ]+)\s*(\d{4})?\b", s_norm):
        d = int(m.group(1))
        mon_str = m.group(2).lower()
        y_str = m.group(3)
        mon_str = mon_str.replace('ё', 'е')
        if mon_str in months:
            mth = months[mon_str]
            if y_str:
                y = int(y_str)
                try:
                    out.append(date(y, mth, d))
                except ValueError:
                    pass
            else:
                # Без года — выберем ближайший к сегодня год
                try:
                    cand = date(today.year, mth, d)
                    delta = (cand - today).days
                    if delta > 183:
                        cand = date(today.year - 1, mth, d)
                    elif delta < -183:
                        cand = date(today.year + 1, mth, d)
                    out.append(cand)
                except ValueError:
                    pass

    # De-duplicate
    uniq = []
    seen = set()
    for dt in out:
        if dt not in seen:
            seen.add(dt)
            uniq.append(dt)
    return uniq


def _extract_best_date_from_file(path: str, sheet_name: Optional[str]) -> Optional[date]:
    """Пытается извлечь дату из имени файла и содержимого листа.
    Возвращает дату, максимально близкую к сегодня, отдавая приоритет последней дате не позже сегодня.
    """
    candidates: List[date] = []
    # from filename
    candidates += _extract_dates_from_text(Path(path).name)

    # from sheet content (scan small top area)
    try:
        if sheet_name:
            vals = read_cell_values(path, sheet_name)
            max_r = min(20, len(vals))
            max_c = 15
            for r in range(max_r):
                row = vals[r] if r < len(vals) else []
                for c in range(min(max_c, len(row))):
                    v = row[c]
                    if v is None:
                        continue
                    # if openpyxl returned datetime/date objects, handle them
                    if isinstance(v, (datetime, date)):
                        d = v.date() if isinstance(v, datetime) else v
                        candidates.append(d)
                    else:
                        candidates += _extract_dates_from_text(str(v))
    except Exception:
        # ignore any parsing errors, fallback to filename-based only
        pass

    if not candidates:
        return None

    today = date.today()
    # Разрешаем даты в пределах недели в будущем (для планов меню)
    near_future_limit = today + timedelta(days=7)
    not_too_future = [d for d in candidates if d <= near_future_limit]
    
    if not_too_future:
        # Предпочитаем даты не позже сегодня, но если их нет - берем ближайшую будущую в пределах недели
        not_future = [d for d in not_too_future if d <= today]
        if not_future:
            return max(not_future)
        else:
            return min(not_too_future)
    
    # если все даты слишком далеко в будущем — возьмём ближайшую
    return min(candidates)


def _find_category_ranges(values: List[List[Optional[str]]], synonyms_map: dict) -> dict:
    """Возвращает {canonical_key: (start_row_inclusive, end_row_inclusive)} в 1-базисе.
    Ищет по синонимам (ключи словаря), возвращает диапазоны по каноническим ключам (значения словаря).
    Пример: 'салаты', 'холодные закуски' -> 'салаты и холодные закуски'.
    """
    marker_rows = []  # list of (row1based, canonical_key)
    for r, row in enumerate(values, start=1):
        row_join = ' '.join([str(c) for c in row if c not in (None, '')])
        s = normalize(row_join, True)
        found = None
        for syn, canon in synonyms_map.items():
            if syn in s:
                found = canon
                break
        if found:
            marker_rows.append((r, found))
    # Сортировка по строкам
    marker_rows.sort(key=lambda x: x[0])
    ranges = {}
    for i, (r, mk) in enumerate(marker_rows):
        start = r + 1  # после заголовка
        end = (marker_rows[i + 1][0] - 1) if i + 1 < len(marker_rows) else len(values)
        if start <= end and mk not in ranges:
            ranges[mk] = (start, end)
    return ranges


def _choose_column_for_block(values: List[List[Optional[str]]], start: int, end: int) -> str:
    """Автоматически выбирает лучший столбец для блока, проверяя столбцы A-F.
    start/end 1-базисные.
    """
    def non_empty_in_col(idx0: int) -> int:
        cnt = 0
        for r in range(start, end + 1):
            row = values[r - 1] if r - 1 < len(values) else []
            v = row[idx0] if idx0 < len(row) else None
            if normalize_dish(v, True):
                cnt += 1
        return cnt
    
    # Проверяем столбцы A, B, C, D, E, F
    columns_to_check = ['A', 'B', 'C', 'D', 'E', 'F']
    best_col = 'A'
    best_count = 0
    
    for col_letter in columns_to_check:
        try:
            col_idx = col_to_index0(col_letter)
            count = non_empty_in_col(col_idx)
            if count > best_count:
                best_count = count
                best_col = col_letter
        except:
            continue
    
    return best_col


def _extract_dishes_from_multiple_columns(values: List[List[Optional[str]]], start: int, end: int, ignore_case: bool, columns: List[str] = None) -> Set[str]:
    """Извлекает блюда из указанных столбцов для указанного диапазона строк.
    start/end 1-базисные.
    """
    if columns is None:
        columns = ['A', 'B', 'C', 'D', 'E', 'F']
    
    dishes = set()
    
    for r in range(start, end + 1):
        row = values[r - 1] if r - 1 < len(values) else []
        
        # Проверяем все указанные столбцы
        for col_letter in columns:
            try:
                col_idx = col_to_index0(col_letter)
                v = row[col_idx] if col_idx < len(row) else None
                variants = normalize_dish_with_variants(v, ignore_case)
                for variant in variants:
                    if variant:
                        dishes.add(variant)
            except:
                continue
            
    return dishes

def _extract_dishes_from_both_columns(values: List[List[Optional[str]]], start: int, end: int, ignore_case: bool) -> Set[str]:
    """Легаси функция для обратной совместимости."""
    return _extract_dishes_from_multiple_columns(values, start, end, ignore_case, ['A', 'D'])


def compare_and_highlight(
    path1: str,
    sheet1: str,
    path2: str,
    sheet2: str,
    col1: str,
    col2: str,
    header_row1: int,
    header_row2: int,
    ignore_case: bool,
    use_fuzzy: bool,
    fuzzy_threshold: int,
    final_choice: int,  # 0=auto by date (implemented), 1=first, 2=second
) -> Tuple[str, int]:
    """
    Сравнение по категориям 'ЗАВТРАКИ' (столбец A) и 'ПЕРВЫЕ БЛЮДА' (столбец E).
    Внутри каждой категории сравниваем значения в соответствующем столбце построчно от заголовка до следующей категории/конца листа.
    При отсутствии маркеров — используем прежнюю логику с одиночной колонкой.
    Возвращает (out_path, matches).
    """
    # determine final/ref based on choice and extract dates
    d1 = _extract_best_date_from_file(path1, sheet1)
    d2 = _extract_best_date_from_file(path2, sheet2)
    
    # Отладочная информация
    print(f"DEBUG: Даты из файлов - d1: {d1}, d2: {d2}")
    
    # Определяем дату для отображения и сохранения: максимум из дат файлов и текущей даты
    files_max_date: Optional[date] = None
    if d1 and d2:
        files_max_date = max(d1, d2)
    elif d1:
        files_max_date = d1
    elif d2:
        files_max_date = d2

    today_dt = date.today()
    display_date: date = max(files_max_date, today_dt) if files_max_date else today_dt
    
    print(f"DEBUG: Максимальная дата из файлов: {files_max_date}")
    print(f"DEBUG: Сегодняшняя дата: {today_dt}")
    print(f"DEBUG: Итоговая дата для отображения: {display_date}")
    
    # ВАЖНО: Независимо от выбора final_choice, используем display_date
    # для названия выходного файла и заголовка листа
    
    if final_choice == 1:
        final_path, final_sheet = path1, sheet1
        ref_path, ref_sheet = path2, sheet2
    elif final_choice == 2:
        final_path, final_sheet = path2, sheet2
        ref_path, ref_sheet = path1, sheet1
    else:
        # Автовыбор: берем файл с более поздней датой как финальный
        if d1 and d2:
            if d2 >= d1:
                final_path, final_sheet = path2, sheet2
                ref_path, ref_sheet = path1, sheet1
            else:
                final_path, final_sheet = path1, sheet1
                ref_path, ref_sheet = path2, sheet2
        elif d1 and not d2:
            final_path, final_sheet = path1, sheet1
            ref_path, ref_sheet = path2, sheet2
        elif d2 and not d1:
            final_path, final_sheet = path2, sheet2
            ref_path, ref_sheet = path1, sheet1
        else:
            # Если дат нет ни в одном файле, берем второй файл по умолчанию
            final_path, final_sheet = path2, sheet2
            ref_path, ref_sheet = path1, sheet1

    # Синонимы -> канонические категории
    synonyms_map = {
        "завтраки": "завтраки",
        "первые блюда": "первые блюда",
        "салаты": "салаты и холодные закуски",
        "холодные закуски": "салаты и холодные закуски",
        "салаты и холодные закуски": "салаты и холодные закуски",
        "блюда из мяса": "блюда из мяса",
        "блюда из птицы": "блюда из птицы",
        "блюда из рыбы": "блюда из рыбы",
        "гарниры": "гарниры",
    }

    # Считаем диапазоны категорий в эталонном и итоговом файлах
    ref_vals = read_cell_values(ref_path, ref_sheet)
    final_xlsx = ensure_xlsx(final_path)
    wb = openpyxl.load_workbook(final_xlsx)
    sh = wb[final_sheet]

    final_vals = []
    for row in sh.iter_rows(values_only=True):
        final_vals.append([None if v is None else str(v) for v in row])

    ref_ranges = _find_category_ranges(ref_vals, synonyms_map)
    final_ranges = _find_category_ranges(final_vals, synonyms_map)

    # Если категории не обнаружены — fallback к логике сравнения по столбцам A и D глобально
    if not ref_ranges or not final_ranges:
        # Собираем референсный набор блюд из столбцов A и D целиком по листу
        ref_cols = ['A', 'D']
        ref_set: Set[str] = set()
        for c_letter in ref_cols:
                c_idx = col_to_index0(c_letter)
                for r in range(max(0, header_row2), len(ref_vals)):
                    row = ref_vals[r] if r < len(ref_vals) else []
                    v = row[c_idx] if c_idx < len(row) else None
                    variants = normalize_dish_with_variants(v, ignore_case)
                    for variant in variants:
                        if variant:
                            ref_set.add(variant)

        def is_match_fallback(dish: str) -> bool:
            # Проверяем все варианты блюда
            dish_variants = normalize_dish_with_variants(dish, ignore_case)
            for dish_variant in dish_variants:
                if not dish_variant:
                    continue
                if not use_fuzzy:
                    if dish_variant in ref_set:
                        return True
                else:
                    best = 0
                    for s in ref_set:
                        sim = sim_percent(dish_variant, s)
                        if sim > best:
                            best = sim
                        if best >= fuzzy_threshold:
                            return True
            return False

        from openpyxl.styles import Font
        red_font = Font(color="FF0000")
        idx_a = col_to_index0('A')
        idx_d = col_to_index0('D')
        matches = 0
        for r in range(1, sh.max_row + 1):
            if r <= max(1, header_row1):
                continue
            # Проверяем столбец A
            cell_a = sh.cell(row=r, column=idx_a + 1)
            text_a = str(cell_a.value) if cell_a.value is not None else ''
            if text_a.strip() and is_match_fallback(text_a):
                cell_a.font = red_font
                matches += 1
            # Проверяем столбец D
            cell_d = sh.cell(row=r, column=idx_d + 1)
            text_d = str(cell_d.value) if cell_d.value is not None else ''
            if text_d.strip() and is_match_fallback(text_d):
                cell_d.font = red_font
                matches += 1
        # Сохраняем файл с корректной датой в имени (без изменения содержимого)
        out_path = make_final_output_path(final_xlsx, display_date)
        wb.save(out_path)
        wb.close()
        return out_path, matches

    # Построим множества по категориям из ref, автоматически определяя лучшие столбцы
    ref_sets: dict = {}
    ref_category_columns: dict = {}  # храним информацию о столбцах для каждой категории
    
    for cat, (start, end) in ref_ranges.items():
        # Автоматически находим лучшие столбцы для категории
        best_cols = []
        columns_to_try = ['A', 'D']
        
        # Находим 2 лучших столбца
        column_scores = []
        for col_letter in columns_to_try:
            try:
                col_idx = col_to_index0(col_letter)
                score = 0
                for r in range(start, end + 1):
                    row = ref_vals[r - 1] if r - 1 < len(ref_vals) else []
                    v = row[col_idx] if col_idx < len(row) else None
                    if normalize_dish(v, ignore_case):
                        score += 1
                column_scores.append((col_letter, score))
            except:
                column_scores.append((col_letter, 0))
        
        # Сортируем по оценке и берём 2 лучших
        column_scores.sort(key=lambda x: x[1], reverse=True)
        best_cols = [col for col, score in column_scores[:2] if score > 0]
        
        if not best_cols:
            best_cols = ['A']  # fallback
        
        ref_category_columns[cat] = best_cols
        items = _extract_dishes_from_multiple_columns(ref_vals, start, end, ignore_case, best_cols)
        ref_sets[cat] = items

    # Также сформируем глобальный набор блюд из референса (все столбцы, весь лист)
    ref_global_set: Set[str] = _extract_dishes_from_multiple_columns(ref_vals, 1, len(ref_vals), ignore_case)

    def is_match_cat(cat: str, dish: str) -> bool:
        # Проверяем все варианты блюда
        dish_variants = normalize_dish_with_variants(dish, ignore_case)
        for dish_variant in dish_variants:
            if not dish_variant:
                continue
            # Сначала пробуем совпадение в пределах категории, затем глобально по всему листу
            sset = ref_sets.get(cat, set())
            if not use_fuzzy:
                if (dish_variant in sset) or (dish_variant in ref_global_set):
                    return True
            else:
                best = 0
                # Проверяем категорию
                for s in sset:
                    sim = sim_percent(dish_variant, s)
                    if sim > best:
                        best = sim
                    if best >= fuzzy_threshold:
                        return True
                # Проверяем глобальный набор
                for s in ref_global_set:
                    sim = sim_percent(dish_variant, s)
                    if sim > best:
                        best = sim
                    if best >= fuzzy_threshold:
                        return True
        return False

    from openpyxl.styles import Font
    red_font = Font(color="FF0000")

    matches = 0
    for cat, (start, end) in final_ranges.items():
        # Определяем лучшие столбцы для категории в итоговом файле
        final_best_cols = []
        columns_to_try = ['A', 'D']
        
        # Находим 2 лучших столбца в итоговом файле
        column_scores = []
        for col_letter in columns_to_try:
            try:
                col_idx = col_to_index0(col_letter)
                score = 0
                for r in range(start, end + 1):
                    row = final_vals[r - 1] if r - 1 < len(final_vals) else []
                    v = row[col_idx] if col_idx < len(row) else None
                    if normalize_dish(v, ignore_case):
                        score += 1
                column_scores.append((col_letter, score))
            except:
                column_scores.append((col_letter, 0))
        
        # Сортируем по оценке и берём 2 лучших
        column_scores.sort(key=lambda x: x[1], reverse=True)
        final_best_cols = [col for col, score in column_scores[:2] if score > 0]
        
        if not final_best_cols:
            final_best_cols = ['A']  # fallback
        
        # Проверяем все лучшие столбцы
        for col_letter in final_best_cols:
            try:
                col_idx = col_to_index0(col_letter)
                for r in range(start, min(end, sh.max_row) + 1):
                    cell = sh.cell(row=r, column=col_idx + 1)
                    text = str(cell.value) if cell.value is not None else ''
                    if text.strip() and is_match_cat(cat, text):
                        cell.font = red_font
                        matches += 1
            except:
                continue

    # Сохраняем файл с корректной датой в имени (без изменения содержимого)
    out_path = make_final_output_path(final_xlsx, display_date)
    wb.save(out_path)
    wb.close()
    return out_path, matches


def make_final_output_path(original_path: str, latest_date: Optional[date] = None) -> str:
    p = Path(original_path)
    if latest_date:
        date_str = latest_date.strftime("%d.%m.%Y")
        return str(p.with_name(p.stem + f"_сравнение_{date_str}" + p.suffix))
    else:
        return str(p.with_name(p.stem + "_сравнение" + p.suffix))


def _add_date_info_to_worksheet(worksheet, latest_date: date) -> None:
    """Добавляет информацию о дате сравнения в верхнюю часть листа."""
    try:
        from openpyxl.styles import Font, Alignment
        
        # Сдвигаем содержимое вниз на одну строку
        worksheet.insert_rows(1)
        
        # Добавляем информацию о дате в первую строку
        date_str = latest_date.strftime("%d.%m.%Y")
        cell = worksheet.cell(row=1, column=1)
        cell.value = f"Результат сравнения меню на {date_str}"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='left')
        
        # Объединяем ячейки для красивого отображения
        worksheet.merge_cells('A1:E1')
        
    except Exception:
        # Если что-то пошло не так, просто игнорируем добавление даты
        pass


def auto_header_row_by_markers(path: str, sheet_name: str) -> int:
    """
    Ищет строки-маркеры (например, 'ЗАВТРАКИ', 'ПЕРВЫЕ БЛЮДА') и возвращает номер строки заголовка,
    после которой начинается сравнение (то есть саму строку маркера).
    Если найдено несколько таких строк, берём первую сверху. Если не найдено — 1.
    """
    markers = ["завтраки", "первые блюда"]
    try:
        vals = read_cell_values(path, sheet_name)
    except Exception:
        return 1
    best = None
    max_scan = min(len(vals), 500)
    for r in range(max_scan):
        row = vals[r]
        for cell in row:
            if cell is None:
                continue
            s = normalize(str(cell), ignore_case=True)
            if any(m in s for m in markers):
                best = r + 1  # 1-based
                break
        if best is not None:
            break
    return best if best is not None else 1

