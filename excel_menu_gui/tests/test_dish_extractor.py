#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Тесты для модуля dish_extractor
"""

import unittest
import tempfile
import os
from pathlib import Path
from datetime import datetime
from unittest.mock import patch, MagicMock

import pandas as pd
import openpyxl

from app.services.dish_extractor import (
    DishItem, ExtractionResult, ExcelDataSource, DishExtractorService,
    extract_categorized_dishes_from_menu, extract_date_from_menu,
    DishExtractionError, FileFormatError
)


class TestDishItem(unittest.TestCase):
    """Тесты для DishItem"""
    
    def test_dish_item_creation(self):
        """Тест создания DishItem"""
        dish = DishItem(name="Борщ", weight="200г", price="150 руб.", category="первое")
        
        self.assertEqual(dish.name, "Борщ")
        self.assertEqual(dish.weight, "200г")
        self.assertEqual(dish.price, "150 руб.")
        self.assertEqual(dish.category, "первое")
    
    def test_dish_item_normalization(self):
        """Тест нормализации данных при создании DishItem"""
        dish = DishItem(name="  Борщ  ", weight=None, price="", category=None)
        
        self.assertEqual(dish.name, "Борщ")
        self.assertEqual(dish.weight, "")
        self.assertEqual(dish.price, "")
        self.assertEqual(dish.category, "")


class TestExtractionResult(unittest.TestCase):
    """Тесты для ExtractionResult"""
    
    def test_extraction_result_creation(self):
        """Тест создания ExtractionResult"""
        dishes = [
            DishItem(name="Борщ", category="первое"),
            DishItem(name="Котлета", category="мясо")
        ]
        categories = {
            "первое": [dishes[0]],
            "мясо": [dishes[1]]
        }
        
        result = ExtractionResult(dishes=dishes, categories=categories)
        
        self.assertEqual(len(result.dishes), 2)
        self.assertEqual(result.total_count, 2)
        self.assertIsNone(result.source_date)
        self.assertEqual(result.categories["первое"][0].name, "Борщ")


class TestExcelDataSource(unittest.TestCase):
    """Тесты для ExcelDataSource"""
    
    def setUp(self):
        """Подготовка к тестам"""
        self.data_source = ExcelDataSource()
    
    def test_parse_date_string_russian(self):
        """Тест парсинга русской даты"""
        # Тест "5 сентября"
        date1 = self.data_source._parse_date_string("5 сентября")
        self.assertIsNotNone(date1)
        self.assertEqual(date1.day, 5)
        self.assertEqual(date1.month, 9)
        
        # Тест "05.09.2024"
        date2 = self.data_source._parse_date_string("05.09.2024")
        self.assertIsNotNone(date2)
        self.assertEqual(date2.day, 5)
        self.assertEqual(date2.month, 9)
        self.assertEqual(date2.year, 2024)
        
        # Тест пустой строки
        date3 = self.data_source._parse_date_string("")
        self.assertIsNone(date3)
    
    def test_should_skip_cell(self):
        """Тест функции пропуска ячеек"""
        # Должны пропустить
        self.assertTrue(self.data_source._should_skip_cell(""))
        self.assertTrue(self.data_source._should_skip_cell("руб"))
        self.assertTrue(self.data_source._should_skip_cell("12:30"))
        self.assertTrue(self.data_source._should_skip_cell("200мл"))
        
        # Не должны пропустить
        self.assertFalse(self.data_source._should_skip_cell("Борщ украинский"))
        self.assertFalse(self.data_source._should_skip_cell("Котлета по-киевски"))
    
    def test_is_valid_dish(self):
        """Тест валидации блюда"""
        existing_dishes = ["Борщ"]
        
        # Валидные блюда
        self.assertTrue(self.data_source._is_valid_dish("Котлета по-киевски", existing_dishes))
        self.assertTrue(self.data_source._is_valid_dish("Винегрет овощной", existing_dishes))  # изменили с "Салат оливье"
        
        # Невалидные блюда
        self.assertFalse(self.data_source._is_valid_dish("Борщ", existing_dishes))  # уже есть
        self.assertFalse(self.data_source._is_valid_dish("руб", existing_dishes))  # короткое
        self.assertFalse(self.data_source._is_valid_dish("салат", existing_dishes))  # заголовок
        self.assertFalse(self.data_source._is_valid_dish("123", existing_dishes))  # только числа
    
    def test_extract_date_nonexistent_file(self):
        """Тест извлечения даты из несуществующего файла"""
        date = self.data_source.extract_date("nonexistent_file.xlsx")
        self.assertIsNone(date)
    
    def test_extract_dishes_nonexistent_file(self):
        """Тест извлечения блюд из несуществующего файла"""
        with self.assertRaises(FileFormatError):
            self.data_source.extract_dishes("nonexistent_file.xlsx")
    
    def test_extract_dishes_wrong_format(self):
        """Тест извлечения блюд из файла неправильного формата"""
        with tempfile.NamedTemporaryFile(suffix=".txt", delete=False) as tmp:
            tmp.write(b"test content")
            tmp_path = tmp.name
        
        try:
            with self.assertRaises(FileFormatError):
                self.data_source.extract_dishes(tmp_path)
        finally:
            os.unlink(tmp_path)


class TestDishExtractorService(unittest.TestCase):
    """Тесты для DishExtractorService"""
    
    def setUp(self):
        """Подготовка к тестам"""
        self.service = DishExtractorService()
    
    def test_detect_source_type(self):
        """Тест определения типа источника"""
        self.assertEqual(self.service._detect_source_type("test.xlsx"), "excel")
        self.assertEqual(self.service._detect_source_type("test.xls"), "excel")
        self.assertEqual(self.service._detect_source_type("test.xlsm"), "excel")
        
        with self.assertRaises(FileFormatError):
            self.service._detect_source_type("test.txt")
    
    @patch('app.services.dish_extractor.ExcelDataSource')
    def test_extract_dishes_with_auto_type(self, mock_excel_source):
        """Тест извлечения блюд с автоопределением типа"""
        # Мокаем источник данных
        mock_source_instance = MagicMock()
        mock_excel_source.return_value = mock_source_instance
        
        # Создаем тестовый результат
        test_result = ExtractionResult(
            dishes=[DishItem(name="Борщ", category="первое")],
            categories={"первое": [DishItem(name="Борщ", category="первое")]}
        )
        mock_source_instance.extract_dishes.return_value = test_result
        
        # Создаем новый сервис с мокнутым источником
        service = DishExtractorService()
        service._sources['excel'] = mock_source_instance
        
        result = service.extract_dishes("test.xlsx")
        
        self.assertIsInstance(result, ExtractionResult)
        self.assertEqual(len(result.dishes), 1)
        self.assertEqual(result.dishes[0].name, "Борщ")
    
    def test_unsupported_source_type(self):
        """Тест обработки неподдерживаемого типа источника"""
        with self.assertRaises(FileFormatError):
            self.service.extract_dishes("test.xlsx", source_type="unsupported")


class TestPublicFunctions(unittest.TestCase):
    """Тесты для публичных функций"""
    
    @patch('app.services.dish_extractor._extractor_service')
    def test_extract_categorized_dishes_from_menu(self, mock_service):
        """Тест функции extract_categorized_dishes_from_menu"""
        mock_service.extract_categorized_dishes.return_value = {
            "первое": ["Борщ", "Щи"],
            "мясо": ["Котлета"]
        }
        
        result = extract_categorized_dishes_from_menu("test.xlsx")
        
        self.assertEqual(len(result), 2)
        self.assertIn("первое", result)
        self.assertIn("мясо", result)
        self.assertEqual(len(result["первое"]), 2)
        mock_service.extract_categorized_dishes.assert_called_once_with("test.xlsx")
    
    @patch('app.services.dish_extractor._extractor_service')
    def test_extract_date_from_menu(self, mock_service):
        """Тест функции extract_date_from_menu"""
        test_date = datetime(2024, 9, 5)
        mock_service.extract_date_from_source.return_value = test_date
        
        result = extract_date_from_menu("test.xlsx")
        
        self.assertEqual(result, test_date)
        mock_service.extract_date_from_source.assert_called_once_with("test.xlsx")


class TestExcelIntegration(unittest.TestCase):
    """Интеграционные тесты с реальными Excel файлами"""
    
    def create_test_xlsx_file(self, file_path: str):
        """Создает тестовый xlsx файл"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "касса 05.09.2024"
        
        # Добавляем заголовки и данные
        ws['A1'] = "ЗАВТРАКИ"
        ws['A2'] = "Каша овсяная"
        ws['A3'] = "Омлет"
        ws['A4'] = ""
        
        ws['E1'] = "ПЕРВЫЕ БЛЮДА"  
        ws['E2'] = "Борщ украинский"
        ws['E3'] = "Щи из капусты"
        ws['E4'] = ""
        
        ws['F1'] = "БЛЮДА ИЗ МЯСА"
        ws['F2'] = "Котлета говяжья"
        ws['F3'] = "Бефстроганов"
        
        wb.save(file_path)
    
    def test_real_xlsx_extraction(self):
        """Тест с реальным xlsx файлом"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            self.create_test_xlsx_file(tmp_path)
            
            service = DishExtractorService()
            result = service.extract_dishes(tmp_path)
            
            self.assertIsInstance(result, ExtractionResult)
            self.assertGreater(result.total_count, 0)
            
            # Проверяем, что извлечена дата
            self.assertIsNotNone(result.source_date)
            self.assertEqual(result.source_date.day, 5)
            self.assertEqual(result.source_date.month, 9)
            self.assertEqual(result.source_date.year, 2024)
            
            # Проверяем категории
            categories = service.extract_categorized_dishes(tmp_path)
            self.assertIn("завтрак", categories)
            self.assertIn("первое", categories)
            self.assertIn("мясо", categories)
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)


if __name__ == '__main__':
    # Настраиваем детальный вывод тестов
    unittest.main(verbosity=2)
