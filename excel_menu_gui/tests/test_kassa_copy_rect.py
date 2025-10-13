#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import tempfile
import unittest
from pathlib import Path

import openpyxl

from app.services.menu_template_filler import MenuTemplateFiller


class TestKassaCopyRect(unittest.TestCase):
    def _create_source(self, path: Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Касса"
        # Поместим текст с датой для извлечения
        ws['A1'] = "12 сентября - пятница"
        # Заголовки
        ws['A6'] = "ЗАВТРАКИ"; ws['B6'] = "Вес/ед.изм."; ws['C6'] = "Цена, руб."
        ws['D6'] = "ПЕРВЫЕ БЛЮДА"; ws['E6'] = "Вес/ед.изм."
        ws['D11'] = "БЛЮДА ИЗ МЯСА"
        ws['D18'] = "БЛЮДА ИЗ ПТИЦЫ"
        ws['D25'] = "БЛЮДА ИЗ РЫБЫ"
        ws['A30'] = "САЛАТЫ и ХОЛОДНЫЕ ЗАКУСКИ"; ws['D30'] = "ГАРНИРЫ"
        # Завтраки
        ws['A7'] = "Блин с курицей"; ws['B7'] = "100г"; ws['C7'] = "105"
        # Супы (специально сдвинуты в E/F, чтобы нормализация вернула в D/E)
        ws['E7'] = "Борщ"; ws['F7'] = "250г"
        # Салаты с Цезарем
        ws['A32'] = "Цезарь с курицей/с креветками"; ws['B32'] = "200г"; ws['C32'] = "275/385"
        wb.save(path)

    def _create_template(self, path: Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Касса"
        # Создаём лист ХЦ для формул
        wb.create_sheet("Хц")
        wb.save(path)

    def test_copy_rect_and_normalize_and_date_and_hc_links(self):
        with tempfile.TemporaryDirectory() as td:
            src = Path(td) / "source.xlsx"
            tpl = Path(td) / "template.xlsx"
            out = Path(td) / "out.xlsx"
            self._create_source(src)
            self._create_template(tpl)

            filler = MenuTemplateFiller()
            ok, msg = filler.copy_kassa_rect_A6_F42(str(tpl), str(src), str(out))
            self.assertTrue(ok, msg)
            self.assertTrue(out.exists())

            wb = openpyxl.load_workbook(out, data_only=False)
            ws = None
            for sh in wb.worksheets:
                if 'касс' in sh.title.lower():
                    ws = sh
                    break
            self.assertIsNotNone(ws)

            # Дата: B2 — день недели, B3 — «12 сентября»
            self.assertEqual(ws['B2'].value, 'пятница')
            self.assertEqual(ws['B3'].value, '12 сентября')

            # Завтрак из источника перенесён
            self.assertEqual(ws['A7'].value, 'Блин с курицей')
            self.assertEqual(ws['B7'].value, '100г')
            self.assertEqual(ws['C7'].value, '105')

            # Нормализация правого блока: супы — D/E
            self.assertEqual(ws['D7'].value, 'Борщ')
            self.assertEqual(ws['E7'].value, '250г')

            # Формулы ХЦ на Цезаря
            hx = wb['Хц']
            f1 = hx['A19'].value
            f2 = hx['A20'].value
            self.assertIsInstance(f1, str)
            self.assertTrue(f1.startswith('='))
            # Ссылается на строку A32 с Цезарем
            self.assertIn('Касса!A32', f1)
            self.assertIsInstance(f2, str)
            self.assertTrue(f2.startswith('='))
            self.assertIn('Касса!A32', f2)


if __name__ == '__main__':
    unittest.main(verbosity=2)