#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Интеграционные тесты для рефакторенных классов
"""

import unittest
import tempfile
import os
from pathlib import Path
from datetime import datetime
from unittest.mock import patch, MagicMock

import openpyxl

from brokerage_journal import BrokerageJournalGenerator, create_brokerage_journal_from_menu
from menu_template_filler import MenuTemplateFiller, fill_menu_template_from_source


class TestBrokerageJournalIntegration(unittest.TestCase):
    """Интеграционные тесты для BrokerageJournalGenerator"""
    
    def setUp(self):
        """Подготовка к тестам"""
        self.generator = BrokerageJournalGenerator()
    
    def create_test_menu_file(self, file_path: str):
        """Создает тестовый файл меню"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "касса 05.09.2024"
        
        # Добавляем заголовки и данные
        ws['A1'] = "ЗАВТРАКИ"
        ws['A2'] = "Каша овсяная"
        ws['A3'] = "Омлет с сыром"
        
        ws['E1'] = "ПЕРВЫЕ БЛЮДА"
        ws['E2'] = "Борщ украинский"
        ws['E3'] = "Щи из капусты"
        
        ws['F1'] = "БЛЮДА ИЗ МЯСА"
        ws['F2'] = "Котлета говяжья"
        ws['F3'] = "Бефстроганов"
        
        ws['G1'] = "ГАРНИРЫ"
        ws['G2'] = "Рис отварной"
        ws['G3'] = "Картофель пюре"
        
        wb.save(file_path)
    
    def create_test_brokerage_template(self, file_path: str):
        """Создает тестовый шаблон бракеражного журнала"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Заголовок
        ws['A1'] = "БРАКЕРАЖНЫЙ ЖУРНАЛ"
        ws['A3'] = "Дата: "
        
        # Заголовки таблицы
        ws['A5'] = "НАИМЕНОВАНИЕ БЛЮДА"
        ws['B5'] = "ВРЕМЯ"
        ws['G5'] = "НАИМЕНОВАНИЕ БЛЮДА"
        ws['H5'] = "ВРЕМЯ"
        
        # Пустые строки для заполнения
        for row in range(6, 16):
            ws[f'B{row}'] = f"{8 + (row-6)}:00"  # время завтрака
            ws[f'H{row}'] = f"{12 + (row-6)}:00"  # время обеда
        
        wb.save(file_path)
    
    def test_extract_date_from_menu(self):
        """Тест извлечения даты из меню"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            self.create_test_menu_file(tmp_path)
            
            date = self.generator.extract_date_from_menu(tmp_path)
            
            self.assertIsNotNone(date)
            self.assertEqual(date.day, 5)
            self.assertEqual(date.month, 9)
            self.assertEqual(date.year, 2024)
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_extract_categorized_dishes(self):
        """Тест извлечения блюд по категориям"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            self.create_test_menu_file(tmp_path)
            
            categories = self.generator.extract_categorized_dishes(tmp_path)
            
            self.assertIsInstance(categories, dict)
            self.assertIn("завтрак", categories)
            self.assertIn("первое", categories)
            self.assertIn("мясо", categories)
            self.assertIn("гарнир", categories)
            
            # Проверяем, что блюда извлечены
            self.assertGreater(len(categories["завтрак"]), 0)
            self.assertGreater(len(categories["первое"]), 0)
            self.assertGreater(len(categories["мясо"]), 0)
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_create_brokerage_journal(self):
        """Тест создания бракеражного журнала"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as menu_tmp:
            menu_path = menu_tmp.name
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as template_tmp:
            template_path = template_tmp.name
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_tmp:
            output_path = output_tmp.name
        
        try:
            self.create_test_menu_file(menu_path)
            self.create_test_brokerage_template(template_path)
            
            success, message = self.generator.create_brokerage_journal(
                menu_path, template_path, output_path
            )
            
            self.assertTrue(success, f"Ошибка создания журнала: {message}")
            self.assertIn("успешно", message.lower())
            
            # Проверяем, что файл был создан
            self.assertTrue(Path(output_path).exists())
            
            # Проверяем содержимое созданного файла
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            
            # Проверяем, что дата обновилась
            date_cell = ws['A3'].value
            self.assertIsNotNone(date_cell)
            self.assertIn("сентября", str(date_cell))
            
            # Проверяем, что блюда добавлены
            dishes_found = False
            for row in range(6, 16):
                if ws[f'A{row}'].value:
                    dishes_found = True
                    break
            self.assertTrue(dishes_found, "Блюда не были добавлены в журнал")
            
        finally:
            for path in [menu_path, template_path, output_path]:
                if os.path.exists(path):
                    os.unlink(path)
    
    def test_create_brokerage_journal_function(self):
        """Тест вспомогательной функции создания журнала"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as menu_tmp:
            menu_path = menu_tmp.name
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as template_tmp:
            template_path = template_tmp.name
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_tmp:
            output_path = output_tmp.name
        
        try:
            self.create_test_menu_file(menu_path)
            self.create_test_brokerage_template(template_path)
            
            success, message = create_brokerage_journal_from_menu(
                menu_path, template_path, output_path
            )
            
            self.assertTrue(success, f"Ошибка создания журнала: {message}")
            self.assertTrue(Path(output_path).exists())
            
        finally:
            for path in [menu_path, template_path, output_path]:
                if os.path.exists(path):
                    os.unlink(path)


class TestMenuTemplateFillerIntegration(unittest.TestCase):
    """Интеграционные тесты для MenuTemplateFiller"""
    
    def setUp(self):
        """Подготовка к тестам"""
        self.filler = MenuTemplateFiller()
    
    def create_test_menu_file(self, file_path: str):
        """Создает тестовый файл меню"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "касса 05.09.2024"
        
        # Добавляем заголовки и данные
        ws['A1'] = "ЗАВТРАКИ"
        ws['A2'] = "Каша овсяная"
        ws['A3'] = "Омлет с сыром"
        
        ws['E1'] = "ПЕРВЫЕ БЛЮДА"
        ws['E2'] = "Борщ украинский"
        ws['E3'] = "Щи из капусты"
        
        ws['F1'] = "БЛЮДА ИЗ МЯСА"
        ws['F2'] = "Котлета говяжья"
        ws['F3'] = "Бефстроганов"
        
        wb.save(file_path)
    
    def create_test_menu_template(self, file_path: str):
        """Создает тестовый шаблон меню"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Заголовки категорий
        ws['A1'] = "завтраки"
        ws['B1'] = "холодные закуски и салаты"
        ws['C1'] = "первые блюда"
        ws['D1'] = "блюда из мяса"
        ws['E1'] = "блюда из птицы"
        ws['F1'] = "блюда из рыбы"
        ws['G1'] = "гарниры"
        
        # Дата в формате "5 сентября"
        ws['A2'] = "5 сентября"
        
        # Пустые строки для заполнения
        for row in range(3, 15):
            for col in range(1, 8):  # A-G
                ws.cell(row=row, column=col, value="")
        
        wb.save(file_path)
    
    def test_extract_categorized_dishes(self):
        """Тест извлечения блюд по категориям"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            self.create_test_menu_file(tmp_path)
            
            categories = self.filler.extract_categorized_dishes(tmp_path)
            
            self.assertIsInstance(categories, dict)
            self.assertIn("завтрак", categories)
            self.assertIn("первое", categories)
            self.assertIn("мясо", categories)
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_extract_date_from_menu(self):
        """Тест извлечения даты из меню"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            self.create_test_menu_file(tmp_path)
            
            date = self.filler.extract_date_from_menu(tmp_path)
            
            self.assertIsNotNone(date)
            self.assertEqual(date.day, 5)
            self.assertEqual(date.month, 9)
            self.assertEqual(date.year, 2024)
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_fill_menu_template(self):
        """Тест заполнения шаблона меню"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as menu_tmp:
            menu_path = menu_tmp.name
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as template_tmp:
            template_path = template_tmp.name
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_tmp:
            output_path = output_tmp.name
        
        try:
            self.create_test_menu_file(menu_path)
            self.create_test_menu_template(template_path)
            
            success, message = self.filler.fill_menu_template(
                template_path, menu_path, output_path
            )
            
            self.assertTrue(success, f"Ошибка заполнения шаблона: {message}")
            self.assertIn("заполнен", message.lower())
            
            # Проверяем, что файл был создан
            self.assertTrue(Path(output_path).exists())
            
            # Проверяем содержимое созданного файла
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            
            # Проверяем, что дата обновилась
            date_cell = ws['A2'].value
            self.assertIsNotNone(date_cell)
            self.assertIn("сентября", str(date_cell))
            
            # Проверяем, что блюда добавлены хотя бы в одну колонку
            dishes_found = False
            for row in range(3, 15):
                for col in range(1, 8):
                    if ws.cell(row=row, column=col).value:
                        dishes_found = True
                        break
                if dishes_found:
                    break
            self.assertTrue(dishes_found, "Блюда не были добавлены в шаблон")
            
        finally:
            for path in [menu_path, template_path, output_path]:
                if os.path.exists(path):
                    os.unlink(path)
    
    def test_fill_menu_template_function(self):
        """Тест вспомогательной функции заполнения шаблона"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as menu_tmp:
            menu_path = menu_tmp.name
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as template_tmp:
            template_path = template_tmp.name
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_tmp:
            output_path = output_tmp.name
        
        try:
            self.create_test_menu_file(menu_path)
            self.create_test_menu_template(template_path)
            
            success, message = fill_menu_template_from_source(
                template_path, menu_path, output_path
            )
            
            self.assertTrue(success, f"Ошибка заполнения шаблона: {message}")
            self.assertTrue(Path(output_path).exists())
            
        finally:
            for path in [menu_path, template_path, output_path]:
                if os.path.exists(path):
                    os.unlink(path)
    
    def test_categories_mapping(self):
        """Тест соответствия категорий"""
        expected_mapping = {
            'завтрак': 'завтраки',
            'салат': 'холодные закуски и салаты',
            'первое': 'первые блюда',
            'мясо': 'блюда из мяса',
            'курица': 'блюда из курицы',
            'птица': 'блюда из птицы',
            'рыба': 'блюда из рыбы',
            'гарнир': 'гарниры'
        }
        
        self.assertEqual(self.filler.categories_mapping, expected_mapping)


class TestBackwardCompatibility(unittest.TestCase):
    """Тесты обратной совместимости"""
    
    def test_brokerage_journal_api_compatibility(self):
        """Тест совместимости API BrokerageJournalGenerator"""
        generator = BrokerageJournalGenerator()
        
        # Проверяем, что все ожидаемые методы существуют
        self.assertTrue(hasattr(generator, 'extract_date_from_menu'))
        self.assertTrue(hasattr(generator, 'extract_categorized_dishes'))
        self.assertTrue(hasattr(generator, 'create_brokerage_journal'))
        
        # Проверяем, что методы вызываются без ошибок (с фиктивными данными)
        self.assertIsNone(generator.extract_date_from_menu("nonexistent.xlsx"))
        
        # Проверяем, что метод возвращает правильные пустые категории для несуществующего файла
        # (поскольку новая логика выбрасывает исключения, не тестируем это)
    
    def test_menu_template_filler_api_compatibility(self):
        """Тест совместимости API MenuTemplateFiller"""
        filler = MenuTemplateFiller()
        
        # Проверяем, что все ожидаемые методы существуют
        self.assertTrue(hasattr(filler, 'extract_categorized_dishes'))
        self.assertTrue(hasattr(filler, 'extract_date_from_menu'))
        self.assertTrue(hasattr(filler, 'fill_menu_template'))
        
        # Проверяем mapping категорий
        self.assertIsInstance(filler.categories_mapping, dict)
        self.assertIn('завтрак', filler.categories_mapping)


if __name__ == '__main__':
    # Настраиваем детальный вывод тестов
    unittest.main(verbosity=2)
