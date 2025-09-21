#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl

def analyze_journal():
    """Анализирует созданный бракеражный журнал"""
    
    file_path = "ФИНАЛЬНЫЙ_бракеражный_журнал.xlsx"
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        print(f"📊 Анализ файла: {file_path}")
        print(f"📋 Лист: {ws.title}")
        print(f"📏 Размер: {ws.max_row} строк, {ws.max_column} столбцов")
        
        # Проверяем дату
        date_cell = ws.cell(row=3, column=1).value
        print(f"📅 Дата в журнале: {date_cell}")
        
        # Проверяем заголовки таблицы
        print(f"\n🔍 Заголовки таблицы (строка 6):")
        for col in range(1, 10):
            header = ws.cell(row=6, column=col).value
            if header:
                print(f"  Колонка {col}: {header}")
        
        # Считаем блюда в левой таблице (колонка A)
        left_dishes = []
        for row in range(7, ws.max_row + 1):
            dish = ws.cell(row=row, column=1).value
            if dish and str(dish).strip():
                left_dishes.append(str(dish).strip())
        
        # Считаем блюда в правой таблице (колонка G)
        right_dishes = []
        for row in range(7, ws.max_row + 1):
            dish = ws.cell(row=row, column=7).value
            if dish and str(dish).strip():
                right_dishes.append(str(dish).strip())
        
        print(f"\n🍽️ Блюда в левой таблице: {len(left_dishes)}")
        if left_dishes[:5]:
            print("  Первые 5:")
            for i, dish in enumerate(left_dishes[:5]):
                print(f"    {i+1}. {dish}")
        
        print(f"\n🍽️ Блюда в правой таблице: {len(right_dishes)}")
        if right_dishes[:5]:
            print("  Первые 5:")
            for i, dish in enumerate(right_dishes[:5]):
                print(f"    {i+1}. {dish}")
        
        total_dishes = len(left_dishes) + len(right_dishes)
        print(f"\n📊 Всего блюд: {total_dishes}")
        
        # Проверяем время в первых строках
        if left_dishes:
            time1 = ws.cell(row=7, column=2).value
            time2 = ws.cell(row=7, column=3).value
            print(f"⏰ Время в левой таблице: {time1}, {time2}")
            
        if right_dishes:
            time3 = ws.cell(row=7, column=8).value
            time4 = ws.cell(row=7, column=9).value
            print(f"⏰ Время в правой таблице: {time3}, {time4}")
        
        print("\n✅ Анализ завершен успешно!")
        
    except Exception as e:
        print(f"❌ Ошибка при анализе: {e}")

if __name__ == "__main__":
    analyze_journal()
