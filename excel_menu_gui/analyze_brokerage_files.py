import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import re

def analyze_menu_file(file_path):
    """Анализирует структуру файла меню"""
    print(f"\n=== Анализ файла меню: {file_path} ===")
    
    try:
        # Попробуем открыть файл с помощью openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print(f"Листы в файле: {wb.sheetnames}")
        
        # Проанализируем первый лист
        ws = wb.active
        print(f"\nАктивный лист: {ws.title}")
        print(f"Размер листа: {ws.max_row} строк, {ws.max_column} колонок")
        
        # Выведем первые несколько строк
        print("\nПервые 10 строк:")
        for row in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(6, ws.max_column + 1)):  # Первые 5 колонок
                cell = ws.cell(row=row, column=col)
                row_data.append(str(cell.value) if cell.value is not None else "")
            print(f"Строка {row}: {row_data}")
        
        # Поиск даты
        print("\n--- Поиск даты ---")
        date_patterns = [
            r'\d{1,2}\s*сентября',
            r'\d{1,2}\.\d{1,2}\.\d{4}',
            r'\d{1,2}/\d{1,2}/\d{4}',
            r'сентября',
            r'пятница'
        ]
        
        for row in range(1, min(21, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell_str = str(cell.value).lower()
                    for pattern in date_patterns:
                        if re.search(pattern, cell_str, re.IGNORECASE):
                            print(f"Найдена дата в ячейке {row},{col}: {cell.value}")
                            break
        
    except Exception as e:
        print(f"Ошибка при анализе файла меню: {e}")
        try:
            # Попробуем с pandas
            df = pd.read_excel(file_path, sheet_name=None)
            print(f"Листы (pandas): {list(df.keys())}")
            for sheet_name, sheet_df in df.items():
                print(f"\nЛист {sheet_name}: {sheet_df.shape}")
                print(sheet_df.head())
        except Exception as e2:
            print(f"Ошибка и с pandas: {e2}")

def analyze_brokerage_file(file_path):
    """Анализирует структуру файла бракеражного журнала"""
    print(f"\n=== Анализ файла бракеражного журнала: {file_path} ===")
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print(f"Листы в файле: {wb.sheetnames}")
        
        ws = wb.active
        print(f"\nАктивный лист: {ws.title}")
        print(f"Размер листа: {ws.max_row} строк, {ws.max_column} колонок")
        
        # Выведем структуру
        print("\nСтруктура файла:")
        for row in range(1, min(21, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(8, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                row_data.append(str(cell.value) if cell.value is not None else "")
            if any(row_data):  # Показывать только непустые строки
                print(f"Строка {row}: {row_data}")
        
        # Поиск места для даты
        print("\n--- Поиск места для даты ---")
        for row in range(1, min(11, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and "дата" in str(cell.value).lower():
                    print(f"Поле даты в ячейке {row},{col}: {cell.value}")
        
        # Поиск таблицы с блюдами
        print("\n--- Поиск таблицы с блюдами ---")
        keywords = ["завтрак", "салат", "первое", "мясо", "курица", "птица", "рыба", "гарнир"]
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell_str = str(cell.value).lower()
                    for keyword in keywords:
                        if keyword in cell_str:
                            print(f"Найдена категория '{keyword}' в ячейке {row},{col}: {cell.value}")
                            break
        
    except Exception as e:
        print(f"Ошибка при анализе файла бракеражного журнала: {e}")

def main():
    # Пути к файлам
    menu_file = r"C:\Users\katya\Downloads\Telegram Desktop\5  сентября - пятница (3).xls"
    brokerage_file = r"C:\Users\katya\Downloads\Telegram Desktop\Бракеражный журнал (2).xlsx"
    
    # Проверяем существование файлов
    if Path(menu_file).exists():
        analyze_menu_file(menu_file)
    else:
        print(f"Файл меню не найден: {menu_file}")
    
    if Path(brokerage_file).exists():
        analyze_brokerage_file(brokerage_file)
    else:
        print(f"Файл бракеражного журнала не найден: {brokerage_file}")

if __name__ == "__main__":
    main()
