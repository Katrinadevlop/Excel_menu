#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Copy breakfast rows from a source Excel to a menu template in a fixed range.
- Detects the breakfast header ("ЗАВТРАК"/"ЗАВТРАКИ") and associated columns
- Reads rows until the next category header (e.g., "САЛАТЫ И ХОЛОДНЫЕ ЗАКУСКИ" or similar)
- Writes to the template on sheet "Касса" (if present) in range A{start_row}..A{end_row}
  and fills weight/price into B/C if available.

Usage:
  python tools/copy_breakfast_fixed_range.py --source <source.xlsx> --template <template.xlsx> --out <output.xlsx> [--start 7] [--end 29]
"""

import sys
import argparse
from pathlib import Path
import openpyxl

# Ensure project root (which contains the 'app' package) is importable when running from tools/
try:
    ROOT = Path(__file__).resolve().parents[1]
    if str(ROOT) not in sys.path:
        sys.path.insert(0, str(ROOT))
except Exception:
    pass

# Try to use advanced extractors if available
try:
    from app.services.dish_extractor import extract_dishes_from_excel, DishItem, extract_date_from_menu  # type: ignore
except Exception:
    extract_dishes_from_excel = None
    extract_date_from_menu = None
    DishItem = None


def detect_breakfast_block(ws):
    """Find header row and columns (name, weight, price) for breakfast on a worksheet.
    Returns (header_row, name_col, weight_col, price_col) or (None, None, None, None).
    """
    header_row = None
    name_col = weight_col = price_col = None

    # Find a row with breakfast header
    for r in range(1, min(100, ws.max_row + 1)):
        # Concatenate row text and also attempt per-cell header matching
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip():
                row_vals.append((c, str(v).strip()))
        row_text = ' '.join(val for _, val in row_vals).lower()
        if not row_vals:
            continue
        if 'завтрак' in row_text:
            # Identify columns based on header keywords
            header_row = r
            for c, val in row_vals:
                v = val.lower()
                if 'завтрак' in v:
                    name_col = c
                if 'вес' in v or 'ед.изм' in v:
                    weight_col = c
                if 'цена' in v or 'руб' in v:
                    price_col = c
            # Some layouts use only 'завтрак' and omit explicit weight/price headers on the header row
            # Try to infer weight/price positions relative to name if needed
            if name_col and (weight_col is None or price_col is None):
                # Common patterns: [name, weight, price] or [weight, name, price]
                # We'll scan immediate neighbors to find 'вес'/'цена' keywords in subsequent rows
                for rr in range(header_row, min(header_row + 5, ws.max_row + 1)):
                    for cc in range(1, ws.max_column + 1):
                        v = ws.cell(row=rr, column=cc).value
                        if v is None:
                            continue
                        sv = str(v).lower()
                        if (weight_col is None) and ('вес' in sv or 'ед.изм' in sv):
                            weight_col = cc
                        if (price_col is None) and ('цена' in sv or 'руб' in sv):
                            price_col = cc
            return header_row, name_col, weight_col, price_col
    return None, None, None, None


def is_next_category_row(ws, r):
    parts = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip():
            parts.append(str(v).strip())
    s = ' '.join(parts).lower()
    if not s:
        return False
    # Typical stop-headers after breakfast
    if 'салат' in s and ('холодн' in s or 'закуск' in s):
        return True
    if 'первые' in s or 'первое' in s:
        return True
    if 'блюда из' in s:
        return True
    if 'гарнир' in s:
        return True
    if 'напит' in s:
        return True
    return False


def collect_breakfast_from_sheet(ws):
    """Collect (name, weight, price) rows from one sheet if it has a breakfast section."""
    hdr, name_col, weight_col, price_col = detect_breakfast_block(ws)
    if hdr is None or name_col is None:
        return []
    rows = []
    for r in range(hdr + 1, ws.max_row + 1):
        if is_next_category_row(ws, r):
            break
        name = ws.cell(row=r, column=name_col).value
        if name is None or str(name).strip() == '':
            # allow sparse; continue
            continue
        # Skip obviously non-dish rows
        sn = str(name).strip()
        sn_low = sn.lower()
        if any(k in sn_low for k in ['вес', 'цена', 'руб']):
            continue
        weight = ws.cell(row=r, column=weight_col).value if weight_col else ''
        price = ws.cell(row=r, column=price_col).value if price_col else ''
        rows.append((sn, '' if weight is None else str(weight).strip(), '' if price is None else str(price).strip()))
    return rows


def collect_breakfast(source_path: Path):
    wb = openpyxl.load_workbook(str(source_path), data_only=True)
    # Prefer a sheet with 'касс' in the name, but scan all
    sheets = sorted(wb.sheetnames, key=lambda n: (0 if 'касс' in n.lower() else 1, n))
    for name in sheets:
        ws = wb[name]
        items = collect_breakfast_from_sheet(ws)
        if items:
            return items
    # If nothing found, try all sheets anyway (maybe non 'касс' sheet has it)
    for name in wb.sheetnames:
        ws = wb[name]
        items = collect_breakfast_from_sheet(ws)
        if items:
            return items
    return []


def _write_block(ws, name_col: int, weight_col: int, price_col: int, start_row: int, end_row: int, items, replace_only_empty: bool = True) -> int:
    """Write (name, weight, price) sequentially into a fixed block. Returns number of written rows.
    Only writes into empty cells when replace_only_empty=True to avoid overwriting existing content.
    """
    r = start_row
    written = 0
    idx = 0
    while r <= end_row and idx < len(items):
        name, weight, price = items[idx]
        cell_name = ws.cell(row=r, column=name_col)
        # Skip merged non-master cells
        if getattr(cell_name, '__class__', object).__name__ == 'MergedCell':
            r += 1
            continue
        can_write = True
        if replace_only_empty:
            can_write = (cell_name.value in (None, ''))
        if can_write:
            cell_name.value = name
            if weight_col:
                cell_w = ws.cell(row=r, column=weight_col)
                if (not replace_only_empty) or (cell_w.value in (None, '')):
                    cell_w.value = weight
            if price_col:
                cell_p = ws.cell(row=r, column=price_col)
                if (not replace_only_empty) or (cell_p.value in (None, '')):
                    cell_p.value = price
            written += 1
            idx += 1
        r += 1
    return written


def write_to_template(template_path: Path, out_path: Path, breakfast_items, bf_start_row: int, bf_end_row: int,
                      soups_items, soups_start: int, soups_end: int,
                      meat_items, meat_start: int, meat_end: int) -> int:
    wb = openpyxl.load_workbook(str(template_path))
    # Choose 'Касса' sheet if exists; else active
    ws = None
    for nm in wb.sheetnames:
        if 'касс' in nm.lower():
            ws = wb[nm]
            break
    if ws is None:
        ws = wb.active

    total = 0
    # Breakfast into A/B/C
    total += _write_block(ws, name_col=1, weight_col=2, price_col=3, start_row=bf_start_row, end_row=bf_end_row, items=breakfast_items, replace_only_empty=True)
    # Soups into D/E/F
    total += _write_block(ws, name_col=4, weight_col=5, price_col=6, start_row=soups_start, end_row=soups_end, items=soups_items, replace_only_empty=True)
    # Meat into D/E/F (same columns), starting within its range; only empty rows are filled, so overlap at D11 won't overwrite soups
    total += _write_block(ws, name_col=4, weight_col=5, price_col=6, start_row=meat_start, end_row=meat_end, items=meat_items, replace_only_empty=True)

    wb.save(str(out_path))
    return total


def collect_category_with_extractor(source: Path, keywords: list[str]):
    """Try to use dish_extractor to collect category dishes with weight/price; fallback to empty list on error."""
    if extract_dishes_from_excel is None:
        return []
    try:
        dishes = extract_dishes_from_excel(str(source), keywords)  # returns List[DishItem]
        items = []
        for d in dishes:
            name = getattr(d, 'name', '')
            weight = getattr(d, 'weight', '')
            price = getattr(d, 'price', '')
            if not name:
                continue
            items.append((str(name).strip(), str(weight or '').strip(), str(price or '').strip()))
        return items
    except Exception:
        return []


def _ru_month_genitive(m: int) -> str:
    return {
        1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля', 5: 'мая', 6: 'июня',
        7: 'июля', 8: 'августа', 9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'
    }.get(m, '')


def _ru_weekday_name(w: int) -> str:
    # Monday=0 .. Sunday=6
    return {
        0: 'понедельник', 1: 'вторник', 2: 'среда', 3: 'четверг', 4: 'пятница', 5: 'суббота', 6: 'воскресенье'
    }.get(w, '')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--source', required=True, help='Path to source Excel with data')
    ap.add_argument('--template', required=True, help='Path to template Excel to fill')
    ap.add_argument('--out', required=False, help='Path to save the filled template. If omitted, file will be saved next to the template as "<d> <месяц> - <день недели>.xlsx"')

    # Breakfast block A
    ap.add_argument('--bf-start', type=int, default=7, help='Breakfast start row (default 7)')
    ap.add_argument('--bf-end', type=int, default=29, help='Breakfast end row inclusive (default 29)')

    # Soups (first courses) block D
    ap.add_argument('--soups-start', type=int, default=6, help='Soups start row in D (default 6)')
    # D11 is a header row; do not write soups into D11. Stop at D10 by default.
    ap.add_argument('--soups-end', type=int, default=10, help='Soups end row inclusive in D (default 10)')

    # Meat block D — start AFTER the header at D11, i.e., from D12
    ap.add_argument('--meat-start', type=int, default=12, help='Meat start row in D (default 12)')
    ap.add_argument('--meat-end', type=int, default=17, help='Meat end row inclusive in D (default 17)')

    args = ap.parse_args()

    source = Path(args.source)
    template = Path(args.template)
    # Allow omitted --out; compute from date (prefer date extracted from source)
    if not source.exists():
        print(f'SOURCE_NOT_FOUND: {source}')
        sys.exit(2)
    if not template.exists():
        print(f'TEMPLATE_NOT_FOUND: {template}')
        sys.exit(2)

    if args.out:
        out = Path(args.out)
    else:
        # Try to extract date from source menu
        dt = None
        if extract_date_from_menu is not None:
            try:
                dt = extract_date_from_menu(str(source))
            except Exception:
                dt = None
        # Fallback to today
        if dt is None:
            from datetime import date
            dt = date.today()
        try:
            day = getattr(dt, 'day') if hasattr(dt, 'day') else dt.day
            month = getattr(dt, 'month') if hasattr(dt, 'month') else dt.month
            weekday = dt.weekday() if hasattr(dt, 'weekday') else 0
        except Exception:
            from datetime import date
            dtn = date.today()
            day, month, weekday = dtn.day, dtn.month, dtn.weekday()
        name = f"{day} {_ru_month_genitive(month)} - {_ru_weekday_name(weekday)}.xlsx"
        out = Path(template).with_name(name)

    # Collect categories
    breakfast_items = collect_breakfast(source)
    soups_items = collect_category_with_extractor(source, ["ПЕРВЫЕ БЛЮДА", "ПЕРВЫЕ"]) or []
    meat_items = collect_category_with_extractor(source, ["БЛЮДА ИЗ МЯСА", "МЯСНЫЕ БЛЮДА"]) or []

    if not breakfast_items:
        print('NO_BREAKFAST_ITEMS_FOUND')
    if not soups_items:
        print('NO_SOUPS_ITEMS_FOUND')
    if not meat_items:
        print('NO_MEAT_ITEMS_FOUND')

    if not (breakfast_items or soups_items or meat_items):
        sys.exit(3)

    total_written = write_to_template(
        template, out,
        breakfast_items, args.bf_start, args.bf_end,
        soups_items, args.soups_start, args.soups_end,
        meat_items, args.meat_start, args.meat_end,
    )
    print(f'WROTE total {total_written} rows across A and D blocks (only into empty cells).')
    print(f'SAVED_TO: {out}')


if __name__ == '__main__':
    main()
