#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys
from pathlib import Path
import openpyxl

def main():
    if len(sys.argv) < 2:
        print("USAGE: inspect_rect.py <xlsx>")
        sys.exit(1)
    p = Path(sys.argv[1])
    if not p.exists():
        print("FILE_NOT_FOUND:", p)
        sys.exit(2)
    wb = openpyxl.load_workbook(str(p), data_only=True)
    ws = None
    for nm in wb.sheetnames:
        if 'касс' in nm.lower():
            ws = wb[nm]
            break
    if ws is None:
        ws = wb.active
    print(f"Sheet={ws.title}")
    r1, r2 = 6, 42
    c1, c2 = 1, 6
    def col_letter(c):
        import string
        letters = ''
        while c:
            c, rem = divmod(c-1, 26)
            letters = string.ascii_uppercase[rem] + letters
        return letters
    # header row
    hdr = []
    for c in range(c1, c2+1):
        hdr.append(f"{col_letter(c)}")
    print("Cols:", "\t".join(hdr))
    for r in range(r1, r2+1):
        vals = []
        for c in range(c1, c2+1):
            v = ws.cell(row=r, column=c).value
            vals.append('' if v is None else str(v))
        print(f"{r:>3}:", "\t".join(vals))

if __name__ == '__main__':
    main()
