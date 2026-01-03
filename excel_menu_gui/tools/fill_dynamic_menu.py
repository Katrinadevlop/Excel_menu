#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Динамическое заполнение шаблона меню по сегментам, чтобы вставки/удаления строк
срабатывали как в Excel и не ломали правую/левую половину листа.
"""

from __future__ import annotations

import argparse
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import List, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

LEFT_HEADERS = [
    "ЗАВТРАКИ",
    "САЛАТЫ и ХОЛОДНЫЕ ЗАКУСКИ",
    "СЭНДВИЧИ",
    "ПЕЛЬМЕНИ",
    "ВАРЕНИКИ",
]

RIGHT_HEADERS = [
    "ПЕРВЫЕ БЛЮДА",
    "БЛЮДА ИЗ МЯСА",
    "БЛЮДА ИЗ ПТИЦЫ",
    "БЛЮДА ИЗ РЫБЫ",
    "ГАРНИРЫ",
    "НАПИТКИ",
    "ХЛЕБ",
    "СОУСЫ",
]

LEFT_LOOKUP = {(" ".join(h.strip().lower().replace("ё", "е").split())): h for h in LEFT_HEADERS}
RIGHT_LOOKUP = {(" ".join(h.strip().lower().replace("ё", "е").split())): h for h in RIGHT_HEADERS}


def norm(text) -> str:
    return " ".join(str(text).strip().lower().replace("ё", "е").split()) if text is not None else ""


def is_blank_value(v) -> bool:
    """True если значение визуально пустое (None/""/пробелы)."""
    if v is None:
        return True
    if v == "":
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def is_header_left(value) -> bool:
    return norm(value) in LEFT_LOOKUP


def is_header_right(value) -> bool:
    return norm(value) in RIGHT_LOOKUP


def is_footer_text(value) -> bool:
    """True если это строка-футер (дисклеймер) внизу меню."""
    if not isinstance(value, str):
        return False
    s = norm(value)
    # В шаблоне/меню обычно 2 строки дисклеймера:
    # 1) начинается с "*ПРИ НАЛИЧИИ У ВАС АЛЛЕРГИИ..."
    # 2) начинается с "Блюда могут содержать следы..."
    return s.startswith("*при наличии у вас аллерг") or s.startswith("блюда могут содержать")


@dataclass
class Segment:
    index: int
    top_row: int
    bottom_row: int  # exclusive
    left_header: str | None
    right_header: str | None
    rows: List[Tuple[str | None, str | None, str | None, str | None, str | None, str | None]]


def copy_row_style(ws: Worksheet, src_row: int, dst_row: int, col_from: int = 1, col_to: int = 6) -> None:
    for col in range(col_from, col_to + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def parse_segments(ws: Worksheet) -> List[Segment]:
    segments: List[Segment] = []
    row = 1
    idx = 0
    last_left = None
    last_right = None
    max_row = ws.max_row

    while row <= max_row:
        left_val = ws[f"A{row}"].value
        right_val = ws[f"D{row}"].value
        left_header = LEFT_LOOKUP.get(norm(left_val))
        right_header = RIGHT_LOOKUP.get(norm(right_val))
        if left_header or right_header:
            # игнорируем «мусорные» повторные заголовки без данных
            has_payload = any(
                ws[f"{col}{row}"].value not in (None, "")
                for col in ("B", "C", "E", "F")
            )
            if (
                not has_payload
                and left_header
                and left_header == last_left
                and not right_header
            ) or (
                not has_payload
                and right_header
                and right_header == last_right
                and not left_header
            ):
                row += 1
                continue

            start = row + 1
            end = start
            while end <= max_row:
                l_val = ws[f"A{end}"].value
                r_val = ws[f"D{end}"].value
                if is_header_left(l_val) or is_header_right(r_val):
                    break
                end += 1

            rows: List[Tuple[str | None, str | None, str | None, str | None, str | None, str | None]] = []
            for r in range(start, end):
                values = tuple(ws[f"{col}{r}"].value for col in "ABCDEF")
                if all(v in (None, "") for v in values):
                    continue
                if norm(values[0]) in LEFT_LOOKUP and all(
                    ws[f"{col}{r}"].value in (None, "") for col in ("B", "C")
                ):
                    continue
                if norm(values[3]) in RIGHT_LOOKUP and all(
                    ws[f"{col}{r}"].value in (None, "") for col in ("E", "F")
                ):
                    continue
                rows.append(values)

            segments.append(
                Segment(
                    index=idx,
                    top_row=row,
                    bottom_row=end,
                    left_header=left_header,
                    right_header=right_header,
                    rows=rows,
                )
            )
            idx += 1
            row = end
            if left_header:
                last_left = left_header
            if right_header:
                last_right = right_header
        else:
            row += 1
    return segments


def align_source_segments(src_segments: List[Segment], tpl_segments: List[Segment]) -> List[Segment]:
    aligned: List[Segment] = []
    tpl_iter = iter(tpl_segments)
    src_iter = iter(src_segments)
    current_src = next(src_iter, None)

    for tpl in tpl_iter:
        desired_left = tpl.left_header or ""
        desired_right = tpl.right_header or ""

        # Продвигаем источник, пока не найдём совпадение по паре заголовков
        while current_src and (
            (current_src.left_header or "") != desired_left
            or (current_src.right_header or "") != desired_right
        ):
            current_src = next(src_iter, None)

        if current_src:
            aligned.append(current_src)
            current_src = next(src_iter, None)
        else:
            aligned.append(
                Segment(
                    index=tpl.index,
                    top_row=tpl.top_row,
                    bottom_row=tpl.bottom_row,
                    left_header=tpl.left_header,
                    right_header=tpl.right_header,
                    rows=[],
                )
            )
    return aligned


def resize_segments_bottom_up(
    ws: Worksheet,
    tpl_segments: List[Segment],
    payloads: List[List[Tuple[str | None, str | None, str | None, str | None, str | None, str | None]]],
) -> None:
    for tpl, rows in sorted(zip(tpl_segments, payloads), key=lambda pair: pair[0].top_row, reverse=True):
        needed = len(rows)
        existing = max(0, tpl.bottom_row - tpl.top_row - 1)
        delta = needed - existing

        if delta > 0:
            ws.insert_rows(tpl.bottom_row, amount=delta)
            sample = tpl.top_row + 1 if tpl.top_row + 1 <= ws.max_row else tpl.top_row
            for offset in range(delta):
                target = tpl.bottom_row + offset
                copy_row_style(ws, sample, target)
                _clear_row(ws, target)
        elif delta < 0:
            ws.delete_rows(tpl.bottom_row + delta, amount=-delta)


def fill_segments_top_down(
    ws: Worksheet,
    tpl_segments: List[Segment],
    payloads: List[List[Tuple[str | None, str | None, str | None, str | None, str | None, str | None]]],
) -> None:
    for tpl, rows in zip(tpl_segments, payloads):
        start = tpl.top_row + 1
        end = tpl.bottom_row

        # Определяем какие стороны реально используем в сегменте
        has_left = tpl.left_header or any(any(v not in (None, "") for v in row[:3]) for row in rows)
        has_right = tpl.right_header or any(any(v not in (None, "") for v in row[3:]) for row in rows)
        cols_to_clear: Tuple[int, ...] = ()
        if has_left and has_right:
            cols_to_clear = (1, 2, 3, 4, 5, 6)
        elif has_left:
            cols_to_clear = (1, 2, 3)
        elif has_right:
            cols_to_clear = (4, 5, 6)

        for r in range(start, end):
            _clear_row(ws, r, cols_to_clear)
        for offset, values in enumerate(rows):
            r = start + offset
            if r >= end:
                break
            _write_row(ws, r, values)


def _ensure_cell_unmerged(ws: Worksheet, row: int, col: int) -> None:
    coord = f"{get_column_letter(col)}{row}"
    for merged in list(ws.merged_cells.ranges):
        if coord in merged:
            ws.unmerge_cells(str(merged))
            break


def _set_cell(ws: Worksheet, row: int, col: int, value) -> None:
    _ensure_cell_unmerged(ws, row, col)
    cell = ws.cell(row=row, column=col)

    # openpyxl insert_rows/delete_rows иногда оставляет «висячие» MergedCell вне merged_cells.ranges.
    # В этом случае просто конвертируем ячейку в обычную Cell и сохраняем её стиль.
    if isinstance(cell, MergedCell):
        saved_font = copy(cell.font)
        saved_border = copy(cell.border)
        saved_fill = copy(cell.fill)
        saved_alignment = copy(cell.alignment)
        saved_number_format = cell.number_format
        saved_protection = copy(cell.protection)

        ws._cells.pop((row, col), None)
        cell = ws.cell(row=row, column=col)

        cell.font = saved_font
        cell.border = saved_border
        cell.fill = saved_fill
        cell.alignment = saved_alignment
        cell.number_format = saved_number_format
        cell.protection = saved_protection

    try:
        cell.value = value
    except AttributeError as exc:
        coord = cell.coordinate
        related = [str(r) for r in list(ws.merged_cells.ranges) if coord in r]
        extra = f" (merged_ranges={related})" if related else " (merged_ranges=[])"
        raise ValueError(f"Не удалось записать в ячейку {coord}: {exc}{extra}") from exc


def _clear_row(ws: Worksheet, row: int, cols: Tuple[int, ...] = (1, 2, 3, 4, 5, 6)) -> None:
    for col in cols:
        _set_cell(ws, row, col, None)


def _write_row(ws: Worksheet, row: int, values: Tuple[str | None, str | None, str | None, str | None, str | None, str | None]) -> None:
    for idx, col in enumerate(range(1, 7)):
        val = values[idx] if idx < len(values) else None
        _set_cell(ws, row, col, val if val not in ("", None) else None)


def normalize_header_cells(ws: Worksheet, segments: List[Segment]) -> None:
    for seg in segments:
        row = seg.top_row
        if seg.left_header:
            _set_cell(ws, row, 1, seg.left_header)
            _set_cell(ws, row, 2, None)
            _set_cell(ws, row, 3, None)
        if seg.right_header:
            _set_cell(ws, row, 4, seg.right_header)
            _set_cell(ws, row, 5, None)
            _set_cell(ws, row, 6, None)


def cleanup_table_merges(ws: Worksheet, segments: List[Segment], min_col: int = 1, max_col: int = 6) -> None:
    """Убирает merge-диапазоны внутри таблицы меню, чтобы insert_rows не создавал 'висячие' MergedCell."""
    if not segments:
        return
    min_row = min(s.top_row for s in segments)
    max_row = max(s.bottom_row for s in segments) - 1

    to_remove = []
    for rng in list(ws.merged_cells.ranges):
        if rng.max_row >= min_row and rng.min_row <= max_row and rng.max_col >= min_col and rng.min_col <= max_col:
            to_remove.append(str(rng))

    # Пробуем корректно разъединить
    for rng in to_remove:
        try:
            ws.unmerge_cells(rng)
        except Exception:
            # при сломанных merge-диапазонах openpyxl может падать — дальше просто зачистим вручную
            pass

    # Удаляем описания merge-диапазонов (openpyxl 3.1+: ranges — это set)
    if to_remove:
        remove_set = set(to_remove)
        to_drop = {r for r in ws.merged_cells.ranges if str(r) in remove_set}
        ws.merged_cells.ranges.difference_update(to_drop)

    # Удаляем висячие MergedCell-объекты в области таблицы
    for key in list(getattr(ws, "_cells", {}).keys()):
        r, c = key
        if min_row <= r <= max_row and min_col <= c <= max_col:
            if isinstance(ws._cells.get(key), MergedCell):
                ws._cells.pop(key, None)


def validate(tpl_segments: List[Segment], payloads: List[List[Tuple]]) -> None:
    for tpl, rows in zip(tpl_segments, payloads):
        actual = tpl.bottom_row - tpl.top_row - 1
        if actual != len(rows):
            raise ValueError(
                f"Сегмент #{tpl.index} ({tpl.left_header or ''} | {tpl.right_header or ''}) "
                f"имеет {actual} строк вместо {len(rows)}"
            )


def _snapshot_cell(cell) -> Tuple:
    # Важно: копируем стили, чтобы «перетаскивание» работало как в Excel при Delete cells -> Shift up
    return (
        cell.value,
        copy(cell.font),
        copy(cell.border),
        copy(cell.fill),
        cell.number_format,
        copy(cell.alignment),
        copy(cell.protection),
    )


def _apply_cell_snapshot(ws: Worksheet, row: int, col: int, snap: Tuple) -> None:
    value, font, border, fill, number_format, alignment, protection = snap

    _set_cell(ws, row, col, None if is_blank_value(value) else value)
    cell = ws.cell(row=row, column=col)

    cell.font = font
    cell.border = border
    cell.fill = fill
    cell.number_format = number_format
    cell.alignment = alignment
    cell.protection = protection

    # На всякий случай очищаем, чтобы «хвосты» не оставались после сдвигов.
    try:
        cell.comment = None
    except Exception:
        pass
    try:
        cell.hyperlink = None
    except Exception:
        pass


def compact_cells_shift_up(ws: Worksheet, cols: Tuple[int, ...], row_from: int, row_to: int) -> None:
    """Удаляет пустые ячейки внутри набора колонок, сдвигая вверх (Excel: Delete cells -> Shift up).

    Важно:
    - НЕ удаляет строки листа целиком (не трогает вторую половину листа)
    - переносит и значения, и стили (иначе заголовки «переедут» без форматирования)
    """
    if row_from > row_to:
        return

    kept_rows: List[List[Tuple]] = []
    for r in range(row_from, row_to + 1):
        snaps: List[Tuple] = []
        non_empty = False
        for c in cols:
            cell = ws.cell(row=r, column=c)
            if not is_blank_value(cell.value):
                non_empty = True
            snaps.append(_snapshot_cell(cell))
        if non_empty:
            kept_rows.append(snaps)

    write_r = row_from
    for snaps in kept_rows:
        for j, c in enumerate(cols):
            _apply_cell_snapshot(ws, write_r, c, snaps[j])
        write_r += 1

    # Остаток диапазона очищаем по значениям (стили оставляем как есть)
    for r in range(write_r, row_to + 1):
        for c in cols:
            _set_cell(ws, r, c, None)


def _table_bounds(segments: List[Segment]) -> Tuple[int, int]:
    row_from = min(s.top_row for s in segments)
    row_to = max(s.bottom_row for s in segments) - 1
    return row_from, row_to


def _last_nonempty_row(ws: Worksheet, cols: Tuple[int, ...], row_from: int, row_to: int) -> int:
    last = row_from - 1
    for r in range(row_from, row_to + 1):
        if any(not is_blank_value(ws.cell(row=r, column=c).value) for c in cols):
            last = r
    return last


def _with_value(snap: Tuple, value) -> Tuple:
    # snap: (value, font, border, fill, number_format, alignment, protection)
    return (value, snap[1], snap[2], snap[3], snap[4], snap[5], snap[6])


def capture_footer_row_styles_from_template(ws: Worksheet) -> List[Tuple[List[Tuple], float | None]]:
    """Снимает стили строк-дисклеймеров из шаблона ДО unmerge/insert_rows.

    Возвращает список (cell_snaps_for_A_to_F, row_height) в порядке появления.
    """
    styles: List[Tuple[List[Tuple], float | None]] = []
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if is_footer_text(a):
            snaps = [_snapshot_cell(ws.cell(row=r, column=c)) for c in range(1, 7)]
            styles.append((snaps, ws.row_dimensions[r].height))
    return styles


def extract_footer_texts_and_clear_left(ws: Worksheet, row_from: int, row_to: int) -> List[str]:
    """Забирает тексты дисклеймера из колонки A и очищает A:C.

    Важно: чистим только A:C, чтобы не трогать правую половину (там может быть 'Хрен' и другие соусы).
    """
    texts: List[str] = []
    for r in range(row_from, row_to + 1):
        a = ws.cell(row=r, column=1).value
        if is_footer_text(a):
            texts.append(str(a))
            for c in (1, 2, 3):
                _set_cell(ws, r, c, None)
    return texts


def place_footer_rows_below_last_item(
    ws: Worksheet,
    footer_texts: List[str],
    footer_styles: List[Tuple[List[Tuple], float | None]],
    row_from: int,
    row_to: int,
) -> None:
    """Ставит строки дисклеймера сразу после последней строки с данными (с учётом обеих половин)."""
    if not footer_texts:
        return

    last_left = _last_nonempty_row(ws, cols=(1, 2, 3), row_from=row_from, row_to=row_to)
    last_right = _last_nonempty_row(ws, cols=(4, 5, 6), row_from=row_from, row_to=row_to)
    insert_at = max(last_left, last_right) + 1

    needed_last = insert_at + len(footer_texts) - 1
    if needed_last > row_to:
        ws.insert_rows(row_to + 1, amount=needed_last - row_to)
        row_to = needed_last

    # Если стилей нет — просто пишем как есть.
    fallback_snaps: List[Tuple] | None = None
    fallback_height: float | None = None
    if footer_styles:
        fallback_snaps, fallback_height = footer_styles[0]

    for i, text in enumerate(footer_texts):
        r = insert_at + i
        snaps, height = (footer_styles[i] if i < len(footer_styles) else (fallback_snaps, fallback_height))

        if snaps:
            # Применяем стили на A:F, но значения оставляем только в A.
            for c in range(1, 7):
                base = snaps[c - 1]
                value = text if c == 1 else None
                _apply_cell_snapshot(ws, r, c, _with_value(base, value))
        else:
            _set_cell(ws, r, 1, text)
            for c in range(2, 7):
                _set_cell(ws, r, c, None)

        if height is not None:
            ws.row_dimensions[r].height = height

        # В шаблоне футеры объединены на всю ширину A:F.
        try:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        except Exception:
            pass


def compact_menu_table_sides(ws: Worksheet, segments: List[Segment]) -> None:
    """Убирает 'дыры' слева/справа в области таблицы меню, не трогая вторую половину.

    Делает то же, что пользователь вручную: выделяет A:C (или D:F) пустые строки и делает Delete cells -> Shift up.
    """
    if not segments:
        return
    row_from, row_to = _table_bounds(segments)
    compact_cells_shift_up(ws, cols=(1, 2, 3), row_from=row_from, row_to=row_to)
    compact_cells_shift_up(ws, cols=(4, 5, 6), row_from=row_from, row_to=row_to)


def fill_dynamic_menu(
    source_path: Path,
    template_path: Path,
    output_path: Path,
    source_sheet: str = "Касса",
    template_sheet: str = "Menu",
) -> None:
    src_wb = load_workbook(source_path, data_only=True)
    tpl_wb = load_workbook(template_path)
    src_ws = src_wb[source_sheet]
    tpl_ws = tpl_wb[template_sheet]

    # Сохраняем стили футера (дисклеймеров) из шаблона ДО unmerge/insert_rows,
    # чтобы потом корректно вернуть объединение A:F и формат.
    footer_styles = capture_footer_row_styles_from_template(tpl_ws)

    src_segments = parse_segments(src_ws)
    tpl_segments = parse_segments(tpl_ws)
    aligned_src = align_source_segments(src_segments, tpl_segments)
    payloads: List[List[Tuple[str | None, str | None, str | None, str | None, str | None, str | None]]] = []
    for tpl_seg, src_seg in zip(tpl_segments, aligned_src):
        payloads.append(src_seg.rows if src_seg.rows else tpl_seg.rows)

    # Критично: убираем merges внутри области таблицы ДО insert_rows/delete_rows
    cleanup_table_merges(tpl_ws, tpl_segments)

    resize_segments_bottom_up(tpl_ws, tpl_segments, payloads)
    tpl_segments = parse_segments(tpl_ws)

    # На всякий случай повторяем очистку merges после вставок/удалений
    cleanup_table_merges(tpl_ws, tpl_segments)

    fill_segments_top_down(tpl_ws, tpl_segments, payloads)
    normalize_header_cells(tpl_ws, tpl_segments)

    # Внутренняя проверка геометрии сегментов — до пост-компактации,
    # потому что компактация может «передвигать» заголовки по строкам внутри половины листа.
    validate(tpl_segments, payloads)

    # Дисклеймеры внизу (2 строки) должны быть ПОСЛЕ всех блюд/соусов.
    # В некоторых файлах последняя позиция справа ('Хрен') попадает на одну строку с дисклеймером слева.
    # Поэтому:
    # 1) забираем тексты дисклеймера из A и чистим A:C (не трогаем D:F)
    # 2) делаем компактацию половин
    # 3) ставим дисклеймеры сразу под последней строкой (max по обеим половинам) и объединяем A:F
    row_from, row_to = _table_bounds(tpl_segments)
    footer_texts = extract_footer_texts_and_clear_left(tpl_ws, row_from=row_from, row_to=row_to)

    # Пост-обработка: убрать «дыры» в левой/правой половине,
    # которые появляются из-за заголовков другой стороны.
    # Это аналог Excel: Delete cells -> Shift up только в A:C и отдельно только в D:F.
    compact_menu_table_sides(tpl_ws, tpl_segments)

    place_footer_rows_below_last_item(
        tpl_ws,
        footer_texts=footer_texts,
        footer_styles=footer_styles,
        row_from=row_from,
        row_to=row_to,
    )

    tpl_wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Динамическое заполнение шаблона меню по сегментам.")
    parser.add_argument("--source", required=True, help="Файл с листом 'Касса' (эталон).")
    parser.add_argument("--template", required=True, help="Excel-шаблон.")
    parser.add_argument("--output", default="menu_ready.xlsx", help="Путь к итоговому файлу.")
    parser.add_argument("--source-sheet", default="Касса", help="Имя листа с исходными данными.")
    parser.add_argument("--template-sheet", default="Menu", help="Имя листа шаблона.")
    args = parser.parse_args()

    fill_dynamic_menu(
        Path(args.source),
        Path(args.template),
        Path(args.output),
        args.source_sheet,
        args.template_sheet,
    )
    print(f"Готово: {args.output}")


if __name__ == "__main__":
    main()
