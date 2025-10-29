#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для вставки салатов и холодных закусок из меню в шаблон (строки A29-A41).
Извлекает салаты из исходного меню и вставляет их в шаблон.
"""

import sys
import openpyxl
from pathlib import Path

# Добавляем родительскую папку в путь для импорта модулей
parent_dir = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(parent_dir))

from app.services.dish_extractor import extract_dishes_from_excel_rows_with_stop


def find_kassa_sheet(wb):
    """Находит лист Касса или возвращает активный лист"""
    for sh in wb.worksheets:
        if 'касс' in sh.title.lower():
            return sh
    return wb.active


def insert_salads_to_template(template_path: str, menu_path: str, output_path: str):
    """
    Извлекает салаты из меню и вставляет их в шаблон в строки A29-A41.
    
    Args:
        template_path: Путь к шаблону
        menu_path: Путь к файлу меню
        output_path: Путь к выходному файлу
    """
    try:
        # Проверяем существование файлов
        if not Path(template_path).exists():
            print(f"❌ Шаблон не найден: {template_path}")
            return False
        
        if not Path(menu_path).exists():
            print(f"❌ Меню не найдено: {menu_path}")
            return False
        
        print(f"📖 Извлекаем салаты из меню: {menu_path}")
        
        # Извлекаем салаты и холодные закуски из меню
        # Используем функцию extract_dishes_from_excel_rows_with_stop, которая остановится
        # перед СЭНДВИЧИ или другими категориями
        salads = extract_dishes_from_excel_rows_with_stop(
            menu_path,
            category_keywords=["САЛАТЫ", "ХОЛОДНЫЕ ЗАКУСКИ", "САЛАТ"],
            stop_keywords=["СЭНДВИЧ", "ПЕРВЫЕ", "БЛЮДА ИЗ", "НАПИТ"]
        )
        
        print(f"✅ Найдено {len(salads)} салатов")
        
        if not salads:
            print("⚠️ Салаты не найдены в меню")
            return False
        
        # Открываем шаблон
        print(f"📝 Открываем шаблон: {template_path}")
        wb = openpyxl.load_workbook(template_path)
        ws = find_kassa_sheet(wb)
        
        # Очищаем диапазон A29-A41, B29-B41, C29-C41
        print(f"🧹 Очищаем диапазон A29:C41 в листе '{ws.title}'")
        for row in range(29, 42):
            for col in range(1, 4):  # A, B, C
                try:
                    ws.cell(row=row, column=col).value = None
                except AttributeError:
                    pass  # Пропускаем объединенные ячейки
        
        # Вставляем салаты в A29-A41 (максимум 13 позиций)
        print(f"📋 Вставляем салаты в диапазон A29:C41")
        inserted_count = 0
        max_salads = 13  # Строки с 29 по 41 (включительно)
        
        for i, salad in enumerate(salads[:max_salads]):
            row = 29 + i
            try:
                # Вставляем название
                ws.cell(row=row, column=1).value = salad.name
                # Вставляем вес
                ws.cell(row=row, column=2).value = salad.weight if salad.weight else ""
                # Вставляем цену
                ws.cell(row=row, column=3).value = salad.price if salad.price else ""
                inserted_count += 1
                print(f"  ✓ A{row}: {salad.name}")
            except AttributeError:
                print(f"  ⚠️ Не удалось вставить в строку {row}")
                pass
        
        # Сохраняем результат
        print(f"💾 Сохраняем результат: {output_path}")
        wb.save(output_path)
        wb.close()
        
        print(f"\n✅ Готово! Вставлено {inserted_count} салатов в строки A29-A41")
        return True
        
    except Exception as e:
        print(f"❌ Ошибка: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Главная функция для запуска из командной строки"""
    if len(sys.argv) != 4:
        print("Использование: python insert_salads_to_template.py <шаблон> <меню> <выходной_файл>")
        print("\nПример:")
        print('  python insert_salads_to_template.py "templates/Шаблон меню пример.xlsx" "меню.xlsx" "результат.xlsx"')
        sys.exit(1)
    
    template_path = sys.argv[1]
    menu_path = sys.argv[2]
    output_path = sys.argv[3]
    
    success = insert_salads_to_template(template_path, menu_path, output_path)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
