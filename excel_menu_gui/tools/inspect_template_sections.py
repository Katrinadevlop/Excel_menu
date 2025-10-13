#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys
from pathlib import Path
import openpyxl

def main():
    if len(sys.argv) < 2:
        print("USAGE: inspect_template_sections.py <path_to_xlsx>")
        sys.exit(1)
    p = Path(sys.argv[1])
    if not p.exists():
        print("FILE_NOT_FOUND:", p)
        sys.exit(2)
    wb = openpyxl.load_workbook(str(p), data_only=True)

    # Choose 'Касса' sheet if exists; else active
    ws = None
    for nm in wb.sheetnames:
        if 'касс' in nm.lower():
            ws = wb[nm]
            break
    if ws is None:
        ws = wb.active

    print(f"Sheet={ws.title}")

    # Breakfast block detection between 'ЗАВТРАКИ' and 'САЛАТЫ И ХОЛОДНЫЕ ЗАКУСКИ'
    def row_text(r):
        parts = []
        for c in range(1, ws.max_column+1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip():
                parts.append(str(v))
        return ' '.join(parts).lower()

    b_hdr = None
    b_end = None
    for r in range(1, min(ws.max_row, 100)+1):
        t = row_text(r)
        if b_hdr is None and 'завтрак' in t:
            b_hdr = r
            continue
        if b_hdr is not None and b_end is None:
            if ('салат' in t) and ('холодн' in t or 'закуск' in t):
                b_end = r - 1
                break
    if b_hdr is None:
        print("BREAKFAST: header not found")
    else:
        start = b_hdr + 1
        if b_end is None:
            b_end = min(ws.max_row, start + 60)
        print(f"BREAKFAST: header_row={b_hdr} data_range=A{start}..A{b_end}")
        non_empty = 0
        for r in range(start, b_end+1):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            c = ws.cell(row=r, column=3).value
            if any(v not in (None, '') for v in (a,b,c)):
                non_empty += 1
                print(f"  {r:>3}: A={a!r} | B={b!r} | C={c!r}")
        print(f"  non_empty_rows={non_empty}")

    # Soups/meat in D/E/F
    def is_header_text(s: str) -> bool:
        s = s.lower()
        return (
            ('первые' in s and 'блюд' in s) or
            ('блюда из' in s) or
            ('мяс' in s) or
            ('суп' in s)
        )

    print("D-BLOCK (D/E/F) rows 6..17:")
    d_hdrs = []
    for r in range(6, min(17, ws.max_row)+1):
        d = ws.cell(row=r, column=4).value
        e = ws.cell(row=r, column=5).value
        f = ws.cell(row=r, column=6).value
        hdr_flag = bool(d and is_header_text(str(d)))
        if r == 11:
            print(f"* {r:>3}: D={d!r} | E={e!r} | F={f!r}  <-- candidate header row")
        else:
            print(f"  {r:>3}: D={d!r} | E={e!r} | F={f!r}{'  [HEADER]' if hdr_flag else ''}")
        if hdr_flag:
            d_hdrs.append(r)
    if d_hdrs:
        print(f"  detected_header_rows={d_hdrs}")

if __name__ == '__main__':
    main()
