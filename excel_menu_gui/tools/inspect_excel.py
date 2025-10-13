#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys
from pathlib import Path
import openpyxl

def main():
    if len(sys.argv) < 2:
        print("USAGE: inspect_excel.py <path_to_xlsx>")
        sys.exit(1)
    p = Path(sys.argv[1])
    if not p.exists():
        print("FILE_NOT_FOUND:", p)
        sys.exit(2)
    try:
        wb = openpyxl.load_workbook(str(p), data_only=True)
    except Exception as e:
        print("OPEN_ERROR:", e)
        sys.exit(3)
    # Scan all sheets; prioritize any sheet with 'касс' in name by ordering
    ordered_sheetnames = sorted(wb.sheetnames, key=lambda n: (0 if 'касс' in n.lower() else 1, n))

    any_sheet_output = False
    for sheet_name in ordered_sheetnames:
        ws = wb[sheet_name]

        def row_text(r):
            parts = []
            for c in range(1, ws.max_column+1):
                v = ws.cell(row=r, column=c).value
                if v is not None and str(v).strip():
                    parts.append(str(v))
            return ' '.join(parts).lower()

        start = None
        end = None
        header_row = None
        max_r = min(1000, ws.max_row)
        for r in range(1, max_r+1):
            t = row_text(r)
            if start is None and 'завтрак' in t:
                header_row = r
                start = r + 1
                continue
            if start is not None and end is None:
                if ('салат' in t) and ('холодн' in t or 'закуск' in t):
                    end = r - 1
                    break
        if start is None:
            # no header found; skip this sheet
            continue
        if end is None:
            end = min(ws.max_row, start + 60)

        print(f"Sheet={ws.title} header_row={header_row} start={start} end={end}")
        if header_row is not None:
            ha = ws.cell(row=header_row, column=1).value
            hb = ws.cell(row=header_row, column=2).value
            hc = ws.cell(row=header_row, column=3).value
            print(f"HDR: A={ha!r} | B={hb!r} | C={hc!r}")

        filled = 0
        for r in range(start, end+1):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            c = ws.cell(row=r, column=3).value
            if any(v not in (None, '') for v in (a,b,c)):
                filled += 1
                print(f"{r:>4}: A={a!r} | B={b!r} | C={c!r}")
        print("non_empty_rows_in_range=", filled)
        any_sheet_output = True

    if not any_sheet_output:
        print("NO_SHEETS_WITH_HEADER_ZAVTRAKI_FOUND")

if __name__ == '__main__':
    main()
