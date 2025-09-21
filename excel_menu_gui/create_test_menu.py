#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Создание тестового Excel файла с правильной структурой завтраков
"""

import openpyxl
from openpyxl.styles import Font, Alignment

def create_test_menu():
    """Создает тестовый файл меню с завтраками до сэндвичей"""
    
    # Создаем новую рабочую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Касса"
    
    # Заголовок
    ws['A1'] = "МЕНЮ СТОЛОВОЙ ПАТРИОТ"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = "5 сентября пятница"
    ws['A2'].font = Font(bold=True)
    
    # Завтраки (все блюда до СЭНДВИЧИ)
    breakfast_dishes = [
        "Блин с ветчиной и сыром",
        "Блин с яблоком", 
        "Блинчик без начинки",
        "Горячий бутерброд с ветчиной и сыром",
        "Дополнительно: ветчина,помидор,сыр",
        "Драники картофельные",
        "Жареный сулугуни с брусничным соусом",
        "Каша геркулесовая молочная",
        "Каша дружба молочная (рис, пшено)",
        "Каша рисовая на кокосовом молоке",
        "Крок-мадам",
        "Оладьи из кабачков",
        "Омлет на заказ",
        "Омлет/Омлет с помидором",
        "Пудинг Рафаэлло",
        "Сосиска жаренная/отварная",
        "Сырники",
        "Хашбраун с яйцом и беконом",
        "Яйцо жареное/отварное"
    ]
    
    # Заполняем завтраки
    row = 4
    for dish in breakfast_dishes:
        ws[f'A{row}'] = dish
        row += 1
    
    # Разделитель - СЭНДВИЧИ (после этого завтраки заканчиваются)
    ws[f'A{row}'] = "СЭНДВИЧИ"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    # Сэндвичи (не должны попасть в завтраки)
    sandwiches = [
        "Сэндвич с бужениной",
        "Сэндвич с ветчиной и сыром", 
        "Сэндвич с куриной грудкой",
        "Сэндвич с форелью"
    ]
    
    for dish in sandwiches:
        ws[f'A{row}'] = dish
        row += 1
    
    # Холодные салаты и закуски
    row += 1
    ws[f'A{row}'] = "ХОЛОДНЫЕ САЛАТЫ И ЗАКУСКИ"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    salads = [
        "Гранатовый браслет",
        "Брокколи с чесноком", 
        "Кальмаровый ароматный",
        "Капуста тушеная свежая",
        "Морковь с сыром"
    ]
    
    for dish in salads:
        ws[f'A{row}'] = dish
        row += 1
    
    # Сохраняем файл
    output_path = r"C:\Users\katya\Desktop\menurepit\excel_menu_gui\test_correct_menu.xlsx"
    wb.save(output_path)
    print(f"Создан тестовый файл: {output_path}")
    print(f"Завтраков: {len(breakfast_dishes)}")
    print(f"Сэндвичей: {len(sandwiches)}")
    print(f"Салатов: {len(salads)}")
    
    return output_path

if __name__ == "__main__":
    create_test_menu()
