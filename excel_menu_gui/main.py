import os
import sys
import shutil
import logging
import hashlib
import json
import subprocess
import calendar
from dataclasses import dataclass
from datetime import datetime, date, time
from pathlib import Path
from typing import Any, List, Optional, Tuple
from urllib.parse import urlsplit, parse_qsl

# Crash diagnostics (fatal + Python exceptions + Qt messages)
# Логи пишутся в файл %TEMP%/excel_menu_gui_crash.log И одновременно дублируются в консоль (stdout/stderr).
_FAULT_LOG_FH = None
_FAULT_LOG_PATH: Optional[Path] = None
# Глобальный реестр активных QThread, чтобы гасить их на выходе даже если closeEvent не сработал
_ACTIVE_THREADS: "set[QThread]" = set()
try:
    import faulthandler
    import tempfile
    import traceback
    import atexit
    from PySide6.QtCore import qInstallMessageHandler, QtMsgType, QMessageLogContext

    _FAULT_LOG_PATH = Path(tempfile.gettempdir()) / "excel_menu_gui_crash.log"
    _FAULT_LOG_FH = open(_FAULT_LOG_PATH, "a", encoding="utf-8", errors="replace")
    faulthandler.enable(file=_FAULT_LOG_FH, all_threads=True)

    def _flush_fault_log():
        try:
            if _FAULT_LOG_FH:
                _FAULT_LOG_FH.flush()
        except Exception:
            pass

    atexit.register(_flush_fault_log)

    def _stop_active_threads_on_exit():
        try:
            for th in list(_ACTIVE_THREADS):
                if th is None:
                    continue
                try:
                    if th.isRunning():
                        th.requestInterruption()
                        th.quit()
                        th.wait(15_000)
                except Exception:
                    pass
        except Exception:
            pass
    atexit.register(_stop_active_threads_on_exit)

    def _write_exc(prefix: str, exc_type, exc_value, exc_tb):
        try:
            ts = datetime.now().isoformat(timespec="seconds")
            line = f"\n[{ts}] {prefix}: {exc_type.__name__}: {exc_value}\n"
            _FAULT_LOG_FH.write(line)
            traceback.print_exception(exc_type, exc_value, exc_tb, file=_FAULT_LOG_FH)
            _FAULT_LOG_FH.flush()

            # дублируем в консоль
            try:
                sys.stderr.write(line)
                traceback.print_exception(exc_type, exc_value, exc_tb, file=sys.stderr)
                sys.stderr.flush()
            except Exception:
                pass
        except Exception:
            pass

    def _excepthook(exc_type, exc_value, exc_tb):
        _write_exc("Uncaught", exc_type, exc_value, exc_tb)
        try:
            sys.__excepthook__(exc_type, exc_value, exc_tb)
        except Exception:
            pass

    sys.excepthook = _excepthook

    def _qt_message_handler(mode: QtMsgType, ctx: QMessageLogContext, msg: str):
        try:
            ts = datetime.now().isoformat(timespec="seconds")
            line = f"\n[{ts}] Qt[{int(mode)}] {msg}\n"
            _FAULT_LOG_FH.write(line)
            _FAULT_LOG_FH.flush()

            try:
                sys.stderr.write(line)
                sys.stderr.flush()
            except Exception:
                pass
        except Exception:
            pass
    try:
        qInstallMessageHandler(_qt_message_handler)
    except Exception:
        pass

except Exception:
    _FAULT_LOG_FH = None
    _FAULT_LOG_PATH = None

from PySide6.QtCore import (
    Qt, QMimeData, QSize, QUrl, QSettings, QEvent, QPoint, QTimer, QDate,
    QObject, Signal, QThread, QCoreApplication, QLockFile,
)
from PySide6.QtGui import QPalette, QColor, QIcon, QPixmap, QPainter, QPen, QBrush, QLinearGradient, QFont, QDesktopServices, QGuiApplication, QTextCharFormat
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QBoxLayout,
    QLabel, QPushButton, QFileDialog, QTextEdit, QComboBox, QLineEdit,
    QGroupBox, QCheckBox, QSpinBox, QRadioButton, QButtonGroup, QMessageBox, QFrame, QSizePolicy, QScrollArea,
    QListWidget, QListWidgetItem, QInputDialog, QDialog, QDialogButtonBox, QCalendarWidget,
)

from app.services.comparator import compare_and_highlight, get_sheet_names, ColumnParseError
# Временная алиас-совместимость для старых импортов внутри процесса
import importlib as _importlib
sys.modules.setdefault('comparator', _importlib.import_module('app.services.comparator'))
sys.modules.setdefault('brokerage_journal', _importlib.import_module('app.reports.brokerage_journal'))
from app.services.template_linker import default_template_path
from app.gui.theme import ThemeMode, apply_theme, start_system_theme_watcher
from app.reports.presentation_handler import create_presentation_with_excel_data
from app.reports.brokerage_journal import create_brokerage_journal_from_menu
from app.reports.pricelist_excel import create_pricelist_xlsx
from app.reports.iikochain_pricetag_merge import export_black_pricetags, TagData
from app.services.dish_extractor import extract_all_dishes_with_details, DishItem
from app.integrations.iiko_rms_client import IikoRmsClient, IikoApiError
from app.integrations.iiko_cloud_client import IikoCloudClient as IikoTransportClient
from app.integrations.iiko_cloud_v1_client import IikoCloudV1Client, IikoOrganization
from app.services.menu_template_filler import MenuTemplateFiller
from tools.fill_dynamic_menu import fill_dynamic_menu
from app.gui.ui_styles import (
    AppStyles, ButtonStyles, LayoutStyles, StyleSheets, ComponentStyles,
    StyleManager, ThemeAwareStyles
)


class DropLineEdit(QLineEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setPlaceholderText("Перетащите файл сюда или нажмите Обзор…")


class PasteOnDoubleClickLineEdit(QLineEdit):
    """QLineEdit, который вставляет буфер обмена по двойному клику."""

    def mouseDoubleClickEvent(self, event):
        try:
            self.paste()
        except Exception:
            pass
        super().mouseDoubleClickEvent(event)

    def dragEnterEvent(self, event):
        md: QMimeData = event.mimeData()
        if md.hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        urls = event.mimeData().urls()
        if urls:
            local = urls[0].toLocalFile()
            if local:
                self.setText(local)
        event.acceptProposedAction()


def label_caption(text: str) -> QLabel:
    lbl = QLabel(text)
    ComponentStyles.style_caption_label(lbl)
    return lbl


def nice_group(title: str, content: QWidget) -> QGroupBox:
    gb = QGroupBox(title)
    lay = QVBoxLayout(gb)
    lay.addWidget(content)
    return gb


class FileDropGroup(QGroupBox):
    def __init__(self, title: str, target_line_edit: QLineEdit, content: QWidget, parent=None):
        super().__init__(title, parent)
        self._target = target_line_edit
        self.setAcceptDrops(True)
        lay = QVBoxLayout(self)
        lay.addWidget(content)

    def dragEnterEvent(self, event):
        md: QMimeData = event.mimeData()
        if md.hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        urls = event.mimeData().urls()
        if urls:
            local = urls[0].toLocalFile()
            if local:
                self._target.setText(local)
        event.acceptProposedAction()


def create_app_icon() -> QIcon:
    """Legacy function for compatibility. Use AppStyles.create_app_icon() instead."""
    return AppStyles.create_app_icon()


def find_template(filename: str) -> Optional[str]:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).parent))
    candidates = [
        base / "excel_menu_gui" / "templates" / filename,
        base / "templates" / filename,
        Path(__file__).parent / "templates" / filename,
    ]
    for p in candidates:
        if p.exists():
            return str(p)
    return None


def find_template_path(rel_path: str) -> Optional[str]:
    """Ищет файл ИЛИ папку внутри templates (поддерживает PyInstaller _MEIPASS)."""
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).parent))
    candidates = [
        base / "excel_menu_gui" / "templates" / rel_path,
        base / "templates" / rel_path,
        Path(__file__).parent / "templates" / rel_path,
    ]
    for p in candidates:
        if p.exists():
            return str(p)
    return None


def _open_schedule_task_name() -> str:
    """Имя задачи Windows Task Scheduler для запланированного открытия блюд.

    Делаем уникальным на случай нескольких копий приложения/папок.
    """
    base = "excel_menu_gui_open_dishes"
    try:
        suffix = hashlib.sha1(str(Path(__file__).resolve()).encode("utf-8")).hexdigest()[:8]
        return f"{base}_{suffix}"
    except Exception:
        return base


def _open_schedule_lock_path() -> str:
    try:
        tmp = os.getenv("TEMP") or os.getenv("TMP") or str(Path.cwd())
        name = _open_schedule_task_name().replace("\\", "_").replace("/", "_")
        return str(Path(tmp) / f"{name}.lock")
    except Exception:
        return "excel_menu_gui_open_schedule.lock"


def _open_schedule_runner_tr() -> str:
    """Команда (/TR) для Task Scheduler: запуск в режиме выполнения расписания."""
    exe_path = Path(sys.executable)
    exe = str(exe_path)

    # Если это python.exe — попробуем pythonw.exe (без консоли)
    try:
        if exe_path.name.lower() == "python.exe":
            pyw = exe_path.with_name("pythonw.exe")
            if pyw.exists():
                exe = str(pyw)
    except Exception:
        pass

    # PyInstaller/frozen: запускаем exe
    if bool(getattr(sys, "frozen", False)):
        return f'"{exe}" --run-open-schedule'

    # dev mode: запускаем скрипт
    try:
        script = str(Path(__file__).resolve())
    except Exception:
        script = str(__file__)
    return f'"{exe}" "{script}" --run-open-schedule'


def _windows_delete_open_schedule_task() -> None:
    if os.name != "nt":
        return

    task_name = _open_schedule_task_name()
    try:
        subprocess.run(
            ["schtasks", "/Delete", "/TN", task_name, "/F"],
            capture_output=True,
            text=True,
        )
    except Exception:
        pass


def _windows_create_open_schedule_task(run_at: datetime) -> Tuple[bool, str]:
    """Создаёт задачу Планировщика Windows для выполнения расписания.

    Возвращает (ok, error_message).
    """
    if os.name != "nt":
        return False, "Windows Task Scheduler доступен только на Windows"

    task_name = _open_schedule_task_name()
    tr = _open_schedule_runner_tr()

    try:
        sd = run_at.strftime("%m/%d/%Y")
        st = run_at.strftime("%H:%M")
    except Exception:
        return False, "Некорректная дата/время"

    # Пытаемся запускать от имени текущего пользователя без сохранения пароля.
    # Это позволяет запускать задачу даже если приложение закрыто.
    ru_user = ""
    try:
        un = (os.getenv("USERNAME") or "").strip()
        ud = (os.getenv("USERDOMAIN") or os.getenv("COMPUTERNAME") or "").strip()
        if un and ud:
            ru_user = f"{ud}\\{un}"
        else:
            ru_user = un
        if not ru_user:
            ru_user = (os.getlogin() or "").strip()
    except Exception:
        ru_user = ""

    cmd = [
        "schtasks",
        "/Create",
        "/F",
        "/SC",
        "ONCE",
        "/TN",
        task_name,
        "/TR",
        tr,
        "/ST",
        st,
        "/SD",
        sd,
        "/RL",
        "LIMITED",
    ]

    # /NP требует /RU; если пользователя не удалось определить — попробуем без него (может сработать на некоторых системах)
    if ru_user:
        cmd += ["/RU", ru_user, "/NP"]

    try:
        res = subprocess.run(cmd, capture_output=True, text=True)
    except Exception as e:
        return False, str(e)

    if int(getattr(res, "returncode", 1)) == 0:
        return True, ""

    err = (getattr(res, "stderr", "") or "").strip()
    out = (getattr(res, "stdout", "") or "").strip()
    return False, err or out or "Не удалось создать задачу Планировщика Windows"


def run_open_schedule_due_silent() -> int:
    """CLI/TaskScheduler entrypoint: выполняет запланированное открытие, если наступило время."""
    # Лок защищает от двойного запуска (если открыт GUI + параллельно сработал schtasks)
    lock = QLockFile(_open_schedule_lock_path())
    try:
        lock.setStaleLockTime(10 * 60 * 1000)
    except Exception:
        pass

    if not lock.tryLock(0):
        return 0

    try:
        settings = QSettings("excel_menu_gui", "excel_menu_gui")

        raw = settings.value("iiko/open_dishes/schedule", "")
        raw_s = str(raw) if raw is not None else ""
        if not raw_s:
            _windows_delete_open_schedule_task()
            return 0

        try:
            job = json.loads(raw_s)
        except Exception:
            # мусор — очищаем
            try:
                settings.remove("iiko/open_dishes/schedule")
            except Exception:
                pass
            _windows_delete_open_schedule_task()
            return 0

        if not isinstance(job, dict):
            try:
                settings.remove("iiko/open_dishes/schedule")
            except Exception:
                pass
            _windows_delete_open_schedule_task()
            return 0

        state = str(job.get("state") or "pending").strip().lower()
        if state in ("done", "failed"):
            _windows_delete_open_schedule_task()
            return 0

        run_at_s = str(job.get("run_at") or "")
        try:
            dt_run = datetime.fromisoformat(run_at_s)
        except Exception:
            return 0

        if datetime.now() < dt_run:
            return 0

        product_ids = job.get("product_ids") or []
        if not isinstance(product_ids, list) or not product_ids:
            try:
                settings.remove("iiko/open_dishes/schedule")
            except Exception:
                pass
            _windows_delete_open_schedule_task()
            return 0

        # читаем REST-параметры
        mode = str(settings.value("iiko/mode", "cloud") or "cloud").strip().lower()
        if mode not in ("rest", "rms", "resto"):
            raise IikoApiError("Открытие блюд по расписанию доступно только в режиме REST (/resto).")

        base_url = str(settings.value("iiko/base_url", "") or "").strip()
        login = str(settings.value("iiko/login", "") or "").strip()
        pass_sha1 = str(settings.value("iiko/pass_sha1", "") or "").strip()
        if not (base_url and login and pass_sha1):
            raise IikoApiError("Не заданы параметры REST (/resto) или не сохранён SHA1 пароля.")

        client = IikoRmsClient(base_url=base_url, login=login, pass_sha1=pass_sha1)

        ok = 0
        for pid in product_ids:
            pid_s = str(pid).strip()
            if not pid_s:
                continue
            client.open_product_from_stoplist(pid_s)
            ok += 1

        # успех -> очищаем расписание
        try:
            settings.remove("iiko/open_dishes/schedule")
        except Exception:
            pass

        _windows_delete_open_schedule_task()
        return ok

    except IikoApiError as e:
        # пометим как failed
        try:
            job = locals().get("job") if "job" in locals() else None
            if isinstance(job, dict):
                job["state"] = "failed"
                job["last_error"] = str(e)
                try:
                    job["failed_at"] = datetime.now().isoformat(timespec="seconds")
                except Exception:
                    pass
                settings.setValue("iiko/open_dishes/schedule", json.dumps(job, ensure_ascii=False))
        except Exception:
            pass

        _windows_delete_open_schedule_task()
        return -1

    except Exception:
        return -1

    finally:
        try:
            lock.unlock()
        except Exception:
            pass


@dataclass
class FileConfig:
    path: str = ""
    sheet: str = ""
    col: str = "A"
    header_row_1based: int = 1


class IikoProductsLoadWorker(QObject):
    """Загружает номенклатуру iiko в фоне, чтобы UI не зависал."""

    finished = Signal(int, object)  # seq, result dict
    failed = Signal(int, str)       # seq, error

    def __init__(self, seq: int, snapshot: dict):
        super().__init__()
        self.seq = int(seq)
        self.snapshot = snapshot or {}

    def run(self) -> None:
        try:
            res = self._do_fetch_and_index()
            self.finished.emit(self.seq, res)
        except BaseException as e:
            # BaseException: чтобы не уронить процесс из фонового потока (SystemExit/KeyboardInterrupt/etc.)
            try:
                self.failed.emit(self.seq, str(e))
            except Exception:
                pass

    def _do_fetch_and_index(self) -> dict:
        mode = str(self.snapshot.get("mode") or "").strip().lower()

        # --- fetch ---
        products: List[Any] = []
        biz_new_token: str = ""

        if mode in ("cloud", "cloud_v1", "cloudv1", "v1"):
            api_url = str(self.snapshot.get("cloud_api_url") or "").strip() or "https://api-ru.iiko.services"
            api_login = str(self.snapshot.get("cloud_api_login") or "").strip()
            org_id = str(self.snapshot.get("cloud_org_id") or "").strip()
            token = str(self.snapshot.get("cloud_access_token") or "").strip()
            if not (api_login and org_id and token):
                raise IikoApiError("Не заданы параметры iikoCloud. Нажмите 'Авторизация точки'.")
            client = IikoCloudV1Client(
                api_url=api_url,
                api_login=api_login,
                organization_id=org_id,
                access_token=token,
            )
            products = client.get_products()

        elif mode in ("biz", "transport", "iikobiz", "iiko.biz"):
            api_url = str(self.snapshot.get("biz_api_url") or "").strip() or "https://iiko.biz:9900"
            user_id = str(self.snapshot.get("biz_user_id") or "").strip()
            user_secret = str(self.snapshot.get("biz_user_secret") or "").strip()
            org_id = str(self.snapshot.get("biz_org_id") or "").strip()
            token = str(self.snapshot.get("biz_access_token") or "").strip()

            if not org_id:
                raise IikoApiError("Не выбрана организация. Нажмите 'Авторизация точки'.")

            if not token and not (user_id and user_secret):
                raise IikoApiError("Не задан user_id/user_secret или access_token iikoTransport.")

            client = IikoTransportClient(
                api_url=api_url,
                user_id=user_id,
                user_secret=user_secret,
                organization_id=org_id,
                access_token=token,
            )

            try:
                products = client.get_products()
            except IikoApiError as e:
                low = str(e).lower()
                if ("http 401" in low or "http 403" in low) and (user_id and user_secret):
                    # обновим токен
                    client2 = IikoTransportClient(
                        api_url=api_url,
                        user_id=user_id,
                        user_secret=user_secret,
                        organization_id=org_id,
                    )
                    products = client2.get_products()
                    try:
                        biz_new_token = (client2.access_token() or "").strip()
                    except Exception:
                        biz_new_token = ""
                else:
                    raise

        else:
            # REST (/resto)
            base_url = str(self.snapshot.get("rest_base_url") or "").strip()
            login = str(self.snapshot.get("rest_login") or "").strip()
            pass_sha1 = str(self.snapshot.get("rest_pass_sha1") or "").strip()
            if not (base_url and login and pass_sha1):
                raise IikoApiError("Не заданы параметры REST (/resto). Нажмите 'Авторизация точки'.")
            client = IikoRmsClient(base_url=base_url, login=login, pass_sha1=pass_sha1)
            products = client.get_products()

        products = list(products or [])

        # --- index for fast search ---
        def norm_sub(s: Any) -> str:
            return str(s or "").lower().replace("ё", "е")

        def pl_key(s: Any) -> str:
            return " ".join(str(s or "").strip().lower().replace("ё", "е").split())

        open_norm: List[Tuple[str, Any]] = []
        open_exact: dict[str, Any] = {}

        pricelist_dishes: List[DishItem] = []
        pricelist_norm: List[Tuple[str, DishItem]] = []
        pricelist_exact: dict[str, DishItem] = {}

        for p in products:
            name = (getattr(p, "name", "") or "").strip()
            if not name:
                continue

            nsub = norm_sub(name)
            open_norm.append((nsub, p))
            k = pl_key(name)
            if k and (k not in open_exact):
                open_exact[k] = p

            w0 = getattr(p, "weight", "")
            pr0 = getattr(p, "price", "")
            desc0 = getattr(p, "description", "")
            d = DishItem(
                name=name,
                weight=("" if w0 is None else w0),
                price=("" if pr0 is None else pr0),
                description=("" if desc0 is None else desc0),
            )
            pricelist_dishes.append(d)
            pricelist_norm.append((norm_sub(d.name), d))
            kd = pl_key(d.name)
            if kd and (kd not in pricelist_exact):
                pricelist_exact[kd] = d

        return {
            "mode": mode,
            "products": products,
            "open_norm": open_norm,
            "open_exact": open_exact,
            "pricelist_dishes": pricelist_dishes,
            "pricelist_norm": pricelist_norm,
            "pricelist_exact": pricelist_exact,
            "biz_new_token": biz_new_token,
        }


class ChainPriceTagsMergeWorker(QObject):
    """Выгружает чёрные ценники по шаблону (Excel COM).

    Заполняем название (+вес), цену и описание (если есть).
    """

    finished = Signal(int, str)  # seq, output_path
    failed = Signal(int, str)    # seq, error

    def __init__(self, seq: int, tags: List[TagData], output_path: str):
        super().__init__()
        self.seq = int(seq)
        self.tags = list(tags or [])
        self.output_path = str(output_path or "")

    def run(self) -> None:
        pythoncom = None
        try:
            try:
                import pythoncom as _pythoncom
                _pythoncom.CoInitialize()
                pythoncom = _pythoncom
            except Exception:
                pythoncom = None

            export_black_pricetags(self.tags, self.output_path)
            self.finished.emit(self.seq, self.output_path)
        except Exception as e:
            self.failed.emit(self.seq, str(e))
        finally:
            try:
                if pythoncom is not None:
                    pythoncom.CoUninitialize()
            except Exception:
                pass


class MainWindow(QMainWindow):
    def _show_access_token_dialog(self, token: str) -> None:
        """Показывает access_token и даёт кнопку 'Скопировать'."""
        try:
            dlg = QDialog(self)
            dlg.setWindowTitle("iiko — access_token")
            lay = QVBoxLayout(dlg)
            lay.addWidget(QLabel("Токен (access_token):"))

            ed = QLineEdit(dlg)
            ed.setText((token or "").strip())
            ed.setReadOnly(True)
            try:
                ed.setCursorPosition(0)
                ed.selectAll()
            except Exception:
                pass
            lay.addWidget(ed)

            btns = QHBoxLayout()
            btn_copy = QPushButton("Скопировать")
            btn_close = QPushButton("Закрыть")

            def _copy():
                try:
                    QApplication.clipboard().setText(ed.text())
                except Exception:
                    pass

            btn_copy.clicked.connect(_copy)
            btn_close.clicked.connect(dlg.accept)
            btns.addWidget(btn_copy)
            btns.addStretch(1)
            btns.addWidget(btn_close)
            lay.addLayout(btns)

            dlg.exec()
        except Exception:
            # если диалог не удалось показать — просто игнорируем
            pass
    def _reset_iiko_auth_settings(self) -> None:
        """Сбрасывает сохранённые данные авторизации iiko (чтобы не подставлялись старые секреты/хэши)."""
        keys = [
            # общие
            "iiko/mode",

            # iikoCloud v1
            "iiko/cloud/api_url",
            "iiko/cloud/api_login",
            "iiko/cloud/access_token",
            "iiko/cloud/org_id",
            "iiko/cloud/org_name",

            # iikoTransport (legacy iiko.biz)
            "iiko/biz/user_secret",
            "iiko/biz/user_id",
            "iiko/biz/access_token",
            "iiko/biz/org_id",
            "iiko/biz/org_name",
            "iiko/biz/api_url",

            # REST (/resto)
            "iiko/base_url",
            "iiko/login",
            "iiko/pass_sha1",
            "iiko/password",
        ]
        for k in keys:
            try:
                self._settings.remove(k)
            except Exception:
                pass

        # сбрасываем кэш-поля в рантайме
        try:
            self._iiko_mode = "cloud"

            self._iiko_cloud_api_url = "https://api-ru.iiko.services"
            self._iiko_cloud_api_login = ""
            self._iiko_cloud_access_token = ""
            self._iiko_cloud_org_id = ""
            self._iiko_cloud_org_name = ""

            # iikoTransport (legacy iiko.biz)
            self._iiko_biz_user_secret = ""
            self._iiko_biz_user_id = "pos_login_f13591ea"
            self._iiko_biz_org_id = ""
            self._iiko_biz_org_name = ""
            self._iiko_biz_api_url = "https://iiko.biz:9900"
            self._iiko_biz_access_token = ""

            # REST (/resto)
            self._iiko_base_url = ""
            self._iiko_login = ""
            self._iiko_pass_sha1_cached = ""
        except Exception:
            pass

        try:
            # инвалидируем возможную фоновую загрузку
            try:
                self._iiko_products_load_seq = int(getattr(self, "_iiko_products_load_seq", 0)) + 1
            except Exception:
                pass

            self._iiko_products_loaded = False
            self._iiko_products_by_key = {}
            self._pricelist_dishes = []
            self._open_iiko_products = []
            self._open_selected_ids = set()

            # индексы для быстрого поиска
            self._open_iiko_products_norm = []
            self._open_iiko_products_exact = {}
            self._pricelist_dishes_norm = []
            self._pricelist_dishes_exact = {}

            self._open_show_all_requested = False
            self._pricelist_show_all_requested = False
        except Exception:
            pass

    def _prompt_text(self, title: str, label: str, echo_mode=QLineEdit.Normal, default_text: str = "") -> Optional[str]:
        """Диалог ввода текста, где двойной клик в поле вставляет буфер обмена."""
        dlg = QDialog(self)
        dlg.setWindowTitle(title)
        lay = QVBoxLayout(dlg)
        lay.addWidget(QLabel(label))

        ed = PasteOnDoubleClickLineEdit(dlg)
        ed.setEchoMode(echo_mode)
        if default_text:
            ed.setText(default_text)
            try:
                ed.selectAll()
            except Exception:
                pass
        lay.addWidget(ed)

        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.accepted.connect(dlg.accept)
        bb.rejected.connect(dlg.reject)
        lay.addWidget(bb)

        ed.setFocus()
        if dlg.exec() != QDialog.Accepted:
            return None
        return (ed.text() or "").strip()

    def __init__(self):
        super().__init__()
        # Локальные настройки (Windows Registry) для запоминания параметров iiko
        self._settings = QSettings("excel_menu_gui", "excel_menu_gui")
        self.setWindowTitle("Работа с меню")
        # Apply centralized styling
        StyleManager.setup_main_window(self)

        central = QWidget()
        self.setCentralWidget(central)
        self.rootLayout = QVBoxLayout(central)
        LayoutStyles.apply_margins(self.rootLayout, LayoutStyles.DEFAULT_MARGINS)
        self.rootLayout.setSpacing(AppStyles.DEFAULT_SPACING)

        # Панель управления (сверху)
        self.topBar = QFrame(); self.topBar.setObjectName("topBar")
        self.layTop = QHBoxLayout(self.topBar)
        LayoutStyles.apply_margins(self.layTop, LayoutStyles.TOPBAR_MARGINS)
        self.layTop.setSpacing(AppStyles.DEFAULT_SPACING)

        lblTheme = QLabel("Тема:")
        self.cmbTheme = QComboBox()
        self.cmbTheme.addItems(["Системная", "Светлая", "Тёмная"])
        self.cmbTheme.setCurrentIndex(0)
        self.cmbTheme.currentIndexChanged.connect(self.on_theme_changed)

        self.btnDownloadTemplate = QPushButton("Скачать шаблон меню")
        self.btnDownloadTemplate.clicked.connect(self.do_download_template)
        self.btnMakePresentation = QPushButton("Сделать презентацию")
        self.btnMakePresentation.clicked.connect(self.do_make_presentation)
        self.btnBrokerage = QPushButton("Бракеражный журнал")
        self.btnBrokerage.clicked.connect(self.do_brokerage_journal)
        self.btnOpenMenu = QPushButton("Авторизация точки")
        self.btnOpenMenu.clicked.connect(self.do_authorize_point)

        self.btnOpenTomorrowDishes = QPushButton("Открыть блюда")
        self.btnOpenTomorrowDishes.clicked.connect(self.do_open_tomorrow_dishes)

        self.btnDocuments = QPushButton("Документы")
        self.btnDocuments.clicked.connect(self.show_documents_tab)

        self.btnDownloadPricelists = QPushButton("Скачать ценники")
        self.btnDownloadPricelists.clicked.connect(self.do_download_pricelists)
        self.btnShowCompare = QPushButton("Сравнить меню")
        self.btnShowCompare.clicked.connect(self.show_compare_sections)

        self.layTop.addWidget(lblTheme)
        self.layTop.addWidget(self.cmbTheme)
        self.layTop.addStretch(1)
        self.layTop.addWidget(self.btnShowCompare)
        self.layTop.addWidget(self.btnDownloadTemplate)
        self.layTop.addWidget(self.btnMakePresentation)
        self.layTop.addWidget(self.btnBrokerage)
        # iiko: авторизация/открыть блюда/ценники
        self.layTop.addWidget(self.btnOpenMenu)
        self.layTop.addWidget(self.btnOpenTomorrowDishes)
        self.layTop.addWidget(self.btnDownloadPricelists)
        self.layTop.addWidget(self.btnDocuments)

        # Apply centralized stylesheet (already set in StyleManager.setup_main_window)

        self.topGroup = nice_group("Панель управления", self.topBar)
        self.topGroup.setObjectName("topGroup")
        LayoutStyles.apply_size_policy(self.topGroup, LayoutStyles.EXPANDING_FIXED)
        self.rootLayout.addWidget(self.topGroup)

        # Область прокрутки для остального содержимого, чтобы панель всегда была наверху
        self.scrollArea = QScrollArea()
        self.scrollArea.setWidgetResizable(True)
        self.scrollArea.setFrameShape(QFrame.NoFrame)
        self.contentContainer = QWidget()
        self.contentLayout = QVBoxLayout(self.contentContainer)
        LayoutStyles.apply_margins(self.contentLayout, LayoutStyles.NO_MARGINS)
        self.contentLayout.setSpacing(AppStyles.CONTENT_SPACING)  # фиксированный вертикальный интервал между компонентами
        self.scrollArea.setWidget(self.contentContainer)
        self.scrollArea.setAlignment(Qt.AlignTop)  # прижимаем контент к верху, если он ниже области
        self.rootLayout.addWidget(self.scrollArea, 1)

        # Excel файл для презентации (используем тот же стиль, что и для файлов сравнения)
        self.edExcelPath = DropLineEdit()
        self.edExcelPath.setPlaceholderText("Выберите Excel файл с меню...")
        self.btnBrowseExcel = QPushButton("Обзор…")
        self.btnBrowseExcel.clicked.connect(lambda: self.browse_excel_file())

        excel_row = QWidget(); excel_layout = QHBoxLayout(excel_row)
        excel_layout.addWidget(self.edExcelPath, 1)
        excel_layout.addWidget(self.btnBrowseExcel)

        excel_group = QWidget(); excel_group_layout = QVBoxLayout(excel_group)
        excel_group_layout.addWidget(label_caption("Выберите Excel файл с меню..."))
        excel_group_layout.addWidget(excel_row)
        
        self.grpExcelFile = FileDropGroup("Выберите Excel файл с меню для презентации...", self.edExcelPath, excel_group)
        # Apply centralized styling for Excel file groups
        ComponentStyles.style_excel_group(self.grpExcelFile)
        self.contentLayout.addWidget(self.grpExcelFile)
        self.grpExcelFile.setVisible(False)
        
        # Панель действий внизу для сравнения (фиксированная)
        self.actionsPanel = QWidget(); self.actionsPanel.setObjectName("actionsPanel")
        self.actionsLayout = QHBoxLayout(self.actionsPanel)
        LayoutStyles.apply_margins(self.actionsLayout, LayoutStyles.MINIMAL_TOP_MARGIN)  # минимальный отступ сверху
        self.btnCompare = QPushButton("Сравнить и подсветить")
        self.btnCompare.clicked.connect(self.do_compare)
        self.actionsLayout.addStretch(1)
        self.actionsLayout.addWidget(self.btnCompare)
        self.rootLayout.addWidget(self.actionsPanel)
        self.actionsPanel.setVisible(False)
        
        # Панель действий внизу для презентаций (фиксированная)
        self.presentationActionsPanel = QWidget(); self.presentationActionsPanel.setObjectName("actionsPanel")
        self.presentationActionsLayout = QHBoxLayout(self.presentationActionsPanel)
        LayoutStyles.apply_margins(self.presentationActionsLayout, LayoutStyles.CONTENT_TOP_MARGIN)  # небольшой отступ сверху
        self.btnCreatePresentationWithData = QPushButton("Скачать презентацию с меню")
        self.btnCreatePresentationWithData.clicked.connect(self.do_create_presentation_with_data)
        self.presentationActionsLayout.addStretch(1)
        self.presentationActionsLayout.addWidget(self.btnCreatePresentationWithData)
        self.rootLayout.addWidget(self.presentationActionsPanel)
        self.presentationActionsPanel.setVisible(False)
        
        # Панель действий внизу для бракеражного журнала (фиксированная)
        self.brokerageActionsPanel = QWidget(); self.brokerageActionsPanel.setObjectName("actionsPanel")
        self.brokerageActionsLayout = QHBoxLayout(self.brokerageActionsPanel)
        LayoutStyles.apply_margins(self.brokerageActionsLayout, LayoutStyles.CONTENT_TOP_MARGIN)  # небольшой отступ сверху
        self.btnCreateBrokerageJournal = QPushButton("Скачать бракеражный журнал")
        self.btnCreateBrokerageJournal.clicked.connect(self.do_create_brokerage_journal_with_data)
        self.brokerageActionsLayout.addStretch(1)
        self.brokerageActionsLayout.addWidget(self.btnCreateBrokerageJournal)
        self.rootLayout.addWidget(self.brokerageActionsPanel)
        self.brokerageActionsPanel.setVisible(False)
        
        # Панель действий внизу для шаблона меню (фиксированная)
        self.templateActionsPanel = QWidget(); self.templateActionsPanel.setObjectName("actionsPanel")
        self.templateActionsLayout = QHBoxLayout(self.templateActionsPanel)
        LayoutStyles.apply_margins(self.templateActionsLayout, LayoutStyles.CONTENT_TOP_MARGIN)  # небольшой отступ сверху
        self.btnFillTemplate = QPushButton("Заполнить и скачать шаблон меню")
        self.btnFillTemplate.clicked.connect(self.do_fill_template_with_data)
        self.templateActionsLayout.addStretch(1)
        self.templateActionsLayout.addWidget(self.btnFillTemplate)
        self.rootLayout.addWidget(self.templateActionsPanel)
        self.templateActionsPanel.setVisible(False)

        # File 1
        self.edPath1 = DropLineEdit()
        self.btnBrowse1 = QPushButton("Обзор…")
        self.btnBrowse1.clicked.connect(lambda: self.browse_file(self.edPath1, which=1))

        row1 = QWidget(); r1 = QHBoxLayout(row1)
        r1.addWidget(self.edPath1, 1)
        r1.addWidget(self.btnBrowse1)

        g1 = QWidget(); l1 = QVBoxLayout(g1)
        l1.addWidget(label_caption("Файл 1"))
        l1.addWidget(row1)
        self.grpFirst = FileDropGroup("Первый файл", self.edPath1, g1)
        # Apply centralized styling for file groups
        ComponentStyles.style_file_group(self.grpFirst)
        self.contentLayout.addWidget(self.grpFirst)
        self.grpFirst.setVisible(False)

        # File 2
        self.edPath2 = DropLineEdit()
        self.btnBrowse2 = QPushButton("Обзор…")
        self.btnBrowse2.clicked.connect(lambda: self.browse_file(self.edPath2, which=2))

        row2 = QWidget(); r2 = QHBoxLayout(row2)
        r2.addWidget(self.edPath2, 1)
        r2.addWidget(self.btnBrowse2)

        g2 = QWidget(); l2 = QVBoxLayout(g2)
        l2.addWidget(label_caption("Файл 2"))
        l2.addWidget(row2)
        self.grpSecond = FileDropGroup("Второй файл", self.edPath2, g2)
        # Apply centralized styling for file groups
        ComponentStyles.style_file_group(self.grpSecond)
        self.contentLayout.addWidget(self.grpSecond)
        self.grpSecond.setVisible(False)

        # (дополнительно) — сворачиваемая группа
        opts = QWidget(); lo = QHBoxLayout(opts)
        LayoutStyles.apply_margins(lo, LayoutStyles.NO_MARGINS)
        lo.setSpacing(AppStyles.CONTENT_SPACING)  # немного больше для удобства чтения
        self.chkIgnoreCase = QCheckBox("Игнорировать регистр")
        self.chkIgnoreCase.setChecked(True)
        self.chkFuzzy = QCheckBox("Похожесть")
        self.spinFuzzy = QSpinBox(); self.spinFuzzy.setRange(0, 100); self.spinFuzzy.setValue(85)
        self.spinFuzzy.setEnabled(False)
        self.chkFuzzy.toggled.connect(self.spinFuzzy.setEnabled)

        lo.addWidget(self.chkIgnoreCase)
        lo.addWidget(self.chkFuzzy)
        lo.addWidget(QLabel("Порог:"))
        lo.addWidget(self.spinFuzzy)
        lo.addStretch(1)

        self.paramsBox = QGroupBox("Параметры (дополнительно)")
        # Apply centralized styling for parameter groups
        ComponentStyles.style_params_group(self.paramsBox)
        lparams = QVBoxLayout(self.paramsBox)
        lparams.setContentsMargins(0, 10, 0, 0)  # полностью убираем отступы
        lparams.setSpacing(AppStyles.CONTENT_SPACING)  # добавляем промежуток между заголовком и содержимым
        # Контейнер параметров без дополнительной рамки
        self._paramsFrame = QFrame(); self._paramsFrame.setObjectName("paramsFrame")
        lf = QHBoxLayout(self._paramsFrame)
        LayoutStyles.apply_margins(lf, LayoutStyles.NO_MARGINS)  # убираем отступы
        lf.setSpacing(0)
        lf.addWidget(opts)
        lparams.addWidget(self._paramsFrame)
        self._paramsFrame.setVisible(False)
        self.paramsBox.toggled.connect(self.on_params_toggled)
        self.contentLayout.addWidget(self.paramsBox)
        self.paramsBox.setVisible(False)

        # ===== ЦЕННИКИ: поиск + выбор =====
        self._pricelist_dishes: List[DishItem] = []
        self._pricelist_selected_keys: set[str] = set()

        # Источник блюд: iiko (REST / iikoCloud)
        self._iiko_mode = str(self._settings.value("iiko/mode", "cloud"))  # cloud | rest

        # REST (resto)
        self._iiko_base_url = str(self._settings.value("iiko/base_url", "https://patriot-co.iiko.it/resto"))
        self._iiko_login = str(self._settings.value("iiko/login", "user"))
        # Храним только sha1-хэш пароля (как требует iikoRMS resto API).
        self._iiko_pass_sha1_cached = str(self._settings.value("iiko/pass_sha1", ""))

        # iikoCloud v1 (api-ru.iiko.services): apiLogin -> access_token
        self._iiko_cloud_api_url = str(self._settings.value("iiko/cloud/api_url", "https://api-ru.iiko.services"))
        self._iiko_cloud_api_login = str(self._settings.value("iiko/cloud/api_login", ""))
        self._iiko_cloud_access_token = str(self._settings.value("iiko/cloud/access_token", ""))
        self._iiko_cloud_org_id = str(self._settings.value("iiko/cloud/org_id", ""))
        self._iiko_cloud_org_name = str(self._settings.value("iiko/cloud/org_name", ""))

        # legacy iiko.biz (оставлено на случай старых настроек)
        self._iiko_biz_api_url = str(self._settings.value("iiko/biz/api_url", "https://iiko.biz:9900"))
        self._iiko_biz_user_id = str(self._settings.value("iiko/biz/user_id", "pos_login_f13591ea"))
        self._iiko_biz_user_secret = str(self._settings.value("iiko/biz/user_secret", ""))
        self._iiko_biz_access_token = str(self._settings.value("iiko/biz/access_token", ""))
        self._iiko_biz_org_id = str(self._settings.value("iiko/biz/org_id", ""))
        self._iiko_biz_org_name = str(self._settings.value("iiko/biz/org_name", ""))

        # Миграция со старой версии: если вдруг сохранён plaintext-пароль — конвертируем и удаляем.
        try:
            legacy_pwd = str(self._settings.value("iiko/password", ""))
            if (not self._iiko_pass_sha1_cached) and legacy_pwd:
                self._iiko_pass_sha1_cached = hashlib.sha1(legacy_pwd.encode("utf-8")).hexdigest()
                self._settings.setValue("iiko/pass_sha1", self._iiko_pass_sha1_cached)
                self._settings.remove("iiko/password")
        except Exception:
            pass

        # ===== Оптимизация: загрузка iiko в фоне + быстрый поиск =====
        self._iiko_products_loaded: bool = False
        self._iiko_products_loading: bool = False
        self._iiko_products_load_seq: int = 0
        self._iiko_products_load_thread: Optional[QThread] = None
        self._iiko_products_load_worker: Optional[IikoProductsLoadWorker] = None
        self._iiko_products_load_origin: str = ""
        self._iiko_products_load_user_initiated: bool = False

        # Индексы для быстрого поиска (строятся после загрузки)
        self._open_iiko_products_norm: List[Tuple[str, Any]] = []
        self._open_iiko_products_exact: dict[str, Any] = {}
        self._pricelist_dishes_norm: List[Tuple[str, DishItem]] = []
        self._pricelist_dishes_exact: dict[str, DishItem] = {}

        # Флаги отложенных действий, если пользователь нажал кнопку до завершения загрузки
        self._open_show_all_requested: bool = False
        self._pricelist_show_all_requested: bool = False

        # ===== iikoChain: объединение выгрузок ценников (.xls) =====
        self._chain_pricetags_merge_loading: bool = False
        self._chain_pricetags_merge_seq: int = 0
        self._chain_pricetags_merge_thread: Optional[QThread] = None
        self._chain_pricetags_merge_worker: Optional[ChainPriceTagsMergeWorker] = None

        src_row = QWidget(); src_layout = QHBoxLayout(src_row)
        LayoutStyles.apply_margins(src_layout, LayoutStyles.NO_MARGINS)
        src_layout.addWidget(QLabel("Источник: iiko"))
        src_layout.addStretch(1)

        self.edDishSearch = QLineEdit()
        self.edDishSearch.setPlaceholderText("Начните вводить название блюда… (Enter — добавить)")
        # debounce, чтобы не фильтровать на каждый символ
        self._pricelist_search_pending_text: str = ""
        self._pricelist_search_timer = QTimer(self)
        self._pricelist_search_timer.setSingleShot(True)
        self._pricelist_search_timer.setInterval(180)
        self._pricelist_search_timer.timeout.connect(self._run_pricelist_search)
        self.edDishSearch.textChanged.connect(self._on_pricelist_search_text_changed)
        self.edDishSearch.returnPressed.connect(self._add_pricelist_from_enter)

        self.lblPricelistInfo = QLabel("Введите название блюда (загрузка из iiko — автоматически)")

        # Кнопку ручной загрузки оставляем, но скрываем: список подтягивается автоматически при поиске.
        self.btnLoadDishes = QPushButton("Загрузить блюда")
        self.btnLoadDishes.clicked.connect(self._load_pricelist_dishes)
        self.btnLoadDishes.setVisible(False)

        self.btnShowAllDishes = QPushButton("Показать все блюда")
        self.btnShowAllDishes.clicked.connect(self._show_all_pricelist_dishes)

        btns_row = QWidget(); btns_layout = QHBoxLayout(btns_row)
        LayoutStyles.apply_margins(btns_layout, LayoutStyles.NO_MARGINS)
        btns_layout.addWidget(self.btnLoadDishes)
        btns_layout.addWidget(self.btnShowAllDishes)
        btns_layout.addStretch(1)

        # Подсказки/список блюд (выпадающий список поверх UI — не в layout)
        self.lstDishSuggestions = QListWidget(self)
        # Важно: Qt.Popup перехватывает клавиатуру (backspace/ввод), поэтому используем Tool-окно.
        # Закрытие по клику вне списка делаем через eventFilter.
        try:
            self.lstDishSuggestions.setWindowFlags(
                Qt.Tool | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.WindowDoesNotAcceptFocus
            )
        except Exception:
            self.lstDishSuggestions.setWindowFlags(Qt.Tool)
        try:
            self.lstDishSuggestions.setAttribute(Qt.WA_ShowWithoutActivating, True)
        except Exception:
            pass
        self.lstDishSuggestions.setFocusPolicy(Qt.NoFocus)
        self.lstDishSuggestions.hide()
        self.lstDishSuggestions.itemClicked.connect(self._on_pricelist_suggestion_clicked)
        self.lstDishSuggestions.itemDoubleClicked.connect(self._on_pricelist_suggestion_clicked)

        self._suppress_pricelist_selected_item_changed = False

        self.lstSelectedDishes = QListWidget()
        self.lstSelectedDishes.setMinimumHeight(160)
        self.lstSelectedDishes.itemChanged.connect(self._on_pricelist_selected_item_changed)

        self.btnClearSelectedDishes = QPushButton("Очистить выбор")
        self.btnClearSelectedDishes.clicked.connect(self._clear_pricelist_selection)

        pricelist_box = QWidget(); pricelist_layout = QVBoxLayout(pricelist_box)
        pricelist_layout.addWidget(src_row)
        pricelist_layout.addWidget(self.edDishSearch)
        pricelist_layout.addWidget(self.lblPricelistInfo)
        pricelist_layout.addWidget(btns_row)
        pricelist_layout.addWidget(label_caption("Выбранные блюда (с галочками)"))
        pricelist_layout.addWidget(self.lstSelectedDishes)
        pricelist_layout.addWidget(self.btnClearSelectedDishes)

        self.grpPricelist = nice_group("Ценники: выбрать блюда", pricelist_box)
        self.contentLayout.addWidget(self.grpPricelist)
        self.grpPricelist.setVisible(False)

        # ===== ОТКРЫТИЕ БЛЮД (поиск в iiko -> снять со стоп-листа) =====
        # Старый сценарий (через Excel) оставлен в коде, но в UI не используется.
        self._tomorrow_menu_dishes: List[DishItem] = []
        self._iiko_products_by_key: dict[str, str] = {}
        self._suppress_tomorrow_item_changed = False

        # Данные для поиска по iiko
        self._open_iiko_products: List[Any] = []
        self._open_selected_ids: set[str] = set()

        # Запланированное открытие (вариант B): сохраняем в настройках, исполняем когда наступит дата.
        self._open_schedule_job: Optional[dict] = None
        self._open_schedule_timer: Optional[QTimer] = None

        # Инфо-строка (показываем только когда есть статус/ошибка; подсказки по UI не показываем)
        self.lblTomorrowInfo = QLabel("")
        self.lblTomorrowInfo.setVisible(False)

        # Календарь выбора даты открытия
        self.calOpenDate = QCalendarWidget()
        self.calOpenDate.setGridVisible(True)
        self.calOpenDate.setFirstDayOfWeek(Qt.Monday)
        # Убираем "лишний" первый столбец (номер недели), чтобы было ровно 7 колонок дней.
        try:
            self.calOpenDate.setVerticalHeaderFormat(QCalendarWidget.NoVerticalHeader)
        except Exception:
            pass

        try:
            # "квадратный" вид: фиксируем размер, чтобы не растягивался
            self.calOpenDate.setFixedSize(320, 260)
        except Exception:
            pass

        # Восстановим выбранную дату из настроек
        try:
            saved = str(self._settings.value("iiko/open_dishes/target_date", ""))
        except Exception:
            saved = ""
        qd_default = QDate.currentDate().addDays(1)  # по умолчанию завтра
        qd = QDate.fromString(saved, "yyyy-MM-dd") if saved else QDate()
        if not qd.isValid():
            qd = qd_default
        self.calOpenDate.setSelectedDate(qd)
        self.calOpenDate.selectionChanged.connect(self._on_open_target_date_changed)
        self._update_open_action_button_label()

        self.edTomorrowSearch = QLineEdit()
        self.edTomorrowSearch.setPlaceholderText("Начните вводить название блюда… (Enter — добавить)")
        # debounce, чтобы не фильтровать на каждый символ
        self._open_search_pending_text: str = ""
        self._open_search_timer = QTimer(self)
        self._open_search_timer.setSingleShot(True)
        self._open_search_timer.setInterval(180)
        self._open_search_timer.timeout.connect(self._run_open_search)
        self.edTomorrowSearch.textChanged.connect(self._on_open_search_text_changed)
        self.edTomorrowSearch.returnPressed.connect(self._add_open_from_enter)

        self.btnShowAllOpenDishes = QPushButton("Показать все блюда")
        self.btnShowAllOpenDishes.clicked.connect(self._show_all_open_dishes)

        btns_row_open = QWidget(); btns_layout_open = QHBoxLayout(btns_row_open)
        LayoutStyles.apply_margins(btns_layout_open, LayoutStyles.NO_MARGINS)
        btns_layout_open.addWidget(self.btnShowAllOpenDishes)
        btns_layout_open.addStretch(1)

        # Подсказки (выпадающий список поверх UI — не в layout)
        self.lstTomorrowDishes = QListWidget(self)
        try:
            self.lstTomorrowDishes.setWindowFlags(
                Qt.Tool | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.WindowDoesNotAcceptFocus
            )
        except Exception:
            self.lstTomorrowDishes.setWindowFlags(Qt.Tool)
        try:
            self.lstTomorrowDishes.setAttribute(Qt.WA_ShowWithoutActivating, True)
        except Exception:
            pass
        self.lstTomorrowDishes.setFocusPolicy(Qt.NoFocus)
        self.lstTomorrowDishes.hide()
        self.lstTomorrowDishes.itemClicked.connect(self._on_open_suggestion_clicked)
        self.lstTomorrowDishes.itemDoubleClicked.connect(self._on_open_suggestion_clicked)

        # Выбранные
        self.lstTomorrowSelectedDishes = QListWidget()
        self.lstTomorrowSelectedDishes.setMinimumHeight(160)

        self.btnClearTomorrowSelection = QPushButton("Очистить выбор")
        self.btnClearTomorrowSelection.clicked.connect(self._clear_open_selection)

        # Левая панель: календарь
        open_left = QWidget(); open_left_layout = QVBoxLayout(open_left)
        LayoutStyles.apply_margins(open_left_layout, LayoutStyles.NO_MARGINS)
        open_left_layout.addWidget(label_caption("Дата открытия"))
        open_left_layout.addWidget(self.calOpenDate)
        open_left_layout.addStretch(1)

        # Правая панель: поиск/выбор
        open_right = QWidget(); open_right_layout = QVBoxLayout(open_right)
        LayoutStyles.apply_margins(open_right_layout, LayoutStyles.NO_MARGINS)
        open_right_layout.addWidget(self.lblTomorrowInfo)
        open_right_layout.addWidget(label_caption("Поиск блюда"))
        open_right_layout.addWidget(self.edTomorrowSearch)
        open_right_layout.addWidget(btns_row_open)
        open_right_layout.addWidget(label_caption("Выбранные блюда (с галочками)"))
        open_right_layout.addWidget(self.lstTomorrowSelectedDishes)
        open_right_layout.addWidget(self.btnClearTomorrowSelection)

        open_root = QWidget(); open_root_layout = QHBoxLayout(open_root)
        LayoutStyles.apply_margins(open_root_layout, LayoutStyles.NO_MARGINS)
        open_root_layout.addWidget(open_left)
        open_root_layout.addWidget(open_right, 1)

        self.grpTomorrowOpen = nice_group("Открыть блюда (iiko стоп-лист)", open_root)
        self.contentLayout.addWidget(self.grpTomorrowOpen)
        self.grpTomorrowOpen.setVisible(False)

        # ===== ДОКУМЕНТЫ: быстрый доступ к файлам =====
        docs_box = QWidget(); docs_layout = QVBoxLayout(docs_box)
        docs_layout.setSpacing(AppStyles.CONTENT_SPACING)

        self.btnVacationStatement = QPushButton("Заявление на отпуск")
        self.btnVacationStatement.clicked.connect(self.open_vacation_statement)
        StyleManager.style_action_button(self.btnVacationStatement)

        self.btnMedicalBooks = QPushButton("Медкнижки")
        self.btnMedicalBooks.clicked.connect(self.open_med_books)
        StyleManager.style_action_button(self.btnMedicalBooks)

        self.btnBirthdayFile = QPushButton("День рождения")
        self.btnBirthdayFile.clicked.connect(self.open_birthday_file)
        StyleManager.style_action_button(self.btnBirthdayFile)

        self.btnHygieneJournal = QPushButton("Гигиенический журнал")
        self.btnHygieneJournal.clicked.connect(self.open_hygiene_journal)
        StyleManager.style_action_button(self.btnHygieneJournal)

        self.btnDirection = QPushButton("Направление")
        self.btnDirection.clicked.connect(self.open_direction_document)
        StyleManager.style_action_button(self.btnDirection)

        self.btnLockerDoc = QPushButton("Раздевалка")
        self.btnLockerDoc.clicked.connect(self.open_locker_document)
        StyleManager.style_action_button(self.btnLockerDoc)

        self.btnFridgeTemp = QPushButton("Температура холодильник")
        self.btnFridgeTemp.clicked.connect(self.open_fridge_temperature)
        StyleManager.style_action_button(self.btnFridgeTemp)

        self.btnFreezerTemp = QPushButton("Температура морозилка")
        self.btnFreezerTemp.clicked.connect(self.open_freezer_temperature)
        StyleManager.style_action_button(self.btnFreezerTemp)

        self.btnBuffetSheet = QPushButton("Буфет бумажка")
        self.btnBuffetSheet.clicked.connect(self.open_buffet_sheet)
        StyleManager.style_action_button(self.btnBuffetSheet)

        self.btnBakerSheet = QPushButton("Пекарь бумажка")
        self.btnBakerSheet.clicked.connect(self.open_baker_sheet)
        StyleManager.style_action_button(self.btnBakerSheet)

        self.btnFryerJournal = QPushButton("Журнал фритюрного масла")
        self.btnFryerJournal.clicked.connect(self.open_fryer_oil_journal)
        StyleManager.style_action_button(self.btnFryerJournal)

        # Новые документы/папки из templates
        self.btnBreakfasts = QPushButton("Завтраки")
        self.btnBreakfasts.clicked.connect(self.open_breakfasts_folder)
        StyleManager.style_action_button(self.btnBreakfasts)

        self.btnDistribution = QPushButton("Раздача")
        self.btnDistribution.clicked.connect(self.open_distribution_sheet)
        StyleManager.style_action_button(self.btnDistribution)

        self.btnPieNames = QPushButton("Название пирогов")
        self.btnPieNames.clicked.connect(self.open_pie_names)
        StyleManager.style_action_button(self.btnPieNames)

        self.btnBakeryPricelist = QPushButton("Прейскурант выпечка")
        self.btnBakeryPricelist.clicked.connect(self.open_bakery_pricelist)
        StyleManager.style_action_button(self.btnBakeryPricelist)

        self.btnCashTemplate = QPushButton("Шаблон наличка")
        self.btnCashTemplate.clicked.connect(self.open_cash_template)
        StyleManager.style_action_button(self.btnCashTemplate)

        # Кнопки "Документы": 3 категории = 3 колонки (1я/2я/3я), внутри — кнопки 2×N одинакового размера
        try:
            # Приведём подписи к коротким названиям (без кавычек/"Excel")
            try:
                self.btnCashTemplate.setText("Наличка")
            except Exception:
                pass

            btn_h = int(getattr(AppStyles, "BUTTON_HEIGHT", 40))

            def _apply_btn_size(btns: List[QPushButton]) -> None:
                for b in btns:
                    try:
                        b.setFixedHeight(btn_h)
                        b.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                    except Exception:
                        pass

            def _make_category(title: str, btns: List[QPushButton]) -> QGroupBox:
                # Одна категория = один столбец (вертикально), кнопки внутри — по алфавиту
                def _btn_key(b: QPushButton) -> str:
                    try:
                        t = str(b.text() or "")
                    except Exception:
                        t = ""
                    return " ".join(t.lower().replace("ё", "е").split())

                btns_sorted = sorted(list(btns or []), key=_btn_key)

                _apply_btn_size(btns_sorted)
                box = QGroupBox(title)
                v = QVBoxLayout(box)
                v.setSpacing(AppStyles.CONTENT_SPACING)
                try:
                    LayoutStyles.apply_margins(v, LayoutStyles.NO_MARGINS)
                except Exception:
                    pass
                for b in btns_sorted:
                    v.addWidget(b)
                v.addStretch(1)
                return box

            # Категории документов
            cat_people = _make_category(
                "Кадры",
                [
                    self.btnVacationStatement,
                    self.btnMedicalBooks,
                    self.btnBirthdayFile,
                    self.btnDirection,
                    self.btnLockerDoc,
                ],
            )
            cat_journals = _make_category(
                "Журналы",
                [
                    self.btnHygieneJournal,
                    self.btnFryerJournal,
                    self.btnFridgeTemp,
                    self.btnFreezerTemp,
                ],
            )
            cat_templates = _make_category(
                "Шаблоны",
                [
                    self.btnBreakfasts,
                    self.btnDistribution,
                    self.btnPieNames,
                    self.btnBakeryPricelist,
                    self.btnCashTemplate,
                    self.btnBuffetSheet,
                    self.btnBakerSheet,
                ],
            )

            cats_row = QWidget()
            cats_grid = QGridLayout(cats_row)
            try:
                LayoutStyles.apply_margins(cats_grid, LayoutStyles.NO_MARGINS)
            except Exception:
                pass
            cats_grid.setHorizontalSpacing(AppStyles.CONTENT_SPACING)
            cats_grid.setVerticalSpacing(AppStyles.CONTENT_SPACING)

            cats_grid.addWidget(cat_people, 0, 0)
            cats_grid.addWidget(cat_journals, 0, 1)
            cats_grid.addWidget(cat_templates, 0, 2)
            cats_grid.setColumnStretch(0, 1)
            cats_grid.setColumnStretch(1, 1)
            cats_grid.setColumnStretch(2, 1)

            docs_layout.addWidget(cats_row)
        except Exception:
            # fallback: если что-то пошло не так — оставим вертикально
            try:
                docs_layout.addWidget(self.btnVacationStatement)
                docs_layout.addWidget(self.btnMedicalBooks)
                docs_layout.addWidget(self.btnBirthdayFile)
                docs_layout.addWidget(self.btnDirection)
                docs_layout.addWidget(self.btnLockerDoc)
                docs_layout.addWidget(self.btnHygieneJournal)
                docs_layout.addWidget(self.btnFridgeTemp)
                docs_layout.addWidget(self.btnFreezerTemp)
                docs_layout.addWidget(self.btnFryerJournal)
                docs_layout.addWidget(self.btnBreakfasts)
                docs_layout.addWidget(self.btnDistribution)
                docs_layout.addWidget(self.btnPieNames)
                docs_layout.addWidget(self.btnBakeryPricelist)
                docs_layout.addWidget(self.btnCashTemplate)
                docs_layout.addWidget(self.btnBuffetSheet)
                docs_layout.addWidget(self.btnBakerSheet)
            except Exception:
                pass

        docs_layout.addStretch(1)

        self.grpDocuments = nice_group("Документы", docs_box)
        self.contentLayout.addWidget(self.grpDocuments)
        self.grpDocuments.setVisible(False)

        # Панель действий внизу для ценников (фиксированная)
        self.pricelistActionsPanel = QWidget(); self.pricelistActionsPanel.setObjectName("actionsPanel")
        self.pricelistActionsLayout = QHBoxLayout(self.pricelistActionsPanel)
        LayoutStyles.apply_margins(self.pricelistActionsLayout, LayoutStyles.CONTENT_TOP_MARGIN)

        # Одна кнопка: выгрузка/объединение чёрных ценников iikoChain ("Большой ценник")
        self.btnMergeChainPricetags = QPushButton("Выгрузить ценники")
        self.btnMergeChainPricetags.clicked.connect(self.do_merge_chain_pricetags)

        self.pricelistActionsLayout.addStretch(1)
        self.pricelistActionsLayout.addWidget(self.btnMergeChainPricetags)
        self.rootLayout.addWidget(self.pricelistActionsPanel)
        self.pricelistActionsPanel.setVisible(False)

        # Панель действий внизу для "Открыть блюда" (фиксированная)
        self.tomorrowOpenActionsPanel = QWidget(); self.tomorrowOpenActionsPanel.setObjectName("actionsPanel")
        self.tomorrowOpenActionsLayout = QHBoxLayout(self.tomorrowOpenActionsPanel)
        LayoutStyles.apply_margins(self.tomorrowOpenActionsLayout, LayoutStyles.CONTENT_TOP_MARGIN)
        self.btnOpenTomorrowChecked = QPushButton("Открыть отмеченные")
        self.btnOpenTomorrowChecked.clicked.connect(self._open_tomorrow_checked)
        self.tomorrowOpenActionsLayout.addStretch(1)
        self.tomorrowOpenActionsLayout.addWidget(self.btnOpenTomorrowChecked)
        self.rootLayout.addWidget(self.tomorrowOpenActionsPanel)
        self.tomorrowOpenActionsPanel.setVisible(False)

        # Добавляем нижний растягивающий элемент, чтобы контент не растягивался равномерно, а был прижат кверху
        self.contentLayout.addStretch(1)

        # Прячем подсказки при клике вне поля поиска (как выпадающий список)
        try:
            QApplication.instance().installEventFilter(self)
        except Exception:
            pass

        # Таймер: проверяем запланированное открытие блюд
        try:
            self._load_open_schedule_job()
            self._open_schedule_timer = QTimer(self)
            self._open_schedule_timer.setInterval(30_000)
            self._open_schedule_timer.timeout.connect(self._check_open_schedule_due)
            self._open_schedule_timer.start()
            # проверим сразу после запуска
            QTimer.singleShot(1500, self._check_open_schedule_due)
        except Exception:
            pass

        # Theming initialization
        self._theme_mode = ThemeMode.SYSTEM  # По умолчанию используем системную тему
        try:
            apply_theme(QApplication.instance(), self._theme_mode)
        except Exception:
            pass

        # Обновим стили календаря после применения темы
        try:
            self._apply_calendar_theme_overrides()
        except Exception:
            pass
        
        # Запускаем таймер для отслеживания изменений системной темы
        self._theme_timer = start_system_theme_watcher(
            lambda is_light: self.handle_system_theme_change(is_light),
            interval_ms=1000  # Проверка каждую секунду
        )

        # Применяем компактный режим при узкой ширине окна (мобильный превью)
        try:
            self._apply_compact_mode(self.width() <= 480)
        except Exception:
            pass

        # Останавливаем фоновые потоки при выходе приложения
        try:
            QApplication.instance().aboutToQuit.connect(self._shutdown_background_threads)
        except Exception:
            pass

    def log(self, msg: str):
        # Лог отключён по запросу — ничего не делаем
        pass

    def on_theme_changed(self, idx: int):
        try:
            # Получаем режим темы из выбора пользователя
            if idx == 0:
                self._theme_mode = ThemeMode.SYSTEM  # Системная тема
            elif idx == 1:
                self._theme_mode = ThemeMode.LIGHT   # Светлая тема
            else:
                self._theme_mode = ThemeMode.DARK    # Тёмная тема

            # Применяем выбранную тему
            apply_theme(QApplication.instance(), self._theme_mode)
        except Exception:
            pass

        try:
            self._apply_calendar_theme_overrides()
        except Exception:
            pass
    def handle_system_theme_change(self, is_light: bool):
        """Обработчик изменения системной темы Windows"""
        try:
            # Обновляем тему только если выбрана "Системная"
            if self._theme_mode == ThemeMode.SYSTEM and self.cmbTheme.currentIndex() == 0:
                # Применяем соответствующую системную тему
                theme = ThemeMode.LIGHT if is_light else ThemeMode.DARK
                apply_theme(QApplication.instance(), theme)
        except Exception:
            pass

        try:
            self._apply_calendar_theme_overrides()
        except Exception:
            pass

    def _is_dark_palette(self) -> bool:
        """Пытаемся определить, тёмная ли сейчас тема по палитре приложения."""
        try:
            p = QApplication.instance().palette()
            c = p.color(QPalette.Window)
            # относительная яркость
            lum = 0.2126 * float(c.red()) + 0.7152 * float(c.green()) + 0.0722 * float(c.blue())
            return lum < 128.0
        except Exception:
            return False

    def _apply_calendar_theme_overrides(self) -> None:
        """Убирает красные выходные и делает бежевое выделение в календаре."""
        try:
            if not hasattr(self, "calOpenDate"):
                return

            dark = self._is_dark_palette()

            # бежевое выделение (в тёмной теме делаем чуть темнее/приглушённее)
            sel_bg = QColor("#e2c9a7") if not dark else QColor("#c2a77d")
            sel_text = QColor("#1b1b1b")

            pal = self.calOpenDate.palette()
            pal.setColor(QPalette.Highlight, sel_bg)
            pal.setColor(QPalette.HighlightedText, sel_text)
            self.calOpenDate.setPalette(pal)

            # Убираем красный цвет выходных (делаем как обычный текст)
            try:
                fmt_weekend = QTextCharFormat()
                fmt_weekend.setForeground(QBrush(self.calOpenDate.palette().color(QPalette.Text)))
                self.calOpenDate.setWeekendTextFormat(fmt_weekend)
                # иногда "красное" задаётся отдельно для дней недели
                try:
                    self.calOpenDate.setWeekdayTextFormat(Qt.Saturday, fmt_weekend)
                    self.calOpenDate.setWeekdayTextFormat(Qt.Sunday, fmt_weekend)
                except Exception:
                    pass
            except Exception:
                pass

        except Exception:
            pass


    def _shutdown_background_threads(self):
        """Безопасно останавливает фоновые QThread, чтобы не было 'QThread destroyed while thread is still running'."""
        for attr in ("_iiko_products_load_thread", "_chain_pricetags_merge_thread"):
            try:
                th = getattr(self, attr, None)
                if th is not None and th.isRunning():
                    try:
                        th.requestInterruption()
                    except Exception:
                        pass
                    try:
                        th.quit()
                    except Exception:
                        pass
                    try:
                        th.wait(15_000)
                    except Exception:
                        pass
                try:
                    _ACTIVE_THREADS.discard(th)
                except Exception:
                    pass
            except Exception:
                pass

    def closeEvent(self, event):
        # При закрытии окна безопасно гасим фоновые потоки.
        self._shutdown_background_threads()

        try:
            if hasattr(self, "_theme_timer") and self._theme_timer is not None:
                self._theme_timer.stop()
        except Exception:
            pass
        super().closeEvent(event)

    def _apply_compact_mode(self, compact: bool) -> None:
        if getattr(self, "_compact", None) == compact:
            return
        self._compact = compact
        try:
            # Корневой layout
            if hasattr(self, "rootLayout"):
                LayoutStyles.apply_margins(self.rootLayout, (8, 8, 8, 8) if compact else LayoutStyles.DEFAULT_MARGINS)
                self.rootLayout.setSpacing(AppStyles.COMPACT_SPACING if compact else AppStyles.DEFAULT_SPACING)
            # Верхняя панель: горизонтально на десктопе, вертикально в компактном режиме
            if hasattr(self, "layTop"):
                self.layTop.setDirection(QBoxLayout.TopToBottom if compact else QBoxLayout.LeftToRight)
                LayoutStyles.apply_margins(self.layTop, (8, 8, 8, 8) if compact else LayoutStyles.TOPBAR_MARGINS)
                self.layTop.setSpacing(AppStyles.COMPACT_SPACING if compact else AppStyles.DEFAULT_SPACING)
            # Контентные отступы
            if hasattr(self, "contentLayout"):
                self.contentLayout.setSpacing(AppStyles.COMPACT_SPACING if compact else AppStyles.CONTENT_SPACING)
        except Exception:
            pass

    def resizeEvent(self, event):
        super().resizeEvent(event)
        try:
            self._apply_compact_mode(self.width() <= 480)
        except Exception:
            pass

    def browse_file(self, target: QLineEdit, which: int):
        path, _ = QFileDialog.getOpenFileName(self, "Выберите файл", str(Path.cwd()), "Excel (*.xls *.xlsx *.xlsm);;Все файлы (*.*)")
        if path:
            target.setText(path)

    def select_default_sheet(self, path: str) -> Optional[str]:
        try:
            names = get_sheet_names(path)
            if not names:
                return None
            for nm in names:
                low = nm.strip().lower()
                if "касс" in low or "kass" in low:
                    return nm
            return names[0]
        except Exception:
            return None


    def do_compare(self):
        try:
            p1 = self.edPath1.text().strip()
            p2 = self.edPath2.text().strip()
            s1 = self.select_default_sheet(p1) if p1 else None
            s2 = self.select_default_sheet(p2) if p2 else None
            if not (p1 and p2):
                QMessageBox.warning(self, "Внимание", "Укажите оба файла.")
                return
            if not (s1 and s2):
                QMessageBox.warning(self, "Листы", "Не удалось определить листы для сравнения.")
                return
            
            # Выбираем место сохранения результата сравнения
            # Получаем дату из Excel файлов для правильного названия
            from comparator import _extract_best_date_from_file
            from datetime import date
            
            # Извлекаем даты из обоих файлов
            d1 = _extract_best_date_from_file(p1, s1)
            d2 = _extract_best_date_from_file(p2, s2)
            
            # Определяем самую позднюю дату для названия файла
            latest_date = None
            if d1 and d2:
                latest_date = max(d1, d2)
            elif d1:
                latest_date = d1
            elif d2:
                latest_date = d2
            
            # Формируем предлагаемое имя с правильной датой
            if latest_date:
                date_str = latest_date.strftime("%d.%m.%Y")
                suggested_name = f"сравнение_меню_{date_str}.xlsx"
            else:
                # Если даты не найдены, используем текущую дату как fallback
                today_str = date.today().strftime("%d.%m.%Y")
                suggested_name = f"сравнение_меню_{today_str}.xlsx"
            
            desktop = Path.home() / "Desktop"
            suggested_path = str(desktop / suggested_name)
            
            save_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Сохранить результат сравнения", 
                suggested_path, 
                "Excel (*.xlsx);;Excel (*.xls);;Все файлы (*.*)"
            )
            
            if not save_path:
                return  # Пользователь отменил сохранение
            
            try:
                # Всегда авто по дате
                temp_out_path, matches = compare_and_highlight(
                    path1=p1, sheet1=s1,
                    path2=p2, sheet2=s2,
                    col1="A",
                    col2="E",
                    header_row1=1,
                    header_row2=1,
                    ignore_case=self.chkIgnoreCase.isChecked(),
                    use_fuzzy=self.chkFuzzy.isChecked(),
                    fuzzy_threshold=int(self.spinFuzzy.value()),
                    final_choice=0,  # авто: финальным будет файл с более поздней датой
                )
                
                # Перемещаем файл в выбранное место
                if temp_out_path != save_path:
                    import shutil
                    shutil.move(temp_out_path, save_path)
                
                self.log(f"Готово. Совпадений: {matches}. Итоговый файл: {save_path}")
                QMessageBox.information(self, "Готово", f"Совпадений: {matches}\nФайл сохранён: {save_path}")
            except ColumnParseError as e:
                QMessageBox.warning(self, "Колонка", str(e))
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def show_compare_sections(self):
        try:
            # Скрываем панели, не относящиеся к сравнению
            if hasattr(self, "grpExcelFile"):
                self.grpExcelFile.setVisible(False)
            if hasattr(self, "presentationActionsPanel"):
                self.presentationActionsPanel.setVisible(False)
            if hasattr(self, "brokerageActionsPanel"):
                self.brokerageActionsPanel.setVisible(False)
            if hasattr(self, "templateActionsPanel"):
                self.templateActionsPanel.setVisible(False)
            if hasattr(self, "grpPricelist"):
                self.grpPricelist.setVisible(False)
            if hasattr(self, "pricelistActionsPanel"):
                self.pricelistActionsPanel.setVisible(False)
            if hasattr(self, "grpTomorrowOpen"):
                self.grpTomorrowOpen.setVisible(False)
            if hasattr(self, "tomorrowOpenActionsPanel"):
                self.tomorrowOpenActionsPanel.setVisible(False)
            if hasattr(self, "grpDocuments"):
                self.grpDocuments.setVisible(False)

            # Показать формы сравнения и панель действий
            if hasattr(self, "grpFirst"):
                self.grpFirst.setVisible(True)
            if hasattr(self, "grpSecond"):
                self.grpSecond.setVisible(True)
            if hasattr(self, "paramsBox"):
                self.paramsBox.setVisible(True)
                self.paramsBox.setChecked(False)
            if hasattr(self, "actionsPanel"):
                self.actionsPanel.setVisible(True)
        except Exception:
            pass

    def show_documents_tab(self) -> None:
        """Показывает вкладку "Документы" с кнопками для открытия файлов."""
        try:
            # Скрываем другие панели
            if hasattr(self, "grpFirst"):
                self.grpFirst.setVisible(False)
            if hasattr(self, "grpSecond"):
                self.grpSecond.setVisible(False)
            if hasattr(self, "paramsBox"):
                self.paramsBox.setVisible(False)
            if hasattr(self, "actionsPanel"):
                self.actionsPanel.setVisible(False)
            if hasattr(self, "grpExcelFile"):
                self.grpExcelFile.setVisible(False)
            if hasattr(self, "presentationActionsPanel"):
                self.presentationActionsPanel.setVisible(False)
            if hasattr(self, "brokerageActionsPanel"):
                self.brokerageActionsPanel.setVisible(False)
            if hasattr(self, "templateActionsPanel"):
                self.templateActionsPanel.setVisible(False)
            if hasattr(self, "grpPricelist"):
                self.grpPricelist.setVisible(False)
            if hasattr(self, "pricelistActionsPanel"):
                self.pricelistActionsPanel.setVisible(False)
            if hasattr(self, "grpTomorrowOpen"):
                self.grpTomorrowOpen.setVisible(False)
            if hasattr(self, "tomorrowOpenActionsPanel"):
                self.tomorrowOpenActionsPanel.setVisible(False)

            # Показываем блок с документами
            if hasattr(self, "grpDocuments"):
                self.grpDocuments.setVisible(True)
                if hasattr(self, "scrollArea"):
                    self.scrollArea.ensureWidgetVisible(self.grpDocuments)
        except Exception:
            pass

    def on_params_toggled(self, checked: bool):
        try:
            if hasattr(self, "_paramsFrame"):
                self._paramsFrame.setVisible(checked)
            if checked and hasattr(self, "scrollArea") and hasattr(self, "_paramsFrame"):
                self.scrollArea.ensureWidgetVisible(self._paramsFrame)
        except Exception:
            pass

    def do_download_template(self):
        """Показывает панель для работы с шаблоном меню"""
        try:
            # Скрываем другие панели
            if hasattr(self, "grpFirst"):
                self.grpFirst.setVisible(False)
            if hasattr(self, "grpSecond"):
                self.grpSecond.setVisible(False)
            if hasattr(self, "paramsBox"):
                self.paramsBox.setVisible(False)
            if hasattr(self, "actionsPanel"):
                self.actionsPanel.setVisible(False)
            if hasattr(self, "presentationActionsPanel"):
                self.presentationActionsPanel.setVisible(False)
            if hasattr(self, "brokerageActionsPanel"):
                self.brokerageActionsPanel.setVisible(False)
            if hasattr(self, "grpPricelist"):
                self.grpPricelist.setVisible(False)
            if hasattr(self, "pricelistActionsPanel"):
                self.pricelistActionsPanel.setVisible(False)
            if hasattr(self, "grpTomorrowOpen"):
                self.grpTomorrowOpen.setVisible(False)
            if hasattr(self, "tomorrowOpenActionsPanel"):
                self.tomorrowOpenActionsPanel.setVisible(False)
            if hasattr(self, "grpDocuments"):
                self.grpDocuments.setVisible(False)
            
            # Показываем панель для работы с шаблоном меню и её панель действий
            if hasattr(self, "grpExcelFile"):
                self.grpExcelFile.setVisible(True)
                self.grpExcelFile.setTitle("Файл меню для заполнения шаблона")
                self.edExcelPath.setPlaceholderText("Выберите Excel файл с меню для заполнения шаблона...")
            if hasattr(self, "templateActionsPanel"):
                self.templateActionsPanel.setVisible(True)
                
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def do_make_presentation(self):
        """Показывает панель для работы с презентациями"""
        try:
            # Скрываем другие панели
            if hasattr(self, "grpFirst"):
                self.grpFirst.setVisible(False)
            if hasattr(self, "grpSecond"):
                self.grpSecond.setVisible(False)
            if hasattr(self, "paramsBox"):
                self.paramsBox.setVisible(False)
            if hasattr(self, "actionsPanel"):
                self.actionsPanel.setVisible(False)
            if hasattr(self, "brokerageActionsPanel"):
                self.brokerageActionsPanel.setVisible(False)
            if hasattr(self, "templateActionsPanel"):
                self.templateActionsPanel.setVisible(False)
            if hasattr(self, "grpPricelist"):
                self.grpPricelist.setVisible(False)
            if hasattr(self, "pricelistActionsPanel"):
                self.pricelistActionsPanel.setVisible(False)
            if hasattr(self, "grpTomorrowOpen"):
                self.grpTomorrowOpen.setVisible(False)
            if hasattr(self, "tomorrowOpenActionsPanel"):
                self.tomorrowOpenActionsPanel.setVisible(False)
            if hasattr(self, "grpDocuments"):
                self.grpDocuments.setVisible(False)
            
            # Показываем панель для работы с презентациями и её панель действий
            if hasattr(self, "grpExcelFile"):
                self.grpExcelFile.setVisible(True)
                self.grpExcelFile.setTitle("Файл меню для презентации")
                self.edExcelPath.setPlaceholderText("Выберите Excel файл с меню (салаты, первые блюда, мясо, птица, рыба, гарниры)...")
            if hasattr(self, "presentationActionsPanel"):
                self.presentationActionsPanel.setVisible(True)
                
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def browse_excel_file(self):
        """Выбор Excel файла для презентации"""
        path, _ = QFileDialog.getOpenFileName(
            self, 
            "Выберите Excel файл с меню", 
            str(Path.cwd()), 
            "Excel (*.xls *.xlsx *.xlsm);;Все файлы (*.*)"
        )
        if path:
            self.edExcelPath.setText(path)

    def do_authorize_point(self):
        """Авторизация точки iiko.

        Поддерживает:
        - iikoRMS REST (/resto)
        - iikoCloud API v1 (api-ru.iiko.services) через apiLogin
        - iikoTransport (iiko.biz:9900) через user_id/user_secret
        """
        try:
            # По умолчанию используем REST (/resto). Чтобы открыть скрытые способы — удерживайте Shift.
            use_advanced = bool(QApplication.keyboardModifiers() & Qt.ShiftModifier)

            # Явно предлагаем "Сбросить" перед вводом, чтобы точно убрать старые данные.
            msg = QMessageBox(self)
            msg.setWindowTitle("iiko")
            msg.setText("Перед авторизацией:")
            msg.setInformativeText(
                "Можно сбросить сохранённые данные, чтобы приложение не использовало старые токены/пароли."
            )
            btn_reset = msg.addButton("Сбросить", QMessageBox.DestructiveRole)
            btn_continue = msg.addButton("Продолжить", QMessageBox.AcceptRole)
            msg.addButton("Отмена", QMessageBox.RejectRole)
            msg.exec()

            clicked = msg.clickedButton()
            if clicked is None:
                return
            if clicked == btn_reset:
                self._reset_iiko_auth_settings()
            elif clicked != btn_continue:
                return

            # По умолчанию используем REST (/resto) — он нужен для снятия со стоп-листа.
            # Остальные способы не удаляем, но скрываем.

            methods = [
                "iikoRMS REST (/resto)",
                "iikoCloud API v1 (apiLogin, api-ru.iiko.services)",
                "iikoTransport (user_id/user_secret, iiko.biz:9900)",
            ]

            if use_advanced:
                cur_mode = (self._iiko_mode or "cloud").strip().lower()
                if cur_mode in ("biz", "transport", "iikobiz", "iiko.biz"):
                    default_idx = 2
                elif cur_mode in ("rest", "rms", "resto"):
                    default_idx = 0
                else:
                    default_idx = 1

                chosen, ok = QInputDialog.getItem(
                    self,
                    "iiko",
                    "Способ подключения:",
                    methods,
                    default_idx,
                    False,
                )
                if not ok:
                    return
            else:
                chosen = "iikoRMS REST (/resto)"

            # ===== iikoRMS REST (/resto) =====
            if chosen.startswith("iikoRMS"):
                base_url_default = (self._iiko_base_url or "").strip() or "https://patriot-co.iiko.it/resto"
                base_url = self._prompt_text(
                    "iiko — REST (/resto)",
                    "Base URL (пример: https://patriot-co.iiko.it/resto)\r\n(двойной клик в поле = вставить из буфера обмена):",
                    QLineEdit.Normal,
                    default_text=base_url_default,
                )
                if not base_url:
                    return

                login_default = (self._iiko_login or "").strip()
                login = self._prompt_text(
                    "iiko — REST (/resto)",
                    "Логин iikoOffice (пользователь)\r\n(двойной клик в поле = вставить из буфера обмена):",
                    QLineEdit.Normal,
                    default_text=login_default,
                )
                if not login:
                    return

                # Сохраняем параметры и получаем/сохраняем sha1 пароля
                self._iiko_mode = "rest"
                self._iiko_base_url = base_url
                self._iiko_login = login

                # попросим пароль (сохраняем только sha1)
                if not self._ensure_iiko_pass_sha1():
                    return

                # Проверка авторизации (быстро): получаем auth key
                client = IikoRmsClient(
                    base_url=self._iiko_base_url,
                    login=self._iiko_login,
                    pass_sha1=self._iiko_pass_sha1_cached,
                )
                client.auth_key()

                # сбросим кэши номенклатуры
                try:
                    self._iiko_products_load_seq = int(getattr(self, "_iiko_products_load_seq", 0)) + 1
                except Exception:
                    pass
                self._iiko_products_loaded = False
                self._iiko_products_by_key = {}
                self._pricelist_dishes = []
                self._open_iiko_products = []
                self._open_selected_ids = set()
                self._open_iiko_products_norm = []
                self._open_iiko_products_exact = {}
                self._pricelist_dishes_norm = []
                self._pricelist_dishes_exact = {}

                try:
                    self._settings.setValue("iiko/mode", "rest")
                    self._settings.setValue("iiko/base_url", self._iiko_base_url)
                    self._settings.setValue("iiko/login", self._iiko_login)
                except Exception:
                    pass

                QMessageBox.information(self, "iiko", "REST подключён. Теперь можно загружать блюда.")
                return

            # ===== iikoTransport (iiko.biz:9900) =====
            if chosen.startswith("iikoTransport"):
                api_url_default = (self._iiko_biz_api_url or "").strip() or "https://iiko.biz:9900"
                api_url = self._prompt_text(
                    "iiko — iikoTransport",
                    "API URL (обычно https://iiko.biz:9900)\r\n(двойной клик в поле = вставить из буфера обмена):",
                    QLineEdit.Normal,
                    default_text=api_url_default,
                )
                if not api_url:
                    return

                user_id_default = (self._iiko_biz_user_id or "").strip() or "pos_login_f13591ea"
                user_id = self._prompt_text(
                    "iiko — iikoTransport",
                    "user_id\r\n(двойной клик в поле = вставить из буфера обмена):",
                    QLineEdit.Normal,
                    default_text=user_id_default,
                )
                if not user_id:
                    return

                user_secret = self._prompt_text(
                    "iiko — iikoTransport",
                    "user_secret\r\n(двойной клик в поле = вставить из буфера обмена):",
                    QLineEdit.Password,
                )
                if not user_secret:
                    return

                client = IikoTransportClient(api_url=api_url, user_id=user_id, user_secret=user_secret)
                access_token = client.access_token()

                # показываем токен (можно скопировать)
                self._show_access_token_dialog(access_token)

                orgs = client.organizations()
                org_id = ""
                org_name = ""
                if len(orgs) == 1:
                    org_id = orgs[0].id
                    org_name = orgs[0].name
                else:
                    labels = [f"{i+1}. {o.name} ({o.id})" for i, o in enumerate(orgs)]
                    chosen_org, ok = QInputDialog.getItem(
                        self,
                        "iiko — iikoTransport",
                        "Выберите организацию:",
                        labels,
                        0,
                        False,
                    )
                    if not ok:
                        return
                    idx = labels.index(chosen_org)
                    org_id = orgs[idx].id
                    org_name = orgs[idx].name

                self._iiko_mode = "biz"
                self._iiko_biz_api_url = api_url
                self._iiko_biz_user_id = user_id
                self._iiko_biz_user_secret = user_secret
                self._iiko_biz_access_token = access_token
                self._iiko_biz_org_id = org_id
                self._iiko_biz_org_name = org_name

                # сбросим кэши номенклатуры
                try:
                    self._iiko_products_load_seq = int(getattr(self, "_iiko_products_load_seq", 0)) + 1
                except Exception:
                    pass
                self._iiko_products_loaded = False
                self._iiko_products_by_key = {}
                self._pricelist_dishes = []
                self._open_iiko_products = []
                self._open_selected_ids = set()
                self._open_iiko_products_norm = []
                self._open_iiko_products_exact = {}
                self._pricelist_dishes_norm = []
                self._pricelist_dishes_exact = {}

                try:
                    self._settings.setValue("iiko/mode", "biz")
                    self._settings.setValue("iiko/biz/api_url", api_url)
                    self._settings.setValue("iiko/biz/user_id", user_id)
                    self._settings.setValue("iiko/biz/user_secret", user_secret)
                    self._settings.setValue("iiko/biz/access_token", access_token)
                    self._settings.setValue("iiko/biz/org_id", org_id)
                    if org_name:
                        self._settings.setValue("iiko/biz/org_name", org_name)
                except Exception:
                    pass

                if org_name:
                    QMessageBox.information(self, "iiko", f"Подключено: {org_name} ({org_id})")
                else:
                    QMessageBox.information(self, "iiko", f"Подключено. organization_id: {org_id}")
                return

            # ===== iikoCloud API v1 (api-ru.iiko.services) =====
            api_url = (self._iiko_cloud_api_url or "").strip() or "https://api-ru.iiko.services"

            api_login_default = (self._iiko_cloud_api_login or "").strip()
            api_login = self._prompt_text(
                "iiko — iikoCloud",
                "apiLogin\r\n(двойной клик в поле = вставить из буфера обмена):",
                QLineEdit.Normal,
                default_text=api_login_default,
            )
            if not api_login:
                return

            client = IikoCloudV1Client(api_url=api_url, api_login=api_login)
            access_token = client.access_token()

            # показываем токен (можно скопировать)
            self._show_access_token_dialog(access_token)

            orgs = client.organizations()
            org_id = ""
            org_name = ""
            if len(orgs) == 1:
                org_id = orgs[0].id
                org_name = orgs[0].name
            else:
                labels = [f"{i+1}. {o.name} ({o.id})" for i, o in enumerate(orgs)]
                chosen_org, ok = QInputDialog.getItem(
                    self,
                    "iiko — iikoCloud",
                    "Выберите организацию:",
                    labels,
                    0,
                    False,
                )
                if not ok:
                    return
                idx = labels.index(chosen_org)
                org_id = orgs[idx].id
                org_name = orgs[idx].name

            self._iiko_mode = "cloud"
            self._iiko_cloud_api_url = api_url
            self._iiko_cloud_api_login = api_login
            self._iiko_cloud_access_token = access_token
            self._iiko_cloud_org_id = org_id
            self._iiko_cloud_org_name = org_name

            # сбросим кэши номенклатуры
            try:
                self._iiko_products_load_seq = int(getattr(self, "_iiko_products_load_seq", 0)) + 1
            except Exception:
                pass
            self._iiko_products_loaded = False
            self._iiko_products_by_key = {}
            self._pricelist_dishes = []
            self._open_iiko_products = []
            self._open_selected_ids = set()
            self._open_iiko_products_norm = []
            self._open_iiko_products_exact = {}
            self._pricelist_dishes_norm = []
            self._pricelist_dishes_exact = {}

            try:
                self._settings.setValue("iiko/mode", "cloud")
                self._settings.setValue("iiko/cloud/api_url", api_url)
                self._settings.setValue("iiko/cloud/api_login", api_login)
                self._settings.setValue("iiko/cloud/access_token", access_token)
                self._settings.setValue("iiko/cloud/org_id", org_id)
                if org_name:
                    self._settings.setValue("iiko/cloud/org_name", org_name)
            except Exception:
                pass

            if org_name:
                QMessageBox.information(self, "iiko", f"Подключено: {org_name} ({org_id})")
            else:
                QMessageBox.information(self, "iiko", f"Подключено. organization_id: {org_id}")

        except IikoApiError as e:
            QMessageBox.critical(self, "iiko", str(e))
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def do_open_menu(self):
        """Открывает файл меню (Excel) в приложении по умолчанию (например, Excel)."""
        try:
            path, _ = QFileDialog.getOpenFileName(
                self,
                "Открыть меню",
                str(Path.cwd()),
                "Excel (*.xls *.xlsx *.xlsm);;Все файлы (*.*)",
            )
            if not path:
                return

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(path))
            if not ok:
                QMessageBox.warning(self, "Не удалось открыть", "Не удалось открыть файл выбранной программой.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def _open_document(self, settings_key: str, dialog_title: str, file_filter: str) -> None:
        """Открывает произвольный файл документа, запоминая его путь в настройках."""
        try:
            path = ""
            try:
                if hasattr(self, "_settings"):
                    value = self._settings.value(settings_key, "")
                    path = str(value) if value is not None else ""
            except Exception:
                path = ""

            if not path or not Path(path).exists():
                path, _ = QFileDialog.getOpenFileName(
                    self,
                    dialog_title,
                    str(Path.cwd()),
                    file_filter,
                )
                if not path:
                    return
                try:
                    if hasattr(self, "_settings"):
                        self._settings.setValue(settings_key, path)
                except Exception:
                    pass

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(path))
            if not ok:
                QMessageBox.warning(self, "Не удалось открыть", "Не удалось открыть файл выбранной программой.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def open_vacation_statement(self) -> None:
        """Кнопка "Заявление на отпуск" — просто открывает шаблон в Word, без копирования на рабочий стол."""
        try:
            template_path = find_template("Заявление на отпуск.doc")
            if not template_path:
                QMessageBox.warning(
                    self,
                    "Шаблон",
                    "Файл 'Заявление на отпуск.doc' не найден. Положите его в папку templates.",
                )
                return

            # Сразу открываем шаблон в приложении по умолчанию (обычно Word)
            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(template_path)))
            if not ok:
                QMessageBox.warning(
                    self,
                    "Открытие",
                    f"Не удалось автоматически открыть файл:\n{template_path}",
                )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def open_med_books(self) -> None:
        """Кнопка "Открыть медкнижки" — открывает шаблон Медкнижки.xlsx из templates."""
        try:
            template_path = find_template("Медкнижки.xlsx")
            if not template_path:
                QMessageBox.warning(
                    self,
                    "Шаблон",
                    "Файл 'Медкнижки.xlsx' не найден. Положите его в папку templates.",
                )
                return

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(template_path)))
            if not ok:
                QMessageBox.warning(
                    self,
                    "Открытие",
                    f"Не удалось автоматически открыть файл:\n{template_path}",
                )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def open_birthday_file(self) -> None:
        """Кнопка "Открыть файл День рождения" — открывает шаблон из templates."""
        try:
            template_path = find_template("День рождения.xlsx")
            if not template_path:
                QMessageBox.warning(
                    self,
                    "Шаблон",
                    "Файл 'День рождения.xlsx' не найден. Положите его в папку templates.",
                )
                return

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(template_path)))
            if not ok:
                QMessageBox.warning(
                    self,
                    "Открытие",
                    f"Не удалось автоматически открыть файл:\n{template_path}",
                )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def open_hygiene_journal(self) -> None:
        """Открывает шаблон гигиенического журнала из папки templates."""
        try:
            template_path = find_template("Гигиенический журнал.xlsx")
            if not template_path:
                QMessageBox.warning(
                    self,
                    "Шаблон",
                    "Файл 'Гигиенический журнал.xlsx' не найден. Положите его в папку templates.",
                )
                return

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(template_path)))
            if not ok:
                QMessageBox.warning(
                    self,
                    "Открытие",
                    f"Не удалось автоматически открыть файл:\n{template_path}",
                )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def open_direction_document(self) -> None:
        """Открывает шаблон направления из папки templates (Направление.doc)."""
        try:
            template_path = find_template("Направление.doc")
            if not template_path:
                QMessageBox.warning(
                    self,
                    "Шаблон",
                    "Файл 'Направление.doc' не найден. Положите его в папку templates.",
                )
                return

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(template_path)))
            if not ok:
                QMessageBox.warning(
                    self,
                    "Открытие",
                    f"Не удалось автоматически открыть файл:\n{template_path}",
                )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def open_locker_document(self) -> None:
        """Открывает шаблон документа "Раздевалка" из папки templates (Раздевалка.xlsx)."""
        try:
            template_path = find_template("Раздевалка.xlsx")
            if not template_path:
                QMessageBox.warning(
                    self,
                    "Шаблон",
                    "Файл 'Раздевалка.xlsx' не найден. Положите его в папку templates.",
                )
                return

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(template_path)))
            if not ok:
                QMessageBox.warning(
                    self,
                    "Открытие",
                    f"Не удалось автоматически открыть файл:\n{template_path}",
                )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def open_buffet_sheet(self) -> None:
        """Открывает шаблон "Бланк для раздачи кофетерий" из templates."""
        try:
            template_path = find_template("Бланк для раздачи кофетерий.xlsx")
            if not template_path:
                QMessageBox.warning(
                    self,
                    "Шаблон",
                    "Файл 'Бланк для раздачи кофетерий.xlsx' не найден. Положите его в папку templates.",
                )
                return

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(template_path)))
            if not ok:
                QMessageBox.warning(
                    self,
                    "Открытие",
                    f"Не удалось автоматически открыть файл:\n{template_path}",
                )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def open_baker_sheet(self) -> None:
        """Открывает шаблон "Акт_приготовления_ВЫПЕЧКАДЕСЕРТЫ" из templates."""
        try:
            template_path = find_template("Акт_приготовления_ВЫПЕЧКАДЕСЕРТЫ.xlsx")
            if not template_path:
                QMessageBox.warning(
                    self,
                    "Шаблон",
                    "Файл 'Акт_приготовления_ВЫПЕЧКАДЕСЕРТЫ.xlsx' не найден. Положите его в папку templates.",
                )
                return

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(template_path)))
            if not ok:
                QMessageBox.warning(
                    self,
                    "Открытие",
                    f"Не удалось автоматически открыть файл:\n{template_path}",
                )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def open_breakfasts_folder(self) -> None:
        """Открывает папку templates/Завтраки."""
        try:
            folder_path = find_template_path("Завтраки")
            if (not folder_path) or (not Path(folder_path).exists()) or (not Path(folder_path).is_dir()):
                QMessageBox.warning(
                    self,
                    "Шаблон",
                    "Папка 'Завтраки' не найдена. Положите папку 'Завтраки' в templates.",
                )
                return

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(folder_path)))
            if not ok:
                QMessageBox.warning(
                    self,
                    "Открытие",
                    f"Не удалось открыть папку:\n{folder_path}",
                )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def open_distribution_sheet(self) -> None:
        """Открывает шаблон "Раздача.xlsx" из templates."""
        try:
            template_path = find_template("Раздача.xlsx")
            if not template_path:
                QMessageBox.warning(
                    self,
                    "Шаблон",
                    "Файл 'Раздача.xlsx' не найден. Положите его в папку templates.",
                )
                return

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(template_path)))
            if not ok:
                QMessageBox.warning(
                    self,
                    "Открытие",
                    f"Не удалось автоматически открыть файл:\n{template_path}",
                )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def open_pie_names(self) -> None:
        """Открывает шаблон "Название пирогов.docx" из templates."""
        try:
            template_path = find_template("Название пирогов.docx")
            if not template_path:
                QMessageBox.warning(
                    self,
                    "Шаблон",
                    "Файл 'Название пирогов.docx' не найден. Положите его в папку templates.",
                )
                return

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(template_path)))
            if not ok:
                QMessageBox.warning(
                    self,
                    "Открытие",
                    f"Не удалось автоматически открыть файл:\n{template_path}",
                )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def open_bakery_pricelist(self) -> None:
        """Открывает шаблон "Прейскурант выпечка.xlsx" из templates."""
        try:
            template_path = find_template("Прейскурант выпечка.xlsx")
            if not template_path:
                QMessageBox.warning(
                    self,
                    "Шаблон",
                    "Файл 'Прейскурант выпечка.xlsx' не найден. Положите его в папку templates.",
                )
                return

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(template_path)))
            if not ok:
                QMessageBox.warning(
                    self,
                    "Открытие",
                    f"Не удалось автоматически открыть файл:\n{template_path}",
                )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def open_cash_template(self) -> None:
        """Открывает шаблон "Наличка.xlsx" из templates, подставляя текущий месяц и год в A1.

        Не создаём копию: меняем сам шаблон и открываем его.
        Если шаблон недоступен для записи — открываем временную копию.
        """
        try:
            template_path = find_template("Наличка.xlsx")
            if not template_path:
                QMessageBox.warning(
                    self,
                    "Шаблон",
                    "Файл 'Наличка.xlsx' не найден. Положите его в папку templates.",
                )
                return

            now = datetime.now()

            month_names = {
                1: "Январь",
                2: "Февраль",
                3: "Март",
                4: "Апрель",
                5: "Май",
                6: "Июнь",
                7: "Июль",
                8: "Август",
                9: "Сентябрь",
                10: "Октябрь",
                11: "Ноябрь",
                12: "Декабрь",
            }
            month_label = month_names.get(int(now.month), str(now.month))
            a1_value = f"{month_label} {now.year}"

            out_path = Path(template_path)

            try:
                from openpyxl import load_workbook
            except Exception:
                QMessageBox.warning(self, "Ошибка", "Не найден модуль openpyxl. Невозможно изменить файл 'Наличка.xlsx'.")
                return

            wb = load_workbook(template_path)
            ws = wb.active
            try:
                ws["A1"].value = a1_value
            except Exception:
                # если лист защищён/ошибка записи
                pass

            # Подставим даты на текущий месяц/год и спрячем лишние дни.
            try:
                days_in_month = int(calendar.monthrange(int(now.year), int(now.month))[1])
                first_day_row = 3  # в шаблоне даты начинаются с A3
                max_days_rows = 31

                for day in range(1, max_days_rows + 1):
                    r = first_day_row + day - 1
                    if day <= days_in_month:
                        ws.cell(row=r, column=1).value = datetime(int(now.year), int(now.month), int(day), 0, 0, 0)
                        try:
                            ws.row_dimensions[r].hidden = False
                        except Exception:
                            pass
                    else:
                        ws.cell(row=r, column=1).value = None
                        try:
                            ws.row_dimensions[r].hidden = True
                        except Exception:
                            pass
            except Exception:
                pass

            try:
                wb.save(str(out_path))
            except Exception as e:
                # Частый случай: шаблон лежит в месте без прав на запись или уже открыт в Excel.
                # Тогда сохраняем во временную копию и открываем её.
                try:
                    tmp_dir = Path(os.getenv("TEMP") or os.getenv("TMP") or str(Path.home()))
                    tmp_name = f"Наличка_{now.month:02d}.{now.year}_временная.xlsx"
                    tmp_path = tmp_dir / tmp_name
                    wb.save(str(tmp_path))
                    out_path = tmp_path
                except Exception:
                    QMessageBox.warning(self, "Ошибка", f"Не удалось сохранить файл 'Наличка': {e}")
                    return

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(out_path)))
            if not ok:
                QMessageBox.warning(
                    self,
                    "Открытие",
                    f"Не удалось автоматически открыть файл:\n{out_path}",
                )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def open_fryer_oil_journal(self) -> None:
        """Открывает шаблон "Журнал замены фритюрного масла" из templates."""
        try:
            template_path = find_template("Журнал замены фритюрного масла.doc")
            if not template_path:
                QMessageBox.warning(
                    self,
                    "Шаблон",
                    "Файл 'Журнал замены фритюрного масла.doc' не найден. Положите его в папку templates.",
                )
                return

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(template_path)))
            if not ok:
                QMessageBox.warning(
                    self,
                    "Открытие",
                    f"Не удалось автоматически открыть файл:\n{template_path}",
                )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def open_fridge_temperature(self) -> None:
        """Открывает шаблон "Температурный режим холодильного оборудования" из templates."""
        try:
            template_path = find_template("Температурный_режим_холодильньного_оборудования.docx")
            if not template_path:
                QMessageBox.warning(
                    self,
                    "Шаблон",
                    "Файл 'Температурный_режим_холодильньного_оборудования.docx' не найден. Положите его в папку templates.",
                )
                return

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(template_path)))
            if not ok:
                QMessageBox.warning(
                    self,
                    "Открытие",
                    f"Не удалось автоматически открыть файл:\n{template_path}",
                )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def open_freezer_temperature(self) -> None:
        """Открывает шаблон "Температурный режим морозильного оборудования" из templates."""
        try:
            template_path = find_template("Температурный_режим_морозильного_оборудования.docx")
            if not template_path:
                QMessageBox.warning(
                    self,
                    "Шаблон",
                    "Файл 'Температурный_режим_морозильного_оборудования.docx' не найден. Положите его в папку templates.",
                )
                return

            ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(template_path)))
            if not ok:
                QMessageBox.warning(
                    self,
                    "Открытие",
                    f"Не удалось автоматически открыть файл:\n{template_path}",
                )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def _set_open_suggestions_visible(self, visible: bool) -> None:
        """Показать/скрыть выпадающие подсказки (popup)."""
        try:
            if not hasattr(self, "lstTomorrowDishes"):
                return

            if not visible:
                self.lstTomorrowDishes.hide()
                return

            # позиционируем под полем поиска
            le = getattr(self, "edTomorrowSearch", None)
            if le is None:
                return

            # ширина = как у поля поиска
            try:
                self.lstTomorrowDishes.setFixedWidth(le.width())
            except Exception:
                pass

            # высота: до 8 строк или максимум 320px
            try:
                cnt = max(1, int(self.lstTomorrowDishes.count()))
                rows = min(cnt, 8)
                row_h = self.lstTomorrowDishes.sizeHintForRow(0)
                if not row_h:
                    row_h = max(24, int(self.lstTomorrowDishes.fontMetrics().height()) + 8)
                h = min(320, rows * row_h + 2 * int(self.lstTomorrowDishes.frameWidth()))
                self.lstTomorrowDishes.setFixedHeight(h)
            except Exception:
                pass

            try:
                gp = le.mapToGlobal(QPoint(0, le.height()))
            except Exception:
                gp = None

            if gp is not None:
                # если внизу не помещается — покажем сверху
                try:
                    screen = QGuiApplication.screenAt(gp) or QGuiApplication.primaryScreen()
                    if screen is not None:
                        geo = screen.availableGeometry()
                        if gp.y() + self.lstTomorrowDishes.height() > geo.bottom():
                            gp = le.mapToGlobal(QPoint(0, -self.lstTomorrowDishes.height()))
                except Exception:
                    pass

            try:
                self.lstTomorrowDishes.move(gp)
            except Exception:
                pass

            self.lstTomorrowDishes.show()
            self.lstTomorrowDishes.raise_()

            # удерживаем фокус в поле ввода, чтобы можно было продолжать печатать
            try:
                anchor.setFocus()
            except Exception:
                pass

        except Exception:
            pass

    def _hide_open_suggestions(self) -> None:
        self._set_open_suggestions_visible(False)

    def _set_tomorrow_info(self, text: str) -> None:
        """Показывает/скрывает строку статуса в разделе "Открыть блюда"."""
        try:
            if not hasattr(self, "lblTomorrowInfo"):
                return
            t = (text or "").strip()
            self.lblTomorrowInfo.setText(t)
            self.lblTomorrowInfo.setVisible(bool(t))
        except Exception:
            pass

    def _on_open_target_date_changed(self) -> None:
        """Сохраняет выбранную дату открытия блюд."""
        try:
            if not hasattr(self, "calOpenDate"):
                return
            qd = self.calOpenDate.selectedDate()
            if not qd.isValid():
                return
            try:
                self._settings.setValue("iiko/open_dishes/target_date", qd.toString("yyyy-MM-dd"))
            except Exception:
                pass
        except Exception:
            pass

        try:
            self._update_open_action_button_label()
        except Exception:
            pass

    def _update_open_action_button_label(self) -> None:
        """Меняет текст кнопки внизу в зависимости от выбранной даты."""
        try:
            if not hasattr(self, "btnOpenTomorrowChecked"):
                return

            target = date.today()
            try:
                if hasattr(self, "calOpenDate"):
                    qd = self.calOpenDate.selectedDate()
                    if qd and qd.isValid():
                        target = date(int(qd.year()), int(qd.month()), int(qd.day()))
            except Exception:
                target = date.today()

            today = date.today()
            if target > today:
                self.btnOpenTomorrowChecked.setText(f"Запланировать на {target.strftime('%d.%m.%Y')}")
            else:
                self.btnOpenTomorrowChecked.setText("Открыть отмеченные")
        except Exception:
            pass

    def _load_open_schedule_job(self) -> None:
        """Читает запланированное открытие блюд из настроек."""
        self._open_schedule_job = None
        try:
            raw = self._settings.value("iiko/open_dishes/schedule", "")
            raw_s = str(raw) if raw is not None else ""
            if not raw_s:
                return
            job = json.loads(raw_s)
            if isinstance(job, dict):
                self._open_schedule_job = job
        except Exception:
            self._open_schedule_job = None

    def _save_open_schedule_job(self, job: Optional[dict]) -> None:
        """Сохраняет/очищает запланированное открытие блюд."""
        self._open_schedule_job = job
        try:
            if job is None:
                self._settings.remove("iiko/open_dishes/schedule")
                return
            self._settings.setValue("iiko/open_dishes/schedule", json.dumps(job, ensure_ascii=False))
        except Exception:
            pass

    def _format_open_schedule_status(self) -> str:
        job = self._open_schedule_job
        if not isinstance(job, dict):
            return ""

        state = (job.get("state") or "pending").strip().lower()
        run_at = str(job.get("run_at") or "")
        try:
            dt = datetime.fromisoformat(run_at)
            when = dt.strftime("%d.%m.%Y %H:%M")
        except Exception:
            when = run_at

        if state == "failed":
            err = (job.get("last_error") or "")
            err_short = (str(err).strip()[:120])
            return f"Запланированное открытие НЕ выполнено ({when}). {err_short}"

        if state == "done":
            return ""

        # pending
        return f"Запланировано открытие: {when}"

    def _check_open_schedule_due(self, silent: bool = False) -> None:
        """Если наступила дата/время — выполняет запланированное открытие блюд."""
        # Лок защищает от двойного запуска (если открыт GUI + параллельно сработал schtasks)
        lock = QLockFile(_open_schedule_lock_path())
        try:
            lock.setStaleLockTime(10 * 60 * 1000)
        except Exception:
            pass

        if not lock.tryLock(0):
            return

        try:
            # обновим из настроек (если запускали в другом месте)
            if self._open_schedule_job is None:
                self._load_open_schedule_job()

            job = self._open_schedule_job
            if not isinstance(job, dict):
                _windows_delete_open_schedule_task()
                return

            state = (job.get("state") or "pending").strip().lower()
            if state in ("done", "failed"):
                _windows_delete_open_schedule_task()
                return

            run_at = str(job.get("run_at") or "")
            try:
                dt_run = datetime.fromisoformat(run_at)
            except Exception:
                return

            now = datetime.now()
            if now < dt_run:
                # можем показать статус, если открыта вкладка
                try:
                    if hasattr(self, "grpTomorrowOpen") and self.grpTomorrowOpen.isVisible():
                        self._set_tomorrow_info(self._format_open_schedule_status())
                except Exception:
                    pass
                return

            product_ids = job.get("product_ids") or []
            if not isinstance(product_ids, list) or not product_ids:
                # нечего выполнять
                self._save_open_schedule_job(None)
                _windows_delete_open_schedule_task()
                return

            # выполняем один раз
            ok = 0
            for pid in product_ids:
                pid_s = str(pid).strip()
                if not pid_s:
                    continue
                self._open_stoplist_product_id(pid_s)
                ok += 1

            # успех -> очищаем расписание
            self._save_open_schedule_job(None)
            _windows_delete_open_schedule_task()

            if not silent:
                try:
                    QMessageBox.information(self, "iiko", f"Запланированное открытие выполнено. Открыто блюд: {ok}.")
                except Exception:
                    pass

        except IikoApiError as e:
            # пометим как failed, чтобы не спамить попытками
            job = self._open_schedule_job or {}
            if isinstance(job, dict):
                job["state"] = "failed"
                job["last_error"] = str(e)
                try:
                    job["failed_at"] = datetime.now().isoformat(timespec="seconds")
                except Exception:
                    pass
                self._save_open_schedule_job(job)
            _windows_delete_open_schedule_task()
            if not silent:
                try:
                    QMessageBox.critical(self, "iiko", f"Не удалось выполнить запланированное открытие: {e}")
                except Exception:
                    pass
        except Exception:
            pass
        finally:
            try:
                lock.unlock()
            except Exception:
                pass

    def _set_pricelist_suggestions_visible(self, visible: bool) -> None:
        """Показать/скрыть выпадающий список блюд для "Ценники" (popup)."""
        try:
            if not hasattr(self, "lstDishSuggestions"):
                return

            if not visible:
                self.lstDishSuggestions.hide()
                return

            anchor = getattr(self, "edDishSearch", None)
            if anchor is None:
                return

            # ширина: как у списка выбранных (или группы)
            w = 0
            try:
                if hasattr(self, "lstSelectedDishes") and self.lstSelectedDishes.isVisible():
                    w = int(self.lstSelectedDishes.width())
            except Exception:
                w = 0
            if not w:
                try:
                    if hasattr(self, "grpPricelist") and self.grpPricelist.isVisible():
                        w = int(self.grpPricelist.width())
                except Exception:
                    w = 0
            if not w:
                w = 520
            try:
                self.lstDishSuggestions.setFixedWidth(max(320, w))
            except Exception:
                pass

            # высота: до 10 строк или максимум 360px
            try:
                cnt = max(1, int(self.lstDishSuggestions.count()))
                rows = min(cnt, 10)
                row_h = self.lstDishSuggestions.sizeHintForRow(0)
                if not row_h:
                    row_h = max(24, int(self.lstDishSuggestions.fontMetrics().height()) + 8)
                h = min(360, rows * row_h + 2 * int(self.lstDishSuggestions.frameWidth()))
                self.lstDishSuggestions.setFixedHeight(h)
            except Exception:
                pass

            try:
                gp = anchor.mapToGlobal(QPoint(0, anchor.height()))
            except Exception:
                gp = None

            if gp is not None:
                # если внизу не помещается — покажем сверху
                try:
                    screen = QGuiApplication.screenAt(gp) or QGuiApplication.primaryScreen()
                    if screen is not None:
                        geo = screen.availableGeometry()

                        x = gp.x()
                        if x + self.lstDishSuggestions.width() > geo.right():
                            x = max(geo.left(), geo.right() - self.lstDishSuggestions.width())

                        y = gp.y()
                        if y + self.lstDishSuggestions.height() > geo.bottom():
                            y = anchor.mapToGlobal(QPoint(0, -self.lstDishSuggestions.height())).y()

                        gp = QPoint(x, y)
                except Exception:
                    pass

                try:
                    self.lstDishSuggestions.move(gp)
                except Exception:
                    pass

            self.lstDishSuggestions.show()
            self.lstDishSuggestions.raise_()

            # удерживаем фокус в поле ввода, чтобы можно было продолжать печатать
            try:
                anchor.setFocus()
            except Exception:
                pass

        except Exception:
            pass

    def _hide_pricelist_suggestions(self) -> None:
        try:
            self._set_pricelist_suggestions_visible(False)
        except Exception:
            pass

    def eventFilter(self, obj, event):
        # Скрываем выпадающие списки (popup) при клике вне них.
        try:
            if event.type() == QEvent.MouseButtonPress:
                # глобальная позиция клика
                try:
                    gp = event.globalPosition().toPoint()  # Qt6
                except Exception:
                    gp = event.globalPos()  # fallback

                def _inside(w) -> bool:
                    try:
                        if (not w) or (not w.isVisible()):
                            return False
                        return w.rect().contains(w.mapFromGlobal(gp))
                    except Exception:
                        return False

                # ===== Открыть блюда =====
                if getattr(self, "grpTomorrowOpen", None) and self.grpTomorrowOpen.isVisible():
                    if getattr(self, "lstTomorrowDishes", None):
                        inside_search = _inside(getattr(self, "edTomorrowSearch", None))
                        inside_suggestions = _inside(getattr(self, "lstTomorrowDishes", None))

                        if inside_search:
                            # если подсказки скрыты, но текст уже введён — покажем снова
                            try:
                                if (not self.lstTomorrowDishes.isVisible()) and len((self.edTomorrowSearch.text() or "").strip()) >= 2:
                                    self._update_open_suggestions(self.edTomorrowSearch.text())
                            except Exception:
                                pass
                        else:
                            # клик вне поиска -> закрываем подсказки (если они показаны)
                            try:
                                if self.lstTomorrowDishes.isVisible() and (not inside_suggestions):
                                    self._hide_open_suggestions()
                            except Exception:
                                pass

                # ===== Ценники =====
                if getattr(self, "grpPricelist", None) and self.grpPricelist.isVisible():
                    if getattr(self, "lstDishSuggestions", None):
                        inside_search = _inside(getattr(self, "edDishSearch", None))
                        inside_suggestions = _inside(getattr(self, "lstDishSuggestions", None))

                        if inside_search:
                            # если список скрыт, но текст уже введён — покажем снова
                            try:
                                if (not self.lstDishSuggestions.isVisible()) and len((self.edDishSearch.text() or "").strip()) >= 2:
                                    self._update_pricelist_suggestions(self.edDishSearch.text())
                            except Exception:
                                pass
                        else:
                            # клик вне поиска -> закрываем подсказки (если они показаны)
                            try:
                                if self.lstDishSuggestions.isVisible() and (not inside_suggestions):
                                    self._hide_pricelist_suggestions()
                            except Exception:
                                pass

        except Exception:
            pass

        return super().eventFilter(obj, event)

    def do_open_tomorrow_dishes(self):
        """Открыть блюда: поиск в iiko -> выбрать -> снять со стоп-листа (REST /resto)."""
        try:
            # Скрываем другие панели
            if hasattr(self, "grpFirst"):
                self.grpFirst.setVisible(False)
            if hasattr(self, "grpSecond"):
                self.grpSecond.setVisible(False)
            if hasattr(self, "paramsBox"):
                self.paramsBox.setVisible(False)
            if hasattr(self, "actionsPanel"):
                self.actionsPanel.setVisible(False)
            if hasattr(self, "presentationActionsPanel"):
                self.presentationActionsPanel.setVisible(False)
            if hasattr(self, "brokerageActionsPanel"):
                self.brokerageActionsPanel.setVisible(False)
            if hasattr(self, "templateActionsPanel"):
                self.templateActionsPanel.setVisible(False)
            if hasattr(self, "grpPricelist"):
                self.grpPricelist.setVisible(False)
            if hasattr(self, "pricelistActionsPanel"):
                self.pricelistActionsPanel.setVisible(False)
            if hasattr(self, "grpDocuments"):
                self.grpDocuments.setVisible(False)

            # Для "Открыть блюда" Excel не нужен
            if hasattr(self, "grpExcelFile"):
                self.grpExcelFile.setVisible(False)

            if hasattr(self, "grpTomorrowOpen"):
                self.grpTomorrowOpen.setVisible(True)
            if hasattr(self, "tomorrowOpenActionsPanel"):
                self.tomorrowOpenActionsPanel.setVisible(True)

            # Сброс состояния поиска/подсказок (выбор блюд сохраняем)
            try:
                self._rebuild_open_selected_ids_from_list()
                self._sort_open_selected_list()
            except Exception:
                pass
            try:
                self.edTomorrowSearch.clear()
            except Exception:
                pass
            try:
                self.lstTomorrowDishes.clear()
            except Exception:
                pass
            self._hide_open_suggestions()

            # Покажем статус расписания (если есть)
            self._set_tomorrow_info(self._format_open_schedule_status())
            try:
                self._update_open_action_button_label()
            except Exception:
                pass

            try:
                self.edTomorrowSearch.setFocus()
            except Exception:
                pass

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def do_download_pricelists(self):
        """Показывает раздел для формирования ценников по выбранным блюдам."""
        try:
            # Скрываем другие панели
            if hasattr(self, "grpFirst"):
                self.grpFirst.setVisible(False)
            if hasattr(self, "grpSecond"):
                self.grpSecond.setVisible(False)
            if hasattr(self, "paramsBox"):
                self.paramsBox.setVisible(False)
            if hasattr(self, "actionsPanel"):
                self.actionsPanel.setVisible(False)
            if hasattr(self, "presentationActionsPanel"):
                self.presentationActionsPanel.setVisible(False)
            if hasattr(self, "brokerageActionsPanel"):
                self.brokerageActionsPanel.setVisible(False)
            if hasattr(self, "templateActionsPanel"):
                self.templateActionsPanel.setVisible(False)
            if hasattr(self, "grpDocuments"):
                self.grpDocuments.setVisible(False)

            # Для ценников НЕ показываем общий блок выбора Excel-файла (он мешает и не нужен при iiko)
            if hasattr(self, "grpExcelFile"):
                self.grpExcelFile.setVisible(False)

            # Скрываем "Открыть блюда" если было открыто
            if hasattr(self, "grpTomorrowOpen"):
                self.grpTomorrowOpen.setVisible(False)
            if hasattr(self, "tomorrowOpenActionsPanel"):
                self.tomorrowOpenActionsPanel.setVisible(False)
            try:
                self._hide_open_suggestions()
            except Exception:
                pass

            # Показываем панель ценников
            if hasattr(self, "grpPricelist"):
                self.grpPricelist.setVisible(True)
            if hasattr(self, "pricelistActionsPanel"):
                self.pricelistActionsPanel.setVisible(True)

            try:
                self._hide_pricelist_suggestions()
            except Exception:
                pass

            # Сброс состояния поиска/подсказок (выбор блюд сохраняем)
            try:
                self._rebuild_pricelist_selected_keys_from_list()
                self._sort_pricelist_selected_list()
            except Exception:
                pass
            try:
                self.edDishSearch.clear()
            except Exception:
                pass
            self.lstDishSuggestions.clear()

            try:
                self.edDishSearch.setFocus()
            except Exception:
                pass

            # Автозагрузка блюд из iiko при входе в раздел (если уже есть авторизация)
            if not self._pricelist_dishes:
                self.lblPricelistInfo.setText("Загружаю блюда из iiko…")
                try:
                    if self._can_autoload_iiko_products():
                        self._load_pricelist_dishes()
                    else:
                        self.lblPricelistInfo.setText(
                            "Введите название блюда (если список пустой — нажмите «Авторизация точки»)"
                        )
                except Exception:
                    self.lblPricelistInfo.setText(
                        "Введите название блюда (если список пустой — нажмите «Авторизация точки»)"
                    )
            else:
                self.lblPricelistInfo.setText(f"Загружено из iiko: {len(self._pricelist_dishes)}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def do_create_presentation_with_data(self):
        """Создает презентацию с данными из Excel файла"""
        try:
            # Получаем путь к Excel файлу
            excel_path = self.edExcelPath.text().strip()
            if not excel_path:
                QMessageBox.warning(self, "Внимание", "Выберите Excel файл с меню.")
                return
            
            # Проверяем существование Excel файла
            if not Path(excel_path).exists():
                QMessageBox.warning(self, "Ошибка", "Указанный Excel файл не найден.")
                return
                
            # Находим шаблон презентации
            template_path = find_template("presentation_template.pptx")
            if not template_path:
                QMessageBox.warning(self, "Шаблон", "Шаблон презентации не найден. Положите файл presentation_template.pptx в папку templates.")
                return
            
            # Выбираем место сохранения презентации
            suggested_name = "menu — копия.pptx"
            desktop = Path.home() / "Desktop"
            suggested_path = str(desktop / suggested_name)
            
            save_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Сохранить презентацию с меню", 
                suggested_path, 
                "PowerPoint (*.pptx);;PowerPoint (*.ppt);;Все файлы (*.*)"
            )
            
            if not save_path:
                return  # Пользователь отменил сохранение
                
            # Создаем презентацию с данными
            success, message = create_presentation_with_excel_data(
                template_path, 
                excel_path, 
                save_path
            )
            
            if success:
                # Убрано информационное окно при успешном сохранении
                pass
            else:
                QMessageBox.warning(self, "Ошибка", f"Не удалось создать презентацию:\n{message}")
                
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def do_brokerage_journal(self):
        """Показывает панель для работы с бракеражным журналом"""
        try:
            # Скрываем другие панели
            if hasattr(self, "grpFirst"):
                self.grpFirst.setVisible(False)
            if hasattr(self, "grpSecond"):
                self.grpSecond.setVisible(False)
            if hasattr(self, "paramsBox"):
                self.paramsBox.setVisible(False)
            if hasattr(self, "actionsPanel"):
                self.actionsPanel.setVisible(False)
            if hasattr(self, "presentationActionsPanel"):
                self.presentationActionsPanel.setVisible(False)
            if hasattr(self, "templateActionsPanel"):
                self.templateActionsPanel.setVisible(False)
            if hasattr(self, "grpPricelist"):
                self.grpPricelist.setVisible(False)
            if hasattr(self, "pricelistActionsPanel"):
                self.pricelistActionsPanel.setVisible(False)
            if hasattr(self, "grpTomorrowOpen"):
                self.grpTomorrowOpen.setVisible(False)
            if hasattr(self, "tomorrowOpenActionsPanel"):
                self.tomorrowOpenActionsPanel.setVisible(False)
            if hasattr(self, "grpDocuments"):
                self.grpDocuments.setVisible(False)
            
            # Показываем панель для работы с бракеражным журналом и её панель действий
            if hasattr(self, "grpExcelFile"):
                self.grpExcelFile.setVisible(True)
                self.grpExcelFile.setTitle("Файл меню для бракеражного журнала")
                self.edExcelPath.setPlaceholderText("Выберите Excel файл с меню для бракеражного журнала...")
            if hasattr(self, "brokerageActionsPanel"):
                self.brokerageActionsPanel.setVisible(True)
                
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def do_create_brokerage_journal_with_data(self):
        """Создает бракеражный журнал с данными из Excel файла"""
        try:
            # Получаем путь к Excel файлу с меню
            excel_path = self.edExcelPath.text().strip()
            if not excel_path:
                QMessageBox.warning(self, "Внимание", "Выберите Excel файл с меню.")
                return
            
            # Проверяем существование Excel файла
            if not Path(excel_path).exists():
                QMessageBox.warning(self, "Ошибка", "Указанный Excel файл не найден.")
                return
                
            # Находим шаблон бракеражного журнала
            template_path = find_template("Бракеражный журнал шаблон.xlsx")
            if not template_path:
                QMessageBox.warning(self, "Шаблон", "Шаблон бракеражного журнала не найден. Положите файл 'Бракеражный журнал шаблон.xlsx' в папку templates.")
                return
            
            # Получаем дату для названия файла
            from app.reports.brokerage_journal import BrokerageJournalGenerator
            from datetime import date
            
            generator = BrokerageJournalGenerator()
            menu_date = generator.extract_date_from_menu(excel_path)
            
            if menu_date:
                date_str = menu_date.strftime("%d.%m.%Y")
                suggested_name = f"бракеражный_журнал_{date_str}.xlsx"
            else:
                today_str = date.today().strftime("%d.%m.%Y")
                suggested_name = f"бракеражный_журнал_{today_str}.xlsx"
            
            # Выбираем место сохранения бракеражного журнала
            desktop = Path.home() / "Desktop"
            suggested_path = str(desktop / suggested_name)
            
            save_path, _ = QFileDialog.getSaveFileName(
                self,
                "Сохранить бракеражный журнал",
                suggested_path,
                "Excel (*.xlsx);;Excel (*.xls);;Все файлы (*.*)"
            )
            
            if not save_path:
                return  # Пользователь отменил сохранение
                
            # Создаем бракеражный журнал с данными
            success, message = create_brokerage_journal_from_menu(
                excel_path, 
                template_path, 
                save_path
            )
            
            if success:
                # Убрано информационное окно при успешном сохранении
                pass
            else:
                QMessageBox.warning(self, "Ошибка", f"Не удалось создать бракеражный журнал:\n{message}")
                
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def do_fill_template_with_data(self):
        """Заполняет шаблон меню сегментным алгоритмом (A-F)."""
        try:
            excel_path = self.edExcelPath.text().strip()
            if not excel_path:
                QMessageBox.warning(self, "Внимание", "Выберите Excel файл с меню.")
                return
            if not Path(excel_path).exists():
                QMessageBox.warning(self, "Ошибка", "Указанный Excel файл не найден.")
                return

            template_path = default_template_path()
            if not template_path or not Path(template_path).exists():
                QMessageBox.warning(self, "Шаблон", "Шаблон меню не найден. Положите файл шаблона в папку templates/.")
                return

            desktop = Path.home() / "Desktop"

            # Имя файла: "<дата месяц> - <день недели>.xlsx" (берём из исходного файла: B3 и B2).
            suggested_name = "menu_ready.xlsx"

            def _sanitize_filename_part(s) -> str:
                text = str(s or "")
                for ch in '<>:"/\\|?*':
                    text = text.replace(ch, " ")
                return " ".join(text.split()).strip()

            try:
                from openpyxl import load_workbook

                wb_in = load_workbook(excel_path, data_only=True)
                if "Касса" in wb_in.sheetnames:
                    ws_in = wb_in["Касса"]
                else:
                    ws_in = wb_in.active

                weekday = _sanitize_filename_part(ws_in["B2"].value)
                day_month = _sanitize_filename_part(ws_in["B3"].value)

                base = ""
                if day_month and weekday:
                    base = f"{day_month} - {weekday}"
                elif day_month:
                    base = day_month
                else:
                    base = _sanitize_filename_part(Path(excel_path).stem)

                if base:
                    # Всегда предлагаем имя без суффикса "- готово".
                    # Если файл уже существует, диалог сохранения сам спросит подтверждение перезаписи.
                    suggested_name = f"{base}.xlsx"
            except Exception:
                pass

            suggested_path = str(desktop / suggested_name)
            save_path, _ = QFileDialog.getSaveFileName(
                self,
                "Сохранить динамически заполненный шаблон",
                suggested_path,
                "Excel (*.xlsx);;Все файлы (*.*)",
            )
            if not save_path:
                return

            fill_dynamic_menu(
                Path(excel_path),
                Path(template_path),
                Path(save_path),
                source_sheet="Касса",
                template_sheet="Касса",
            )
            QMessageBox.information(self, "Готово", "Шаблон успешно заполнен по сегментам.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    # ===== iiko: общие хелперы =====
    def _ensure_iiko_pass_sha1(self) -> bool:
        """Если sha1(pass) не сохранён — спрашиваем пароль один раз и сохраняем только sha1."""
        if self._iiko_pass_sha1_cached:
            return True

        pwd = self._prompt_text(
            "iiko",
            "Введите пароль iiko (сохранится только SHA1-хэш на этом компьютере)\n(двойной клик в поле = вставить из буфера обмена):",
            QLineEdit.Password,
        )
        if not pwd:
            return False

        self._iiko_pass_sha1_cached = hashlib.sha1(pwd.encode("utf-8")).hexdigest()
        try:
            self._settings.setValue("iiko/pass_sha1", self._iiko_pass_sha1_cached)
            # На всякий случай удалим возможный legacy plaintext
            self._settings.remove("iiko/password")
        except Exception:
            pass

        return True

    def _can_autoload_iiko_products(self) -> bool:
        """Есть ли сохранённые данные, чтобы на вкладках не просить ввод заново."""
        mode = (self._iiko_mode or "cloud").strip().lower()

        # iikoCloud API v1
        if mode in ("cloud", "cloud_v1", "cloudv1", "v1"):
            return bool(
                (self._iiko_cloud_api_login or "").strip()
                and (self._iiko_cloud_org_id or "").strip()
                and (self._iiko_cloud_access_token or "").strip()
            )

        # iikoTransport (iiko.biz:9900)
        if mode in ("biz", "transport", "iikobiz", "iiko.biz"):
            org_ok = bool((self._iiko_biz_org_id or "").strip())
            token_ok = bool((self._iiko_biz_access_token or "").strip())
            creds_ok = bool(
                (self._iiko_biz_user_id or "").strip()
                and (self._iiko_biz_user_secret or "").strip()
            )
            return org_ok and (token_ok or creds_ok)

        # REST (/resto)
        return bool(
            (self._iiko_base_url or "").strip()
            and (self._iiko_login or "").strip()
            and (self._iiko_pass_sha1_cached or "").strip()
        )

    def _iiko_products_snapshot(self) -> dict:
        """Снимок настроек авторизации iiko для фонового воркера."""
        mode = (self._iiko_mode or "cloud").strip().lower()
        snap: dict[str, Any] = {"mode": mode}

        if mode in ("cloud", "cloud_v1", "cloudv1", "v1"):
            snap["cloud_api_url"] = (self._iiko_cloud_api_url or "").strip() or "https://api-ru.iiko.services"
            snap["cloud_api_login"] = (self._iiko_cloud_api_login or "").strip()
            snap["cloud_org_id"] = (self._iiko_cloud_org_id or "").strip()
            snap["cloud_access_token"] = (self._iiko_cloud_access_token or "").strip()
            return snap

        if mode in ("biz", "transport", "iikobiz", "iiko.biz"):
            snap["biz_api_url"] = (self._iiko_biz_api_url or "").strip() or "https://iiko.biz:9900"
            snap["biz_user_id"] = (self._iiko_biz_user_id or "").strip()
            snap["biz_user_secret"] = (self._iiko_biz_user_secret or "").strip()
            snap["biz_org_id"] = (self._iiko_biz_org_id or "").strip()
            snap["biz_access_token"] = (self._iiko_biz_access_token or "").strip()
            return snap

        # REST (/resto)
        if not (self._iiko_pass_sha1_cached or "").strip():
            if not self._ensure_iiko_pass_sha1():
                raise IikoApiError("Не задан SHA1-хэш пароля для REST. Нажмите 'Авторизация точки'.")

        snap["rest_base_url"] = (self._iiko_base_url or "").strip()
        snap["rest_login"] = (self._iiko_login or "").strip()
        snap["rest_pass_sha1"] = (self._iiko_pass_sha1_cached or "").strip()
        return snap

    def _set_iiko_load_status(self, origin: str, text: str) -> None:
        """Показывает статус загрузки iiko в нужной части UI."""
        try:
            if origin in ("pricelist", "both"):
                if hasattr(self, "lblPricelistInfo"):
                    self.lblPricelistInfo.setText(text)
        except Exception:
            pass

        try:
            if origin in ("open", "both"):
                self._set_tomorrow_info(text)
        except Exception:
            pass

    def _start_iiko_products_load(self, origin: str, user_initiated: bool = False) -> None:
        """Запускает загрузку блюд iiko в фоне (QThread), чтобы UI не зависал."""
        if getattr(self, "_iiko_products_loading", False):
            return

        self._iiko_products_loading = True
        self._iiko_products_load_origin = str(origin or "")
        self._iiko_products_load_user_initiated = bool(user_initiated)

        self._iiko_products_load_seq = int(getattr(self, "_iiko_products_load_seq", 0)) + 1
        seq = int(self._iiko_products_load_seq)

        # статус
        self._set_iiko_load_status(origin, "Загружаю блюда из iiko…")

        try:
            snapshot = self._iiko_products_snapshot()
        except Exception as e:
            self._iiko_products_loading = False
            self._set_iiko_load_status(origin, f"Ошибка загрузки iiko: {e}")
            if user_initiated:
                try:
                    QMessageBox.critical(self, "iiko", str(e))
                except Exception:
                    pass
            return

        # Важно: не делаем QThread дочерним виджету. Если окно закроют во время загрузки,
        # дочерний QThread может быть уничтожен "на ходу" и приложение вылетит.
        thread = QThread()
        worker = IikoProductsLoadWorker(seq, snapshot)
        worker.moveToThread(thread)

        thread.started.connect(worker.run)
        worker.finished.connect(self._on_iiko_products_loaded)
        worker.failed.connect(self._on_iiko_products_load_failed)

        worker.finished.connect(thread.quit)
        worker.failed.connect(thread.quit)

        worker.finished.connect(worker.deleteLater)
        worker.failed.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        # учёт активных потоков
        try:
            _ACTIVE_THREADS.add(thread)
            thread.finished.connect(lambda: _ACTIVE_THREADS.discard(thread))
        except Exception:
            pass

        self._iiko_products_load_thread = thread
        self._iiko_products_load_worker = worker

        thread.start()

    def _ensure_iiko_products_loaded_async(self, origin: str, user_initiated: bool = False) -> bool:
        """Гарантирует, что номенклатура загружается/загружена.

        Возвращает True если уже загружено, иначе False.
        """
        try:
            if getattr(self, "_iiko_products_loaded", False):
                return True

            if getattr(self, "_iiko_products_loading", False):
                # уже грузим — просто выходим
                self._set_iiko_load_status(origin, "Загружаю блюда из iiko…")
                return False

            if not self._can_autoload_iiko_products():
                self._set_iiko_load_status(
                    origin,
                    "Список блюд не загружен. Нажмите «Авторизация точки» и попробуйте снова.",
                )
                return False

            self._start_iiko_products_load(origin=origin, user_initiated=user_initiated)
            return False
        except Exception as e:
            self._set_iiko_load_status(origin, f"Ошибка загрузки iiko: {e}")
            return False

    def _on_iiko_products_loaded(self, seq: int, result: object) -> None:
        """Применяет загруженную номенклатуру к UI (в главном потоке)."""
        try:
            self._iiko_products_loading = False

            # сброс ссылок (поток/воркер сами удалятся через deleteLater)
            self._iiko_products_load_thread = None
            self._iiko_products_load_worker = None

            # если это устаревший результат — не применяем
            if int(seq) != int(getattr(self, "_iiko_products_load_seq", 0)):
                return

            if not isinstance(result, dict):
                raise IikoApiError("Некорректный результат загрузки номенклатуры.")

            products = list(result.get("products") or [])
            self._open_iiko_products = products
            self._pricelist_dishes = list(result.get("pricelist_dishes") or [])

            self._open_iiko_products_norm = list(result.get("open_norm") or [])
            self._open_iiko_products_exact = dict(result.get("open_exact") or {})

            self._pricelist_dishes_norm = list(result.get("pricelist_norm") or [])
            self._pricelist_dishes_exact = dict(result.get("pricelist_exact") or {})

            self._iiko_products_loaded = True

            # если воркер обновил токен (iikoTransport) — сохраним
            try:
                new_tok = (result.get("biz_new_token") or "").strip()
            except Exception:
                new_tok = ""
            if new_tok:
                try:
                    self._iiko_biz_access_token = new_tok
                    self._settings.setValue("iiko/biz/access_token", new_tok)
                except Exception:
                    pass

            # обновляем UI по отложенным флагам/тексту
            try:
                if getattr(self, "_pricelist_show_all_requested", False):
                    self._pricelist_show_all_requested = False
                    self._show_all_pricelist_dishes()
                else:
                    self._run_pricelist_search()
            except Exception:
                pass

            try:
                if getattr(self, "_open_show_all_requested", False):
                    self._open_show_all_requested = False
                    self._show_all_open_dishes()
                else:
                    self._run_open_search()
            except Exception:
                pass

            # если ничего не ищем — просто покажем "загружено"
            try:
                if hasattr(self, "grpPricelist") and self.grpPricelist.isVisible():
                    if len((self.edDishSearch.text() or "").strip()) < 2 and not self.lstDishSuggestions.isVisible():
                        self.lblPricelistInfo.setText(f"Загружено из iiko: {len(self._pricelist_dishes)}")
            except Exception:
                pass

            try:
                if hasattr(self, "grpTomorrowOpen") and self.grpTomorrowOpen.isVisible():
                    if len((self.edTomorrowSearch.text() or "").strip()) < 2 and not self.lstTomorrowDishes.isVisible():
                        self._set_tomorrow_info(self._format_open_schedule_status())
            except Exception:
                pass

        except Exception as e:
            self._on_iiko_products_load_failed(int(seq), str(e))

    def _on_iiko_products_load_failed(self, seq: int, error: str) -> None:
        try:
            self._iiko_products_loading = False
            self._iiko_products_load_thread = None
            self._iiko_products_load_worker = None

            # устаревший результат
            if int(seq) != int(getattr(self, "_iiko_products_load_seq", 0)):
                return

            origin = str(getattr(self, "_iiko_products_load_origin", "") or "")
            if origin not in ("open", "pricelist", "both"):
                origin = "both"

            msg = (error or "").strip() or "Ошибка загрузки iiko"
            self._set_iiko_load_status(origin, f"Ошибка загрузки iiko: {msg}")

            if bool(getattr(self, "_iiko_products_load_user_initiated", False)):
                try:
                    QMessageBox.critical(self, "iiko", msg)
                except Exception:
                    pass

        except Exception:
            pass

    def _on_pricelist_search_text_changed(self, text: str) -> None:
        self._pricelist_search_pending_text = text or ""
        if len((text or "").strip()) < 2:
            try:
                self._hide_pricelist_suggestions()
            except Exception:
                pass

            # обновим строку статуса, чтобы не оставалось "Найдено..."
            try:
                if getattr(self, "_iiko_products_loaded", False):
                    self.lblPricelistInfo.setText(f"Загружено из iiko: {len(self._pricelist_dishes)}")
                else:
                    self.lblPricelistInfo.setText("Введите название блюда (загрузка из iiko — автоматически)")
            except Exception:
                pass
            return

        try:
            self._pricelist_search_timer.start()
        except Exception:
            pass

    def _run_pricelist_search(self) -> None:
        try:
            self._update_pricelist_suggestions(self._pricelist_search_pending_text)
        except Exception:
            pass

    def _on_open_search_text_changed(self, text: str) -> None:
        self._open_search_pending_text = text or ""
        if len((text or "").strip()) < 2:
            try:
                self._hide_open_suggestions()
                # вернём статус расписания, если есть
                self._set_tomorrow_info(self._format_open_schedule_status())
            except Exception:
                pass
            return
        try:
            self._open_search_timer.start()
        except Exception:
            pass

    def _run_open_search(self) -> None:
        try:
            self._update_open_suggestions(self._open_search_pending_text)
        except Exception:
            pass

    def _get_iiko_products(self) -> List[Any]:
        """Возвращает список продуктов iiko в зависимости от выбранного режима.

        Поддерживается:
        - iikoCloud API v1 (api-ru.iiko.services) через apiLogin/access_token
        - iikoTransport (iiko.biz:9900) через user_id/user_secret (или сохранённый access_token)
        - REST (/resto)

        Для Cloud/Transport работаем с уже сохранённым access_token.
        Если он протух — попросим переавторизоваться.
        """
        mode = (self._iiko_mode or "cloud").strip().lower()

        # ===== iikoCloud API v1 =====
        if mode in ("cloud", "cloud_v1", "cloudv1", "v1"):
            api_url = (self._iiko_cloud_api_url or "").strip() or "https://api-ru.iiko.services"
            api_login = (self._iiko_cloud_api_login or "").strip()
            org_id = (self._iiko_cloud_org_id or "").strip()

            if not org_id:
                raise IikoApiError("Не выбрана организация. Нажмите 'Авторизация точки'.")

            token = (self._iiko_cloud_access_token or "").strip()
            if not token:
                raise IikoApiError(
                    "Не найден access_token iikoCloud. Сначала нажмите 'Авторизация точки'."
                )

            client = IikoCloudV1Client(
                api_url=api_url,
                api_login=api_login,
                organization_id=org_id,
                access_token=token,
            )
            return client.get_products()

        # ===== iikoTransport (iiko.biz:9900) =====
        if mode in ("biz", "transport", "iikobiz", "iiko.biz"):
            api_url = (self._iiko_biz_api_url or "").strip() or "https://iiko.biz:9900"
            user_id = (self._iiko_biz_user_id or "").strip()
            user_secret = (self._iiko_biz_user_secret or "").strip()
            org_id = (self._iiko_biz_org_id or "").strip()
            token = (self._iiko_biz_access_token or "").strip()

            if not org_id:
                raise IikoApiError("Не выбрана организация. Нажмите 'Авторизация точки'.")

            if not token and not (user_id and user_secret):
                raise IikoApiError(
                    "Не задан user_id/user_secret или access_token iikoTransport. Сначала нажмите 'Авторизация точки'."
                )

            client = IikoTransportClient(
                api_url=api_url,
                user_id=user_id,
                user_secret=user_secret,
                organization_id=org_id,
                access_token=token,
            )
            try:
                products = client.get_products()
            except IikoApiError as e:
                # Частый случай: протух токен. Если есть user_secret — попробуем обновить токен один раз.
                low = str(e).lower()
                if ("http 401" in low or "http 403" in low) and (user_id and user_secret):
                    client2 = IikoTransportClient(
                        api_url=api_url,
                        user_id=user_id,
                        user_secret=user_secret,
                        organization_id=org_id,
                    )
                    products = client2.get_products()
                    # Сохраним новый токен
                    try:
                        new_token = client2.access_token()
                        if new_token:
                            self._iiko_biz_access_token = new_token
                            self._settings.setValue("iiko/biz/access_token", new_token)
                    except Exception:
                        pass
                else:
                    raise

            # Сохраним токен (если он получился в процессе) — чтобы быстрее работало дальше.
            try:
                tok = client.access_token()
                if tok and tok != token:
                    self._iiko_biz_access_token = tok
                    self._settings.setValue("iiko/biz/access_token", tok)
            except Exception:
                pass

            return products

        # ===== REST (/resto) =====
        if not self._ensure_iiko_pass_sha1():
            raise IikoApiError("Не задан SHA1-хэш пароля для REST. Нажмите 'Авторизация точки'.")

        base_url = (self._iiko_base_url or "").strip()
        login = (self._iiko_login or "").strip()
        pass_sha1 = (self._iiko_pass_sha1_cached or "").strip()
        if not base_url or not login or not pass_sha1:
            raise IikoApiError("Не заданы параметры подключения iiko.")

        client = IikoRmsClient(base_url=base_url, login=login, pass_sha1=pass_sha1)
        return client.get_products()

    def _ensure_iiko_products_index(self) -> bool:
        """Подгружает номенклатуру iiko и строит индекс name->productId (для сопоставления с Excel)."""
        if self._iiko_products_by_key:
            return True

        products = self._get_iiko_products()

        idx: dict[str, str] = {}
        for p in products:
            key = self._pl_key(getattr(p, 'name', ''))
            if not key:
                continue
            pid = (getattr(p, 'product_id', '') or "").strip()
            if not pid:
                continue
            idx.setdefault(key, pid)

        self._iiko_products_by_key = idx
        return True

    def _pl_key(self, name: str) -> str:
        return " ".join(str(name).strip().lower().replace('ё', 'е').split())

    def _rebuild_open_selected_ids_from_list(self) -> None:
        """Синхронизирует self._open_selected_ids с UI-списком выбранных блюд."""
        try:
            ids: set[str] = set()
            if hasattr(self, "lstTomorrowSelectedDishes"):
                for i in range(self.lstTomorrowSelectedDishes.count()):
                    it = self.lstTomorrowSelectedDishes.item(i)
                    pid = str(it.data(Qt.UserRole) or "").strip()
                    if pid:
                        ids.add(pid)
            self._open_selected_ids = ids
        except Exception:
            self._open_selected_ids = set()

    def _rebuild_pricelist_selected_keys_from_list(self) -> None:
        """Синхронизирует self._pricelist_selected_keys с UI-списком выбранных блюд."""
        try:
            keys: set[str] = set()
            if hasattr(self, "lstSelectedDishes"):
                for i in range(self.lstSelectedDishes.count()):
                    it = self.lstSelectedDishes.item(i)
                    d = it.data(Qt.UserRole)
                    if isinstance(d, DishItem):
                        k = self._pl_key(d.name)
                        if k:
                            keys.add(k)
            self._pricelist_selected_keys = keys
        except Exception:
            self._pricelist_selected_keys = set()

    def _sort_open_selected_list(self) -> None:
        """Сортирует список выбранных блюд в "Открыть блюда" по названию."""
        try:
            if not hasattr(self, "lstTomorrowSelectedDishes"):
                return

            rows = []
            for i in range(self.lstTomorrowSelectedDishes.count()):
                it = self.lstTomorrowSelectedDishes.item(i)

                pid = str(it.data(Qt.UserRole) or "").strip()
                status = str(it.data(Qt.UserRole + 1) or "ready")
                base_line = str(it.data(Qt.UserRole + 2) or it.text() or "").strip()
                sort_key = str(it.data(Qt.UserRole + 3) or "").strip()

                if not sort_key:
                    name = (base_line.split(" — ", 1)[0] if base_line else "")
                    sort_key = self._pl_key(name)

                try:
                    check_state = it.checkState()
                except Exception:
                    check_state = Qt.Checked

                try:
                    fg = it.foreground()
                except Exception:
                    fg = None

                try:
                    flags = it.flags()
                except Exception:
                    flags = Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsUserCheckable

                rows.append((sort_key, base_line, pid, status, check_state, fg, flags))

            rows.sort(key=lambda r: ((r[0] or ""), (r[1] or "")))

            self.lstTomorrowSelectedDishes.setUpdatesEnabled(False)
            try:
                self.lstTomorrowSelectedDishes.clear()
                for sort_key, base_line, pid, status, check_state, fg, flags in rows:
                    it2 = QListWidgetItem("")
                    it2.setData(Qt.UserRole, pid)
                    it2.setData(Qt.UserRole + 1, status)
                    it2.setData(Qt.UserRole + 2, base_line)
                    it2.setData(Qt.UserRole + 3, sort_key)

                    it2.setFlags(flags | Qt.ItemIsUserCheckable)
                    it2.setCheckState(check_state)

                    if fg is not None:
                        try:
                            it2.setForeground(fg)
                        except Exception:
                            pass

                    # финальный текст с учётом статуса
                    self._apply_open_selected_item_text(it2, base_line)
                    self.lstTomorrowSelectedDishes.addItem(it2)
            finally:
                self.lstTomorrowSelectedDishes.setUpdatesEnabled(True)

            self._rebuild_open_selected_ids_from_list()
        except Exception:
            pass

    def _sort_pricelist_selected_list(self) -> None:
        """Сортирует список выбранных блюд в "Ценники" по названию."""
        try:
            if not hasattr(self, "lstSelectedDishes"):
                return

            rows = []
            for i in range(self.lstSelectedDishes.count()):
                it = self.lstSelectedDishes.item(i)
                d = it.data(Qt.UserRole)

                if isinstance(d, DishItem):
                    sort_key = self._pl_key(d.name)
                    line = self._format_dish_line(d)
                else:
                    line = str(it.text() or "")
                    sort_key = self._pl_key(line.split(" — ", 1)[0] if line else "")

                try:
                    check_state = it.checkState()
                except Exception:
                    check_state = Qt.Checked

                # по новому правилу: сняли галочку = удаляем из списка
                if check_state != Qt.Checked:
                    continue

                rows.append((sort_key, line, d, check_state))

            rows.sort(key=lambda r: ((r[0] or ""), (r[1] or "")))

            self._suppress_pricelist_selected_item_changed = True
            self.lstSelectedDishes.setUpdatesEnabled(False)
            try:
                self.lstSelectedDishes.clear()
                for _, line, d, check_state in rows:
                    it2 = QListWidgetItem(line)
                    it2.setData(Qt.UserRole, d)
                    it2.setFlags(it2.flags() | Qt.ItemIsUserCheckable)
                    it2.setCheckState(check_state)
                    self.lstSelectedDishes.addItem(it2)
            finally:
                self.lstSelectedDishes.setUpdatesEnabled(True)
                self._suppress_pricelist_selected_item_changed = False

            self._rebuild_pricelist_selected_keys_from_list()
        except Exception:
            try:
                self._suppress_pricelist_selected_item_changed = False
            except Exception:
                pass

    # ===== Открыть блюда: поиск в iiko -> выбрать -> снять со стоп-листа =====
    def _format_iiko_product_line(self, p: Any) -> str:
        name = (getattr(p, "name", "") or "").strip()
        weight = (getattr(p, "weight", "") or "").strip()
        price = (getattr(p, "price", "") or "").strip()

        parts = [name]
        if weight:
            parts.append(weight)
        if price:
            parts.append(price)
        return " — ".join([x for x in parts if x])

    def _load_open_iiko_products(self) -> None:
        """Ручной запуск загрузки блюд iiko для вкладки "Открыть блюда" (в фоне)."""
        try:
            if getattr(self, "_iiko_products_loaded", False):
                # уже есть — просто обновим подсказки
                try:
                    self._run_open_search()
                except Exception:
                    pass
                return
            self._ensure_iiko_products_loaded_async(origin="open", user_initiated=True)
        except Exception:
            pass

    def _show_all_open_dishes(self):
        """Показывает подсказки без фильтра (с ограничением по количеству)."""
        try:
            # если ещё не загружено — запускаем в фоне и запоминаем, что хотели показать все
            if not getattr(self, "_iiko_products_loaded", False):
                self._open_show_all_requested = True
                self._ensure_iiko_products_loaded_async(origin="open", user_initiated=True)
                return

            if getattr(self, "_iiko_products_loading", False):
                self._open_show_all_requested = True
                self._set_tomorrow_info("Загружаю блюда из iiko…")
                return

            self.lstTomorrowDishes.setUpdatesEnabled(False)
            try:
                self.lstTomorrowDishes.clear()
                if not self._open_iiko_products:
                    self._set_tomorrow_info("Список блюд пуст")
                    self._hide_open_suggestions()
                    return

                limit = 500
                for p in self._open_iiko_products[:limit]:
                    item = QListWidgetItem(self._format_iiko_product_line(p))
                    item.setData(Qt.UserRole, p)
                    self.lstTomorrowDishes.addItem(item)
            finally:
                self.lstTomorrowDishes.setUpdatesEnabled(True)

            self._set_open_suggestions_visible(self.lstTomorrowDishes.count() > 0)

            if len(self._open_iiko_products) > limit:
                self._set_tomorrow_info(
                    f"Загружено из iiko: {len(self._open_iiko_products)} (показаны первые {limit}). "
                    "Введите 2+ символа для поиска."
                )
            else:
                self._set_tomorrow_info(f"Загружено из iiko: {len(self._open_iiko_products)}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def _update_open_suggestions(self, text: str):
        """Обновляет подсказки по вводу (вкладка "Открыть блюда")."""
        try:
            self.lstTomorrowDishes.clear()
        except Exception:
            pass

        q = (text or "").strip().lower().replace("ё", "е")
        if len(q) < 2:
            self._hide_open_suggestions()
            # вернём статус расписания, чтобы не висело "Загружаю…"
            try:
                self._set_tomorrow_info(self._format_open_schedule_status())
            except Exception:
                pass
            return

        # если ещё не загружено — запускаем фоновую загрузку
        if not getattr(self, "_iiko_products_loaded", False):
            self._set_tomorrow_info("Загружаю блюда из iiko…")
            self._ensure_iiko_products_loaded_async(origin="open", user_initiated=False)
            return

        if not self._open_iiko_products_norm:
            self._hide_open_suggestions()
            self._set_tomorrow_info("Список блюд пуст")
            return

        shown = 0
        self.lstTomorrowDishes.setUpdatesEnabled(False)
        try:
            for name_norm, p in self._open_iiko_products_norm:
                if q in name_norm:
                    item = QListWidgetItem(self._format_iiko_product_line(p))
                    item.setData(Qt.UserRole, p)
                    self.lstTomorrowDishes.addItem(item)
                    shown += 1
                    if shown >= 30:
                        break
        finally:
            self.lstTomorrowDishes.setUpdatesEnabled(True)

        if shown:
            self._set_open_suggestions_visible(True)
            self._set_tomorrow_info(f"Найдено: {shown} (показаны первые 30)")
        else:
            self._hide_open_suggestions()
            self._set_tomorrow_info("Совпадений не найдено")

    def _apply_open_selected_item_text(self, it: QListWidgetItem, base_line: str) -> None:
        """Обновляет отображение выбранного блюда с учётом статуса (ОТКРЫТО/ошибка/...)."""
        try:
            status = str(it.data(Qt.UserRole + 1) or "ready")
        except Exception:
            status = "ready"

        suffix = ""
        if status == "opened":
            suffix = "  (ОТКРЫТО)"
        elif status == "failed":
            suffix = "  (ошибка открытия)"
        elif status == "opening":
            suffix = "  (открываю…)"

        try:
            it.setText(f"{base_line}{suffix}" if suffix else base_line)
        except Exception:
            pass

    def _add_open_selected(self, p: Any):
        pid = (getattr(p, "product_id", "") or "").strip()
        base_line = self._format_iiko_product_line(p)
        if not pid or not base_line:
            return

        name = (getattr(p, "name", "") or "").strip()
        sort_key = self._pl_key(name or base_line)

        # Если уже есть в списке — не удаляем старое, а обновляем строку (вес/цена и т.п.)
        if pid in self._open_selected_ids:
            try:
                for i in range(self.lstTomorrowSelectedDishes.count()):
                    it = self.lstTomorrowSelectedDishes.item(i)
                    if str(it.data(Qt.UserRole) or "").strip() == pid:
                        it.setData(Qt.UserRole + 2, base_line)
                        it.setData(Qt.UserRole + 3, sort_key)
                        self._apply_open_selected_item_text(it, base_line)
                        break
            except Exception:
                pass
            try:
                self._sort_open_selected_list()
            except Exception:
                pass
            return

        it = QListWidgetItem(base_line)
        it.setData(Qt.UserRole, pid)
        it.setData(Qt.UserRole + 1, "ready")
        # base text (без статуса), чтобы после открытия не терять вес/цену
        it.setData(Qt.UserRole + 2, base_line)
        it.setData(Qt.UserRole + 3, sort_key)
        it.setFlags(it.flags() | Qt.ItemIsUserCheckable)
        it.setCheckState(Qt.Checked)

        self.lstTomorrowSelectedDishes.addItem(it)
        self._open_selected_ids.add(pid)
        try:
            self._sort_open_selected_list()
        except Exception:
            pass

    def _on_open_suggestion_clicked(self, item: QListWidgetItem):
        try:
            p = item.data(Qt.UserRole)
            if p is not None:
                self._add_open_selected(p)
            # после выбора — прячем подсказки как выпадающий список
            self._hide_open_suggestions()
        except Exception:
            pass

    def _add_open_from_enter(self):
        """Enter в поле поиска: добавляем точное совпадение или первый пункт из подсказок."""
        try:
            q_raw = (self.edTomorrowSearch.text() or "").strip()
            if not q_raw:
                return

            # если ещё не загружено — запускаем загрузку в фоне
            if not getattr(self, "_iiko_products_loaded", False):
                self._set_tomorrow_info("Загружаю блюда из iiko…")
                self._ensure_iiko_products_loaded_async(origin="open", user_initiated=False)
                return

            q = self._pl_key(q_raw)

            # 1) точное совпадение по названию (O(1) через индекс)
            p = None
            try:
                p = self._open_iiko_products_exact.get(q)
            except Exception:
                p = None
            if p is not None:
                self._add_open_selected(p)
                self._hide_open_suggestions()
                return

            # 2) иначе — первый элемент подсказок (обновим подсказки по текущему тексту)
            try:
                self._update_open_suggestions(q_raw)
            except Exception:
                pass

            if self.lstTomorrowDishes.count() > 0:
                it = self.lstTomorrowDishes.item(0)
                p = it.data(Qt.UserRole)
                if p is not None:
                    self._add_open_selected(p)
                    self._hide_open_suggestions()
                    return

            QMessageBox.information(self, "Не найдено", "По вашему запросу нет совпадений в iiko.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def _clear_open_selection(self):
        try:
            self.lstTomorrowSelectedDishes.clear()
        except Exception:
            pass
        self._open_selected_ids = set()

    # ===== Открытие блюд на завтра (Excel -> iiko stoplist) =====
    def _load_tomorrow_dishes_from_excel(self):
        try:
            excel_path = self.edExcelPath.text().strip()
            if not excel_path:
                QMessageBox.warning(self, "Внимание", "Выберите Excel файл меню на завтра.")
                return
            if not Path(excel_path).exists():
                QMessageBox.warning(self, "Ошибка", "Указанный Excel файл не найден.")
                return

            # 1) Забираем блюда из Excel
            dishes = extract_all_dishes_with_details(excel_path)
            if not dishes:
                QMessageBox.warning(self, "Не найдено", "В Excel не удалось найти блюда (проверьте формат меню).")
                return

            self._tomorrow_menu_dishes = dishes

            # 2) Готовим индекс iiko, чтобы сопоставить имена -> productId
            if not self._ensure_iiko_products_index():
                return

            self._populate_tomorrow_dish_list()

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def _populate_tomorrow_dish_list(self):
        self._suppress_tomorrow_item_changed = True
        try:
            self.lstTomorrowDishes.clear()

            not_found = 0
            for d in self._tomorrow_menu_dishes:
                name = (d.name or "").strip()
                if not name:
                    continue

                key = self._pl_key(name)
                pid = self._iiko_products_by_key.get(key, "")

                it = QListWidgetItem(name)
                it.setData(Qt.UserRole, pid)
                it.setData(Qt.UserRole + 1, "ready" if pid else "not_found")
                it.setData(Qt.UserRole + 2, name)

                if pid:
                    it.setFlags(it.flags() | Qt.ItemIsUserCheckable)
                    it.setCheckState(Qt.Unchecked)
                else:
                    not_found += 1
                    # не даём поставить галочку
                    it.setFlags(it.flags() & ~Qt.ItemIsUserCheckable)
                    it.setForeground(QBrush(QColor("#888888")))
                    it.setText(f"{name}  (не найдено в iiko)")

                self.lstTomorrowDishes.addItem(it)

            total = self.lstTomorrowDishes.count()
            self._set_tomorrow_info(
                f"Загружено из Excel: {total}. Не найдено в iiko: {not_found}. "
                "Поставьте галочку у найденных — блюдо откроется (снимем со стоп-листа)."
            )

        finally:
            self._suppress_tomorrow_item_changed = False

    def _filter_tomorrow_dishes(self, text: str):
        q = self._pl_key(text)
        for i in range(self.lstTomorrowDishes.count()):
            it = self.lstTomorrowDishes.item(i)
            base_name = (it.data(Qt.UserRole + 2) or it.text() or "")
            show = (not q) or (q in self._pl_key(base_name))
            it.setHidden(not show)

    def _open_stoplist_product_id(self, pid: str) -> None:
        """Снять блюдо со стоп-листа по product_id (REST /resto)."""
        pid = (pid or "").strip()
        if not pid:
            raise IikoApiError("Не задан product_id.")

        mode = (self._iiko_mode or "cloud").strip().lower()
        if mode not in ("rest", "rms", "resto"):
            raise IikoApiError(
                "Открытие блюда (снятие со стоп-листа) доступно только через REST (/resto). "
                "Нажмите 'Авторизация точки' и выполните вход (REST)."
            )

        # REST: убедимся, что есть sha1 пароля
        if not self._ensure_iiko_pass_sha1():
            raise IikoApiError("Не задан SHA1-хэш пароля для REST. Нажмите 'Авторизация точки'.")

        base_url = (self._iiko_base_url or "").strip()
        login = (self._iiko_login or "").strip()
        pass_sha1 = (self._iiko_pass_sha1_cached or "").strip()
        if not base_url or not login or not pass_sha1:
            raise IikoApiError("Не заданы параметры REST (/resto). Нажмите 'Авторизация точки'.")

        client = IikoRmsClient(base_url=base_url, login=login, pass_sha1=pass_sha1)
        client.open_product_from_stoplist(pid)

    def _open_one_tomorrow_item(self, it: QListWidgetItem) -> bool:
        """Снять блюдо со стоп-листа.

        Реализация: через iikoRMS REST (/resto).
        """
        pid = (it.data(Qt.UserRole) or "").strip()
        if not pid:
            return False

        self._open_stoplist_product_id(pid)
        return True

    def _on_tomorrow_item_changed(self, item: QListWidgetItem):
        if self._suppress_tomorrow_item_changed:
            return

        try:
            status = item.data(Qt.UserRole + 1)
            if status == "not_found":
                return

            if item.checkState() != Qt.Checked:
                return

            # Поставили галочку -> пробуем "открыть" (снять со стоп-листа)
            self._suppress_tomorrow_item_changed = True
            try:
                item.setData(Qt.UserRole + 1, "opening")
                base_name = (item.data(Qt.UserRole + 2) or item.text() or "")
                item.setText(f"{base_name}  (открываю…)")
            finally:
                self._suppress_tomorrow_item_changed = False

            try:
                self._open_one_tomorrow_item(item)
                self._suppress_tomorrow_item_changed = True
                try:
                    base_name = (item.data(Qt.UserRole + 2) or item.text() or "")
                    item.setData(Qt.UserRole + 1, "opened")
                    item.setForeground(QBrush(QColor("#2e7d32")))
                    item.setText(f"{base_name}  (ОТКРЫТО)")
                    item.setCheckState(Qt.Checked)
                finally:
                    self._suppress_tomorrow_item_changed = False

            except IikoApiError as e:
                self._suppress_tomorrow_item_changed = True
                try:
                    base_name = (item.data(Qt.UserRole + 2) or item.text() or "")
                    item.setData(Qt.UserRole + 1, "failed")
                    item.setForeground(QBrush(QColor("#b71c1c")))
                    item.setText(f"{base_name}  (ошибка открытия)")
                    item.setCheckState(Qt.Unchecked)
                finally:
                    self._suppress_tomorrow_item_changed = False

                QMessageBox.critical(self, "iiko", str(e))

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def _open_tomorrow_checked(self):
        """Открывает (снимает со стоп-листа) отмеченные блюда.

        Если выбрана дата в будущем — планирует открытие на эту дату (вариант B).
        """
        try:
            if not hasattr(self, "lstTomorrowSelectedDishes"):
                return

            # собираем выбранные блюда
            items: List[QListWidgetItem] = []
            product_ids: List[str] = []
            titles: List[str] = []
            for i in range(self.lstTomorrowSelectedDishes.count()):
                it = self.lstTomorrowSelectedDishes.item(i)
                if it.checkState() != Qt.Checked:
                    continue
                if it.data(Qt.UserRole + 1) == "opened":
                    continue
                pid = (it.data(Qt.UserRole) or "").strip()
                if not pid:
                    continue
                items.append(it)
                product_ids.append(pid)
                titles.append(str(it.data(Qt.UserRole + 2) or it.text() or "").strip())

            if not product_ids:
                QMessageBox.information(self, "Нет выбора", "Отметьте блюда галочками, которые нужно открыть.")
                return

            # выбранная дата
            qd = None
            try:
                if hasattr(self, "calOpenDate"):
                    qd = self.calOpenDate.selectedDate()
            except Exception:
                qd = None

            if qd is None or (hasattr(qd, "isValid") and not qd.isValid()):
                target = date.today()
            else:
                target = date(int(qd.year()), int(qd.month()), int(qd.day()))

            today = date.today()

            # будущее: планируем
            if target > today:
                run_at = datetime.combine(target, time(0, 1))
                added_count: Optional[int] = None

                # если уже есть pending-расписание — дадим выбрать: добавить к нему или заменить
                if isinstance(self._open_schedule_job, dict) and (str(self._open_schedule_job.get("state") or "pending")).lower() == "pending":
                    old = dict(self._open_schedule_job or {})

                    old_run_at_s = str(old.get("run_at") or "")
                    old_dt: Optional[datetime] = None
                    old_when = old_run_at_s
                    try:
                        old_dt = datetime.fromisoformat(old_run_at_s)
                        old_when = old_dt.strftime("%d.%m.%Y %H:%M")
                    except Exception:
                        old_dt = None

                    new_when = run_at.strftime("%d.%m.%Y %H:%M")

                    old_pids = old.get("product_ids") or []
                    if not isinstance(old_pids, list):
                        old_pids = []
                    old_pids_clean = [str(x or "").strip() for x in old_pids if str(x or "").strip()]

                    old_titles = old.get("titles") or []
                    if not isinstance(old_titles, list):
                        old_titles = []
                    old_titles_clean = [str(t or "").strip() for t in old_titles if str(t or "").strip()]

                    old_count = len(old_titles_clean) or len(old_pids_clean)
                    new_count = len(product_ids)

                    msg = QMessageBox(self)
                    msg.setIcon(QMessageBox.Question)
                    msg.setWindowTitle("iiko")
                    msg.setText("Уже есть запланированное открытие.")
                    msg.setInformativeText(
                        f"Сейчас запланировано: {old_when}. Блюд: {old_count}.\n"
                        f"Выбрано сейчас: {new_count}. (Дата в календаре: {target.strftime('%d.%m.%Y')}, будет {new_when})\n\n"
                        "Добавить — добавит блюда к уже запланированным (дата останется прежней).\n"
                        "Заменить — перезапишет список (и дату возьмёт из календаря)."
                    )
                    try:
                        msg.setTextFormat(Qt.PlainText)
                    except Exception:
                        pass

                    if old_titles_clean:
                        try:
                            msg.setDetailedText("\n".join(old_titles_clean))
                        except Exception:
                            pass

                    btn_add = msg.addButton("Добавить", QMessageBox.AcceptRole)
                    btn_replace = msg.addButton("Заменить", QMessageBox.DestructiveRole)
                    msg.addButton("Отмена", QMessageBox.RejectRole)
                    try:
                        msg.setDefaultButton(btn_add)
                    except Exception:
                        pass

                    msg.exec()
                    clicked = msg.clickedButton()

                    if clicked == btn_add:
                        # Merge: сначала старые, потом новые (без дублей)
                        old_title_by_pid: dict[str, str] = {}
                        try:
                            for idx, pid in enumerate(old_pids_clean):
                                if not pid:
                                    continue
                                t = ""
                                if idx < len(old_titles):
                                    t = str(old_titles[idx] or "").strip()
                                if pid not in old_title_by_pid:
                                    old_title_by_pid[pid] = t
                        except Exception:
                            old_title_by_pid = {}

                        new_title_by_pid: dict[str, str] = {}
                        try:
                            for idx, pid in enumerate(product_ids):
                                pid_s = str(pid or "").strip()
                                if not pid_s:
                                    continue
                                t = ""
                                if idx < len(titles):
                                    t = str(titles[idx] or "").strip()
                                if pid_s not in new_title_by_pid:
                                    new_title_by_pid[pid_s] = t
                        except Exception:
                            new_title_by_pid = {}

                        merged_pids: List[str] = []
                        seen: set[str] = set()
                        for pid in (old_pids_clean + list(product_ids)):
                            pid_s = str(pid or "").strip()
                            if not pid_s or pid_s in seen:
                                continue
                            merged_pids.append(pid_s)
                            seen.add(pid_s)

                        merged_titles: List[str] = []
                        for pid in merged_pids:
                            t = old_title_by_pid.get(pid)
                            if t is None:
                                t = new_title_by_pid.get(pid)
                            merged_titles.append(str(t or "").strip())

                        added_count = max(0, len(merged_pids) - len(old_pids_clean))
                        product_ids = merged_pids
                        titles = merged_titles

                        # Дату оставляем прежнюю, если она корректно распарсилась
                        run_at = old_dt or run_at

                    elif clicked != btn_replace:
                        return

                job = {
                    "state": "pending",
                    "run_at": run_at.isoformat(timespec="seconds"),
                    "product_ids": product_ids,
                    "titles": titles,
                    "created_at": datetime.now().isoformat(timespec="seconds"),
                }
                if added_count is not None:
                    try:
                        job["updated_at"] = datetime.now().isoformat(timespec="seconds")
                        job["added_count"] = int(added_count)
                    except Exception:
                        pass

                self._save_open_schedule_job(job)

                # Пытаемся создать задачу Планировщика Windows, чтобы сработало даже если приложение закрыто.
                sched_ok = False
                sched_err = ""
                try:
                    sched_ok, sched_err = _windows_create_open_schedule_task(run_at)
                except Exception as e:
                    sched_ok, sched_err = False, str(e)

                prefix = ""
                try:
                    if added_count is not None:
                        prefix = f"Добавлено блюд: {int(added_count)}. Всего блюд: {len(product_ids)}.\n"
                except Exception:
                    prefix = ""

                try:
                    when_date = run_at.strftime('%d.%m.%Y')
                    when_time = run_at.strftime('%H:%M')
                except Exception:
                    when_date = target.strftime('%d.%m.%Y')
                    when_time = "00:01"

                if sched_ok:
                    task_name = ""
                    try:
                        task_name = _open_schedule_task_name()
                    except Exception:
                        task_name = ""

                    info = (
                        f"{prefix}Запланировано на {when_date} ({when_time}).\n"
                        "Сработает автоматически (Планировщик Windows).\n"
                        + (f"Задача: {task_name}\n" if task_name else "")
                        + "Компьютер должен быть включён."
                    )
                else:
                    info = (
                        f"{prefix}Запланировано на {when_date} ({when_time}).\n"
                        "Не удалось создать задачу Планировщика Windows — откроется при следующем запуске приложения после наступления времени.\n"
                        f"Причина: {sched_err}"
                    )

                self._set_tomorrow_info(self._format_open_schedule_status())
                QMessageBox.information(self, "Запланировано", info)
                return

            if target < today:
                QMessageBox.warning(self, "Дата", "Вы выбрали дату в прошлом.")
                return

            # сегодня: открываем сейчас
            any_done = False
            for it in items:
                try:
                    self._open_one_tomorrow_item(it)
                    it.setData(Qt.UserRole + 1, "opened")
                    it.setForeground(QBrush(QColor("#2e7d32")))
                    base_line = (it.data(Qt.UserRole + 2) or it.text() or "")
                    it.setText(f"{base_line}  (ОТКРЫТО)")
                    any_done = True
                except IikoApiError as e:
                    it.setData(Qt.UserRole + 1, "failed")
                    it.setForeground(QBrush(QColor("#b71c1c")))
                    base_line = (it.data(Qt.UserRole + 2) or it.text() or "")
                    it.setText(f"{base_line}  (ошибка открытия)")
                    QMessageBox.critical(self, "iiko", str(e))
                    return

            if any_done:
                QMessageBox.information(self, "Готово", "Открытие отмеченных блюд завершено.")
            else:
                QMessageBox.information(self, "Нет выбора", "Отметьте блюда галочками, которые нужно открыть.")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def _format_dish_line(self, d: DishItem) -> str:
        parts = [d.name]
        if d.weight:
            parts.append(str(d.weight))
        if d.price:
            parts.append(str(d.price))
        return " — ".join(parts)

    def _load_pricelist_dishes(self):
        """Ручной запуск загрузки блюд iiko для вкладки "Ценники" (в фоне)."""
        try:
            if getattr(self, "_iiko_products_loaded", False):
                try:
                    self.lblPricelistInfo.setText(f"Загружено из iiko: {len(self._pricelist_dishes)}")
                except Exception:
                    pass
                try:
                    self._run_pricelist_search()
                except Exception:
                    pass
                return

            self._ensure_iiko_products_loaded_async(origin="pricelist", user_initiated=True)
        except Exception:
            pass

    def _show_all_pricelist_dishes(self):
        """Показывает список блюд без фильтра (с ограничением по количеству)."""
        try:
            # если ещё не загружено — запускаем в фоне и запоминаем, что хотели показать все
            if not getattr(self, "_iiko_products_loaded", False):
                self._pricelist_show_all_requested = True
                self._ensure_iiko_products_loaded_async(origin="pricelist", user_initiated=True)
                return

            if getattr(self, "_iiko_products_loading", False):
                self._pricelist_show_all_requested = True
                try:
                    self.lblPricelistInfo.setText("Загружаю блюда из iiko…")
                except Exception:
                    pass
                return

            self.lstDishSuggestions.setUpdatesEnabled(False)
            try:
                self.lstDishSuggestions.clear()
                if not self._pricelist_dishes:
                    self._hide_pricelist_suggestions()
                    return

                limit = 500
                for d in self._pricelist_dishes[:limit]:
                    item = QListWidgetItem(self._format_dish_line(d))
                    item.setData(Qt.UserRole, d)
                    self.lstDishSuggestions.addItem(item)
            finally:
                self.lstDishSuggestions.setUpdatesEnabled(True)

            self._set_pricelist_suggestions_visible(self.lstDishSuggestions.count() > 0)

            if len(self._pricelist_dishes) > limit:
                self.lblPricelistInfo.setText(
                    f"Загружено из iiko: {len(self._pricelist_dishes)} (показаны первые {limit})"
                )
            else:
                self.lblPricelistInfo.setText(f"Загружено из iiko: {len(self._pricelist_dishes)}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def _update_pricelist_suggestions(self, text: str):
        try:
            self.lstDishSuggestions.clear()
        except Exception:
            pass

        q = (text or "").strip().lower().replace('ё', 'е')
        if len(q) < 2:
            self._hide_pricelist_suggestions()
            # чтобы не оставалось "Загружаю…"
            try:
                if getattr(self, "_iiko_products_loaded", False):
                    self.lblPricelistInfo.setText(f"Загружено из iiko: {len(self._pricelist_dishes)}")
            except Exception:
                pass
            return

        # если ещё не загружено — запускаем фоновую загрузку
        if not getattr(self, "_iiko_products_loaded", False):
            try:
                self.lblPricelistInfo.setText("Загружаю блюда из iiko…")
            except Exception:
                pass
            self._ensure_iiko_products_loaded_async(origin="pricelist", user_initiated=False)
            return

        if not self._pricelist_dishes_norm:
            self._hide_pricelist_suggestions()
            self.lblPricelistInfo.setText("Список блюд пуст")
            return

        shown = 0
        self.lstDishSuggestions.setUpdatesEnabled(False)
        try:
            for name_norm, d in self._pricelist_dishes_norm:
                if q in name_norm:
                    item = QListWidgetItem(self._format_dish_line(d))
                    item.setData(Qt.UserRole, d)
                    self.lstDishSuggestions.addItem(item)
                    shown += 1
                    if shown >= 30:
                        break
        finally:
            self.lstDishSuggestions.setUpdatesEnabled(True)

        if shown:
            self._set_pricelist_suggestions_visible(True)
            self.lblPricelistInfo.setText(f"Найдено: {shown} (показаны первые 30)")
        else:
            self._hide_pricelist_suggestions()
            self.lblPricelistInfo.setText("Совпадений не найдено")

    def _add_pricelist_selected(self, d: DishItem):
        key = self._pl_key(d.name)
        if not key:
            return

        # Если блюдо уже было выбрано — обновляем строку (вес/цена), но не удаляем из списка.
        if key in self._pricelist_selected_keys:
            try:
                for i in range(self.lstSelectedDishes.count()):
                    it = self.lstSelectedDishes.item(i)
                    old = it.data(Qt.UserRole)
                    if isinstance(old, DishItem) and self._pl_key(old.name) == key:
                        prev_state = it.checkState()
                        it.setData(Qt.UserRole, d)
                        it.setText(self._format_dish_line(d))
                        it.setCheckState(prev_state)
                        break
            except Exception:
                pass
            try:
                self._sort_pricelist_selected_list()
            except Exception:
                pass
            return

        it = QListWidgetItem(self._format_dish_line(d))
        it.setData(Qt.UserRole, d)
        it.setFlags(it.flags() | Qt.ItemIsUserCheckable)
        it.setCheckState(Qt.Checked)
        self.lstSelectedDishes.addItem(it)
        self._pricelist_selected_keys.add(key)
        try:
            self._sort_pricelist_selected_list()
        except Exception:
            pass

    def _on_pricelist_suggestion_clicked(self, item: QListWidgetItem):
        try:
            d = item.data(Qt.UserRole)
            if isinstance(d, DishItem):
                self._add_pricelist_selected(d)
            self._hide_pricelist_suggestions()
        except Exception:
            pass

    def _add_pricelist_from_enter(self):
        """Enter в поле поиска: добавляем точное совпадение или первый пункт из подсказок."""
        try:
            q_raw = (self.edDishSearch.text() or "").strip()
            if not q_raw:
                return

            # если ещё не загружено — запускаем загрузку в фоне
            if not getattr(self, "_iiko_products_loaded", False):
                try:
                    self.lblPricelistInfo.setText("Загружаю блюда из iiko…")
                except Exception:
                    pass
                self._ensure_iiko_products_loaded_async(origin="pricelist", user_initiated=False)
                return

            q = self._pl_key(q_raw)

            # 1) точное совпадение по названию (O(1) через индекс)
            d = None
            try:
                d = self._pricelist_dishes_exact.get(q)
            except Exception:
                d = None
            if isinstance(d, DishItem):
                self._add_pricelist_selected(d)
                self._hide_pricelist_suggestions()
                return

            # 2) иначе — первый элемент подсказок
            try:
                self._update_pricelist_suggestions(q_raw)
            except Exception:
                pass

            if self.lstDishSuggestions.count() > 0:
                it = self.lstDishSuggestions.item(0)
                d2 = it.data(Qt.UserRole)
                if isinstance(d2, DishItem):
                    self._add_pricelist_selected(d2)
                    self._hide_pricelist_suggestions()
                    return

            QMessageBox.information(self, "Не найдено", "По вашему запросу нет совпадений в загруженном меню.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def _on_pricelist_selected_item_changed(self, item: QListWidgetItem) -> None:
        if getattr(self, "_suppress_pricelist_selected_item_changed", False):
            return

        try:
            if item.checkState() == Qt.Checked:
                return

            # сняли галочку => удаляем из списка
            key = ""
            try:
                d = item.data(Qt.UserRole)
                if isinstance(d, DishItem):
                    key = self._pl_key(d.name)
                else:
                    line = (item.text() or "").strip()
                    key = self._pl_key(line.split(" — ", 1)[0] if line else "")
            except Exception:
                key = ""

            if key:
                try:
                    self._pricelist_selected_keys.discard(key)
                except Exception:
                    pass

            try:
                row = self.lstSelectedDishes.row(item)
                if row >= 0:
                    self.lstSelectedDishes.takeItem(row)
            except Exception:
                pass

        except Exception:
            pass

    def _clear_pricelist_selection(self):
        self.lstSelectedDishes.clear()
        self._pricelist_selected_keys = set()

    def do_create_pricelist_excel(self):
        """Сформировать ценники (xlsx) и сохранить на Рабочий стол.

        Без диалога выбора файла.
        """
        try:
            # берем только отмеченные галочкой
            selected: List[DishItem] = []
            for i in range(self.lstSelectedDishes.count()):
                it = self.lstSelectedDishes.item(i)
                if it.checkState() != Qt.Checked:
                    continue
                d = it.data(Qt.UserRole)
                if isinstance(d, DishItem):
                    selected.append(d)

            if not selected:
                QMessageBox.warning(self, "Внимание", "Выберите хотя бы одно блюдо (поставьте галочку).")
                return

            desktop = Path.home() / "Desktop"
            out_path = str(desktop / "ценники.xlsx")

            create_pricelist_xlsx(selected, out_path)

            try:
                QMessageBox.information(self, "Готово", f"Файл создан:\n{out_path}")
            except Exception:
                pass

            try:
                QDesktopServices.openUrl(QUrl.fromLocalFile(out_path))
            except Exception:
                pass

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def do_merge_chain_pricetags(self):
        """Выгрузить чёрные ценники по шаблону в один файл.

        Сохраняем на Рабочий стол как "ценники.xls".
        """
        try:
            # берем только отмеченные галочкой
            selected_tags: List[TagData] = []
            for i in range(self.lstSelectedDishes.count()):
                it = self.lstSelectedDishes.item(i)
                if it.checkState() != Qt.Checked:
                    continue
                d = it.data(Qt.UserRole)

                if isinstance(d, DishItem):
                    nm = (d.name or "").strip()
                    if not nm:
                        continue

                    # на всякий случай обновим данные из загруженной номенклатуры (если выбранный объект старый)
                    w = (d.weight if d.weight is not None else "")
                    pr = (d.price if d.price is not None else "")
                    desc = (d.description if d.description is not None else "")
                    try:
                        k = self._pl_key(nm)
                        d2 = self._pricelist_dishes_exact.get(k)
                        if isinstance(d2, DishItem):
                            if not (w or "").strip():
                                w = (d2.weight or "")
                            if not (pr or "").strip():
                                pr = (d2.price or "")
                            if not (desc or "").strip():
                                desc = (d2.description or "")
                    except Exception:
                        pass

                    selected_tags.append(
                        TagData(
                            name=nm,
                            weight=w,
                            composition=desc,
                            price=pr,
                        )
                    )
                else:
                    # fallback: строка вида "Название — вес — цена"
                    raw = (it.text() or "").strip()
                    if not raw:
                        continue
                    nm = raw.split(" — ", 1)[0].strip()
                    if nm:
                        selected_tags.append(TagData(name=nm))

            if not selected_tags:
                QMessageBox.warning(self, "Внимание", "Выберите хотя бы одно блюдо (поставьте галочку).")
                return

            desktop = Path.home() / "Desktop"
            out_path = str(desktop / "ценники.xls")

            self._start_chain_pricetags_merge(selected_tags, out_path)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def _start_chain_pricetags_merge(self, tags: List[TagData], output_path: str) -> None:
        """Старт формирования чёрных ценников в фоне (QThread)."""
        try:
            if getattr(self, "_chain_pricetags_merge_loading", False):
                return
            self._chain_pricetags_merge_loading = True

            self._chain_pricetags_merge_seq = int(getattr(self, "_chain_pricetags_merge_seq", 0)) + 1
            seq = int(self._chain_pricetags_merge_seq)

            try:
                missing_price = 0
                try:
                    for t in list(tags or []):
                        pr = getattr(t, "price", None)
                        pr_txt = "" if pr is None else str(pr).strip()
                        if pr_txt == "":
                            missing_price += 1
                except Exception:
                    missing_price = 0

                if missing_price:
                    self.lblPricelistInfo.setText(f"Формирую чёрные ценники… (без цены: {missing_price})")
                else:
                    self.lblPricelistInfo.setText("Формирую чёрные ценники…")
            except Exception:
                pass

            try:
                if hasattr(self, "btnMergeChainPricetags"):
                    self.btnMergeChainPricetags.setEnabled(False)
            except Exception:
                pass

            # Аналогично: поток без parent, чтобы безопаснее переживать закрытие окна.
            thread = QThread()
            worker = ChainPriceTagsMergeWorker(seq, list(tags or []), str(output_path or ""))
            worker.moveToThread(thread)

            thread.started.connect(worker.run)
            worker.finished.connect(self._on_chain_pricetags_merge_finished)
            worker.failed.connect(self._on_chain_pricetags_merge_failed)

            worker.finished.connect(thread.quit)
            worker.failed.connect(thread.quit)
            worker.finished.connect(worker.deleteLater)
            worker.failed.connect(worker.deleteLater)
            thread.finished.connect(thread.deleteLater)
            try:
                _ACTIVE_THREADS.add(thread)
                thread.finished.connect(lambda: _ACTIVE_THREADS.discard(thread))
            except Exception:
                pass

            self._chain_pricetags_merge_thread = thread
            self._chain_pricetags_merge_worker = worker

            thread.start()
        except Exception as e:
            self._chain_pricetags_merge_loading = False
            try:
                if hasattr(self, "btnMergeChainPricetags"):
                    self.btnMergeChainPricetags.setEnabled(True)
            except Exception:
                pass
            raise e

    def _on_chain_pricetags_merge_finished(self, seq: int, output_path: str) -> None:
        try:
            self._chain_pricetags_merge_loading = False
            self._chain_pricetags_merge_thread = None
            self._chain_pricetags_merge_worker = None

            # устаревший результат — не применяем
            if int(seq) != int(getattr(self, "_chain_pricetags_merge_seq", 0)):
                return

            try:
                if hasattr(self, "btnMergeChainPricetags"):
                    self.btnMergeChainPricetags.setEnabled(True)
            except Exception:
                pass

            try:
                self.lblPricelistInfo.setText("Готово: файл создан")
            except Exception:
                pass

            try:
                QMessageBox.information(self, "Готово", f"Файл создан:\n{output_path}")
            except Exception:
                pass

            try:
                if output_path:
                    QDesktopServices.openUrl(QUrl.fromLocalFile(output_path))
            except Exception:
                pass

        except Exception:
            pass

    def _on_chain_pricetags_merge_failed(self, seq: int, error: str) -> None:
        try:
            self._chain_pricetags_merge_loading = False
            self._chain_pricetags_merge_thread = None
            self._chain_pricetags_merge_worker = None

            try:
                if hasattr(self, "btnMergeChainPricetags"):
                    self.btnMergeChainPricetags.setEnabled(True)
            except Exception:
                pass

            try:
                self.lblPricelistInfo.setText(f"Ошибка: {error}")
            except Exception:
                pass

            # показываем только если это актуальный запуск
            if int(seq) == int(getattr(self, "_chain_pricetags_merge_seq", 0)):
                try:
                    QMessageBox.critical(self, "Ошибка", str(error))
                except Exception:
                    pass
        except Exception:
            pass


def main():
    # Режим для Планировщика Windows: выполнить расписание и выйти
    if "--run-open-schedule" in sys.argv:
        try:
            QCoreApplication(sys.argv)
        except Exception:
            pass
        try:
            run_open_schedule_due_silent()
        except Exception:
            pass
        return

    app = QApplication(sys.argv)

    # Единый шрифт во всём приложении
    try:
        font = QFont()
        font.setFamily(str(getattr(AppStyles, "DEFAULT_FONT_FAMILY", "Segoe UI")))
        font.setPointSize(int(getattr(AppStyles, "DEFAULT_FONT_SIZE", 14)))
        app.setFont(font)
    except Exception:
        pass

    w = MainWindow()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()

