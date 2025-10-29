#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Модуль для заполнения шаблона меню данными из другого файла меню
"""

import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import re
import tempfile
import xlrd

from app.services.dish_extractor import (
    extract_categorized_dishes_from_menu,
    extract_date_from_menu,
    extract_dishes_from_excel,
    extract_dishes_from_excel_rows_with_stop,
    DishItem,
)
from app.services.comparator import _find_category_ranges, _extract_dishes_from_multiple_columns, read_cell_values, normalize_dish


def convert_xls_to_xlsx(xls_path: str) -> str:
    """Конвертирует старый формат .xls в .xlsx и возвращает путь к временному файлу"""
    try:
        # Читаем .xls файл с помощью xlrd
        xls_book = xlrd.open_workbook(xls_path, formatting_info=False)
        
        # Создаем временный .xlsx файл
        temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.xlsx', delete=False)
        temp_path = temp_file.name
        temp_file.close()
        
        # Создаем новый .xlsx workbook
        xlsx_book = openpyxl.Workbook()
        xlsx_book.remove(xlsx_book.active)  # Удаляем стандартный лист
        
        # Копируем каждый лист
        for sheet_idx in range(xls_book.nsheets):
            xls_sheet = xls_book.sheet_by_index(sheet_idx)
            xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)
            
            # Копируем все ячейки
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row_idx, col_idx)
                    # xlrd использует 0-based индексы, openpyxl - 1-based
                    xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
        
        # Сохраняем временный файл
        xlsx_book.save(temp_path)
        return temp_path
        
    except Exception as e:
        raise Exception(f"Ошибка конвертации .xls в .xlsx: {str(e)}")


class MenuTemplateFiller:
    """Заполняет шаблон меню данными из другого файла"""
    
    def __init__(self):
        self.categories_mapping = {
            'завтрак': 'завтраки',
            'салат': 'холодные закуски и салаты', 
            'первое': 'первые блюда',
            'мясо': 'блюда из мяса',
            'курица': 'блюда из курицы',
            'птица': 'блюда из птицы',  # Добавляем птицу
            'рыба': 'блюда из рыбы',
            'гарнир': 'гарниры'
        }
    
    def extract_categorized_dishes(self, menu_path: str) -> Dict[str, List[str]]:
        """Извлекает блюда по категориям из файла меню (используем уже готовую логику)"""
        return extract_categorized_dishes_from_menu(menu_path)
    
    def extract_categorized_dishes_advanced(self, menu_path: str) -> Dict[str, List[str]]:
        """Улучшенное извлечение блюд по категориям используя логику из comparator.py"""
        try:
            # Определяем какой лист использовать
            from app.services.comparator import get_sheet_names
            sheets = get_sheet_names(menu_path)
            if not sheets:
                return {}
            
            # Ищем лист с приоритетом "касс" или первый доступный
            sheet_name = None
            for name in sheets:
                if 'касс' in name.lower():
                    sheet_name = name
                    break
            if not sheet_name:
                sheet_name = sheets[0]
            
            # Читаем данные из файла
            values = read_cell_values(menu_path, sheet_name)
            
            # Определяем синонимы для поиска категорий (как в comparator.py)
            synonyms_map = {
                'завтрак': 'завтраки',
                'салат': 'салаты и холодные закуски', 
                'холодн': 'салаты и холодные закуски',
                'закуск': 'салаты и холодные закуски',
                'перв': 'первые блюда',
                'мяс': 'блюда из мяса',
                'птиц': 'блюда из птицы',
                'курин': 'блюда из птицы',
                'рыб': 'блюда из рыбы',
                'гарнир': 'гарниры'
            }
            
            # Находим диапазоны категорий
            ranges = _find_category_ranges(values, synonyms_map)
            
            # Извлекаем блюда из каждой категории
            result = {
                'завтрак': [],
                'салат': [], 
                'первое': [],
                'мясо': [],
                'курица': [],
                'птица': [],
                'рыба': [],
                'гарнир': []
            }
            
            # Мапинг категорий из comparator в наши категории
            category_mapping = {
                'завтраки': 'завтрак',
                'салаты и холодные закуски': 'салат',
                'первые блюда': 'первое',
                'блюда из мяса': 'мясо',
                'блюда из птицы': 'птица', 
                'блюда из рыбы': 'рыба',
                'гарниры': 'гарнир'
            }
            
            # Извлекаем блюда для каждой найденной категории
            for comp_category, (start, end) in ranges.items():
                our_category = category_mapping.get(comp_category)
                if our_category:
                    # Извлекаем блюда из столбцов A и D для этого диапазона
                    dishes_set = _extract_dishes_from_multiple_columns(values, start, end, True, ['A', 'D', 'E'])
                    
                    # Фильтруем и очищаем блюда
                    clean_dishes = []
                    for dish in dishes_set:
                        if dish and len(dish.strip()) > 3:
                            # Пропускаем заголовки
                            dish_lower = dish.lower()
                            if not any(header in dish_lower for header in ['блюда', 'салаты', 'гарниры', 'первые', 'вес', 'цена', 'руб']):
                                clean_dishes.append(dish.strip())
                    
                    result[our_category] = clean_dishes
                    print(f"Категория {comp_category} -> {our_category}: найдено {len(clean_dishes)} блюд")
            
            return result
            
        except Exception as e:
            print(f"Ошибка в продвинутом извлечении: {e}")
            # Fallback к старому методу
            return self.extract_categorized_dishes(menu_path)
    
    def extract_dishes_with_details(self, menu_path: str) -> Dict[str, List[Dict[str, str]]]:
        """Извлекает блюда с деталями (название, вес, цена) из файла меню"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(menu_path)
            ws = wb.active
            
            result = {
                'завтрак': [],
                'салат': [], 
                'первое': [],
                'мясо': [],
                'курица': [],
                'птица': [],
                'рыба': [],
                'гарнир': []
            }
            
            # Ищем категории и извлекаем блюда с деталями
            current_category = None
            
            for row in range(1, min(100, ws.max_row + 1)):
                # Проверяем заголовки категорий
                cell_val = ws.cell(row=row, column=1).value
                if cell_val:
                    cell_text = str(cell_val).lower().strip()
                    
                    # Проверяем заголовки категорий с учётом приоритета и контекста
                    # Важно: завтрак должен быть в начале (до строки 15), салаты - после (строки 20+)
                    if 'завтрак' in cell_text and 'салат' not in cell_text and row < 15:
                        current_category = 'завтрак'
                        print(f"Найдена категория завтрак в строке {row}: {cell_val}")
                        continue
                    elif ('салат' in cell_text or 'закуск' in cell_text or 'холодн' in cell_text) and row >= 20:
                        # Это заголовок салатов (после завтраков)
                        if 'салат' in cell_text or ('холодн' in cell_text and 'закуск' in cell_text):
                            current_category = 'салат'
                            print(f"Найдена категория салат в строке {row}: {cell_val}")
                            continue
                    elif 'перв' in cell_text:
                        current_category = 'первое'
                        print(f"Найдена категория первое в строке {row}: {cell_val}")
                        continue
                    elif 'мяс' in cell_text:
                        current_category = 'мясо'
                        print(f"Найдена категория мясо в строке {row}: {cell_val}")
                        continue
                    elif 'птиц' in cell_text or 'курин' in cell_text:
                        current_category = 'птица'
                        print(f"Найдена категория птица в строке {row}: {cell_val}")
                        continue
                    elif 'рыб' in cell_text:
                        current_category = 'рыба'
                        print(f"Найдена категория рыба в строке {row}: {cell_val}")
                        continue
                    elif 'гарнир' in cell_text:
                        current_category = 'гарнир'
                        print(f"Найдена категория гарнир в строке {row}: {cell_val}")
                        continue
                
                # Если нашли категорию, собираем блюда
                if current_category and cell_val and current_category in result:
                    cell_text = str(cell_val).lower().strip()
                    # Проверяем, не является ли это заголовком
                    if not any(header in cell_text for header in ['блюда', 'салаты', 'гарниры', 'первые', 'вес', 'цена', 'руб', 'завтрак']):
                        # Получаем вес и цену из соседних колонок
                        weight = ws.cell(row=row, column=2).value
                        price = ws.cell(row=row, column=3).value
                        
                        dish_info = {
                            'name': str(cell_val).strip(),
                            'weight': str(weight).strip() if weight else '',
                            'price': str(price).strip() if price else ''
                        }
                        
                        result[current_category].append(dish_info)
                        print(f"Добавлено блюдо в {current_category}: {dish_info}")
            
            wb.close()
            
            # Отчет о найденных блюдах
            total = sum(len(dishes) for dishes in result.values())
            print(f"\nИзвлечено блюд с деталями:")
            for cat, dishes in result.items():
                if dishes:
                    print(f"  {cat}: {len(dishes)} блюд")
            print(f"Всего блюд: {total}")
            
            return result
            
        except Exception as e:
            print(f"Ошибка при извлечении блюд с деталями: {e}")
            return {}
    
    def find_column_by_header(self, ws, header_text: str) -> Optional[int]:
        """Находит номер колонки по заголовку"""
        header_variations = {
            'завтраки': ['завтрак', 'завтраки'],
            'холодные закуски и салаты': ['салат', 'холодн', 'закуск'],  # Работает с "САЛАТЫ и ХОЛОДНЫЕ ЗАКУСКИ"
            'первые блюда': ['первые', 'первое'],
            'блюда из мяса': ['мяса', 'мясо', 'мясн'],  # Работает с "БЛЮДА ИЗ МЯСА"
            'блюда из курицы': ['курица', 'курин', 'птица', 'птицы'],
            'блюда из птицы': ['птицы', 'птица', 'курица', 'курин'],  # Работает с "БЛЮДА ИЗ ПТИЦЫ"
            'блюда из рыбы': ['рыбы', 'рыба', 'рыбн'],  # Работает с "БЛЮДА ИЗ РЫБЫ"
            'гарниры': ['гарниры']
        }
        
        # Ищем в первых 50 строках (некоторые заголовки могут быть на 30-й строке)
        for row in range(1, min(51, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val:
                    cell_text = str(cell_val).lower().strip()
                    
                    # Проверяем точное совпадение
                    if header_text.lower() in cell_text:
                        return col
                    
                    # Проверяем вариации
                    if header_text.lower() in header_variations:
                        for variation in header_variations[header_text.lower()]:
                            if variation in cell_text:
                                return col
        
        return None
    
    def find_data_start_row(self, ws, col: int, category: str = None) -> int:
        """Находит строку начала данных в колонке (после заголовков)"""
        # Специальная логика для гарниров - ищем именно заголовок "ГАРНИРЫ"
        if category == 'гарнир':
            for row in range(20, min(50, ws.max_row + 1)):  # Ищем с 20 по 50 строку
                cell_val = ws.cell(row=row, column=col).value
                if cell_val:
                    cell_text = str(cell_val).lower().strip()
                    if 'гарнир' in cell_text and ('гарниры' in cell_text or 'гарнир' == cell_text):
                        print(f"Найден заголовок гарниров в строке {row}: {cell_val}")
                        return row + 1  # Начинаем с следующей строки после заголовка
        
        # Специальная логика для завтраков - ищем только "ЗАВТРАК" в начале файла (до строки 10)
        if category == 'завтрак':
            for row in range(1, min(15, ws.max_row + 1)):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val:
                    cell_text = str(cell_val).lower().strip()
                    # Ищем именно заголовок "ЗАВТРАК", но не "САЛАТЫ"
                    if 'завтрак' in cell_text and 'салат' not in cell_text:
                        print(f"Найден заголовок завтрака в строке {row}: {cell_val}")
                        return row + 1
        
        # Специальная логика для салатов - ищем "САЛАТЫ" после строки 20 (после завтраков)
        if category == 'салат':
            for row in range(20, min(50, ws.max_row + 1)):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val:
                    cell_text = str(cell_val).lower().strip()
                    # Ищем именно заголовок "САЛАТЫ" с "ХОЛОДН" или "ЗАКУСК"
                    if 'салат' in cell_text and ('холодн' in cell_text or 'закуск' in cell_text):
                        print(f"Найден заголовок салатов в строке {row}: {cell_val}")
                        return row + 1
        
        # Для остальных категорий - находим строку с соответствующим заголовком
        header_keywords = {
            'первое': ['первые'],
            'мясо': ['мяса', 'мясо'],
            'курица': ['курица', 'птица'],
            'птица': ['птица', 'курица'],
            'рыба': ['рыба', 'рыбы']
        }
        
        if category in header_keywords:
            keywords = header_keywords[category]
            for row in range(1, min(50, ws.max_row + 1)):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val:
                    cell_text = str(cell_val).lower().strip()
                    if any(keyword in cell_text for keyword in keywords):
                        print(f"Найден заголовок {category} в строке {row}: {cell_val}")
                        return row + 1
        
        # Общий поиск заголовка категории
        header_row = None
        for row in range(1, min(50, ws.max_row + 1)):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val:
                cell_text = str(cell_val).lower().strip()
                # Ищем заголовки категорий
                if any(header in cell_text for header in ['завтрак', 'салат', 'мясо', 'курица', 'птица', 'рыба', 'гарнир', 'первые', 'блюда']):
                    header_row = row
                    break
        
        # Если нашли заголовок, начинаем с следующей строки
        if header_row:
            return header_row + 1
        
        # Иначе ищем строку с "Вес/ед.изм." и начинаем после неё
        for row in range(1, min(50, ws.max_row + 1)):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val and 'вес' in str(cell_val).lower():
                return row + 1
                
        return 7  # По умолчанию начинаем с 7-й строки
    
    def handle_caesar_salad(self, dish_name: str) -> List[str]:
        """Обрабатывает Цезарь салат, разделяя по слэшу (устарело)."""
        if 'цезарь' in dish_name.lower() and '/' in dish_name:
            parts = dish_name.split('/')
            result = []
            base = parts[0].strip()
            for i, part in enumerate(parts):
                if i == 0:
                    result.append(part.strip())
                else:
                    result.append(f"{base.split()[0]} {part.strip()}")
            return result
        return [dish_name]

    def expand_variants_with_details(self, name: str, weight: str, price: str) -> List[Dict[str, str]]:
        """
        Разворачивает запись блюда с вариантами, разделёнными '/', сохраняя соответствие
        для веса и цены, если они также содержат варианты через '/'.

        Примеры:
            name:  "Омлет/Омлет с брокколи", weight: "170г/200г", price: "140/170"
              -> [
                   {name: "Омлет", weight: "170г", price: "140"},
                   {name: "Омлет с брокколи", weight: "200г", price: "170"}
                 ]

            name:  "Яйцо отварное/жареное", weight: "1шт", price: "50/55"
              -> [
                   {name: "Яйцо отварное", weight: "1шт", price: "50"},
                   {name: "Яйцо жареное",  weight: "1шт", price: "55"}
                 ]
        """
        if not name or '/' not in name:
            return [{"name": (name or '').strip(), "weight": (weight or '').strip(), "price": (price or '').strip()}]

        name_parts = [p.strip() for p in name.split('/') if p.strip()]
        w_parts = [p.strip() for p in (weight or '').split('/') if p.strip()] if weight else []
        p_parts = [p.strip() for p in (price or '').split('/') if p.strip()] if price else []

        # Восстановление названий для случаев типа "Яйцо отварное/жареное"
        def rebuild_names(parts: List[str]) -> List[str]:
            if len(parts) < 2:
                return parts
            first = parts[0]
            if ' ' not in first:
                # Первый вариант — одно слово, второй может быть полным названием
                return parts
            base = first.rsplit(' ', 1)[0]  # всё до последнего слова как общая основа
            rebuilt = [first]
            for p in parts[1:]:
                if ' ' in p:
                    rebuilt.append(p)
                else:
                    rebuilt.append(f"{base} {p}")
            return rebuilt

        name_parts = rebuild_names(name_parts)

        variants = []
        n = len(name_parts)
        for i in range(n):
            ni = name_parts[i]
            wi = w_parts[i] if i < len(w_parts) else (weight or '').strip()
            pi = p_parts[i] if i < len(p_parts) else (price or '').strip()
            variants.append({"name": ni, "weight": wi, "price": pi})
        return variants
    
    def find_category_end_row(self, ws, col: int, start_row: int, category: str) -> int:
        """Находит конечную строку для категории, ища следующий заголовок"""
        # Особая логика для гарниров - ищем до "НАПИТКИ"
        if category == 'гарнир':
            for row in range(start_row, min(200, ws.max_row + 1)):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val and 'напит' in str(cell_val).lower():
                    return row - 1  # Возвращаем последнюю строку перед "Напитки"
        
        # Особая логика для салатов - ищем до "ПЕРВЫЕ БЛЮДА" или других заголовков
        if category == 'салат':
            for row in range(start_row, min(200, ws.max_row + 1)):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val:
                    cell_text = str(cell_val).lower().strip()
                    # Останавливаемся перед следующими категориями
                    if any(header in cell_text for header in ['первые', 'напит', 'хлеб', 'блюда из']):
                        return row - 1
        
        # Для остальных категорий ищем до следующего заголовка
        for row in range(start_row, min(200, ws.max_row + 1)):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val:
                cell_text = str(cell_val).lower().strip()
                if any(header in cell_text for header in ['блюда', 'салаты', 'гарниры', 'первые', 'напит', 'хлеб']):
                    return row - 1  # Возвращаем последнюю строку перед новым заголовком
        
        # Если не нашли следующий заголовок, возвращаем конец листа
        return ws.max_row

    def find_end_row_until_salads(self, ws, start_row: int) -> int:
        """Ищет строку конца блока завтраков до заголовка "САЛАТЫ И ХОЛОДНЫЕ ЗАКУСКИ".
        Склеивает текст строки по всем колонкам и проверяет наличие ключевых слов.
        Возвращает номер строки-1 перед заголовком; если заголовок не найден — возвращает конец листа.
        """
        def row_text(r: int) -> str:
            parts = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None and str(v).strip():
                    parts.append(str(v).strip())
            return ' '.join(parts).lower()

        for row in range(start_row, min(200, ws.max_row + 1)):
            txt = row_text(row)
            # Заголовок может быть записан по-разному, проверяем сочетания
            if ('салат' in txt) and ('холодн' in txt or 'закуск' in txt):
                return row - 1
        return ws.max_row
    
    def fill_template_column(self, ws, col: int, dishes: List[str], start_row: int) -> int:
        """Заполняет колонку блюдами, начиная с указанной строки"""
        current_row = start_row
        filled_count = 0
        
        for dish in dishes:
            # Обрабатываем Цезарь салат
            dish_variants = self.handle_caesar_salad(dish)
            
            for variant in dish_variants:
                if current_row > ws.max_row:
                    break
                    
                # Получаем ячейку
                cell = ws.cell(row=current_row, column=col)
                
                # Проверяем, является ли ячейка объединенной
                if hasattr(cell, 'coordinate'):
                    # Проверяем, есть ли эта ячейка в списке объединенных ячеек
                    is_merged = False
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            is_merged = True
                            break
                    
                    # Пропускаем объединенные ячейки, кроме главной
                    if is_merged and cell.__class__.__name__ == 'MergedCell':
                        current_row += 1
                        continue
                
                # Заполняем только пустые ячейки
                current_val = cell.value
                if not current_val or str(current_val).strip() == '':
                    try:
                        cell.value = variant
                        filled_count += 1
                    except AttributeError:
                        # Если не можем записать в эту ячейку, пропускаем
                        pass
                
                current_row += 1
        
        return filled_count
    
    def fill_template_column_limited(self, ws, col: int, dishes: List[str], start_row: int, end_row: int, category: str) -> int:
        """Заполняет колонку блюдами в ограниченном диапазоне строк"""
        current_row = start_row
        filled_count = 0
        
        for dish in dishes:
            # Обрабатываем Цезарь салат
            dish_variants = self.handle_caesar_salad(dish)
            
            for variant in dish_variants:
                if current_row > end_row:  # Ограничиваем по концу диапазона
                    print(f"Превышен лимит для {category}: конец на строке {end_row}")
                    break
                    
                # Получаем ячейку
                cell = ws.cell(row=current_row, column=col)
                
                # Проверяем, является ли ячейка объединенной
                if hasattr(cell, 'coordinate'):
                    # Проверяем, есть ли эта ячейка в списке объединенных ячеек
                    is_merged = False
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            is_merged = True
                            break
                    
                    # Пропускаем объединенные ячейки, кроме главной
                    if is_merged and cell.__class__.__name__ == 'MergedCell':
                        current_row += 1
                        continue
                
                # Заполняем только пустые ячейки
                current_val = cell.value
                if not current_val or str(current_val).strip() == '':
                    try:
                        cell.value = variant
                        filled_count += 1
                        print(f"Заполнено {category} в строку {current_row}: {variant}")
                    except AttributeError:
                        # Если не можем записать в эту ячейку, пропускаем
                        pass
                else:
                    print(f"Ячейка {current_row}:{col} для {category} уже заполнена: {current_val}")
                
                current_row += 1
            
            # Прерываем, если превысили лимит
            if current_row > end_row:
                break
        
        return filled_count
    
    def fill_template_with_details(self, ws, name_col: int, weight_col: int, price_col: int,
                                  dishes_with_details: List[Dict[str, str]], start_row: int) -> int:
        """Заполняет шаблон блюдами с деталями (название, вес, цена) через общий движок вставки."""
        from app.services.excel_inserter import fill_cells_sequential, TargetColumns
        from app.services.dish_extractor import DishItem
        from math import inf

        items_raw: List[Dict[str, str]] = []
        for d in dishes_with_details:
            items_raw.extend(self.expand_variants_with_details(d.get('name', ''), d.get('weight', ''), d.get('price', '')))
        items = [DishItem(name=dr.get('name',''), weight=dr.get('weight',''), price=dr.get('price','')) for dr in items_raw]
        # До конца листа
        stop_row = ws.max_row + 1
        return fill_cells_sequential(
            ws,
            start_row=start_row,
            stop_row=stop_row,
            columns=TargetColumns(name_col=name_col, weight_col=weight_col, price_col=price_col),
            dishes=items,
            replace_only_empty=True,
        )
    
    def fill_template_with_details_limited(self, ws, name_col: int, weight_col: int, price_col: int,
                                          dishes_with_details: List[Dict[str, str]], start_row: int, end_row: int, category: str) -> int:
        """Заполняет шаблон блюдами с деталями в ограниченном диапазоне через общий движок вставки."""
        from app.services.excel_inserter import fill_cells_sequential, TargetColumns
        from app.services.dish_extractor import DishItem

        items_raw: List[Dict[str, str]] = []
        for d in dishes_with_details:
            items_raw.extend(self.expand_variants_with_details(d.get('name', ''), d.get('weight', ''), d.get('price', '')))
        items = [DishItem(name=dr.get('name',''), weight=dr.get('weight',''), price=dr.get('price','')) for dr in items_raw]
        # end_row включителен в прежней логике, общий движок ожидает исключающую верхнюю границу
        stop_row = end_row + 1
        return fill_cells_sequential(
            ws,
            start_row=start_row,
            stop_row=stop_row,
            columns=TargetColumns(name_col=name_col, weight_col=weight_col, price_col=price_col),
            dishes=items,
            replace_only_empty=True,
        )
    
    def extract_date_from_menu(self, menu_path: str) -> Optional[datetime]:
        """Извлекает дату из файла меню (используем уже готовую логику)"""
        return extract_date_from_menu(menu_path)
    
    def update_template_date(self, ws, menu_date: Optional[datetime], include_weekday: bool = True):
        """Обновляет дату (и при необходимости день недели) в верхних строках шаблона.
        По требованиям: день недели пишется в B2, а «число» (день месяца) — в B3.
        Если эти ячейки недоступны, используем поиск по верхнему блоку.
        """
        if not menu_date:
            menu_date = datetime.now()
        russian_months = {
            1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля', 5: 'мая', 6: 'июня',
            7: 'июля', 8: 'августа', 9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'
        }
        weekday_names = {
            0: 'понедельник', 1: 'вторник', 2: 'среда', 3: 'четверг', 4: 'пятница', 5: 'суббота', 6: 'воскресенье'
        }
        # 1) Пытаемся записать строго в B2 (день недели) и B3 (число)
        try:
            ws.cell(row=2, column=2).value = weekday_names.get(menu_date.weekday(), '')  # B2: день недели
            ws.cell(row=3, column=2).value = f"{menu_date.day} {russian_months.get(menu_date.month, 'сентября')}"  # B3: число и месяц
            return
        except Exception:
            pass
        # 2) Фолбек: перезапись ближайшей ячейки с месяцем в верхнем блоке
        base = f"{menu_date.day} {russian_months.get(menu_date.month, 'сентября')}"
        full = f"{base} - {weekday_names.get(menu_date.weekday(), '')}" if include_weekday else base
        for row in range(1, min(6, ws.max_row + 1)):
            for col in range(1, min(11, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                cell_val = cell.value
                if not cell_val:
                    continue
                cell_text = str(cell_val).lower()
                if any(month in cell_text for month in ['январ', 'феврал', 'март', 'апрел', 'май', 'июн',
                                                        'июл', 'август', 'сентябр', 'октябр', 'ноябр', 'декабр']):
                    if cell.__class__.__name__ == 'MergedCell':
                        continue
                    try:
                        cell.value = full
                        return
                    except AttributeError:
                        continue
    
    def fill_menu_template(self, template_path: str, source_menu_path: str, output_path: str) -> Tuple[bool, str]:
        """
        Заполняет шаблон меню данными из исходного файла
        
        Args:
            template_path: Путь к шаблону меню
            source_menu_path: Путь к файлу-источнику данных  
            output_path: Путь к выходному файлу
            
        Returns:
            (success, message): Кортеж с результатом операции
        """
        try:
            # Проверяем существование файлов
            if not Path(template_path).exists():
                return False, f"Шаблон не найден: {template_path}"
            
            if not Path(source_menu_path).exists():
                return False, f"Исходный файл не найден: {source_menu_path}"
            
            # Извлекаем данные из исходного файла (с продвинутой логикой)
            print(f"Извлекаем данные из файла: {source_menu_path}")
            categories = self.extract_categorized_dishes_advanced(source_menu_path)
            
            # Отладочный вывод
            total_dishes = sum(len(dishes) for dishes in categories.values())
            print(f"Извлечено блюд по категориям:")
            for cat, dishes in categories.items():
                print(f"  {cat}: {len(dishes)} блюд")
            print(f"Всего блюд: {total_dishes}")
            
            if total_dishes == 0:
                return False, "Не удалось извлечь блюда из исходного файла"
            
            # Извлекаем дату из исходного файла
            menu_date = self.extract_date_from_menu(source_menu_path)
            
            # Открываем шаблон
            wb = openpyxl.load_workbook(template_path)
            
            # Находим лист "Касса" для заполнения данными (там хранятся сами блюда)
            ws = None
            for sheet in wb.worksheets:
                if 'касс' in sheet.title.lower():
                    ws = sheet
                    break
            
            if not ws:
                # Fallback к основному листу
                ws = wb.active
                print("Предупреждение: Лист 'Касса' не найден, используем основной лист")
            else:
                print(f"Используем лист: {ws.title}")
            
            # Обновляем дату в шаблоне
            self.update_template_date(ws, menu_date)
            
            # Заполняем колонки данными
            total_filled = 0
            categories_filled = 0
            
            for category, dishes in categories.items():
                if not dishes:
                    continue
                
                # Находим соответствующую колонку в шаблоне
                template_header = self.categories_mapping.get(category)
                if not template_header:
                    print(f"Предупреждение: Не найдено соответствие для категории '{category}'")
                    continue
                
                col = self.find_column_by_header(ws, template_header)
                # Требование: завтраки всегда пишем в колонку A (1)
                if category == 'завтрак':
                    col = 1
                if not col:
                    print(f"Предупреждение: Не найдена колонка для '{template_header}'")
                    continue
                
                # Находим начальную строку для данных
                start_row = self.find_data_start_row(ws, col, category)
                
                # Особая логика ограничений по диапазону
                if category == 'гарнир':
                    end_row = self.find_category_end_row(ws, col, start_row, category)
                    print(f"Гарниры: ограничиваем диапазон {start_row}-{end_row}")
                    filled_count = self.fill_template_column_limited(ws, col, dishes, start_row, end_row, category)
                elif category == 'завтрак':
                    # Завтраки должны попадать строго между заголовками "ЗАВТРАКИ" и "САЛАТЫ И ХОЛОДНЫЕ ЗАКУСКИ"
                    end_row = self.find_end_row_until_salads(ws, start_row)
                    print(f"Завтраки: ограничиваем диапазон {start_row}-{end_row}")
                    filled_count = self.fill_template_column_limited(ws, col, dishes, start_row, end_row, category)
                else:
                    # Обычное заполнение для остальных категорий
                    filled_count = self.fill_template_column(ws, col, dishes, start_row)
                total_filled += filled_count
                categories_filled += 1
                
                print(f"Категория '{category}' -> колонка {col}: добавлено {filled_count} блюд")
            
            # Сохраняем результат
            wb.save(output_path)
            
            date_str = menu_date.strftime("%d.%m.%Y") if menu_date else "текущая дата"
            message = f"Шаблон меню заполнен для даты {date_str}\n"
            message += f"Заполнено категорий: {categories_filled}\n"
            message += f"Всего добавлено блюд: {total_filled}"
            
            return True, message
            
        except Exception as e:
            return False, f"Ошибка при заполнении шаблона: {str(e)}"


    def copy_from_source_sheets_to_template(self, template_path: str, source_menu_path: str, output_path: str,
                                            breakfast_range: Tuple[int, int] = (7, 27),
                                            lunch_range: Optional[Tuple[int, int]] = None) -> Tuple[bool, str]:
        """
        Копирует данные из источника (листы «Завтрак», «Обед») в одноимённые листы шаблона.
        
        Маппинг колонок (источник -> шаблон):
          - источник: col0=вес/выход, col1=название, col2=цена
          - шаблон:  A=название,       B=вес/выход, C=цена
        
        Диапазоны:
          - «Завтрак»: A7..A27; B7..B27; C7..C27
          - «Обед»:    те же A/B/C, начиная со стартовой строки списка (если не указано иначе — как у «Завтрак»)
        
        Правила:
          - перед заполнением очищаем целевой диапазон (значения), оформление не трогаем
          - переносим блюда сверху вниз до конца диапазона
          - пустые/битые строки пропускаем; если цена отсутствует — C оставляем пустой
          - листы «Касса», «Гц», «Хц», «Раздача» не трогаем
        """
        import openpyxl
        import logging
        from app.services.dish_extractor import DishItem
        logger = logging.getLogger("menu.template")

        def sanitize_price(raw) -> str:
            if raw is None:
                return ""
            s = str(raw)
            parts = [p.strip() for p in s.split('/')]
            out = []
            import re as _re
            for p in parts:
                m = _re.search(r"(\d+(?:[\.,]\d{1,2})?)", p)
                if not m:
                    continue
                num = m.group(1).replace('.', ',')
                out.append(num)
            return "/".join(out)

        def find_sheet_case_insensitive(wb, keyword: str):
            for sh in wb.worksheets:
                if keyword.lower() in sh.title.lower():
                    return sh
            return None

        def clear_range(ws, start_row: int, end_row: int, cols: Tuple[int, int, int]):
            c1, c2, c3 = cols
            for r in range(start_row, end_row + 1):
                for c in (c1, c2, c3):
                    cell = ws.cell(row=r, column=c)
                    try:
                        cell.value = None
                    except AttributeError:
                        # пропускаем MergedCell
                        pass

        def write_items(ws, items: List[DishItem], start_row: int, end_row: int, cols: Tuple[int, int, int]) -> int:
            c_name, c_weight, c_price = cols
            r = start_row
            written = 0
            for it in items:
                if r > end_row:
                    break
                name = (it.name or '').strip()
                if not name:
                    continue
                weight = (it.weight or '').strip()
                price = (it.price or '').strip()
                # Пишем, не трогая оформление; пропускаем MergedCell-ячейки
                try:
                    ws.cell(row=r, column=c_name).value = name
                except AttributeError:
                    pass
                try:
                    ws.cell(row=r, column=c_weight).value = weight
                except AttributeError:
                    pass
                try:
                    ws.cell(row=r, column=c_price).value = price
                except AttributeError:
                    pass
                r += 1
                written += 1
            return written

        try:
            if not Path(template_path).exists():
                return False, f"Шаблон не найден: {template_path}"
            if not Path(source_menu_path).exists():
                return False, f"Источник не найден: {source_menu_path}"

            # Открываем книги
            src_wb = openpyxl.load_workbook(source_menu_path, data_only=True)
            tpl_wb = openpyxl.load_workbook(template_path)

            # Готовим список задач: (имя_листа, диапозон)
            br_start, br_end = breakfast_range
            ln_start, ln_end = lunch_range if lunch_range else breakfast_range
            tasks = [
                ("завтрак", br_start, br_end),
                ("обед", ln_start, ln_end),
            ]

            total_written = 0
            for key, r_start, r_end in tasks:
                # Ищем листы (источник/шаблон)
                src_ws = find_sheet_case_insensitive(src_wb, key)
                tpl_ws = find_sheet_case_insensitive(tpl_wb, key)
                if src_ws is None or tpl_ws is None:
                    logger.info(f"Лист '{key}' отсутствует в одном из файлов, пропускаю")
                    continue
                if tpl_ws.title.lower() in ["касса", "гц", "хц", "раздача"]:
                    logger.info(f"Лист '{tpl_ws.title}' нельзя менять, пропускаю")
                    continue

                # Читаем данные из источника: col0=вес, col1=название, col2=цена
                items: List[DishItem] = []
                header_tokens = ['завтрак', 'вес', 'ед.изм', 'ед. изм', 'цена', 'руб']
                for r in range(1, src_ws.max_row + 1):
                    w = src_ws.cell(row=r, column=1).value  # вес
                    n = src_ws.cell(row=r, column=2).value  # название
                    p = src_ws.cell(row=r, column=3).value  # цена
                    n_txt = (str(n).strip() if n else '')
                    if not n_txt:
                        continue
                    # Пропускаем заголовки/шапки
                    n_low = n_txt.lower()
                    w_low = (str(w).lower() if isinstance(w, str) else str(w).lower()) if w not in (None, '') else ''
                    p_low = (str(p).lower() if isinstance(p, str) else str(p).lower()) if p not in (None, '') else ''
                    if any(tok in n_low for tok in header_tokens) or any(tok in w_low for tok in header_tokens) or any(tok in p_low for tok in header_tokens):
                        continue
                    w_txt = str(w).strip() if w else ''
                    p_txt = sanitize_price(p)
                    items.append(DishItem(name=n_txt, weight=w_txt, price=p_txt))

                # Очищаем диапазон в шаблоне
                clear_range(tpl_ws, r_start, r_end, (1, 2, 3))

                # Записываем сверху вниз
                written = write_items(tpl_ws, items, r_start, r_end, (1, 2, 3))
                total_written += written

            # Сохраняем результат как новый файл
            tpl_wb.save(output_path)
            return True, f"Скопировано строк: {total_written}"
        except Exception as e:
            return False, f"Ошибка при копировании из источника в шаблон: {str(e)}"

    def fill_menu_template_with_details(self, template_path: str, source_menu_path: str, output_path: str) -> Tuple[bool, str]:
        """
        Заполняет шаблон меню с деталями (название, вес, цена)
        
        Args:
            template_path: Путь к шаблону меню
            source_menu_path: Путь к файлу-источнику данных  
            output_path: Путь к выходному файлу
            
        Returns:
            (success, message): Кортеж с результатом операции
        """
        try:
            # Проверяем существование файлов
            if not Path(template_path).exists():
                return False, f"Шаблон не найден: {template_path}"
            
            if not Path(source_menu_path).exists():
                return False, f"Исходный файл не найден: {source_menu_path}"
            
            # Извлекаем данные с деталями из исходного файла
            print(f"Извлекаем данные с деталями из файла: {source_menu_path}")
            categories_detailed = self.extract_dishes_with_details(source_menu_path)
            
            # Отладочный вывод
            total_dishes = sum(len(dishes) for dishes in categories_detailed.values())
            if total_dishes == 0:
                return False, "Не удалось извлечь блюда с деталями из исходного файла"
            
            # Извлекаем дату из исходного файла
            menu_date = self.extract_date_from_menu(source_menu_path)
            
            # Открываем шаблон
            wb = openpyxl.load_workbook(template_path)
            
            # Находим лист "Касса" для заполнения данными (там хранятся сами блюда)
            ws = None
            for sheet in wb.worksheets:
                if 'касс' in sheet.title.lower():
                    ws = sheet
                    break
            
            if not ws:
                ws = wb.active
                print("Предупреждение: Лист 'Касса' не найден, используем основной лист")
            else:
                print(f"Используем лист: {ws.title}")
            
            # Обновляем дату в шаблоне
            self.update_template_date(ws, menu_date)
            
            # Определяем мапинг колонок для разных категорий
            column_mapping = {
                # Левая сторона: завтраки, салаты
                'завтрак': {'name': 1, 'weight': 2, 'price': 3},
                'салат': {'name': 1, 'weight': 2, 'price': 3},
                # Правая сторона: первые, мясо, птица, рыба, гарниры
                'первое': {'name': 4, 'weight': 5, 'price': 6},
                'мясо': {'name': 4, 'weight': 5, 'price': 6},
                'птица': {'name': 4, 'weight': 5, 'price': 6},
                'рыба': {'name': 4, 'weight': 5, 'price': 6},
                'гарнир': {'name': 4, 'weight': 5, 'price': 6}
            }
            
            # Заполняем колонки данными
            total_filled = 0
            categories_filled = 0
            
            for category, dishes_details in categories_detailed.items():
                if not dishes_details:
                    continue
                
                # Получаем колонки для этой категории
                if category not in column_mapping:
                    print(f"Предупреждение: Не найдено соответствие колонок для категории '{category}'")
                    continue
                
                cols = column_mapping[category]
                name_col, weight_col, price_col = cols['name'], cols['weight'], cols['price']
                
                # Находим начальную строку для данных
                start_row = self.find_data_start_row(ws, name_col, category)
                
                # Особая логика для гарниров и завтраков — ограничиваем диапазон по следующему заголовку
                if category == 'гарнир':
                    end_row = self.find_category_end_row(ws, name_col, start_row, category)
                    print(f"Гарниры: ограничиваем диапазон {start_row}-{end_row}")
                    filled_count = self.fill_template_with_details_limited(
                        ws, name_col, weight_col, price_col, dishes_details, start_row, end_row, category
                    )
                elif category == 'завтрак':
                    end_row = self.find_end_row_until_salads(ws, start_row)
                    print(f"Завтраки: ограничиваем диапазон {start_row}-{end_row}")
                    filled_count = self.fill_template_with_details_limited(
                        ws, name_col, weight_col, price_col, dishes_details, start_row, end_row, category
                    )
                else:
                    # Обычное заполнение для остальных категорий
                    filled_count = self.fill_template_with_details(
                        ws, name_col, weight_col, price_col, dishes_details, start_row
                    )
                
                total_filled += filled_count
                categories_filled += 1
                
                print(f"Категория '{category}' -> колонки {name_col}/{weight_col}/{price_col}: добавлено {filled_count} блюд")
            
            # Сохраняем результат
            wb.save(output_path)
            
            date_str = menu_date.strftime("%d.%m.%Y") if menu_date else "текущая дата"
            message = f"Шаблон меню заполнен (с деталями) для даты {date_str}\n"
            message += f"Заполнено категорий: {categories_filled}\n"
            message += f"Всего добавлено блюд: {total_filled}"
            
            return True, message
        except Exception as e:
            return False, f"Ошибка при заполнении шаблона с деталями: {str(e)}"

    def fill_menu_template_fixed_ranges(
        self,
        template_path: str,
        source_menu_path: str,
        output_path: str,
        breakfast_range: Tuple[int, int] = (7, 27),
        soups_range: Tuple[int, int] = (6, 10),
        meat_range: Tuple[int, int] = (12, 17),
    ) -> Tuple[bool, str]:
        """
        Заполняет шаблон в фиксированные диапазоны:
        - Завтраки: A/B/C, строки A7..A27
        - Первые блюда (супы): D/E/F, строки D6..D10
        - Мясные блюда: D/E/F, строки D12..D17
        D11 считается заголовком и не заполняется.
        """
        try:
            from app.services.excel_inserter import fill_cells_sequential, TargetColumns
            import openpyxl
            from pathlib import Path

            if not Path(template_path).exists():
                return False, f"Шаблон не найден: {template_path}"
            if not Path(source_menu_path).exists():
                return False, f"Исходный файл не найден: {source_menu_path}"

            # Логгер для шаблона
            logger = None  # логирование отключено

            # Извлекаем дату для обновления в шаблоне
            menu_date = self.extract_date_from_menu(source_menu_path)

            # Извлекаем блюда
            # Завтраки — читаем до салатов/холодных закусок (детально)
            breakfasts: List[DishItem] = []
            try:
                breakfasts = extract_dishes_from_excel_rows_with_stop(
                    source_menu_path,
                    ["ЗАВТРАК"],
                    ["САЛАТЫ", "ХОЛОДНЫЕ", "ЗАКУСКИ"],
                )
            except Exception as e:
                logger.warning(f"Не удалось извлечь завтраки: {e}")
                breakfasts = []
            # Fallback: берём детальные завтраки из активного листа (если построчно не нашлось)
            if not breakfasts:
                try:
                    details = self.extract_dishes_with_details(source_menu_path)
                    br = details.get('завтрак', []) if isinstance(details, dict) else []
                    breakfasts = [DishItem(name=d.get('name',''), weight=d.get('weight',''), price=d.get('price','')) for d in br if d.get('name')]
                    logger.info(f"Завтраки (fallback по деталям): {len(breakfasts)}")
                except Exception as e:
                    logger.warning(f"Fallback завтраков не удался: {e}")

            # Первые блюда (супы) — детально
            soups: List[DishItem] = []
            try:
                soups = extract_dishes_from_excel(source_menu_path, ["ПЕРВЫЕ БЛЮДА", "ПЕРВЫЕ"])
            except Exception as e:
                logger.warning(f"Не удалось извлечь первые блюда: {e}")
                soups = []

            # Мясные блюда — детально
            meats: List[DishItem] = []
            try:
                meats = extract_dishes_from_excel(source_menu_path, ["БЛЮДА ИЗ МЯСА", "МЯСНЫЕ БЛЮДА"])
            except Exception as e:
                logger.warning(f"Не удалось извлечь мясные блюда: {e}")
                meats = []

            # Блюда из птицы — детально
            logger.info("Извлечение (птица): функция=extract_dishes_from_excel, keywords=['БЛЮДА ИЗ ПТИЦЫ','БЛЮДА ИЗ КУРИЦЫ','КУРИНЫЕ БЛЮДА']")
            poultry: List[DishItem] = []
            try:
                poultry = extract_dishes_from_excel(source_menu_path, ["БЛЮДА ИЗ ПТИЦЫ", "БЛЮДА ИЗ КУРИЦЫ", "КУРИНЫЕ БЛЮДА"])
                if logger.isEnabledFor(logging.DEBUG):
                    logger.debug("Птица (raw): " + ", ".join([getattr(x, 'name', '') for x in poultry]))
                logger.info(f"Найдено блюд из птицы (raw): {len(poultry)}")
            except Exception as e:
                logger.warning(f"Не удалось извлечь блюда из птицы: {e}")
                poultry = []

            # Рыба — детально
            fish: List[DishItem] = []
            try:
                fish = extract_dishes_from_excel(source_menu_path, ["БЛЮДА ИЗ РЫБЫ", "РЫБНЫЕ БЛЮДА"])
            except Exception as e:
                logger.warning(f"Не удалось извлечь рыбные блюда: {e}")
                fish = []

            # Гарниры — детально
            garnirs: List[DishItem] = []
            try:
                garnirs = extract_dishes_from_excel(source_menu_path, ["ГАРНИРЫ", "ГАРНИР"])
            except Exception as e:
                logger.warning(f"Не удалось извлечь гарниры: {e}")
                garnirs = []

            # Салаты — детально (построчно до следующей категории)
            salads: List[DishItem] = []
            try:
                salads = extract_dishes_from_excel_rows_with_stop(
                    source_menu_path,
                    ["САЛАТ"],
                    ["ПЕРВЫЕ", "БЛЮДА ИЗ", "ГАРНИР", "НАПИТ"]
                )
            except Exception as e:
                logger.warning(f"Не удалось извлечь салаты: {e}")
                salads = []
            # Fallback по деталям, если не нашли построчно
            if not salads:
                try:
                    details = self.extract_dishes_with_details(source_menu_path)
                    sd = details.get('салат', []) if isinstance(details, dict) else []
                    salads = [DishItem(name=d.get('name',''), weight=d.get('weight',''), price=d.get('price','')) for d in sd if d.get('name')]
                    logger.info(f"Салаты (fallback по деталям): {len(salads)}")
                except Exception as e:
                    logger.warning(f"Fallback салатов не удался: {e}")

            # Выравниваем категории по логике бракеражного журнала
            try:
                from app.reports.brokerage_journal import BrokerageJournalGenerator
                gen = BrokerageJournalGenerator()
                cat_names = gen.extract_categorized_dishes(source_menu_path)
            except Exception as e:
                logger.warning(f"Не удалось получить категории из бракеражного журнала: {e}")
                cat_names = {}

            def _norm(s: str) -> str:
                return ' '.join(str(s).lower().replace('ё', 'е').split())

            def _tokenize(s: str) -> List[str]:
                s = _norm(s)
                # Разбиваем по пробелам и символам '/','-','(',')',','
                for ch in ['/','-','(',')',',',';']:
                    s = s.replace(ch, ' ')
                toks = [t for t in s.split() if t]
                return toks

            def _best_match(detailed: List[DishItem], name: str) -> Optional[DishItem]:
                # Точное совпадение
                norm = _norm(name)
                by_norm = { _norm(getattr(di, 'name', '')): di for di in detailed }
                if norm in by_norm:
                    return by_norm[norm]
                # Включение одной строки в другую
                for di in detailed:
                    dn = _norm(getattr(di, 'name', ''))
                    if not dn:
                        continue
                    if norm in dn or dn in norm:
                        return di
                # По токенам — берём с наибольшим пересечением
                name_toks = set(_tokenize(name))
                best = None
                best_score = 0.0
                for di in detailed:
                    dn = getattr(di, 'name', '')
                    toks = set(_tokenize(dn))
                    if not toks:
                        continue
                    inter = len(name_toks & toks)
                    if not name_toks:
                        continue
                    score = inter / max(1, len(name_toks))
                    if score > best_score:
                        best_score = score
                        best = di
                if best_score >= 0.6:
                    return best
                return None

            def _align(detailed: List[DishItem], names: List[str]) -> List[DishItem]:
                """Выравнивает по порядку names, стараясь сохранить вес/цену через нестрогое сопоставление."""
                if not names:
                    return detailed
                aligned: List[DishItem] = []
                for n in names:
                    if not n or len(str(n).strip()) < 2:
                        continue
                    di = _best_match(detailed, n)
                    if di:
                        aligned.append(di)
                    else:
                        aligned.append(DishItem(name=str(n).strip()))
                return aligned

            # Завтраки и салаты оставляем как есть (детальные), без выравнивания — чтобы не потерять вес/цену
            # Остальные категории выравниваем по журналу
            soups = _align(soups, cat_names.get('первое', []))
            meats = _align(meats, cat_names.get('мясо', []))
            poultry = _align(poultry, cat_names.get('птица', []) + cat_names.get('курица', []))
            fish = _align(fish, cat_names.get('рыба', []))
            garnirs = _align(garnirs, cat_names.get('гарнир', []))

            if logger.isEnabledFor(logging.DEBUG):
                logger.debug("Завтраки: " + ", ".join([getattr(x, 'name', '') for x in breakfasts]))
                logger.debug("Супы: " + ", ".join([getattr(x, 'name', '') for x in soups]))
                logger.debug("Мясо: " + ", ".join([getattr(x, 'name', '') for x in meats]))
                logger.debug("Птица: " + ", ".join([getattr(x, 'name', '') for x in poultry]))
                logger.debug("Рыба: " + ", ".join([getattr(x, 'name', '') for x in fish]))
                logger.debug("Гарниры: " + ", ".join([getattr(x, 'name', '') for x in garnirs]))
                logger.debug("Салаты: " + ", ".join([getattr(x, 'name', '') for x in salads]))

            # Открываем шаблон и находим лист
            wb = openpyxl.load_workbook(template_path)
            ws = None
            for sheet in wb.worksheets:
                if 'касс' in sheet.title.lower():
                    ws = sheet
                    break
            if ws is None:
                ws = wb.active

            # Обновляем дату в шаблоне (если удаётся)
            self.update_template_date(ws, menu_date)

            total_inserted = 0

            # Завтраки — A/B/C
            b_start, b_end = breakfast_range
            # Вставка завтраков: лист, диапазон, колонки
            if logger.isEnabledFor(logging.DEBUG):
                logger.debug("Завтраки (final): " + ", ".join([f"{getattr(x,'name','')} [{getattr(x,'weight','')}] [{getattr(x,'price','')}]" for x in breakfasts]))
            total_inserted += fill_cells_sequential(
                ws,
                start_row=b_start,
                stop_row=b_end + 1,
                columns=TargetColumns(name_col=1, weight_col=2, price_col=3),
                dishes=breakfasts,
                replace_only_empty=False,
                logger=logger,
                log_context=f"завтраки A{b_start}..A{b_end}"
            )

            # Супы — D/E/F (до строки 10 включительно). Если D6 — заголовок, начинаем с D7.
            s_start, s_end = soups_range
            s_write_start = s_start
            try:
                cell_at_start = ws.cell(row=s_start, column=4).value
                if cell_at_start and ('перв' in str(cell_at_start).lower()):
                    s_write_start = s_start + 1
            except Exception:
                pass
            # Вставка супов: лист, диапазон, колонки
            if logger.isEnabledFor(logging.DEBUG):
                logger.debug("Супы (final): " + ", ".join([getattr(x, 'name', '') for x in soups]))
            total_inserted += fill_cells_sequential(
                ws,
                start_row=s_write_start,
                stop_row=s_end + 1,
                columns=TargetColumns(name_col=4, weight_col=5, price_col=6),
                dishes=soups,
                replace_only_empty=False,
                logger=logger,
                log_context=f"первые блюда D{s_write_start}..D{s_end}"
            )

            # Мясо — D/E/F (после заголовка на D11)
            m_start, m_end = meat_range
            # Вставка мяса: лист, диапазон, колонки
            if logger.isEnabledFor(logging.DEBUG):
                logger.debug("Мясо (final): " + ", ".join([getattr(x, 'name', '') for x in meats]))
            total_inserted += fill_cells_sequential(
                ws,
                start_row=m_start,
                stop_row=m_end + 1,
                columns=TargetColumns(name_col=4, weight_col=5, price_col=6),
                dishes=meats,
                replace_only_empty=False,
                logger=logger,
                log_context=f"мясные блюда D{m_start}..D{m_end}"
            )

            # Птица — найдём диапазон по заголовку "БЛЮДА ИЗ ПТИЦЫ/КУРИЦЫ"; если не нашли — ставим после мяса
            try:
                poultry_start = self.find_data_start_row(ws, 4, 'птица')
                # Проверяем, что действительно заголовок птицы, иначе fallback
                header_cell = ws.cell(row=poultry_start-1, column=4).value if poultry_start > 1 else None
                if not header_cell or not any(k in str(header_cell).lower() for k in ['птиц', 'куриц']):
                    poultry_start = m_end + 2
                poultry_end = self.find_category_end_row(ws, 4, poultry_start, 'птица')
                logger.info(f"Вставка (птица): лист={ws.title}, диапазон=D{poultry_start}..D{poultry_end}, колонки=4/5/6, позиций={len(poultry)}, overwrite=True")
                total_inserted += fill_cells_sequential(
                    ws,
                    start_row=poultry_start,
                    stop_row=poultry_end + 1,
                    columns=TargetColumns(name_col=4, weight_col=5, price_col=6),
                    dishes=poultry,
                    replace_only_empty=False,
                    logger=logger,
                    log_context=f"птица D{poultry_start}..D{poultry_end}"
                )
            except Exception as e:
                logger.warning(f"Пропуск вставки блюд из птицы из-за ошибки: {e}")

            # Рыба — найдём диапазон по заголовку "БЛЮДА ИЗ РЫБЫ"
            try:
                fish_start = self.find_data_start_row(ws, 4, 'рыба')
                fish_end = self.find_category_end_row(ws, 4, fish_start, 'рыба')
                total_inserted += fill_cells_sequential(
                    ws,
                    start_row=fish_start,
                    stop_row=fish_end + 1,
                    columns=TargetColumns(name_col=4, weight_col=5, price_col=6),
                    dishes=fish,
                    replace_only_empty=False,
                    logger=logger,
                    log_context=f"рыба D{fish_start}..D{fish_end}"
                )
            except Exception as e:
                logger.warning(f"Пропуск вставки рыбы из-за ошибки: {e}")

            # Гарниры — найдём диапазон по заголовку "ГАРНИРЫ"
            try:
                garn_start = self.find_data_start_row(ws, 4, 'гарнир')
                garn_end = self.find_category_end_row(ws, 4, garn_start, 'гарнир')
                total_inserted += fill_cells_sequential(
                    ws,
                    start_row=garn_start,
                    stop_row=garn_end + 1,
                    columns=TargetColumns(name_col=4, weight_col=5, price_col=6),
                    dishes=garnirs,
                    replace_only_empty=False,
                    logger=logger,
                    log_context=f"гарниры D{garn_start}..D{garn_end}"
                )
            except Exception as e:
                logger.warning(f"Пропуск вставки гарниров из-за ошибки: {e}")

            # Салаты — вставляем в левую колонку A после заголовка "САЛАТЫ..." и ниже блока завтраков
            try:
                salads_start = self.find_data_start_row(ws, 1, 'салат')
                salads_end = self.find_category_end_row(ws, 1, salads_start, 'салат')
                safe_start = max(salads_start, b_end + 1)
                total_inserted += fill_cells_sequential(
                    ws,
                    start_row=safe_start,
                    stop_row=salads_end + 1,
                    columns=TargetColumns(name_col=1, weight_col=2, price_col=3),
                    dishes=salads,
                    replace_only_empty=False,
                    logger=logger,
                    log_context=f"салаты A{safe_start}..A{salads_end}"
                )
            except Exception as e:
                logger.warning(f"Пропуск вставки салатов из-за ошибки: {e}")

            wb.save(output_path)

            # Итоговое сообщение
            msg_lines = [
                "Шаблон меню заполнен по фиксированным диапазонам:",
                f"  Завтраки: A{b_start}..A{b_end}",
                f"  Первые блюда: D{s_start}..D{s_end}",
                f"  Мясные блюда: D{m_start}..D{m_end}",
                f"Всего добавлено строк: {total_inserted}",
            ]
            return True, "\n".join(msg_lines)

        except Exception as e:
            return False, f"Ошибка при заполнении фиксированных диапазонов: {str(e)}"

    def _sort_block(self, ws, start_row: int, end_row: int, name_col: int, weight_col: int, price_col: int) -> int:
        """Сортирует блок строк по названию блюда (по алфавиту), перенося вместе вес и цену.
        Пустые строки опускает вниз в пределах блока.
        """
        # Считываем элементы
        items = []
        for r in range(start_row, end_row + 1):
            try:
                n = ws.cell(row=r, column=name_col).value
            except Exception:
                n = None
            try:
                w = ws.cell(row=r, column=weight_col).value
            except Exception:
                w = None
            try:
                p = ws.cell(row=r, column=price_col).value
            except Exception:
                p = None
            if n is not None and str(n).strip() != "":
                items.append((str(n).strip(), w, p))

        # Нормализация для сравнения (без регистра, 'ё' -> 'е')
        def _norm_name(s: str) -> str:
            return str(s).strip().lower().replace('ё', 'е')

        items.sort(key=lambda t: _norm_name(t[0]))

        # Записываем обратно и очищаем остаток
        i = 0
        written = 0
        for r in range(start_row, end_row + 1):
            if i < len(items):
                name, weight, price = items[i]
                try:
                    ws.cell(row=r, column=name_col).value = name
                except Exception:
                    pass
                try:
                    ws.cell(row=r, column=weight_col).value = weight
                except Exception:
                    pass
                try:
                    ws.cell(row=r, column=price_col).value = price
                except Exception:
                    pass
                i += 1
                written += 1
            else:
                # Очистка оставшихся строк
                try:
                    ws.cell(row=r, column=name_col).value = None
                except Exception:
                    pass
                try:
                    ws.cell(row=r, column=weight_col).value = None
                except Exception:
                    pass
                try:
                    ws.cell(row=r, column=price_col).value = None
                except Exception:
                    pass
        return written

    def sort_kassa_ranges(self, ws) -> None:
        """Сортирует все категории в фиксированных диапазонах на листе «Касса»:
        - Завтраки: A7..A27
        - Салаты: A28..A41 (A42 - заголовок СЭНДВИЧИ)
        - Супы: D7..D10
        - Мясо: D12..D17
        - Птица: D19..D24
        - Рыба: D26..D29
        - Гарниры: D31..D38
        """
        # Левая колонка (имя/вес/цена в A/B/C)
        self._sort_block(ws, start_row=7, end_row=27, name_col=1, weight_col=2, price_col=3)  # Завтраки
        self._sort_block(ws, start_row=28, end_row=41, name_col=1, weight_col=2, price_col=3)  # Салаты
        # Правая колонка (имя/вес/цена в D/E/F)
        self._sort_block(ws, start_row=7, end_row=10, name_col=4, weight_col=5, price_col=6)
        self._sort_block(ws, start_row=12, end_row=17, name_col=4, weight_col=5, price_col=6)
        self._sort_block(ws, start_row=19, end_row=24, name_col=4, weight_col=5, price_col=6)
        self._sort_block(ws, start_row=26, end_row=29, name_col=4, weight_col=5, price_col=6)
        self._sort_block(ws, start_row=31, end_row=38, name_col=4, weight_col=5, price_col=6)

    def copy_kassa_rect_A6_F42(self, template_path: str, source_menu_path: str, output_path: str) -> Tuple[bool, str]:
        """
        Копирует прямоугольник значений A6..F42 из файла пользователя в такой же прямоугольник шаблона
        только на листах «Касса/Касс». Оформление и объединения не меняем.
        """
        import openpyxl
        import os

        temp_file_path = None
        try:
            if not Path(template_path).exists():
                return False, f"Шаблон не найден: {template_path}"
            if not Path(source_menu_path).exists():
                return False, f"Источник не найден: {source_menu_path}"

            # Проверяем формат исходного файла
            source_path_to_use = source_menu_path
            if source_menu_path.lower().endswith('.xls'):
                # Конвертируем .xls в .xlsx
                try:
                    temp_file_path = convert_xls_to_xlsx(source_menu_path)
                    source_path_to_use = temp_file_path
                except Exception as conv_err:
                    return False, f"Не удалось конвертировать .xls файл: {str(conv_err)}"

            src_wb = openpyxl.load_workbook(source_path_to_use, data_only=True)
            tpl_wb = openpyxl.load_workbook(template_path)

            def find_kassa(ws_list):
                for sh in ws_list:
                    if 'касс' in sh.title.lower():
                        return sh
                return None

            src_ws = find_kassa(src_wb.worksheets) or src_wb.active
            tpl_ws = find_kassa(tpl_wb.worksheets) or tpl_wb.active

            # Диапазон
            r1, r2 = 6, 42
            c1, c2 = 1, 6  # A..F

            # Очистка целевого диапазона (только значения)
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    try:
                        tpl_ws.cell(row=r, column=c).value = None
                    except AttributeError:
                        pass  # пропускаем MergedCell

            # Проверяем, где находится заголовок ЗАВТРАК в источнике
            breakfast_row_in_source = None
            for r in range(1, 15):  # Ищем в первых 15 строках
                v = src_ws.cell(row=r, column=1).value
                if v and 'ЗАВТРАК' in str(v).upper():
                    breakfast_row_in_source = r
                    break
            
            # Вычисляем сдвиг (если ЗАВТРАК не в A6, нужно сдвинуть)
            shift = 0
            if breakfast_row_in_source and breakfast_row_in_source != 6:
                shift = breakfast_row_in_source - 6
                print(f"Обнаружен сдвиг: заголовок ЗАВТРАК в строке {breakfast_row_in_source}, сдвигаем на {shift} строк")
            
            # Копирование значений из источника со сдвигом
            copied = 0
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    # Читаем из источника с учетом сдвига
                    source_row = r + shift
                    v = src_ws.cell(row=source_row, column=c).value
                    try:
                        tpl_ws.cell(row=r, column=c).value = v
                        copied += 1
                    except AttributeError:
                        # если целевая ячейка — не мастер объединения, просто пропускаем
                        pass
            
            # Данные уже скопированы из источника в правильные места
            # Просто применяем сортировку к каждой категории
            print(f"Данные скопированы. Применяем сортировку...")

            # Нормализация правой части: имена блюд должны быть в колонке D, вес — в E, цена — в F
            import re as _re
            def _is_header(val: str) -> bool:
                if not val:
                    return False
                u = str(val).upper()
                return any(k in u for k in [
                    'ПЕРВЫЕ', 'БЛЮДА ИЗ МЯСА', 'БЛЮДА ИЗ ПТИЦЫ', 'БЛЮДА ИЗ РЫБЫ', 'ГАРНИРЫ', 'НАПИТК'
                ])
            def _looks_weight(val: str) -> bool:
                if not val:
                    return False
                s = str(val).lower().strip()
                # 1) Явные единицы измерения
                if _re.search(r'(к?кал|ккал|г|гр|грам|шт|мл|л|кг)', s):
                    return True
                # 2) Чисто числовой формат веса, возможно с вариантами через '/': 250 или 250/20 или 0,25
                if _re.fullmatch(r"\d+(?:[\.,]\d+)?(?:\s*/\s*\d+(?:[\.,]\d+)?)*", s):
                    return True
                return False

            # Находим все заголовки в колонке D
            header_rows = []
            for rr in range(r1, r2 + 1):
                dv = tpl_ws.cell(row=rr, column=4).value  # D
                if _is_header(dv):
                    header_rows.append(rr)
            header_rows.append(r2 + 1)  # хвост

            # Для каждого блока после заголовка приводим D/E/F к схеме: D=название, E=вес, F=цена (если есть)
            def _is_dish_like(val) -> bool:
                if val in (None, ''):
                    return False
                s = str(val).strip()
                if not s:
                    return False
                if _is_header(s):
                    return False
                if _looks_weight(s):
                    return False
                return True
            def _is_price_like(val) -> bool:
                if val in (None, ''):
                    return False
                s = str(val)
                # число или варианты через '/'
                return _re.search(r"^\s*\d+(?:[\.,]\d{1,2})?(?:\s*/\s*\d+(?:[\.,]\d{1,2})?)*\s*$", s) is not None

            for i in range(len(header_rows) - 1):
                start = header_rows[i] + 1
                end = header_rows[i+1] - 1
                if start < r1:
                    start = r1
                if end > r2:
                    end = r2
                for rr in range(start, end + 1):
                    d = tpl_ws.cell(row=rr, column=4).value
                    e = tpl_ws.cell(row=rr, column=5).value
                    f = tpl_ws.cell(row=rr, column=6).value
                    d_is_dish = _is_dish_like(d)
                    e_is_dish = _is_dish_like(e)
                    d_is_w = _looks_weight(d)
                    e_is_w = _looks_weight(e)
                    f_is_w = _looks_weight(f)
                    f_is_p = _is_price_like(f)

                    # Если D не блюдо, а E похоже на блюдо — переносим E->D
                    if (not d_is_dish) and e_is_dish:
                        new_d = e
                        # Вес берем из D (если там вес) или из F (если там вес), иначе оставляем как есть из E, если это вес
                        new_e = ''
                        if d_is_w:
                            new_e = d
                        elif f_is_w:
                            new_e = f
                        elif e_is_w:
                            new_e = e
                        # Цена — если F выглядит как цена, оставляем в F; иначе не трогаем
                        try:
                            tpl_ws.cell(row=rr, column=4).value = new_d
                        except AttributeError:
                            pass
                        try:
                            tpl_ws.cell(row=rr, column=5).value = new_e
                        except AttributeError:
                            pass
                        # F оставляем как есть (может быть ценой)
                        continue

                    # Если D пусто, а E пусто, но F похоже на блюдо — F->D
                    if (d in (None, '')) and (e in (None, '')) and _is_dish_like(f):
                        try:
                            tpl_ws.cell(row=rr, column=4).value = f
                        except AttributeError:
                            pass
                        try:
                            tpl_ws.cell(row=rr, column=5).value = ''
                        except AttributeError:
                            pass
                        continue

                    # Если D — вес, а E — не вес (скорее блюдо) — обмен местами
                    if d_is_w and not e_is_w and e not in (None, ''):
                        try:
                            tpl_ws.cell(row=rr, column=4).value = e
                        except AttributeError:
                            pass
                        try:
                            tpl_ws.cell(row=rr, column=5).value = d
                        except AttributeError:
                            pass
                        continue

            # Обновляем дату в верхней части «Кассы» из источника (добавляем день недели)
            try:
                menu_date = self.extract_date_from_menu(source_menu_path)
                self.update_template_date(tpl_ws, menu_date, include_weekday=True)
            except Exception:
                pass

            # Ссылки для ХЦ: A19/A20 должны указывать на «Цезарь …» из салатов на «Касса» и делиться по «/»
            try:
                # Находим строку с «цезар» в колонке A (салаты) на Кассе в диапазоне A6..A42
                caesar_row = None
                for rr in range(6, 43):
                    val = tpl_ws.cell(row=rr, column=1).value  # A
                    if val and 'цезар' in str(val).lower():
                        caesar_row = rr
                        break
                if caesar_row:
                    # Ищем лист «Хц/ХЦ»
                    hx_ws = None
                    for sh in tpl_wb.worksheets:
                        if 'хц' in sh.title.lower():
                            hx_ws = sh
                            break
                    if hx_ws is not None:
                        ref = f"Касса!A{caesar_row}"
                        # A19: часть до «/»
                        f1 = f"=IFERROR(TRIM(LEFT({ref}, IFERROR(FIND(\"/\", {ref})-1, LEN({ref})))), {ref})"
                        # A20: часть после «/»
                        f2 = f"=IFERROR(TRIM(MID({ref}, FIND(\"/\", {ref})+1, 999)), \"\")"
                        try:
                            hx_ws.cell(row=19, column=1).value = f1
                            hx_ws.cell(row=20, column=1).value = f2
                        except Exception:
                            pass
            except Exception:
                pass

            # Постобработка: добавить единицы измерения в колонки B и E согласно диапазонам
            # - Блюда (г): A7..A59 => B7..B59; D7..D38 и D45..D54 => E7..E38, E45..E54
            # - Напитки (мл): D40..D43 => E40..E43
            import re as _re

            def _has_unit(text: str) -> bool:
                if not text:
                    return False
                s = str(text).lower()
                return ('г' in s) or ('мл' in s)

            def _append_unit_to_part(part: str, unit: str) -> str:
                p = part.strip()
                if not p:
                    return p
                # если уже есть нужные единицы — оставляем
                if _has_unit(p):
                    return p
                # добавляем только если заканчивается на цифру
                if _re.search(r"\d\s*$", p):
                    return f"{p}{unit}"
                return p

            def _ensure_unit(cell_value, unit: str) -> str:
                if cell_value in (None, ""):
                    return cell_value
                s = str(cell_value)
                # если уже есть г/мл где-либо — не меняем
                if _has_unit(s):
                    return s
                parts = [pp for pp in s.split('/')]
                new_parts = [_append_unit_to_part(pp, unit) for pp in parts]
                return '/'.join(new_parts)

            # Левая колонка весов: блюда B7..B59 -> г
            for rr in range(7, 60):
                try:
                    cell = tpl_ws.cell(row=rr, column=2)
                    new_val = _ensure_unit(cell.value, 'г')
                    if new_val != cell.value:
                        cell.value = new_val
                except Exception:
                    pass

            # Правая колонка весов: блюда E7..E38 и E45..E54 -> г
            for rr in list(range(7, 39)) + list(range(45, 55)):
                try:
                    cell = tpl_ws.cell(row=rr, column=5)
                    new_val = _ensure_unit(cell.value, 'г')
                    if new_val != cell.value:
                        cell.value = new_val
                except Exception:
                    pass

            # Правая колонка весов: напитки E40..E43 -> мл
            for rr in range(40, 44):
                try:
                    cell = tpl_ws.cell(row=rr, column=5)
                    new_val = _ensure_unit(cell.value, 'мл')
                    if new_val != cell.value:
                        cell.value = new_val
                except Exception:
                    pass

            # Сортировка категорий в указанных диапазонах
            try:
                self.sort_kassa_ranges(tpl_ws)
            except Exception:
                pass

            tpl_wb.save(output_path)
            return True, f"Скопировано {copied} ячеек в A6..F42; дата обновлена; ссылки ХЦ установлены; единицы добавлены; категории отсортированы"
        except Exception as e:
            return False, f"Ошибка при копировании A6..F42: {str(e)}"
        finally:
            # Удаляем временный файл, если он был создан
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                except Exception:
                    pass

    def fill_kassa_with_counts(
        self,
        template_path: str,
        source_menu_path: str,
        output_path: str,
        counts: Optional[Dict[str, int]] = None,
        breakfast_range: Tuple[int, int] = (7, 27),
    ) -> Tuple[bool, str]:
        """
        Заполняет ТОЛЬКО лист «Касса» фиксированным количеством блюд по категориям.
        Количества по умолчанию:
          - супы: 4; рыба: 4; мясо: 6; курица/птица: 6; гарниры: 9; салаты: 13.
        Завтраки — диапазон A7..A27 (A/B/C).
        """
        import openpyxl
        import logging
        logger = logging.getLogger("menu.template")
        from app.services.excel_inserter import fill_cells_sequential, TargetColumns

        # Значения по умолчанию
        if counts is None:
            counts = {
                'soups': 4,
                'fish': 4,
                'meat': 6,
                'poultry': 6,
                'garnirs': 9,
                'salads': 13,
            }
        
        # 1) Извлекаем блюда детально (как в фиксированных диапазонах)
        temp_file_path = None
        try:
            import openpyxl as _ox
            import os
            
            # Проверяем формат исходного файла
            source_path_to_use = source_menu_path
            if source_menu_path.lower().endswith('.xls'):
                # Конвертируем .xls в .xlsx
                try:
                    temp_file_path = convert_xls_to_xlsx(source_menu_path)
                    source_path_to_use = temp_file_path
                except Exception as conv_err:
                    return False, f"Не удалось конвертировать .xls файл: {str(conv_err)}"
            
            src_wb = _ox.load_workbook(source_path_to_use, data_only=True)

            def _find_sheet(wb, keywords: List[str]):
                for sh in wb.worksheets:
                    title = sh.title.lower()
                    if any(k in title for k in [kw.lower() for kw in keywords]):
                        return sh
                return None

            def _sanitize_price(raw) -> str:
                if raw is None:
                    return ""
                s = str(raw)
                parts = [p.strip() for p in s.split('/')]
                out = []
                import re as _re
                for p in parts:
                    m = _re.search(r"(\d+(?:[\.,]\d{1,2})?)", p)
                    if not m:
                        continue
                    num = m.group(1).replace('.', ',')
                    out.append(num)
                return "/".join(out)

            def _read_items_from_sheet(sh, limit: int) -> List[DishItem]:
                items: List[DishItem] = []
                header_tokens = ['завтрак', 'салат', 'вес', 'ед.изм', 'ед. изм', 'цена', 'руб']
                for r in range(1, sh.max_row + 1):
                    w = sh.cell(row=r, column=1).value
                    n = sh.cell(row=r, column=2).value
                    p = sh.cell(row=r, column=3).value
                    name = (str(n).strip() if n else '')
                    if not name:
                        continue
                    n_low = name.lower()
                    w_low = (str(w).lower() if isinstance(w, str) else str(w).lower()) if w not in (None, '') else ''
                    p_low = (str(p).lower() if isinstance(p, str) else str(p).lower()) if p not in (None, '') else ''
                    # Пропускаем шапки/заголовки
                    if any(tok in n_low for tok in header_tokens) or any(tok in w_low for tok in header_tokens) or any(tok in p_low for tok in header_tokens):
                        continue
                    weight = str(w).strip() if w else ''
                    price = _sanitize_price(p)
                    items.append(DishItem(name=name, weight=weight, price=price))
                    if len(items) >= limit:
                        break
                return items

            # Завтраки — сначала пробуем читать напрямую с листа «Завтрак» (col0/1/2), затем извлекатели
            breakfasts: List[DishItem] = []
            try:
                br_sh = _find_sheet(src_wb, ["завтрак"])
                if br_sh is not None:
                    br_limit = max(0, breakfast_range[1] - breakfast_range[0] + 1)
                    breakfasts = _read_items_from_sheet(br_sh, br_limit)
            except Exception:
                breakfasts = []
            if not breakfasts:
                try:
                    breakfasts = extract_dishes_from_excel_rows_with_stop(
                        source_menu_path,
                        ["ЗАВТРАК"],
                        ["САЛАТЫ", "ХОЛОДНЫЕ", "ЗАКУСКИ"],
                    )
                except Exception:
                    breakfasts = []
            if not breakfasts:
                try:
                    details = self.extract_dishes_with_details(source_menu_path)
                    br = details.get('завтрак', []) if isinstance(details, dict) else []
                    breakfasts = [DishItem(name=d.get('name',''), weight=d.get('weight',''), price=d.get('price','')) for d in br if d.get('name')]
                except Exception:
                    breakfasts = []

            # Остальные категории (базово — извлекатели)
            soups = extract_dishes_from_excel(source_menu_path, ["ПЕРВЫЕ БЛЮДА", "ПЕРВЫЕ"]) or []
            meat = extract_dishes_from_excel(source_menu_path, ["БЛЮДА ИЗ МЯСА", "МЯСНЫЕ БЛЮДА"]) or []
            fish = extract_dishes_from_excel(source_menu_path, ["БЛЮДА ИЗ РЫБЫ", "РЫБНЫЕ БЛЮДА"]) or []

            # «По такому же принципу»: если в источнике есть листы для этих категорий — читаем напрямую из листов (col0/1/2)
            poultry_sh = _find_sheet(src_wb, ["птиц", "куриц"])
            garnir_sh = _find_sheet(src_wb, ["гарнир"])
            salads_sh = _find_sheet(src_wb, ["салат"])

            poultry = _read_items_from_sheet(poultry_sh, 6) if poultry_sh is not None else extract_dishes_from_excel(source_menu_path, ["БЛЮДА ИЗ ПТИЦЫ", "БЛЮДА ИЗ КУРИЦЫ", "КУРИНЫЕ БЛЮДА"]) or []
            garnirs = _read_items_from_sheet(garnir_sh, 9) if garnir_sh is not None else extract_dishes_from_excel(source_menu_path, ["ГАРНИРЫ", "ГАРНИР"]) or []
            salads = _read_items_from_sheet(salads_sh, 13) if salads_sh is not None else extract_dishes_from_excel_rows_with_stop(
                source_menu_path,
                ["САЛАТ"],
                ["ПЕРВЫЕ", "БЛЮДА ИЗ", "ГАРНИР", "НАПИТ"]
            ) or []
        except Exception as e:
            return False, f"Ошибка извлечения блюд: {e}"
        finally:
            # Удаляем временный файл, если он был создан
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                except Exception:
                    pass

        # Ограничиваем по количеству
        def take(lst: List[DishItem], n: int) -> List[DishItem]:
            return lst[:max(0, n)] if lst else []

        soups = take(soups, counts.get('soups', 4))
        fish = take(fish, counts.get('fish', 4))
        meat = take(meat, counts.get('meat', 6))
        poultry = take(poultry, counts.get('poultry', 6))
        garnirs = take(garnirs, counts.get('garnirs', 9))
        salads = take(salads, counts.get('salads', 13))

        # 2) Загружаем шаблон и работаем ТОЛЬКО с листом «Касса»
        wb = openpyxl.load_workbook(template_path)
        ws = None
        for sh in wb.worksheets:
            if 'касс' in sh.title.lower():
                ws = sh
                break
        if ws is None:
            ws = wb.active
        
        # Обновляем дату
        self.update_template_date(ws, self.extract_date_from_menu(source_menu_path))

        # Утилиты
        def clear_rect(start_row: int, rows: int, name_col: int, weight_col: int, price_col: int):
            end_row = start_row + rows - 1
            for r in range(start_row, end_row + 1):
                for c in (name_col, weight_col, price_col):
                    try:
                        ws.cell(row=r, column=c).value = None
                    except AttributeError:
                        pass
            return end_row

        def write_list(start_row: int, items: List[DishItem], name_col: int, weight_col: int, price_col: int):
            return fill_cells_sequential(
                ws,
                start_row=start_row,
                stop_row=start_row + len(items),
                columns=TargetColumns(name_col=name_col, weight_col=weight_col, price_col=price_col),
                dishes=items,
                replace_only_empty=False,
                logger=logger,
            )

        inserted_total = 0

        # 3) Завтраки A/B/C: A7..A27 (по правилу)
        br_start, br_end = breakfast_range
        # Очистка
        clear_rect(br_start, br_end - br_start + 1, 1, 2, 3)
        # Запись (обрезаем по диапазону)
        if breakfasts:
            max_rows = br_end - br_start + 1
            breakfasts = breakfasts[:max_rows]
            inserted_total += write_list(br_start, breakfasts, 1, 2, 3)

        # 4) Супы D/E/F — начиная от заголовка «ПЕРВЫЕ БЛЮДА», строго 4
        try:
            s_start = self.find_data_start_row(ws, 4, 'первое')
            clear_rect(s_start, len(soups), 4, 5, 6)
            inserted_total += write_list(s_start, soups, 4, 5, 6)
        except Exception:
            pass

        # 5) Мясо D/E/F — от заголовка «БЛЮДА ИЗ МЯСА», 6
        try:
            m_start = self.find_data_start_row(ws, 4, 'мясо')
            clear_rect(m_start, len(meat), 4, 5, 6)
            inserted_total += write_list(m_start, meat, 4, 5, 6)
        except Exception:
            pass

        # 6) Птица D/E/F — от заголовка «БЛЮДА ИЗ ПТИЦЫ/КУРИЦЫ», 6
        try:
            p_start = self.find_data_start_row(ws, 4, 'птица')
            clear_rect(p_start, len(poultry), 4, 5, 6)
            inserted_total += write_list(p_start, poultry, 4, 5, 6)
        except Exception:
            pass

        # 7) Рыба D/E/F — от заголовка «БЛЮДА ИЗ РЫБЫ», 4
        try:
            f_start = self.find_data_start_row(ws, 4, 'рыба')
            clear_rect(f_start, len(fish), 4, 5, 6)
            inserted_total += write_list(f_start, fish, 4, 5, 6)
        except Exception:
            pass

        # 8) Гарниры D/E/F — от заголовка «ГАРНИРЫ», 9
        try:
            g_start = self.find_data_start_row(ws, 4, 'гарнир')
            clear_rect(g_start, len(garnirs), 4, 5, 6)
            inserted_total += write_list(g_start, garnirs, 4, 5, 6)
        except Exception:
            pass

        # 9) Салаты A/B/C — от заголовка «САЛАТЫ...», 13, но не раньше, чем после завтраков
        try:
            sl_start = self.find_data_start_row(ws, 1, 'салат')
            sl_start = max(sl_start, br_end + 1)
            clear_rect(sl_start, len(salads), 1, 2, 3)
            inserted_total += write_list(sl_start, salads, 1, 2, 3)
        except Exception:
            pass

        wb.save(output_path)
        return True, f"Заполнено строк: {inserted_total} (Касса, фиксированные количества)"


def fill_menu_template_from_source(template_path: str, source_menu_path: str, output_path: str) -> Tuple[bool, str]:
    """Удобная функция для заполнения шаблона меню"""
    filler = MenuTemplateFiller()
    return filler.fill_menu_template(template_path, source_menu_path, output_path)

def fill_menu_template_with_details_from_source(template_path: str, source_menu_path: str, output_path: str) -> Tuple[bool, str]:
    """Удобная функция для заполнения шаблона меню с деталями"""
    filler = MenuTemplateFiller()
    return filler.fill_menu_template_with_details(template_path, source_menu_path, output_path)

def fill_breakfast_only(template_path: str, source_menu_path: str, output_path: str) -> Tuple[bool, str]:
    """Копирует только завтраки с деталями (название, вес, цена)"""
    try:
        import openpyxl
        from pathlib import Path
        
        # Проверяем существование файлов
        if not Path(template_path).exists():
            return False, f"Шаблон не найден: {template_path}"
        
        if not Path(source_menu_path).exists():
            return False, f"Исходный файл не найден: {source_menu_path}"
        
        # Открываем исходный файл и ищем завтраки
        print(f"Извлекаем завтраки из {source_menu_path}")
        source_wb = openpyxl.load_workbook(source_menu_path)
        source_ws = source_wb.active
        
        breakfast_dishes = []
        in_breakfast_section = False
        
        for row in range(1, min(100, source_ws.max_row + 1)):
            name_cell = source_ws.cell(row=row, column=1).value
            if not name_cell:
                continue
                
            name_text = str(name_cell).lower().strip()
            
            # Начало секции завтраков
            if 'завтрак' in name_text:
                in_breakfast_section = True
                print(f"Начало секции завтраков в строке {row}: {name_cell}")
                continue
            
            # Конец секции завтраков
            if in_breakfast_section and any(word in name_text for word in ['салат', 'холодн', 'перв', 'блюда']):
                print(f"Конец секции завтраков в строке {row}: {name_cell}")
                break
            
            # Собираем блюда завтраков
            if in_breakfast_section:
                # Пропускаем служебные строки
                if any(word in name_text for word in ['вес', 'цена', 'руб']):
                    continue
                
                # Получаем вес и цену
                weight = source_ws.cell(row=row, column=2).value
                price = source_ws.cell(row=row, column=3).value
                
                dish = {
                    'name': str(name_cell).strip(),
                    'weight': str(weight).strip() if weight else '',
                    'price': str(price).strip() if price else ''
                }
                
                breakfast_dishes.append(dish)
                print(f"Найдено блюдо: {dish}")
        
        source_wb.close()
        
        if not breakfast_dishes:
            return False, "Не найдено блюд для завтрака"
        
        print(f"Найдено {len(breakfast_dishes)} блюд для завтрака")
        
        # Открываем шаблон и заполняем завтраки
        template_wb = openpyxl.load_workbook(template_path)
        
        # Находим лист Касса
        template_ws = None
        for sheet in template_wb.worksheets:
            if 'касс' in sheet.title.lower():
                template_ws = sheet
                break
        
        if not template_ws:
            template_ws = template_wb.active
            print("Предупреждение: Лист 'Касса' не найден")
        
        # Находим заголовок завтраков в шаблоне
        start_row = None
        for row in range(1, 20):
            cell_val = template_ws.cell(row=row, column=1).value
            if cell_val and 'завтрак' in str(cell_val).lower():
                start_row = row + 1
                print(f"Найден заголовок завтраков в строке {row}, начинаем с {start_row}")
                break
        
        if not start_row:
            start_row = 7  # По умолчанию
            print(f"Заголовок не найден, начинаем с строки {start_row}")
        
        # Заполняем завтраки
        filled_count = 0
        current_row = start_row
        
        for dish in breakfast_dishes:
            # Пропускаем занятые строки
            while current_row <= template_ws.max_row:
                existing_name = template_ws.cell(row=current_row, column=1).value
                if not existing_name or str(existing_name).strip() == '':
                    break
                current_row += 1
            
            if current_row > template_ws.max_row:
                print(f"Достигнут конец листа")
                break
            
            # Заполняем данные
            template_ws.cell(row=current_row, column=1).value = dish['name']      # Название
            template_ws.cell(row=current_row, column=2).value = dish['weight']    # Вес
            template_ws.cell(row=current_row, column=3).value = dish['price']     # Цена
            
            filled_count += 1
            print(f"Заполнена строка {current_row}: {dish['name']} | {dish['weight']} | {dish['price']}")
            current_row += 1
        
        # Сохраняем
        template_wb.save(output_path)
        template_wb.close()
        
        message = f"Успешно скопировано {filled_count} завтраков с деталями"
        return True, message
        
    except Exception as e:
        return False, f"Ошибка: {str(e)}"
