#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Общий модуль для извлечения блюд из различных источников данных.
Централизованная логика для работы с меню в разных форматах.
"""

import pandas as pd
import openpyxl
import re
from pathlib import Path
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple, Union
from dataclasses import dataclass
from abc import ABC, abstractmethod


# Исключения
class DishExtractionError(Exception):
    """Базовая ошибка извлечения блюд"""
    pass


class FileFormatError(DishExtractionError):
    """Ошибка неподдерживаемого формата файла"""
    pass


class SheetNotFoundError(DishExtractionError):
    """Ошибка отсутствия листа в файле"""
    pass


# Модели данных
@dataclass
class DishItem:
    """Представляет отдельное блюдо с базовой информацией.
    
    Данный класс содержит основную информацию о блюде:
    название, вес, цену и категорию. Автоматически очищает
    и нормализует данные при создании объекта.
    
    Attributes:
        name (str): Название блюда
        weight (str): Вес или порция блюда (например, "150г", "1 шт")
        price (str): Цена блюда (например, "120 руб")
        category (str): Категория блюда (завтрак, салат, первое, мясо, курица, рыба, гарнир)
    """
    name: str
    weight: str = ""
    price: str = ""
    category: str = ""
    
    def __post_init__(self):
        # Очистка и нормализация данных при создании
        self.name = str(self.name).strip()
        self.weight = str(self.weight).strip() if self.weight else ""
        self.price = str(self.price).strip() if self.price else ""
        self.category = str(self.category).strip() if self.category else ""


@dataclass
class ExtractionResult:
    """Результат извлечения блюд с метаданными.
    
    Содержит результаты извлечения блюд из источника данных,
    включая общий список, группировку по категориям и дополнительную информацию.
    
    Attributes:
        dishes (List[DishItem]): Список всех извлеченных блюд
        categories (Dict[str, List[DishItem]]): Блюда, сгруппированные по категориям
        source_date (Optional[datetime]): Дата, найденная в источнике данных
        total_count (int): Общее количество блюд (вычисляется автоматически)
    """
    dishes: List[DishItem]
    categories: Dict[str, List[DishItem]]
    source_date: Optional[datetime] = None
    total_count: int = 0
    
    def __post_init__(self):
        self.total_count = len(self.dishes)


# Абстракция для источников данных
class DataSource(ABC):
    """Абстрактный базовый класс для источников данных о блюдах.
    
    Определяет интерфейс для всех типов источников данных (Excel, CSV, базы данных).
    Наследники должны реализовать методы для извлечения блюд и дат.
    """
    
    @abstractmethod
    def extract_dishes(self, source_path: str, **kwargs) -> ExtractionResult:
        """Извлекает блюда из источника данных.
        
        Args:
            source_path: Путь к источнику данных
            **kwargs: Дополнительные параметры извлечения
            
        Returns:
            ExtractionResult: Результат извлечения с блюдами и метаданными
            
        Raises:
            DishExtractionError: При ошибках извлечения данных
        """
        pass
    
    @abstractmethod
    def extract_date(self, source_path: str, **kwargs) -> Optional[datetime]:
        """Извлекает дату из источника данных.
        
        Args:
            source_path: Путь к источнику данных
            **kwargs: Дополнительные параметры извлечения
            
        Returns:
            Optional[datetime]: Найденная дата или None
        """
        pass


class ExcelDataSource(DataSource):
    """Источник данных для работы с Excel файлами меню.
    
    Обрабатывает файлы форматов .xls, .xlsx, .xlsm.
    Извлекает блюда по категориям и находит даты в содержимом файлов.
    Поддерживает различные структуры Excel-документов.
    """
    
    def __init__(self):
        self.categories_mapping = {
            'завтрак': 'завтрак',
            'салат': 'салат',
            'первое': 'первое',
            'мясо': 'мясо',
            'курица': 'курица',
            'птица': 'птица',
            'рыба': 'рыба',
            'гарнир': 'гарнир'
        }
    
    def extract_dishes(self, source_path: str, **kwargs) -> ExtractionResult:
        """Извлекает все блюда из Excel файла"""
        if not Path(source_path).exists():
            raise FileFormatError(f"Файл не найден: {source_path}")
        
        # Определяем формат файла
        ext = Path(source_path).suffix.lower()
        if ext not in ['.xls', '.xlsx', '.xlsm']:
            raise FileFormatError(f"Неподдерживаемый формат файла: {ext}")
        
        # Извлекаем блюда по категориям
        categorized_dishes = self._extract_categorized_dishes(source_path)
        
        # Преобразуем в список DishItem
        all_dishes = []
        categories_result = {}
        
        for category, dish_names in categorized_dishes.items():
            category_dishes = []
            for dish_name in dish_names:
                dish_item = DishItem(name=dish_name, category=category)
                all_dishes.append(dish_item)
                category_dishes.append(dish_item)
            categories_result[category] = category_dishes
        
        # Извлекаем дату
        source_date = self.extract_date(source_path)
        
        return ExtractionResult(
            dishes=all_dishes,
            categories=categories_result,
            source_date=source_date
        )
    
    def extract_date(self, source_path: str, **kwargs) -> Optional[datetime]:
        """Извлекает дату из Excel файла"""
        try:
            if source_path.endswith('.xls'):
                return self._extract_date_from_xls(source_path)
            else:
                return self._extract_date_from_xlsx(source_path)
        except Exception as e:
            print(f"Ошибка при извлечении даты: {e}")
            return None
    
    def _extract_date_from_xlsx(self, path: str) -> Optional[datetime]:
        """Извлекает дату из xlsx файла"""
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            
            # Проверяем название файла
            filename = Path(path).stem
            date_from_filename = self._parse_date_string(filename)
            if date_from_filename:
                return date_from_filename
            
            # Ищем дату в листах
            for ws in wb.worksheets:
                # Проверяем название листа
                date_from_sheet = self._parse_date_string(ws.title)
                if date_from_sheet:
                    return date_from_sheet
                
                # Ищем дату в первых 20 строках
                for row in range(1, min(21, ws.max_row + 1)):
                    for col in range(1, min(ws.max_column + 1, 10)):
                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            date_from_cell = self._parse_date_string(str(cell.value))
                            if date_from_cell:
                                return date_from_cell
        except Exception:
            pass
        return None
    
    def _extract_date_from_xls(self, path: str) -> Optional[datetime]:
        """Извлекает дату из xls файла"""
        try:
            df_dict = pd.read_excel(path, sheet_name=None)
            
            for sheet_name, df in df_dict.items():
                # Проверяем название листа
                date_from_sheet = self._parse_date_string(sheet_name)
                if date_from_sheet:
                    return date_from_sheet
                
                # Ищем дату в содержимом листа
                for col in df.columns:
                    if pd.notna(col):
                        date_from_col = self._parse_date_string(str(col))
                        if date_from_col:
                            return date_from_col
                
                # Ищем дату в первых строках
                for _, row in df.head(10).iterrows():
                    for cell in row:
                        if pd.notna(cell):
                            date_from_cell = self._parse_date_string(str(cell))
                            if date_from_cell:
                                return date_from_cell
        except Exception:
            pass
        return None
    
    def _parse_date_string(self, text: str) -> Optional[datetime]:
        """Парсит дату из строки"""
        if not text:
            return None
            
        text = text.strip().lower()
        
        # Русские месяцы
        months_ru = {
            'января': 1, 'февраля': 2, 'марта': 3, 'апреля': 4, 'мая': 5, 'июня': 6,
            'июля': 7, 'августа': 8, 'сентября': 9, 'октября': 10, 'ноября': 11, 'декабря': 12
        }
        
        # Паттерны для поиска даты
        patterns = [
            r'(\d{1,2})\s+(\w+)',  # "5 сентября"
            r'(\d{1,2})\.(\d{1,2})\.(\d{4})',  # "05.09.2024"
            r'(\d{1,2})/(\d{1,2})/(\d{4})',  # "05/09/2024"
            r'(\d{1,2})\s+(\w+)\s+(\d{4})',  # "5 сентября 2024"
            r'(\d{2})\.(\d{2})\.(\d{2})',  # "05.09.24"
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                try:
                    if len(match.groups()) == 2:
                        # Формат "день месяц"
                        day = int(match.group(1))
                        month_str = match.group(2)
                        month = months_ru.get(month_str)
                        if month:
                            year = datetime.now().year
                            return datetime(year, month, day)
                    elif len(match.groups()) == 3:
                        if match.group(2) in months_ru:
                            # Формат "день месяц год"
                            day = int(match.group(1))
                            month = months_ru[match.group(2)]
                            year = int(match.group(3))
                        else:
                            # Формат "день.месяц.год"
                            day = int(match.group(1))
                            month = int(match.group(2))
                            year = int(match.group(3))
                            if year < 100:  # Двузначный год
                                year += 2000
                        return datetime(year, month, day)
                except (ValueError, TypeError):
                    continue
        
        return None
    
    def _extract_categorized_dishes(self, source_path: str) -> Dict[str, List[str]]:
        """Извлекает блюда по категориям"""
        result: Dict[str, List[str]] = {k: [] for k in self.categories_mapping.keys()}
        
        try:
            if source_path.endswith('.xls'):
                self._extract_from_xls(source_path, result)
            else:
                self._extract_from_xlsx(source_path, result)
        except Exception as e:
            print(f"Ошибка при извлечении категорий: {e}")
        
        return result
    
    def _extract_from_xlsx(self, path: str, result: Dict[str, List[str]]):
        """Извлекает блюда из xlsx файла"""
        wb = openpyxl.load_workbook(path, data_only=True)
        if wb.worksheets:
            self._extract_from_worksheet(wb.worksheets[0], result)
    
    def _extract_from_xls(self, path: str, result: Dict[str, List[str]]):
        """Извлекает блюда из xls файла"""
        df_dict = pd.read_excel(path, sheet_name=None)
        
        # Ищем лист с приоритетом
        preferred = None
        for name, df in df_dict.items():
            nm = str(name).lower()
            if 'касс' in nm:
                preferred = df
                break
        
        if preferred is None:
            for name, df in df_dict.items():
                nm = str(name).lower()
                if 'меню' in nm:
                    preferred = df
                    break
        
        if preferred is None and df_dict:
            preferred = list(df_dict.values())[0]
        
        if preferred is not None:
            self._extract_from_dataframe(preferred, result)
    
    def _extract_from_worksheet(self, ws, result: Dict[str, List[str]]):
        """Извлекает блюда из worksheet"""
        # Находим строку заголовков
        header_row = None
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_text = ''
            for col_idx in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    row_text += str(cell_val).strip() + ' '
            
            if 'ЗАВТРАКИ' in row_text.upper():
                header_row = row_idx
                break
        
        if header_row is None:
            return
        
        # Извлекаем завтраки и салаты из первого столбца
        current_category = 'завтрак'
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value  # Столбец A
            if cell_value:
                cell_str = str(cell_value).strip()
                
                # Проверяем заголовки
                if 'САЛАТ' in cell_str.upper() and 'ХОЛОДН' in cell_str.upper():
                    current_category = 'салат'
                    continue
                
                # Добавляем блюдо
                if not self._should_skip_cell(cell_str) and self._is_valid_dish(cell_str, []):
                    result[current_category].append(cell_str)
        
        # Извлекаем остальные блюда из правых столбцов
        current_category = None
        for col_idx in range(4, ws.max_column + 1):
            for row_idx in range(header_row, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip()
                    
                    # Определяем категорию
                    if 'ПЕРВЫЕ' in cell_str.upper() and 'БЛЮДА' in cell_str.upper():
                        current_category = 'первое'
                    elif 'БЛЮДА ИЗ МЯСА' in cell_str.upper():
                        current_category = 'мясо'
                    elif 'БЛЮДА ИЗ ПТИЦЫ' in cell_str.upper():
                        current_category = 'птица'
                    elif 'БЛЮДА ИЗ РЫБЫ' in cell_str.upper():
                        current_category = 'рыба'
                    elif 'ГАРНИРЫ' in cell_str.upper():
                        current_category = 'гарнир'
                    elif 'НАПИТКИ' in cell_str.upper():
                        return  # Останавливаемся на напитках
                    elif current_category and not self._should_skip_cell(cell_str) and self._is_valid_dish(cell_str, []):
                        result[current_category].append(cell_str)
    
    def _extract_from_dataframe(self, df: pd.DataFrame, result: Dict[str, List[str]]):
        """Извлекает блюда из DataFrame"""
        # Находим строку заголовков
        header_row = None
        for idx in range(min(10, len(df))):
            row = df.iloc[idx]
            row_text = ' '.join([str(cell).strip() for cell in row if pd.notna(cell) and str(cell).strip()])
            if 'ЗАВТРАКИ' in row_text.upper():
                header_row = idx
                break
        
        if header_row is None:
            return
        
        # Извлекаем завтраки из первого столбца
        for row_idx in range(header_row + 1, len(df)):
            cell_value = df.iloc[row_idx, 0]  # Столбец A
            if pd.notna(cell_value):
                cell_str = str(cell_value).strip()
                if not self._should_skip_cell(cell_str) and self._is_valid_dish(cell_str, []):
                    result['завтрак'].append(cell_str)
            elif len(result['завтрак']) > 0:
                break
        
        # Извлекаем остальные блюда
        current_category = None
        for col_idx in range(3, len(df.columns)):
            for row_idx in range(header_row, len(df)):
                cell_value = df.iloc[row_idx, col_idx]
                if pd.notna(cell_value):
                    cell_str = str(cell_value).strip()
                    
                    # Определяем категорию
                    if 'ПЕРВЫЕ' in cell_str.upper() and 'БЛЮДА' in cell_str.upper():
                        current_category = 'первое'
                    elif 'БЛЮДА ИЗ МЯСА' in cell_str.upper():
                        current_category = 'мясо'
                    elif 'БЛЮДА ИЗ ПТИЦЫ' in cell_str.upper():
                        current_category = 'птица'
                    elif 'БЛЮДА ИЗ РЫБЫ' in cell_str.upper():
                        current_category = 'рыба'
                    elif 'ГАРНИРЫ' in cell_str.upper():
                        current_category = 'гарнир'
                    elif 'НАПИТКИ' in cell_str.upper():
                        return
                    elif current_category and not self._should_skip_cell(cell_str) and self._is_valid_dish(cell_str, []):
                        result[current_category].append(cell_str)
    
    def _should_skip_cell(self, cell_str: str) -> bool:
        """Проверяет, нужно ли пропустить ячейку"""
        cell_lower = cell_str.lower().strip()
        
        if len(cell_str) < 4:
            return True
        
        skip_words = [
            'вес', 'цена', 'руб', 'ед.изм', 'утверждаю', 'директор', 
            'меню', 'столовой', 'патриот', 'москва', 'наб', 'стр',
            'попова', 'сентября', 'пятница', '_____', 'понедельник', 'вторник',
            'среда', 'четверг', 'пятница', 'суббота', 'воскресенье',
            'январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль',
            'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь',
            'завтрак', 'обед', 'ужин', 'полдник', 'время', 'дата',
            'напит', 'сок', 'чай', 'кофе', 'смузи', 'фреш',
            'соус', 'майонез', 'кетчуп',
            'наименование', 'блюда', 'час', 'мин'
        ]
        
        # Проверяем время
        if re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', cell_str):
            return True
        
        for skip_word in skip_words:
            if skip_word in cell_lower:
                return True
        
        # Пропускаем объемы напитков
        if re.search(r'\d+/\d+\s*мл|\d+\s*мл|200/300мл|300мл|225\s*мл', cell_str):
            return True
        
        # Пропускаем числа и символы
        if re.match(r'^[\d\s\.,/г-]+$', cell_str):
            return True
        
        # Пропускаем адреса
        if 'ул.' in cell_lower or 'д.' in cell_lower or 'овчинниковская' in cell_lower:
            return True
            
        return False
    
    def _is_valid_dish(self, dish_name: str, existing_dishes: List[str]) -> bool:
        """Проверяет, является ли строка валидным названием блюда"""
        if dish_name in existing_dishes:
            return False
        
        if len(dish_name) < 5:
            return False
        
        # Проверяем заголовки категорий
        dish_lower = dish_name.lower().strip()
        category_titles = [
            'завтрак', 'салат', 'холодн', 'закуск', 
            'первое', 'первы', 'блюд', 'гарнир', 
            'мясо', 'курица', 'птица', 'рыба'
        ]
        
        if len(dish_name) < 35:
            for title in category_titles:
                if dish_lower == title or dish_lower.startswith(title + ' ') or dish_lower.endswith(' ' + title):
                    return False
        
        # Только числа и символы
        if re.match(r'^[\d\s\.,/г-]+$', dish_name):
            return False
            
        return True


class DishExtractorService:
    """Основной сервис для извлечения блюд"""
    
    def __init__(self):
        self._sources = {
            'excel': ExcelDataSource(),
        }
    
    def extract_dishes(self, source_path: str, source_type: str = 'auto', **kwargs) -> ExtractionResult:
        """
        Извлекает блюда из источника данных
        
        Args:
            source_path: Путь к источнику данных
            source_type: Тип источника ('excel', 'auto')
            **kwargs: Дополнительные параметры
        
        Returns:
            ExtractionResult: Результат извлечения
        """
        if source_type == 'auto':
            source_type = self._detect_source_type(source_path)
        
        if source_type not in self._sources:
            raise FileFormatError(f"Неподдерживаемый тип источника: {source_type}")
        
        source = self._sources[source_type]
        return source.extract_dishes(source_path, **kwargs)
    
    def extract_categorized_dishes(self, source_path: str, source_type: str = 'auto', **kwargs) -> Dict[str, List[str]]:
        """
        Извлекает блюда, сгруппированные по категориям
        
        Args:
            source_path: Путь к источнику данных
            source_type: Тип источника
            **kwargs: Дополнительные параметры
        
        Returns:
            Dict[str, List[str]]: Словарь категория -> список блюд
        """
        result = self.extract_dishes(source_path, source_type, **kwargs)
        return {category: [dish.name for dish in dishes] for category, dishes in result.categories.items()}
    
    def extract_dishes_by_category(self, source_path: str, category: str, source_type: str = 'auto', **kwargs) -> List[str]:
        """
        Извлекает блюда определенной категории
        
        Args:
            source_path: Путь к источнику данных
            category: Категория блюд
            source_type: Тип источника
            **kwargs: Дополнительные параметры
        
        Returns:
            List[str]: Список названий блюд
        """
        categorized = self.extract_categorized_dishes(source_path, source_type, **kwargs)
        return categorized.get(category, [])
    
    def extract_date_from_source(self, source_path: str, source_type: str = 'auto', **kwargs) -> Optional[datetime]:
        """
        Извлекает дату из источника данных
        
        Args:
            source_path: Путь к источнику данных
            source_type: Тип источника
            **kwargs: Дополнительные параметры
        
        Returns:
            Optional[datetime]: Дата из источника или None
        """
        if source_type == 'auto':
            source_type = self._detect_source_type(source_path)
        
        if source_type not in self._sources:
            raise FileFormatError(f"Неподдерживаемый тип источника: {source_type}")
        
        source = self._sources[source_type]
        return source.extract_date(source_path, **kwargs)
    
    def _detect_source_type(self, source_path: str) -> str:
        """Определяет тип источника данных по расширению файла"""
        ext = Path(source_path).suffix.lower()
        if ext in ['.xls', '.xlsx', '.xlsm']:
            return 'excel'
        else:
            raise FileFormatError(f"Не удалось определить тип источника для файла: {source_path}")


# Глобальный экземпляр сервиса для удобства использования
_extractor_service = DishExtractorService()


# Публичные функции для совместимости с существующим кодом
def extract_categorized_dishes_from_menu(menu_path: str) -> Dict[str, List[str]]:
    """Извлекает блюда по категориям из файла меню (совместимость)"""
    return _extractor_service.extract_categorized_dishes(menu_path)


def extract_dishes_by_category(menu_path: str, category: str) -> List[str]:
    """Извлекает блюда определенной категории (совместимость)"""
    return _extractor_service.extract_dishes_by_category(menu_path, category)


def extract_date_from_menu(menu_path: str) -> Optional[datetime]:
    """Извлекает дату из файла меню (совместимость)"""
    return _extractor_service.extract_date_from_source(menu_path)


def get_dish_extractor() -> DishExtractorService:
    """Возвращает экземпляр сервиса извлечения блюд"""
    return _extractor_service


# ===== Дополнительные функции извлечения блюд (перенесены из presentation_handler) =====

def _upper_no_yo(s: str) -> str:
    return s.upper().replace('Ё', 'Е') if isinstance(s, str) else str(s).upper().replace('Ё', 'Е')


def detect_category_columns(df, category_row: int, category_name: str) -> List[int]:
    """
    Определяет индексы столбцов для указанной категории в таблице DataFrame.

    Args:
        df (pandas.DataFrame): Таблица с исходными данными Excel без заголовков.
        category_row (int): Номер строки (0-базовый индекс) с заголовком категории.
        category_name (str): Текст заголовка категории (в верхнем регистре), например "САЛАТЫ".

    Returns:
        List[int]: Список из трёх индексов столбцов [name_col, weight_col, price_col].
                   Если определить не удалось, возвращает правый набор [3, 4, 5].
    """
    try:
        row = df.iloc[category_row]
        category_column = None
        for col_idx in range(len(df.columns)):
            if pd.notna(df.iloc[category_row, col_idx]):
                cell_content = str(df.iloc[category_row, col_idx]).upper().replace('Ё', 'Е')
                if category_name in cell_content:
                    category_column = col_idx
                    break
        if category_column is None:
            return [3, 4, 5]
        if category_column <= 2:
            return [0, 1, 2]
        else:
            return [3, 4, 5]
    except Exception:
        return [3, 4, 5]


def _sanitize_price_string(raw: str) -> str:
    """Возвращает цену без валюты (руб/р/₽), только число (и ','), поддерживает варианты через '/'."""
    if not raw:
        return ""
    s = str(raw)
    # Разбиваем по '/', обрабатываем каждую часть отдельно
    parts = [p.strip() for p in s.split('/')]
    cleaned_parts = []
    for p in parts:
        # Ищем первое число с опциональной дробной частью
        m = re.search(r"(\d+(?:[\.,]\d{1,2})?)", p)
        if not m:
            cleaned_parts.append("")
            continue
        num = m.group(1).replace('.', ',')
        cleaned_parts.append(num)
    # Склеиваем назад, убираем пустые хвосты
    cleaned = [c for c in cleaned_parts if c != ""]
    return "/".join(cleaned)


def extract_dishes_from_excel_column(excel_path: str, category_keywords: List[str]) -> List[DishItem]:
    """
    Извлекает блюда из Excel при колоночной структуре данных (Название | Вес | Цена).

    Args:
        excel_path (str): Путь к Excel-файлу (.xlsx/.xlsm/.xls).
        category_keywords (List[str]): Список ключевых фраз для идентификации нужной категории
            (например: ['САЛАТЫ', 'ХОЛОДНЫЕ ЗАКУСКИ']). Поиск ведётся по строке заголовка.

    Returns:
        List[DishItem]: Список блюд с заполненными полями name/weight/price.
    """
    try:
        xls = pd.ExcelFile(excel_path)
        sheet_name = None
        for nm in xls.sheet_names:
            if 'касс' in str(nm).strip().lower():
                sheet_name = nm
                break
        if sheet_name is None and xls.sheet_names:
            sheet_name = xls.sheet_names[0]

        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, dtype=object)

        def row_text(row) -> str:
            parts = []
            for v in row:
                if pd.notna(v):
                    parts.append(str(v))
            return ' '.join(parts).strip()

        header_row = None
        category_columns = {}
        for i in range(min(20, len(df))):
            row_content = row_text(df.iloc[i]).upper().replace('Ё', 'Е')
            found_categories = 0
            for keyword_set in category_keywords:
                keywords = keyword_set.upper().split(' ')
                if any(kw.upper() in row_content for kw in keywords if len(kw) > 2):
                    found_categories += 1
            if found_categories > 0:
                header_row = i
                for col_idx, cell_value in enumerate(df.iloc[i]):
                    if pd.notna(cell_value):
                        cell_text = str(cell_value).upper().replace('Ё', 'Е')
                        for keyword_set in category_keywords:
                            keywords = keyword_set.upper().split(' ')
                            if any(kw.upper() in cell_text for kw in keywords if len(kw) > 2):
                                category_columns[col_idx] = keyword_set
                                break
                break

        if header_row is None or not category_columns:
            return []

        dishes: List[DishItem] = []
        for row_idx in range(header_row + 1, len(df)):
            row = df.iloc[row_idx]
            for col_idx, category in category_columns.items():
                category_matches = False
                for keyword_set in category_keywords:
                    keywords = keyword_set.upper().split(' ')
                    if any(kw.upper() in category.upper() for kw in keywords if len(kw) > 2):
                        category_matches = True
                        break
                if not category_matches:
                    continue
                if col_idx < len(row) and pd.notna(row.iloc[col_idx]):
                    dish_name = str(row.iloc[col_idx]).strip()
                    if dish_name and not dish_name.isupper() and len(dish_name) > 3:
                        weight = ""
                        price = ""
                        for offset in [1, 2, 3]:
                            if col_idx + offset < len(row) and pd.notna(row.iloc[col_idx + offset]):
                                cell_value = str(row.iloc[col_idx + offset]).strip()
                                if not weight and re.search(r'\d+.*?(?:г|шт|мл|л)', cell_value, re.IGNORECASE):
                                    weight = cell_value
                                if not price and re.search(r'\d+', cell_value) and not re.search(r'(?:г|шт|мл|л)', cell_value):
                                    price = _sanitize_price_string(cell_value)
                        dishes.append(DishItem(name=dish_name, weight=weight, price=price))
            if not any(pd.notna(cell) for cell in row):
                break
        return dishes
    except Exception as e:
        print(f"Ошибка при извлечении блюд категории {category_keywords}: {e}")
        return []


def extract_dishes_from_excel(excel_path: str, category_keywords: List[str]) -> List[DishItem]:
    """
    Унифицированный метод извлечения блюд по ключевым словам категории.
    Сначала пробует колоночную структуру, затем — построчную.

    Args:
        excel_path (str): Путь к Excel-файлу.
        category_keywords (List[str]): Ключевые слова для заголовка категории.

    Returns:
        List[DishItem]: Найденные блюда (может быть пустым списком).
    """
    dishes = extract_dishes_from_excel_column(excel_path, category_keywords)
    if dishes:
        return dishes
    return extract_dishes_from_excel_rows(excel_path, category_keywords)


def extract_dishes_from_excel_rows(excel_path: str, category_keywords: List[str]) -> List[DishItem]:
    """
    Извлекает блюда при построчной структуре: ищет строку заголовка категории,
    затем читает последующие строки до следующей категории/пустых строк.

    Args:
        excel_path (str): Путь к Excel-файлу.
        category_keywords (List[str]): Ключевые слова заголовка (например, ['ПЕРВЫЕ БЛЮДА']).

    Returns:
        List[DishItem]: Список найденных блюд с весом и ценой (если удалось распознать).
    """
    try:
        try:
            xls = pd.ExcelFile(excel_path)
            sheet_name = None
            for nm in xls.sheet_names:
                if 'касс' in str(nm).strip().lower():
                    sheet_name = nm
                    break
            if sheet_name is None and xls.sheet_names:
                sheet_name = xls.sheet_names[0]
        except Exception:
            sheet_name = 0

        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, dtype=object)

        def row_text(row) -> str:
            parts = []
            for v in row:
                if pd.notna(v):
                    parts.append(str(v))
            return ' '.join(parts).strip()

        category_row = None
        for i in range(min(50, len(df))):
            s = row_text(df.iloc[i]).upper().replace('Ё', 'Е')
            if not s:
                continue
            for keyword_set in category_keywords:
                keywords = keyword_set.upper().split(' ')
                if any(kw.upper() in s for kw in keywords if len(kw) > 2):
                    category_row = i
                    break
                elif all(kw.upper() in s for kw in keywords):
                    category_row = i
                    break
            if category_row is not None:
                break
        if category_row is None:
            return []

        units_pattern = r'(?:к?кал|ккал|г|гр|грамм(?:а|ов)?|мл|л|кг)'
        price_pattern = r'(?<!\\d)(\\d{1,6}(?:[\\.,]\\d{1,2})?)\\s*(?:руб\\w*|р\\.?|₽)?'

        known_cats = [
            'ЗАВТРАК', 'ПЕРВЫЕ БЛЮДА', 'ВТОРЫЕ БЛЮДА', 'ГАРНИР', 'НАПИТК', 'ДЕСЕРТ',
            'БЛЮДА ИЗ МЯСА', 'БЛЮДА ИЗ ПТИЦЫ', 'БЛЮДА ИЗ РЫБЫ', 'САЛАТЫ', 'ХОЛОДНЫЕ ЗАКУСКИ',
            'МЯСНЫЕ БЛЮДА', 'РЫБНЫЕ БЛЮДА', 'ГАРНИРЫ'
        ]

        def is_category_row(row) -> bool:
            s = row_text(row).upper()
            if not s:
                return False
            if any(k in s for k in known_cats):
                return True
            letters = ''.join(ch for ch in s if ch.isalpha())
            if letters and letters == letters.upper() and len(letters) >= 4:
                return True
            return False

        def extract_weight_from_row(row) -> str:
            for v in row:
                if pd.isna(v):
                    continue
                s = str(v)
                m = re.search(rf'(\d+[\.,]?\d*)\s*{units_pattern}', s, flags=re.IGNORECASE)
                if m:
                    qty = m.group(1).replace(',', '.')
                    unit_m = re.search(rf'{units_pattern}', s, flags=re.IGNORECASE)
                    unit = unit_m.group(0) if unit_m else ''
                    return f"{qty.replace('.', ',')} {unit}"
            return ''

        def is_weight_like(s: str) -> bool:
            return re.search(rf'{units_pattern}', s, flags=re.IGNORECASE) is not None

        def extract_price_from_row(row) -> Optional[str]:
            candidates = []
            for v in row:
                if pd.isna(v):
                    continue
                s = str(v)
                if is_weight_like(s):
                    continue
                # Ищем все числа (возможна последняя цена в строке)
                for m in re.finditer(r"(\d+(?:[\.,]\d{1,2})?)", s, flags=re.IGNORECASE):
                    num = m.group(1).replace(',', '.')
                    try:
                        val = float(num)
                    except ValueError:
                        continue
                    candidates.append(val)
            if not candidates:
                return None
            val = candidates[-1]
            if abs(val - int(val)) < 1e-6:
                return f"{int(val)}"
            else:
                return f"{str(val).replace('.', ',')}"

        dishes: List[DishItem] = []
        current_row = category_row + 1
        empty_streak = 0
        while current_row < len(df):
            row = df.iloc[current_row]
            s_join = row_text(row)
            if is_category_row(row):
                break
            if not s_join:
                empty_streak += 1
                if empty_streak >= 3:
                    break
                current_row += 1
                continue
            else:
                empty_streak = 0

            name = ''
            for v in row:
                if pd.notna(v):
                    t = str(v).strip()
                    if t:
                        name = t
                        break
            weight = extract_weight_from_row(row)
            price = extract_price_from_row(row)
            if name and not name.isupper():
                dishes.append(DishItem(name=name, weight=weight, price=price or ""))
            current_row += 1
        return dishes
    except Exception as e:
        print(f"Ошибка при извлечении блюд категории {category_keywords}: {e}")
        return []


def extract_dishes_from_excel_rows_with_stop(excel_path: str, category_keywords: List[str], stop_keywords: List[str]) -> List[DishItem]:
    """
    Построчное извлечение блюд до появления следующей категории-стоп слова.

    Args:
        excel_path (str): Путь к Excel-файлу.
        category_keywords (List[str]): Ключевые слова целевой категории (начало диапазона).
        stop_keywords (List[str]): Ключевые слова категорий, на которых нужно остановиться.

    Returns:
        List[DishItem]: Список блюд в пределах найденного диапазона.
    """
    try:
        xls = pd.ExcelFile(excel_path)
        sheet_name = None
        for nm in xls.sheet_names:
            if 'касс' in str(nm).strip().lower():
                sheet_name = nm
                break
        if sheet_name is None and xls.sheet_names:
            sheet_name = xls.sheet_names[0]
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, dtype=object)

        def row_text(row) -> str:
            parts = []
            for v in row:
                if pd.notna(v):
                    parts.append(str(v))
            return ' '.join(parts).strip()

        header_row = None
        for i in range(min(80, len(df))):
            s = _upper_no_yo(row_text(df.iloc[i]))
            if not s:
                continue
            for keyword_set in category_keywords:
                keywords = _upper_no_yo(keyword_set).split(' ')
                if any(kw and kw in s for kw in keywords if len(kw) > 2):
                    header_row = i
                    break
            if header_row is not None:
                break
        if header_row is None:
            return []

        units_pattern = r'(?:к?кал|ккал|г|гр|грамм(?:а|ов)?|мл|л|кг)'
        price_pattern = r'(?<!\\d)(\\d{1,6}(?:[\\.,]\\d{1,2})?)\\s*(?:руб\\w*|р\\.?|₽)?'

        def is_category_row(row) -> bool:
            s = _upper_no_yo(row_text(row))
            if not s:
                return False
            letters = ''.join(ch for ch in s if ch.isalpha())
            if letters and letters == letters.upper() and len(letters) >= 4:
                return True
            return False

        def extract_weight_from_row(row) -> str:
            for v in row:
                if pd.isna(v):
                    continue
                s = str(v)
                m = re.search(rf'(\\d+[\\.,]?\\d*)\\s*{units_pattern}', s, flags=re.IGNORECASE)
                if m:
                    qty = m.group(1).replace(',', '.')
                    unit_m = re.search(rf'{units_pattern}', s, flags=re.IGNORECASE)
                    unit = unit_m.group(0) if unit_m else ''
                    return f"{qty.replace('.', ',')} {unit}"
            return ''

        def is_weight_like(s: str) -> bool:
            return re.search(rf'{units_pattern}', s, flags=re.IGNORECASE) is not None

        def extract_price_from_row(row) -> Optional[str]:
            candidates = []
            for v in row:
                if pd.isna(v):
                    continue
                s = str(v)
                if is_weight_like(s):
                    continue
                for m in re.finditer(price_pattern, s, flags=re.IGNORECASE):
                    num = m.group(1).replace(',', '.')
                    try:
                        val = float(num)
                    except ValueError:
                        continue
                    candidates.append(val)
            if not candidates:
                return None
            val = candidates[-1]
            if abs(val - int(val)) < 1e-6:
                txt = f"{int(val)} руб."
            else:
                txt = f"{str(val).replace('.', ',')} руб."
            return txt

        dishes: List[DishItem] = []
        current_row = header_row + 1
        empty_streak = 0
        stop_upper = [_upper_no_yo(x) for x in stop_keywords]
        while current_row < len(df):
            row = df.iloc[current_row]
            s_join = row_text(row)
            s_upper = _upper_no_yo(s_join)
            if is_category_row(row) and any(st in s_upper for st in stop_upper):
                break
            if is_category_row(row) and s_join:
                break
            if not s_join:
                empty_streak += 1
                if empty_streak >= 3:
                    break
                current_row += 1
                continue
            else:
                empty_streak = 0
            name = ''
            for v in row:
                if pd.notna(v):
                    t = str(v).strip()
                    if t:
                        name = t
                        break
            weight = extract_weight_from_row(row)
            price = extract_price_from_row(row)
            if name and not name.isupper():
                dishes.append(DishItem(name=name, weight=weight, price=price or ""))
            current_row += 1
        return dishes
    except Exception as e:
        print(f"Ошибка при выборочном извлечении (до стоп-категории) {category_keywords}: {e}")
        return []


def extract_dishes_from_multiple_sheets(excel_path: str, sheet_names: List[str]) -> List[DishItem]:
    """
    Извлекает блюда, последовательно просматривая несколько листов по именам.

    Args:
        excel_path (str): Путь к Excel-файлу.
        sheet_names (List[str]): Приоритетный список имён листов для чтения.

    Returns:
        List[DishItem]: Все найденные блюда на перечисленных листах.
    """
    all_dishes: List[DishItem] = []
    try:
        xls = pd.ExcelFile(excel_path)
        for sheet_name in sheet_names:
            if sheet_name in xls.sheet_names:
                try:
                    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, dtype=object)
                    for i in range(len(df)):
                        for j in range(len(df.columns)):
                            if pd.notna(df.iloc[i, j]):
                                dish_name = str(df.iloc[i, j]).strip()
                                if (dish_name and not dish_name.isupper() and len(dish_name) > 3 and not dish_name.replace(' ', '').isdigit()):
                                    weight = ""
                                    price = ""
                                    for di in [-1, 0, 1]:
                                        for dj in [1, 2, 3]:
                                            try:
                                                if (i + di >= 0 and j + dj < len(df.columns) and i + di < len(df) and pd.notna(df.iloc[i + di, j + dj])):
                                                    cell_value = str(df.iloc[i + di, j + dj]).strip()
                                                    if not weight and re.search(r'\d+.*?(?:г|шт|мл|л)', cell_value, re.IGNORECASE):
                                                        weight = cell_value
                                                    if not price and re.search(r'\d+', cell_value) and not re.search(r'(?:г|шт|мл|л)', cell_value):
                                                        if cell_value.isdigit():
                                                            price = f"{cell_value} руб."
                                                        else:
                                                            price = cell_value
                                            except Exception:
                                                continue
                                    all_dishes.append(DishItem(name=dish_name, weight=weight, price=price))
                except Exception as e:
                    print(f"Ошибка при чтении листа {sheet_name}: {e}")
    except Exception as e:
        print(f"Ошибка при открытии файла: {e}")
    return all_dishes


def extract_salads_by_range(excel_path: str) -> List[DishItem]:
    """
    Извлекает салаты по диапазону: от 'САЛАТЫ И ХОЛОДНЫЕ ЗАКУСКИ' до 'СЭНДВИЧИ' или ближайшей следующей категории.

    Args:
        excel_path (str): Путь к Excel-файлу.

    Returns:
        List[DishItem]: Найденные салаты (название/вес/цена, если доступны).
    """
    try:
        xls = pd.ExcelFile(excel_path)
        sheet_name = None
        for nm in xls.sheet_names:
            if 'касс' in str(nm).strip().lower():
                sheet_name = nm
                break
        if sheet_name is None and xls.sheet_names:
            sheet_name = xls.sheet_names[0]
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, dtype=object)

        def row_text(row) -> str:
            parts = []
            for v in row:
                if pd.notna(v):
                    parts.append(str(v))
            return ' '.join(parts).strip()

        start_row = None
        end_row = None
        for i in range(len(df)):
            row_content = row_text(df.iloc[i]).upper().replace('Ё', 'Е')
            if start_row is None:
                if ('САЛАТ' in row_content and 'ХОЛОДН' in row_content and 'ЗАКУСК' in row_content) or \
                   ('САЛАТЫ' in row_content and ('ХОЛОДНЫЕ' in row_content or 'ЗАКУСКИ' in row_content)):
                    start_row = i
                    print(f"Найден заголовок салатов в строке {i + 1}: {row_content}")
                    continue
            if start_row is not None and end_row is None:
                if 'СЭНДВИЧ' in row_content:
                    end_row = i
                    print(f"Найден заголовок сэндвичей в строке {i + 1}: {row_content}")
                    break
        if start_row is None:
            print("Заголовок 'САЛАТЫ И ХОЛОДНЫЕ ЗАКУСКИ' не найден")
            return []
        if end_row is None:
            end_row = len(df)
            for i in range(start_row + 1, len(df)):
                row_content = row_text(df.iloc[i]).upper().replace('Ё', 'Е')
                if any(category in row_content for category in [
                    'ПЕРВЫЕ БЛЮДА', 'ВТОРЫЕ БЛЮДА', 'ГОРЯЧИЕ БЛЮДА', 
                    'МЯСНЫЕ БЛЮДА', 'РЫБНЫЕ БЛЮДА', 'ГАРНИРЫ', 'НАПИТКИ'
                ]):
                    end_row = i
                    break
        print(f"Обрабатываем салаты от строки {start_row + 1} до строки {end_row}")
        dishes: List[DishItem] = []
        for i in range(start_row + 1, end_row):
            if i >= len(df):
                break
            row = df.iloc[i]
            row_content = row_text(row)
            if not row_content.strip():
                continue
            if row_content.isupper() and len(row_content) > 10:
                continue
            dish_name = ""
            dish_weight = ""
            dish_price = ""
            for j, cell_value in enumerate(row):
                if pd.notna(cell_value):
                    cell_text = str(cell_value).strip()
                    if not cell_text:
                        continue
                    if not dish_name and not cell_text.isupper() and len(cell_text) > 3:
                        if not re.match(r'^\d+([.,]\d+)?\s*(руб|₽|р\.?)?$', cell_text) and \
                           not re.search(r'\d+\s*(г|гр|мл|л|шт)', cell_text, re.IGNORECASE):
                            dish_name = cell_text
                            continue
                    if not dish_weight and re.search(r'\d+.*?(г|гр|грамм|мл|л|кг|шт)', cell_text, re.IGNORECASE):
                        dish_weight = cell_text
                        continue
                    if not dish_price and re.search(r'\d+', cell_text):
                        if not re.search(r'г|гр|грамм|мл|л|кг|шт', cell_text, re.IGNORECASE):
                            if cell_text.replace('.', '').replace(',', '').isdigit():
                                dish_price = f"{cell_text} руб."
                            elif re.search(r'\d+.*?(руб|₽|р\.?)', cell_text, re.IGNORECASE):
                                dish_price = cell_text
                            else:
                                number_match = re.search(r'\d+([.,]\d+)?', cell_text)
                                if number_match:
                                    dish_price = f"{number_match.group()} руб."
            if dish_name and len(dish_name) > 2:
                dishes.append(DishItem(name=dish_name, weight=dish_weight, price=dish_price))
                print(f"Найден салат: {dish_name} | {dish_weight} | {dish_price}")
        print(f"Всего найдено салатов: {len(dishes)}")
        return dishes
    except Exception as e:
        print(f"Ошибка при извлечении салатов по диапазону: {e}")
        return []


def extract_salads_from_excel(excel_path: str) -> List[DishItem]:
    """
    Высокоуровневое извлечение салатов: сначала по точному диапазону, затем по альтернативным листам.

    Args:
        excel_path (str): Путь к Excel-файлу.

    Returns:
        List[DishItem]: Список салатов.
    """
    try:
        salads = extract_salads_by_range(excel_path)
        if salads:
            return salads
        return extract_dishes_from_multiple_sheets(excel_path, ['Хц', 'Холодные', 'Салаты', 'касса '])
    except Exception as e:
        print(f"Ошибка при извлечении салатов: {e}")
        return []


def extract_first_courses_by_range(excel_path: str) -> List[DishItem]:
    """
    Извлекает первые блюда по диапазону от 'ПЕРВЫЕ БЛЮДА' до следующей категории.

    Args:
        excel_path (str): Путь к Excel-файлу.

    Returns:
        List[DishItem]: Список первых блюд.
    """
    try:
        xls = pd.ExcelFile(excel_path)
        sheet_name = None
        for nm in xls.sheet_names:
            if 'касс' in str(nm).strip().lower():
                sheet_name = nm
                break
        if sheet_name is None and xls.sheet_names:
            sheet_name = xls.sheet_names[0]
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, dtype=object)

        def row_text(row) -> str:
            parts = []
            for v in row:
                if pd.notna(v):
                    parts.append(str(v))
            return ' '.join(parts).strip()

        start_row = None
        end_row = None
        for i in range(len(df)):
            row_content = row_text(df.iloc[i]).upper().replace('Ё', 'Е')
            if start_row is None:
                if 'ПЕРВЫЕ БЛЮДА' in row_content or ('ПЕРВЫЕ' in row_content and 'БЛЮДА' in row_content):
                    start_row = i
                    print(f"Найден заголовок первых блюд в строке {i + 1}: {row_content}")
                    continue
            if start_row is not None and end_row is None:
                if any(category in row_content for category in [
                    'САЛАТЫ', 'ХОЛОДНЫЕ ЗАКУСКИ', 'БЛЮДА ИЗ МЯСА', 'МЯСНЫЕ',
                    'БЛЮДА ИЗ ПТИЦЫ', 'ПТИЦ', 'БЛЮДА ИЗ РЫБЫ', 'РЫБНЫЕ',
                    'ГАРНИРЫ', 'НАПИТКИ', 'ДЕСЕРТЫ', 'ВТОРЫЕ БЛЮДА'
                ]):
                    end_row = i
                    print(f"Найден конец секции первых блюд в строке {i + 1}: {row_content}")
                    break
        if start_row is None:
            print("Заголовок 'ПЕРВЫЕ БЛЮДА' не найден")
            return []
        if end_row is None:
            end_row = min(start_row + 50, len(df))
        print(f"Обрабатываем первые блюда от строки {start_row + 1} до строки {end_row}")
        dishes: List[DishItem] = []
        for i in range(start_row + 1, end_row):
            if i >= len(df):
                break
            row = df.iloc[i]
            row_content = row_text(row)
            if not row_content.strip():
                continue
            if row_content.isupper() and len(row_content) > 10:
                continue
            dish_name = ""
            dish_weight = ""
            dish_price = ""
            if len(df.columns) > 4 and pd.notna(df.iloc[i, 4]):
                name_text = str(df.iloc[i, 4]).strip()
                if name_text and not name_text.isupper() and len(name_text) > 2:
                    dish_name = name_text
            if not dish_name:
                continue
            if len(df.columns) > 5 and pd.notna(df.iloc[i, 5]):
                weight_text = str(df.iloc[i, 5]).strip()
                if weight_text:
                    dish_weight = weight_text
            if len(df.columns) > 6 and pd.notna(df.iloc[i, 6]):
                price_text = str(df.iloc[i, 6]).strip()
                if price_text:
                    if price_text.replace('.', '').replace(',', '').isdigit():
                        dish_price = f"{price_text} руб."
                    else:
                        dish_price = price_text
            if dish_name and len(dish_name) > 2:
                dishes.append(DishItem(name=dish_name, weight=dish_weight, price=dish_price))
                print(f"Найдено первое блюдо: {dish_name} | {dish_weight} | {dish_price}")
        print(f"Всего найдено первых блюд: {len(dishes)}")
        return dishes
    except Exception as e:
        print(f"Ошибка при извлечении первых блюд по диапазону: {e}")
        return []


def extract_first_courses_from_excel(excel_path: str) -> List[DishItem]:
    """
    Высокоуровневое извлечение первых блюд: сначала точный диапазон, затем общий поиск по ключевым словам.

    Args:
        excel_path (str): Путь к Excel-файлу.

    Returns:
        List[DishItem]: Список первых блюд.
    """
    try:
        dishes = extract_first_courses_by_range(excel_path)
        if dishes:
            return dishes
        print("Поиск первых блюд через общую функцию...")
        keywords = ['ПЕРВЫЕ БЛЮДА', 'ПЕРВЫЕ']
        all_dishes = extract_dishes_from_excel(excel_path, keywords)
        print(f"Найдено {len(all_dishes)} первых блюд")
        return all_dishes
    except Exception as e:
        print(f"Ошибка при извлечении первых блюд: {e}")
        return []


def extract_meat_dishes_by_range(excel_path: str) -> List[DishItem]:
    """
    Извлекает мясные блюда по диапазону от 'БЛЮДА ИЗ МЯСА' до 'БЛЮДА ИЗ ПТИЦЫ' (или ближайшей следующей категории).

    Args:
        excel_path (str): Путь к Excel-файлу.

    Returns:
        List[DishItem]: Список мясных блюд.
    """
    try:
        xls = pd.ExcelFile(excel_path)
        sheet_name = None
        for nm in xls.sheet_names:
            if 'касс' in str(nm).strip().lower():
                sheet_name = nm
                break
        if sheet_name is None and xls.sheet_names:
            sheet_name = xls.sheet_names[0]
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, dtype=object)

        def row_text(row) -> str:
            parts = []
            for v in row:
                if pd.notna(v):
                    parts.append(str(v))
            return ' '.join(parts).strip()

        start_row = None
        end_row = None
        for i in range(len(df)):
            row_content = row_text(df.iloc[i]).upper().replace('Ё', 'Е')
            if start_row is None:
                if 'БЛЮДА ИЗ МЯСА' in row_content or 'МЯСНЫЕ БЛЮДА' in row_content:
                    start_row = i
                    print(f"Найден заголовок мясных блюд в строке {i + 1}: {row_content}")
                    continue
            if start_row is not None and end_row is None:
                if 'БЛЮДА ИЗ ПТИЦЫ' in row_content or ('ПТИЦ' in row_content and 'БЛЮДА' in row_content):
                    end_row = i
                    print(f"Найден заголовок блюд из птицы в строке {i + 1}: {row_content}")
                    break
        if start_row is None:
            print("Заголовок 'БЛЮДА ИЗ МЯСА' не найден")
            return []
        if end_row is None:
            end_row = len(df)
            for i in range(start_row + 1, len(df)):
                row_content = row_text(df.iloc[i]).upper().replace('Ё', 'Е')
                if any(category in row_content for category in [
                    'РЫБНЫЕ БЛЮДА', 'БЛЮДА ИЗ РЫБЫ', 'ГАРНИРЫ', 'НАПИТКИ', 
                    'ДЕСЕРТЫ', 'САЛАТЫ', 'СЭНДВИЧ'
                ]):
                    end_row = i
                    break
        print(f"Обрабатываем мясные блюда от строки {start_row + 1} до строки {end_row}")
        dishes: List[DishItem] = []
        for i in range(start_row + 1, end_row):
            if i >= len(df):
                break
            row = df.iloc[i]
            row_content = row_text(row)
            if not row_content.strip():
                continue
            if row_content.isupper() and len(row_content) > 10:
                continue
            dish_name = ""
            dish_weight = ""
            dish_price = ""
            if len(df.columns) > 4 and pd.notna(df.iloc[i, 4]):
                name_text = str(df.iloc[i, 4]).strip()
                if name_text and not name_text.isupper() and len(name_text) > 2:
                    dish_name = name_text
            if not dish_name:
                continue
            if len(df.columns) > 5 and pd.notna(df.iloc[i, 5]):
                weight_text = str(df.iloc[i, 5]).strip()
                if weight_text:
                    dish_weight = weight_text
            if len(df.columns) > 6 and pd.notna(df.iloc[i, 6]):
                price_text = str(df.iloc[i, 6]).strip()
                if price_text:
                    if price_text.replace('.', '').replace(',', '').isdigit():
                        dish_price = f"{price_text} руб."
                    else:
                        dish_price = price_text
            if dish_name and len(dish_name) > 2:
                dishes.append(DishItem(name=dish_name, weight=dish_weight, price=dish_price))
                print(f"Найдено блюдо из мяса: {dish_name} | {dish_weight} | {dish_price}")
        print(f"Всего найдено мясных блюд: {len(dishes)}")
        return dishes
    except Exception as e:
        print(f"Ошибка при извлечении мясных блюд по диапазону: {e}")
        return []


def extract_meat_dishes_from_excel(excel_path: str) -> List[DishItem]:
    """
    Высокоуровневое извлечение мясных блюд: диапазон, затем построчный поиск до стоп-категории,
    затем общий поиск с фильтрацией супов.

    Args:
        excel_path (str): Путь к Excel-файлу.

    Returns:
        List[DishItem]: Список мясных блюд (без первых блюд/супов).
    """
    try:
        dishes = extract_meat_dishes_by_range(excel_path)
        if dishes:
            return dishes
        stop_keywords = ['БЛЮДА ИЗ ПТИЦЫ', 'ПТИЦА', 'РЫБНЫЕ БЛЮДА', 'БЛЮДА ИЗ РЫБЫ']
        keywords = ['БЛЮДА ИЗ МЯСА', 'МЯСНЫЕ БЛЮДА', 'МЯСО']
        dishes = extract_dishes_from_excel_rows_with_stop(excel_path, keywords, stop_keywords)
        if dishes:
            print(f"Найдено {len(dishes)} мясных блюд через построчный поиск с остановкой")
            return dishes
        print("Поиск мясных блюд через колоночную структуру...")
        all_keywords = ['БЛЮДА ИЗ МЯСА', 'МЯСНЫЕ БЛЮДА', 'МЯСО']
        dishes = extract_dishes_from_excel(excel_path, all_keywords)
        if dishes:
            filtered_dishes: List[DishItem] = []
            for dish in dishes:
                if not any(soup_word in dish.name.lower() for soup_word in ['суп', 'бульон', 'солянка', 'борщ', 'щи', 'окрошка', 'харчо']):
                    filtered_dishes.append(dish)
            if filtered_dishes:
                return filtered_dishes
        return []
    except Exception as e:
        print(f"Ошибка при извлечении мясных блюд: {e}")
        return []


def extract_poultry_dishes_by_range(excel_path: str) -> List[DishItem]:
    """
    Извлекает блюда из птицы по диапазону от 'БЛЮДА ИЗ ПТИЦЫ' до 'БЛЮДА ИЗ РЫБЫ'.

    Args:
        excel_path (str): Путь к Excel-файлу.

    Returns:
        List[DishItem]: Список блюд из птицы.
    """
    try:
        xls = pd.ExcelFile(excel_path)
        sheet_name = None
        for nm in xls.sheet_names:
            if 'касс' in str(nm).strip().lower():
                sheet_name = nm
                break
        if sheet_name is None and xls.sheet_names:
            sheet_name = xls.sheet_names[0]
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, dtype=object)
        print(f"Размер файла: {len(df)} строк, {len(df.columns)} столбцов")

        def row_text(row) -> str:
            parts = []
            for v in row:
                if pd.notna(v):
                    parts.append(str(v))
            return ' '.join(parts).strip()

        if len(df.columns) <= 6:
            name_col, weight_col, price_col = 3, 4, 5
            print("Определена структура с 6 столбцами: D(название), E(вес), F(цена)")
        else:
            name_col, weight_col, price_col = 4, 5, 6
            print("Определена структура с 7+ столбцами: E(название), F(вес), G(цена)")

        start_row = None
        end_row = None
        for i in range(len(df)):
            row_content = row_text(df.iloc[i]).upper().replace('Ё', 'Е')
            if start_row is None:
                if 'БЛЮДА ИЗ ПТИЦЫ' in row_content or ('ПТИЦ' in row_content and 'БЛЮДА' in row_content):
                    start_row = i
                    print(f"Найден заголовок блюд из птицы в строке {i + 1}: {row_content}")
                    continue
            if start_row is not None and end_row is None:
                if 'БЛЮДА ИЗ РЫБЫ' in row_content or ('РЫБ' in row_content and 'БЛЮДА' in row_content):
                    end_row = i
                    print(f"Найден заголовок блюд из рыбы в строке {i + 1}: {row_content}")
                    break
        if start_row is None:
            print("Заголовок 'БЛЮДА ИЗ ПТИЦЫ' не найден")
            return []
        if end_row is None:
            end_row = len(df)
            for i in range(start_row + 1, len(df)):
                row_content = row_text(df.iloc[i]).upper().replace('Ё', 'Е')
                if any(category in row_content for category in ['ГАРНИРЫ', 'НАПИТКИ', 'ДЕСЕРТЫ', 'САЛАТЫ', 'СЭНДВИЧ']):
                    end_row = i
                    break
        print(f"Обрабатываем блюда из птицы от строки {start_row + 1} до строки {end_row}")
        print(f"Используем столбцы: {chr(65+name_col)}(название), {chr(65+weight_col)}(вес), {chr(65+price_col)}(цена)")
        dishes: List[DishItem] = []
        for i in range(start_row + 1, end_row):
            if i >= len(df):
                break
            row = df.iloc[i]
            row_content = row_text(row)
            if not row_content.strip():
                continue
            if row_content.isupper() and len(row_content) > 10:
                continue
            dish_name = ""
            dish_weight = ""
            dish_price = ""
            if len(df.columns) > name_col and pd.notna(df.iloc[i, name_col]):
                name_text = str(df.iloc[i, name_col]).strip()
                if (name_text and not name_text.isupper() and len(name_text) > 2 and 'БЛЮДА ИЗ' not in name_text.upper()):
                    dish_name = name_text
            if not dish_name:
                continue
            if len(df.columns) > weight_col and pd.notna(df.iloc[i, weight_col]):
                weight_text = str(df.iloc[i, weight_col]).strip()
                if weight_text:
                    dish_weight = weight_text
            if len(df.columns) > price_col and pd.notna(df.iloc[i, price_col]):
                price_text = str(df.iloc[i, price_col]).strip()
                if price_text:
                    if price_text.replace('.', '').replace(',', '').isdigit():
                        dish_price = f"{price_text} руб."
                    else:
                        dish_price = price_text
            if dish_name and len(dish_name) > 2:
                dishes.append(DishItem(name=dish_name, weight=dish_weight, price=dish_price))
                print(f"Найдено блюдо из птицы: {dish_name} | {dish_weight} | {dish_price}")
        print(f"Всего найдено блюд из птицы: {len(dishes)}")
        return dishes
    except Exception as e:
        print(f"Ошибка при извлечении блюд из птицы по диапазону: {e}")
        return []


def extract_poultry_dishes_from_excel(excel_path: str) -> List[DishItem]:
    """
    Высокоуровневое извлечение блюд из птицы: точный диапазон, затем перебор листов.

    Args:
        excel_path (str): Путь к Excel-файлу.

    Returns:
        List[DishItem]: Список блюд из птицы.
    """
    try:
        dishes = extract_poultry_dishes_by_range(excel_path)
        if dishes:
            return dishes
        print("Поиск блюд из птицы через листы...")
        return extract_dishes_from_multiple_sheets(excel_path, ['Раздача', 'Обед', 'Гц', 'Птица', 'касса '])
    except Exception as e:
        print(f"Ошибка при извлечении блюд из птицы: {e}")
        return []


def extract_fish_dishes_by_range(excel_path: str) -> List[DishItem]:
    """
    Извлекает рыбные блюда по диапазону от 'БЛЮДА ИЗ РЫБЫ' до следующей категории (обычно 'ГАРНИРЫ').

    Args:
        excel_path (str): Путь к Excel-файлу.

    Returns:
        List[DishItem]: Список рыбных блюд.
    """
    try:
        xls = pd.ExcelFile(excel_path)
        sheet_name = None
        for nm in xls.sheet_names:
            if 'обед' in str(nm).strip().lower():
                sheet_name = nm
                break
        if sheet_name is None:
            for nm in xls.sheet_names:
                if 'касс' in str(nm).strip().lower():
                    sheet_name = nm
                    break
        if sheet_name is None and xls.sheet_names:
            sheet_name = xls.sheet_names[0]
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, dtype=object)
        print(f"Используем лист: {sheet_name}")
        print(f"Размер листа: {len(df)} строк, {len(df.columns)} столбцов")

        def row_text(row) -> str:
            parts = []
            for v in row:
                if pd.notna(v):
                    parts.append(str(v))
            return ' '.join(parts).strip()

        start_row = None
        end_row = None
        for i in range(len(df)):
            row_content = row_text(df.iloc[i]).upper().replace('Ё', 'Е')
            if start_row is None:
                if 'БЛЮДА ИЗ РЫБЫ' in row_content or 'РЫБНЫЕ БЛЮДА' in row_content:
                    start_row = i
                    print(f"Найден заголовок рыбных блюд в строке {i + 1}: {row_content}")
                    continue
            if start_row is not None and end_row is None:
                if any(category in row_content for category in ['ГАРНИРЫ']):
                    end_row = i
                    print(f"Найден конец секции рыбных блюд в строке {i + 1}: {row_content}")
                    break
        if start_row is None:
            print("Заголовок 'БЛЮДА ИЗ РЫБЫ' не найден")
            return []
        if end_row is None:
            end_row = min(start_row + 50, len(df))
        print(f"Обрабатываем рыбные блюда от строки {start_row + 1} до строки {end_row}")
        dishes: List[DishItem] = []
        for i in range(start_row + 1, end_row):
            if i >= len(df):
                break
            row = df.iloc[i]
            row_content = row_text(row)
            if not row_content.strip():
                continue
            if row_content.isupper() and len(row_content) > 10:
                continue
            dish_name = ""
            dish_weight = ""
            dish_price = ""
            if len(df.columns) > 4 and pd.notna(df.iloc[i, 4]):
                name_text = str(df.iloc[i, 4]).strip()
                if name_text and not name_text.isupper() and len(name_text) > 2:
                    dish_name = name_text
            if not dish_name:
                continue
            if len(df.columns) > 5 and pd.notna(df.iloc[i, 5]):
                weight_text = str(df.iloc[i, 5]).strip()
                if weight_text:
                    dish_weight = weight_text
            if len(df.columns) > 6 and pd.notna(df.iloc[i, 6]):
                price_text = str(df.iloc[i, 6]).strip()
                if price_text:
                    if price_text.replace('.', '').replace(',', '').isdigit():
                        dish_price = f"{price_text} руб."
                    else:
                        dish_price = price_text
            if dish_name and len(dish_name) > 2:
                dishes.append(DishItem(name=dish_name, weight=dish_weight, price=dish_price))
                print(f"Найдено блюдо из рыбы: {dish_name} | {dish_weight} | {dish_price}")
        print(f"Всего найдено рыбных блюд: {len(dishes)}")
        return dishes
    except Exception as e:
        print(f"Ошибка при извлечении рыбных блюд по диапазону: {e}")
        return []


def extract_fish_dishes_from_excel(excel_path: str) -> List[DishItem]:
    """
    Высокоуровневое извлечение рыбных блюд: точный диапазон, затем перебор листов.

    Args:
        excel_path (str): Путь к Excel-файлу.

    Returns:
        List[DishItem]: Список рыбных блюд.
    """
    try:
        dishes = extract_fish_dishes_by_range(excel_path)
        if dishes:
            return dishes
        print("Поиск блюд из рыбы через листы...")
        return extract_dishes_from_multiple_sheets(excel_path, ['Обед', 'Гц', 'Рыба', 'касса '])
    except Exception as e:
        print(f"Ошибка при извлечении блюд из рыбы: {e}")
        return []


def extract_side_dishes_by_range(excel_path: str) -> List[DishItem]:
    """
    Извлекает гарниры по диапазону от 'ГАРНИРЫ' до следующей категории.

    Args:
        excel_path (str): Путь к Excel-файлу.

    Returns:
        List[DishItem]: Список гарниров.
    """
    try:
        xls = pd.ExcelFile(excel_path)
        sheet_name = None
        for nm in xls.sheet_names:
            if 'касс' in str(nm).strip().lower():
                sheet_name = nm
                break
        if sheet_name is None and xls.sheet_names:
            sheet_name = xls.sheet_names[0]
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, dtype=object)
        print(f"Размер файла: {len(df)} строк, {len(df.columns)} столбцов")

        def row_text(row) -> str:
            parts = []
            for v in row:
                if pd.notna(v):
                    parts.append(str(v))
            return ' '.join(parts).strip()

        gourmet_column = None
        start_row = None
        end_row = None
        for i in range(len(df)):
            row_content = row_text(df.iloc[i]).upper().replace('Ё', 'Е')
            if start_row is None:
                if 'ГАРНИРЫ' in row_content or 'ГАРНИР' in row_content:
                    start_row = i
                    print(f"Найден заголовок гарниров в строке {i + 1}: {row_content}")
                    for col_idx in range(len(df.columns)):
                        if pd.notna(df.iloc[i, col_idx]):
                            cell_content = str(df.iloc[i, col_idx]).upper().replace('Ё', 'Е')
                            if 'ГАРНИР' in cell_content:
                                gourmet_column = col_idx
                                print(f"Заголовок гарниров найден в столбце {chr(65 + col_idx)}")
                                break
                    continue
            if start_row is not None and end_row is None:
                if any(category in row_content for category in [
                    'НАПИТКИ', 'ДЕСЕРТЫ', 'САЛАТЫ', 'СЭНДВИЧ', 'ЗАКУСКИ',
                    'ПЕРВЫЕ БЛЮДА', 'ВТОРЫЕ БЛЮДА', 'БЛЮДА ИЗ', 'ВЫПЕЧКА'
                ]):
                    end_row = i
                    print(f"Найден конец секции гарниров в строке {i + 1}: {row_content}")
                    break
        if start_row is None:
            print("Заголовок 'ГАРНИРЫ' не найден")
            return []
        if end_row is None:
            end_row = min(start_row + 30, len(df))
        if gourmet_column is not None:
            if gourmet_column == 3:
                name_col, weight_col, price_col = 0, 1, 2
                print("Определена структура: гарниры в левой части A(название), B(вес), C(цена)")
            else:
                name_col, weight_col, price_col = 4, 5, 6
                print("Определена структура: гарниры в правой части E(название), F(вес), G(цена)")
        else:
            if len(df.columns) <= 6:
                name_col, weight_col, price_col = 0, 1, 2
                print("Определена структура по умолчанию для 6-столбцов: A(название), B(вес), C(цена)")
            else:
                name_col, weight_col, price_col = 4, 5, 6
                print("Определена структура по умолчанию для 7+-столбцов: E(название), F(вес), G(цена)")
        print(f"Обрабатываем гарниры от строки {start_row + 1} до строки {end_row}")
        print(f"Используем столбцы: {chr(65+name_col)}(название), {chr(65+weight_col)}(вес), {chr(65+price_col)}(цена)")
        dishes: List[DishItem] = []
        for i in range(start_row + 1, end_row):
            if i >= len(df):
                break
            row = df.iloc[i]
            row_content = row_text(row)
            if not row_content.strip():
                continue
            if row_content.isupper() and len(row_content) > 10:
                continue
            dish_name = ""
            dish_weight = ""
            dish_price = ""
            if len(df.columns) > name_col and pd.notna(df.iloc[i, name_col]):
                name_text = str(df.iloc[i, name_col]).strip()
                if (name_text and not name_text.isupper() and len(name_text) > 2 and
                    'ГАРНИР' not in name_text.upper() and 
                    not re.match(r'^\d+.*?(г|мл|л|шт)', name_text) and
                    not name_text.replace('.', '').replace(',', '').isdigit()):
                    dish_name = name_text
            if not dish_name:
                continue
            if len(df.columns) > weight_col and pd.notna(df.iloc[i, weight_col]):
                weight_text = str(df.iloc[i, weight_col]).strip()
                if weight_text:
                    dish_weight = weight_text
            if len(df.columns) > price_col and pd.notna(df.iloc[i, price_col]):
                price_text = str(df.iloc[i, price_col]).strip()
                if price_text:
                    if price_text.replace('.', '').replace(',', '').isdigit():
                        dish_price = f"{price_text} руб."
                    else:
                        dish_price = price_text
            if dish_name and len(dish_name) > 2:
                dishes.append(DishItem(name=dish_name, weight=dish_weight, price=dish_price))
                print(f"Найден гарнир: {dish_name} | {dish_weight} | {dish_price}")
        print(f"Всего найдено гарниров: {len(dishes)}")
        return dishes
    except Exception as e:
        print(f"Ошибка при извлечении гарниров по диапазону: {e}")
        return []


def extract_side_dishes_from_excel(excel_path: str) -> List[DishItem]:
    """
    Высокоуровневое извлечение гарниров: точный диапазон, затем общий/альтернативный поиск.

    Args:
        excel_path (str): Путь к Excel-файлу.

    Returns:
        List[DishItem]: Список гарниров.
    """
    try:
        dishes = extract_side_dishes_by_range(excel_path)
        if dishes:
            return dishes
        print("Поиск гарниров через общую функцию...")
        keywords = ['ГАРНИРЫ', 'ГАРНИР']
        dishes = extract_dishes_from_excel(excel_path, keywords)
        if dishes:
            print(f"Найдено {len(dishes)} гарниров через общую функцию")
            return dishes
        print("Последняя попытка - поиск гарниров через листы...")
        return extract_dishes_from_multiple_sheets(excel_path, ['Раздача', 'Обед', 'Гц', 'Гарниры', 'касса '])
    except Exception as e:
        print(f"Ошибка при извлечении гарниров: {e}")
        return []


def extract_fish_dishes_from_column_e(excel_path: str) -> List[DishItem]:
    """
    Извлекает рыбные блюда только из диапазона 'БЛЮДА ИЗ РЫБЫ' до 'ГАРНИРЫ',
    используя колонки, в которых расположен заголовок раздела.

    Args:
        excel_path (str): Путь к Excel-файлу.

    Returns:
        List[DishItem]: Список рыбных блюд указанного диапазона.
    """
    try:
        xls = pd.ExcelFile(excel_path)
        sheet_name = None
        for nm in xls.sheet_names:
            if 'касс' in str(nm).strip().lower():
                sheet_name = nm
                break
        if sheet_name is None and xls.sheet_names:
            sheet_name = xls.sheet_names[0]
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, dtype=object)

        def row_text(row) -> str:
            parts = []
            for v in row:
                if pd.notna(v):
                    parts.append(str(v))
            return ' '.join(parts).strip()

        start_row = None
        end_row = None
        fish_columns = None
        for i in range(len(df)):
            row_content = row_text(df.iloc[i]).upper().replace('Ё', 'Е')
            if start_row is None:
                if 'БЛЮДА ИЗ РЫБЫ' in row_content or ('РЫБН' in row_content and 'БЛЮДА' in row_content):
                    start_row = i
                    print(f"Найден заголовок рыбных блюд в строке {i + 1}: {row_content}")
                    for col_idx in range(len(df.columns)):
                        if pd.notna(df.iloc[i, col_idx]):
                            cell_content = str(df.iloc[i, col_idx]).upper().replace('Ё', 'Е')
                            if 'БЛЮДА ИЗ РЫБЫ' in cell_content or ('РЫБН' in cell_content and 'БЛЮДА' in cell_content):
                                fish_columns = [col_idx, col_idx + 1, col_idx + 2] if col_idx + 2 < len(df.columns) else [col_idx]
                                print(f"Рыбные блюда находятся в столбцах: {fish_columns}")
                                break
                    continue
            if start_row is not None and end_row is None:
                if 'ГАРНИРЫ' in row_content or 'ГАРНИР' in row_content:
                    end_row = i
                    print(f"Найден заголовок гарниров в строке {i + 1}: {row_content}")
                    print(f"Останавливаемся на гарнирах, не включая их")
                    break
                if any(category in row_content for category in [
                    'НАПИТКИ', 'ДЕСЕРТЫ', 'САЛАТЫ', 'СЭНДВИЧ', 'ЗАКУСКИ',
                    'ПЕРВЫЕ БЛЮДА', 'ВТОРЫЕ БЛЮДА', 'ВЫПЕЧКА', 'ХЛЕБ'
                ]):
                    end_row = i
                    print(f"Найден конец секции в строке {i + 1}: {row_content}")
                    break
        if start_row is None or fish_columns is None:
            print("Заголовок 'БЛЮДА ИЗ РЫБЫ' не найден или не удалось определить столбцы")
            return []
        if end_row is None:
            end_row = min(start_row + 100, len(df))
        print(f"Извлекаем данные от строки {start_row + 1} до строки {end_row}")
        dishes: List[DishItem] = []
        for i in range(start_row + 1, end_row):
            if i >= len(df):
                break
            row = df.iloc[i]
            row_content = row_text(row)
            if not row_content.strip():
                continue
            if row_content.isupper() and len(row_content) > 10:
                continue
            dish_name = ""
            dish_weight = ""
            dish_price = ""
            row_values = []
            for col_idx in fish_columns:
                if col_idx < len(df.columns) and pd.notna(df.iloc[i, col_idx]):
                    cell_text = str(df.iloc[i, col_idx]).strip()
                    if cell_text:
                        row_values.append(cell_text)
            if not row_values:
                continue
            if row_values:
                potential_name = row_values[0]
                if (not potential_name.isupper() and len(potential_name) > 2 and
                    not re.match(r'^\d+([.,]\d+)?\s*(руб|₽|р\.?)?$', potential_name) and
                    not re.search(r'\d+\s*(г|гр|грамм|мл|л|кг|шт)', potential_name, re.IGNORECASE)):
                    dish_name = potential_name
            if not dish_name:
                continue
            for value in row_values[1:]:
                if not dish_weight and re.search(r'\d+.*?(г|гр|грамм|мл|л|кг|шт)', value, re.IGNORECASE):
                    dish_weight = value
                    continue
                if not dish_price and re.search(r'\d+', value):
                    if not re.search(r'г|гр|грамм|мл|л|кг|шт', value, re.IGNORECASE):
                        if value.replace('.', '').replace(',', '').isdigit():
                            dish_price = f"{value} руб."
                        elif re.search(r'\d+.*?(руб|₽|р\.?)', value, re.IGNORECASE):
                            dish_price = value
                        else:
                            number_match = re.search(r'\d+([.,]\d+)?', value)
                            if number_match:
                                dish_price = f"{number_match.group()} руб."
            if dish_name and len(dish_name) > 2:
                dishes.append(DishItem(name=dish_name, weight=dish_weight, price=dish_price))
                print(f"Найдено блюдо: {dish_name} | {dish_weight} | {dish_price}")
        print(f"Всего найдено рыбных блюд: {len(dishes)}")
        return dishes
    except Exception as e:
        print(f"Ошибка при извлечении рыбных блюд из столбца E: {e}")
        return []
