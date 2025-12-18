#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from typing import Iterable, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

from app.services.dish_extractor import DishItem


def _normalize_price(price: str) -> str:
    if price is None:
        return ""
    s = str(price).strip()
    if not s:
        return ""
    # если просто число — добавим ₽
    if s.replace(" ", "").replace(",", ".").replace(".", "", 1).isdigit():
        return f"{s} ₽"
    return s


def _normalize_weight(weight: str) -> str:
    if weight is None:
        return ""
    return str(weight).strip()


def create_pricelist_xlsx(dishes: List[DishItem], output_path: str, *, title: Optional[str] = None) -> None:
    """Создает Excel с ценниками.

    Формат одного ценника:
      - 1 строка: название (объединение на 3 столбца)
      - 2 строка: "Вес" + значение (значение объединено на 2 столбца)
      - 3 строка: "Цена" + значение (значение объединено на 2 столбца)
      - 4 строка: пустая (разделитель)

    Каждый ценник идет ниже предыдущего.
    """

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ценники"

    # Базовые ширины (ценник = 3 столбца)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    name_font = Font(name="Calibri", size=14, bold=True)
    label_font = Font(name="Calibri", size=11, bold=True)
    value_font = Font(name="Calibri", size=12)
    price_font = Font(name="Calibri", size=14, bold=True)

    align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if title:
        # Небольшой заголовок сверху (не обязателен)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        c = ws.cell(row=1, column=1)
        c.value = title
        c.font = Font(name="Calibri", size=14, bold=True)
        c.alignment = align_center_wrap
        start_row = 3
    else:
        start_row = 1

    block_height = 4

    for idx, dish in enumerate(dishes):
        top = start_row + idx * (block_height + 1)  # +1 строка разделителя между ценниками

        # 1) Название (merge 3 columns)
        ws.merge_cells(start_row=top, start_column=1, end_row=top, end_column=3)
        name_cell = ws.cell(row=top, column=1)
        name_cell.value = dish.name
        name_cell.font = name_font
        name_cell.alignment = align_center_wrap

        # 2) Вес
        ws.cell(row=top + 1, column=1, value="Вес").font = label_font
        ws.cell(row=top + 1, column=1).alignment = align_left
        ws.merge_cells(start_row=top + 1, start_column=2, end_row=top + 1, end_column=3)
        w_cell = ws.cell(row=top + 1, column=2)
        w_cell.value = _normalize_weight(dish.weight)
        w_cell.font = value_font
        w_cell.alignment = align_left

        # 3) Цена
        ws.cell(row=top + 2, column=1, value="Цена").font = label_font
        ws.cell(row=top + 2, column=1).alignment = align_left
        ws.merge_cells(start_row=top + 2, start_column=2, end_row=top + 2, end_column=3)
        p_cell = ws.cell(row=top + 2, column=2)
        p_cell.value = _normalize_price(dish.price)
        p_cell.font = price_font
        p_cell.alignment = align_left

        # 4) пустая строка внутри блока (чтобы было "поле" как в ценнике)
        ws.merge_cells(start_row=top + 3, start_column=1, end_row=top + 3, end_column=3)
        ws.cell(row=top + 3, column=1).value = ""

        # Рамки: проставим border на области 3x3 (первые 3 строки) + 4я строка (merged) тоже с рамкой
        for r in range(top, top + 4):
            for c in range(1, 4):
                cell = ws.cell(row=r, column=c)
                cell.border = border

        # Высоты строк под печать
        ws.row_dimensions[top].height = 32
        ws.row_dimensions[top + 1].height = 18
        ws.row_dimensions[top + 2].height = 22
        ws.row_dimensions[top + 3].height = 10

    wb.save(output_path)
