import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path
from datetime import datetime
import re
from typing import List, Dict, Optional, Tuple

from app.services.dish_extractor import extract_categorized_dishes_from_menu, extract_date_from_menu, extract_dishes_from_column_d7_d38, extract_column_a7_a43_skip_a30

class BrokerageJournalGenerator:
    """Генератор бракеражного журнала на основе меню"""
    
    def __init__(self):
        pass
    
    def extract_date_from_menu(self, menu_path: str) -> Optional[datetime]:
        """Извлекает дату из файла меню"""
        return extract_date_from_menu(menu_path)
    

    def extract_categorized_dishes(self, menu_path: str) -> Dict[str, List[str]]:
        """Извлекает блюда из меню, распределяя по категориям:
        завтраки, салаты и холодные закуски, первые блюда, блюда из мяса, блюда из птицы, блюда из рыбы, гарниры
        """
        return extract_categorized_dishes_from_menu(menu_path)


    def _is_section_header(self, text: str) -> bool:
        """Определяет, является ли строка заголовком раздела (не блюдом)."""
        if not text:
            return False
        txt = str(text).strip()
        lower = txt.lower()
        
        # Точные заголовки разделов, которые точно нужно исключить
        exact_headers = [
            "завтраки", "салаты и холодные закуски", "первые блюда", "блюда из мяса",
            "блюда из птицы", "блюда из рыбы", "гарниры", "сэндвичи", "сендвичи"
        ]
        
        # Проверяем точное совпадение с заголовками (без учета регистра)
        if lower in exact_headers:
            return True
            
        # Строки полностью в верхнем регистре и очень короткие (вероятно заголовки)
        # Важно: не отбрасываем обычные блюда типа "Салат ..." — проверяем только формы разделов
        if txt.isupper() and len(txt) <= 40 and any(h in lower for h in [
            'блюда', 'салаты', 'закуск', 'первые', 'вторые', 'гарниры', 'напит', 'сэндвич']):
            return True
            
        return False
    
    def _should_exclude_by_name(self, name: str) -> bool:
        """Фильтр блюд, которые не нужно вставлять независимо от категории."""
        nm = str(name).strip().lower()
        if not nm:
            return True
        banned_parts = ['пельмен', 'варени', 'сэндвич', 'сендвич']
        return any(bp in nm for bp in banned_parts)
    
    def create_brokerage_journal(self, menu_path: str, template_path: str, output_path: str) -> Tuple[bool, str]:
        """Создает бракеражный журнал на основе меню. Левый столбец (A) формируется копированием A6..A* до заголовка "СЭНДВИЧИ", без строк-заголовков "ЗАВТРАКИ" и "САЛАТЫ И ХОЛОДНЫЕ ЗАКУСКИ". Правый столбец (G) — остальные категории."""
        try:
            # Проверяем существование шаблона
            if not Path(template_path).exists():
                return False, f"Шаблон бракеражного журнала не найден: {template_path}"
            
            # Извлекаем дату из меню
            menu_date = self.extract_date_from_menu(menu_path)
            if not menu_date:
                menu_date = datetime.now()
            
            # Извлекаем блюда по категориям (для правого столбца)
            categories = self.extract_categorized_dishes(menu_path)
            
            # Отладочный вывод результата
            print(f"\nРезультат извлечения категорий:")
            for category, dishes in categories.items():
                print(f"{category}: {len(dishes)} блюд - {dishes}")
            
            # Левый столбец (A): используем новую функцию для извлечения блюд из D7:D38
            print("\nИзвлекаем блюда для левого столбца из диапазона D7:D38...")
            # Извлекаем сырые блюда для правой колонки G прямо из D7:D38 (без 4 заголовков)
            right_source = extract_dishes_from_column_d7_d38(menu_path)
            
            # Для левого столбца используем прежнюю логику фильтрации и дополнения категориями
            filtered_left = []
            for dish_name in right_source:
                if not self._is_section_header(dish_name) and not self._should_exclude_by_name(dish_name):
                    filtered_left.append(dish_name)
            cat_left = []
            cat_left.extend(categories.get('завтрак', []))
            cat_left.extend(categories.get('салат', []))
            merged: List[str] = []
            seen = set()
            for n in (filtered_left + cat_left):
                if not n:
                    continue
                if self._is_section_header(n) or self._should_exclude_by_name(n):
                    continue
                key = n.strip()
                if key in seen:
                    continue
                seen.add(key)
                merged.append(key)
            left_list = merged[:37]
            print(f"\nКоличество блюд для левого столбца (A): {len(left_list)}")
            
            # Также готовим данные для левого столбца A7..A43 напрямую из меню (пропуская A30)
            a_map = extract_column_a7_a43_skip_a30(menu_path)
            
            # Открываем шаблон и выбираем целевой лист
            wb = openpyxl.load_workbook(template_path)
            # Ищем лист, где в первых строчках встречается "наименование блюда"; иначе active
            ws = None
            for ws_candidate in wb.worksheets:
                found = False
                for r in range(1, min(21, ws_candidate.max_row + 1)):
                    for c in range(1, min(10, ws_candidate.max_column + 1)):
                        v = ws_candidate.cell(row=r, column=c).value
                        if v and ('наимен' in str(v).lower()) and ('блюд' in str(v).lower()):
                            found = True
                            break
                    if found:
                        break
                if found:
                    ws = ws_candidate
                    break
            if ws is None:
                ws = wb.active
            
            # Обновляем дату в шаблоне (строка 3, колонка 1)
            russian_months = {
                1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля', 5: 'мая', 6: 'июня',
                7: 'июля', 8: 'августа', 9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'
            }
            date_str_display = f"{menu_date.day} {russian_months.get(menu_date.month, 'unknown')}"
            ws.cell(row=3, column=1, value=date_str_display)
            
            # Форматируем дату для названия листа
            date_str = menu_date.strftime('%d.%m.%y')
            ws.title = date_str
            
            # Пытаемся найти строку заголовков таблицы ("НАИМЕНОВАНИЕ БЛЮДА")
            # Сканируем до 20 первых строк, по всем колонкам A..I
            header_row = None
            for r in range(1, min(21, ws.max_row + 1)):
                for c in range(1, 10):
                    v = ws.cell(row=r, column=c).value
                    if v and ('наимен' in str(v).lower()) and ('блюд' in str(v).lower()):
                        header_row = r
                        break
                if header_row is not None:
                    break
            if header_row is None:
                # Фолбэк: предполагаем, что данные начинаются с 6-й строки (A6)
                header_row = 5
            
            start_row = header_row + 1  # первая строка для блюд
            
            # Диапазон для записи берём до конца листа (чтобы писать и в пустые строки)
            stop_row = ws.max_row + 1
            
            # Применяем дополнительный фильтр по названиям (не вставлять пельмени/сэндвичи)
            left_list = [n for n in left_list if not self._should_exclude_by_name(n)]
            # Для правого столбца используем исходный список из D7:D38 без добавок (right_source)

            # Левый столбец A7..A43:
            # - если меню уже в формате "Касса" (A7..A43) — берём напрямую оттуда,
            # - иначе используем извлеченные категории (left_list).
            values_a = [a_map[r] for r in sorted(a_map.keys())] if a_map else left_list
            inserted_left = 0
            for idx, name in enumerate(values_a[:37]):  # максимум A7..A43
                ws.cell(row=7 + idx, column=1, value=name)
                inserted_left += 1
            # Очищаем хвост до A43
            for r in range(7 + inserted_left, 44):
                ws.cell(row=r, column=1, value=None)

            # Правый столбец (G7..G34): жёсткая перезапись списком из D7:D38 (без 4 заголовков)
            inserted_right = 0
            for idx, name in enumerate(right_source[:28]):  # G7..G34 — 28 строк
                ws.cell(row=7 + idx, column=7, value=name)
                inserted_right += 1

            wb.save(output_path)
            return True, f"Бракеражный журнал успешно создан: {date_str} (A: {inserted_left}, G: {inserted_right})"
        except Exception as e:
            return False, f"Ошибка при создании бракеражного журнала: {str(e)}"
    
    def _create_header(self, ws, date: datetime):
        """Создает заголовок бракеражного журнала"""
        # Дата
        date_str = date.strftime("%d %B").replace('September', 'сентября').replace('August', 'августа')
        ws.cell(row=3, column=1, value=date_str)
        
        # Заголовок
        ws.cell(row=1, column=1, value="БРАКЕРАЖНЫЙ ЖУРНАЛ")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        
        # Объединяем ячейки для заголовка
        ws.merge_cells('A1:G1')


def create_brokerage_journal_from_menu(menu_path: str, template_path: str, output_path: str) -> Tuple[bool, str]:
    """Удобная функция для создания бракеражного журнала"""
    generator = BrokerageJournalGenerator()
    return generator.create_brokerage_journal(menu_path, template_path, output_path)
