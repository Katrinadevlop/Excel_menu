#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Тестовый скрипт для проверки правильного разделения категорий завтраков и салатов
"""

import sys
from pathlib import Path

# Добавляем корневую директорию проекта в путь
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from app.services.menu_template_filler import MenuTemplateFiller
import openpyxl

def test_category_detection():
    """Тестирует определение категорий в шаблоне"""
    print("=" * 60)
    print("Тест определения категорий завтраков и салатов")
    print("=" * 60)
    
    # Путь к шаблону
    template_path = project_root / "templates" / "Шаблон меню пример.xlsx"
    
    if not template_path.exists():
        print(f"❌ Шаблон не найден: {template_path}")
        return
    
    # Открываем шаблон
    wb = openpyxl.load_workbook(template_path)
    
    # Ищем лист "Касса"
    ws = None
    for sheet in wb.worksheets:
        if 'касс' in sheet.title.lower():
            ws = sheet
            break
    
    if not ws:
        ws = wb.active
        print(f"⚠️ Лист 'Касса' не найден, используем: {ws.title}")
    else:
        print(f"✓ Используем лист: {ws.title}")
    
    # Создаём экземпляр MenuTemplateFiller
    filler = MenuTemplateFiller()
    
    print("\n" + "=" * 60)
    print("Проверка поиска заголовков")
    print("=" * 60)
    
    # Тест 1: Поиск заголовка завтраков
    print("\n1. Поиск заголовка ЗАВТРАКОВ в колонке A:")
    breakfast_start = filler.find_data_start_row(ws, col=1, category='завтрак')
    print(f"   Начало данных завтраков: строка {breakfast_start}")
    if breakfast_start > 0:
        header_row = breakfast_start - 1
        header_val = ws.cell(row=header_row, column=1).value
        print(f"   Заголовок в строке {header_row}: '{header_val}'")
    
    # Тест 2: Поиск конца завтраков (до салатов)
    print("\n2. Поиск конца блока ЗАВТРАКОВ (до заголовка салатов):")
    breakfast_end = filler.find_end_row_until_salads(ws, breakfast_start)
    print(f"   Конец данных завтраков: строка {breakfast_end}")
    if breakfast_end > 0:
        next_header_row = breakfast_end + 1
        next_header_val = ws.cell(row=next_header_row, column=1).value
        print(f"   Следующий заголовок в строке {next_header_row}: '{next_header_val}'")
    
    # Тест 3: Поиск заголовка салатов
    print("\n3. Поиск заголовка САЛАТОВ в колонке A:")
    salad_start = filler.find_data_start_row(ws, col=1, category='салат')
    print(f"   Начало данных салатов: строка {salad_start}")
    if salad_start > 0:
        header_row = salad_start - 1
        header_val = ws.cell(row=header_row, column=1).value
        print(f"   Заголовок в строке {header_row}: '{header_val}'")
    
    # Тест 4: Поиск конца салатов
    print("\n4. Поиск конца блока САЛАТОВ:")
    salad_end = filler.find_category_end_row(ws, col=1, start_row=salad_start, category='салат')
    print(f"   Конец данных салатов: строка {salad_end}")
    
    # Проверка результатов
    print("\n" + "=" * 60)
    print("Результаты проверки")
    print("=" * 60)
    
    success = True
    
    # Проверка 1: Завтраки должны быть в начале (до строки 15)
    if breakfast_start < 15:
        print(f"✓ Завтраки найдены в начале файла (строка {breakfast_start})")
    else:
        print(f"❌ ОШИБКА: Завтраки найдены слишком поздно (строка {breakfast_start})")
        success = False
    
    # Проверка 2: Салаты должны быть после завтраков (после строки 20)
    if salad_start >= 20:
        print(f"✓ Салаты найдены после завтраков (строка {salad_start})")
    else:
        print(f"❌ ОШИБКА: Салаты найдены слишком рано (строка {salad_start})")
        success = False
    
    # Проверка 3: Конец завтраков должен быть до начала салатов
    if breakfast_end < salad_start:
        print(f"✓ Завтраки заканчиваются до салатов (строки {breakfast_start}-{breakfast_end})")
    else:
        print(f"❌ ОШИБКА: Завтраки перекрывают салаты (конец={breakfast_end}, начало салатов={salad_start})")
        success = False
    
    # Проверка 4: Диапазоны не должны пересекаться
    breakfast_range = set(range(breakfast_start, breakfast_end + 1))
    salad_range = set(range(salad_start, salad_end + 1))
    overlap = breakfast_range & salad_range
    
    if not overlap:
        print(f"✓ Диапазоны завтраков и салатов не пересекаются")
    else:
        print(f"❌ ОШИБКА: Диапазоны пересекаются в строках: {sorted(overlap)}")
        success = False
    
    print("\n" + "=" * 60)
    if success:
        print("✅ ВСЕ ПРОВЕРКИ ПРОШЛИ УСПЕШНО!")
    else:
        print("❌ ОБНАРУЖЕНЫ ОШИБКИ В РАЗДЕЛЕНИИ КАТЕГОРИЙ")
    print("=" * 60)
    
    wb.close()

if __name__ == "__main__":
    test_category_detection()
