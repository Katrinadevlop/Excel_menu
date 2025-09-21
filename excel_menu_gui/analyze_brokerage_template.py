import pandas as pd
import openpyxl
from pathlib import Path

def analyze_brokerage_template():
    """Анализирует структуру шаблона бракеражного журнала"""
    template_path = r"C:\Users\katya\Desktop\menurepit\excel_menu_gui\templates\Бракеражный журнал шаблон.xlsx"
    
    if not Path(template_path).exists():
        print(f"❌ Шаблон не найден: {template_path}")
        return
    
    print(f"📊 Анализ шаблона: {Path(template_path).name}")
    
    try:
        # Открываем с openpyxl
        wb = openpyxl.load_workbook(template_path, data_only=True)
        print(f"📋 Листы в файле: {wb.sheetnames}")
        
        # Анализируем первый лист
        ws = wb.active
        print(f"\n📄 Активный лист: {ws.title}")
        print(f"📏 Размер: {ws.max_row} строк, {ws.max_column} колонок")
        
        print("\n🔍 Структура шаблона:")
        for row in range(1, min(31, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(10, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    row_data.append(str(cell.value))
                else:
                    row_data.append("")
            
            # Показываем только строки с содержимым
            if any(row_data):
                print(f"Строка {row:2d}: {row_data}")
        
        # Ищем места для даты
        print("\n📅 Поиск места для даты:")
        for row in range(1, min(11, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and "дата" in str(cell.value).lower():
                    print(f"  Поле даты в {row},{col}: '{cell.value}'")
        
        # Ищем таблицу с блюдами
        print("\n🍽️ Поиск таблицы блюд:")
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell_str = str(cell.value).lower()
                    if "наименование" in cell_str and "блюд" in cell_str:
                        print(f"  Заголовок таблицы в {row},{col}: '{cell.value}'")
                        
                        # Показываем несколько строк после заголовка
                        print("  Структура таблицы:")
                        for r in range(row, min(row + 10, ws.max_row + 1)):
                            table_row = []
                            for c in range(1, min(8, ws.max_column + 1)):
                                table_cell = ws.cell(row=r, column=c)
                                table_row.append(str(table_cell.value) if table_cell.value else "")
                            print(f"    Строка {r}: {table_row}")
                        break
        
    except Exception as e:
        print(f"❌ Ошибка при анализе: {e}")

if __name__ == "__main__":
    analyze_brokerage_template()
