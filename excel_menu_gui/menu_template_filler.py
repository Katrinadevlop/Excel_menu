#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Модуль для заполнения шаблона меню данными из другого файла меню
"""

import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import re

from dish_extractor import extract_categorized_dishes_from_menu, extract_date_from_menu


class MenuTemplateFiller:
    """Заполняет шаблон меню данными из другого файла"""
    
    def __init__(self):
        self.categories_mapping = {
            'завтрак': 'завтраки',
            'салат': 'холодные закуски и салаты', 
            'первое': 'первые блюда',
            'мясо': 'блюда из мяса',
            'курица': 'блюда из курицы',
            'птица': 'блюда из птицы',  # Добавляем птицу
            'рыба': 'блюда из рыбы',
            'гарнир': 'гарниры'
        }
    
    def extract_categorized_dishes(self, menu_path: str) -> Dict[str, List[str]]:
        """Извлекает блюда по категориям из файла меню (используем уже готовую логику)"""
        return extract_categorized_dishes_from_menu(menu_path)
    
    def find_column_by_header(self, ws, header_text: str) -> Optional[int]:
        """Находит номер колонки по заголовку"""
        header_variations = {
            'завтраки': ['завтрак', 'завтраки'],
            'холодные закуски и салаты': ['салат', 'холодн', 'закуск'],
            'первые блюда': ['первые', 'первое', 'блюда'],
            'блюда из мяса': ['мясо', 'мясн'],
            'блюда из курицы': ['курица', 'курин'],
            'блюда из птицы': ['птица', 'птицы'],
            'блюда из рыбы': ['рыба', 'рыбн'],
            'гарниры': ['гарнир']
        }
        
        # Ищем в первых 10 строках
        for row in range(1, min(11, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val:
                    cell_text = str(cell_val).lower().strip()
                    
                    # Проверяем точное совпадение
                    if header_text.lower() in cell_text:
                        return col
                    
                    # Проверяем вариации
                    if header_text.lower() in header_variations:
                        for variation in header_variations[header_text.lower()]:
                            if variation in cell_text:
                                return col
        
        return None
    
    def find_data_start_row(self, ws, col: int) -> int:
        """Находит строку начала данных в колонке (после заголовков)"""
        # Ищем первую строку с данными (не заголовком и не пустой)
        for row in range(1, min(20, ws.max_row + 1)):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val:
                cell_text = str(cell_val).lower().strip()
                # Пропускаем заголовки
                if any(header in cell_text for header in ['блюда', 'завтрак', 'салат', 'мясо', 'курица', 'птица', 'рыба', 'гарнир', 'первые']):
                    continue
                # Пропускаем служебные строки
                if any(skip in cell_text for skip in ['вес', 'цена', 'руб', 'изм']):
                    continue
                # Это похоже на блюдо
                if len(cell_text) > 8:
                    return row
        
        return 6  # По умолчанию начинаем с 6-й строки
    
    def handle_caesar_salad(self, dish_name: str) -> List[str]:
        """Обрабатывает Цезарь салат, разделяя по слэшу"""
        if 'цезарь' in dish_name.lower() and '/' in dish_name:
            parts = dish_name.split('/')
            result = []
            base = parts[0].strip()
            for i, part in enumerate(parts):
                if i == 0:
                    result.append(part.strip())
                else:
                    result.append(f"{base.split()[0]} {part.strip()}")
            return result
        return [dish_name]
    
    def fill_template_column(self, ws, col: int, dishes: List[str], start_row: int) -> int:
        """Заполляет колонку блюдами, начиная с указанной строки"""
        current_row = start_row
        filled_count = 0
        
        for dish in dishes:
            # Обрабатываем Цезарь салат
            dish_variants = self.handle_caesar_salad(dish)
            
            for variant in dish_variants:
                if current_row > ws.max_row:
                    break
                    
                # Получаем ячейку
                cell = ws.cell(row=current_row, column=col)
                
                # Проверяем, является ли ячейка объединенной
                if hasattr(cell, 'coordinate'):
                    # Проверяем, есть ли эта ячейка в списке объединенных ячеек
                    is_merged = False
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            is_merged = True
                            break
                    
                    # Пропускаем объединенные ячейки, кроме главной
                    if is_merged and cell.__class__.__name__ == 'MergedCell':
                        current_row += 1
                        continue
                
                # Заполняем только пустые ячейки
                current_val = cell.value
                if not current_val or str(current_val).strip() == '':
                    try:
                        cell.value = variant
                        filled_count += 1
                    except AttributeError:
                        # Если не можем записать в эту ячейку, пропускаем
                        pass
                
                current_row += 1
        
        return filled_count
    
    def extract_date_from_menu(self, menu_path: str) -> Optional[datetime]:
        """Извлекает дату из файла меню (используем уже готовую логику)"""
        return extract_date_from_menu(menu_path)
    
    def update_template_date(self, ws, menu_date: Optional[datetime]):
        """Обновляет дату в шаблоне"""
        if not menu_date:
            menu_date = datetime.now()
        
        # Ищем ячейки с датой в первых строках
        for row in range(1, min(6, ws.max_row + 1)):
            for col in range(1, min(5, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                cell_val = cell.value
                if cell_val:
                    cell_text = str(cell_val).lower()
                    # Если в ячейке есть упоминание месяца или это похоже на дату
                    if any(month in cell_text for month in ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 
                                                            'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']):
                        # Обновляем дату
                        russian_months = {
                            1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля', 5: 'мая', 6: 'июня',
                            7: 'июля', 8: 'августа', 9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'
                        }
                        date_str = f"{menu_date.day} {russian_months.get(menu_date.month, 'сентября')}"
                        
                        # Проверяем, не является ли ячейка объединенной MergedCell
                        if cell.__class__.__name__ == 'MergedCell':
                            continue  # Пропускаем объединенные ячейки
                        
                        try:
                            cell.value = date_str
                            return
                        except AttributeError:
                            # Если не можем записать, продолжаем поиск
                            continue
    
    def fill_menu_template(self, template_path: str, source_menu_path: str, output_path: str) -> Tuple[bool, str]:
        """
        Заполняет шаблон меню данными из исходного файла
        
        Args:
            template_path: Путь к шаблону меню
            source_menu_path: Путь к файлу-источнику данных  
            output_path: Путь к выходному файлу
            
        Returns:
            (success, message): Кортеж с результатом операции
        """
        try:
            # Проверяем существование файлов
            if not Path(template_path).exists():
                return False, f"Шаблон не найден: {template_path}"
            
            if not Path(source_menu_path).exists():
                return False, f"Исходный файл не найден: {source_menu_path}"
            
            # Извлекаем данные из исходного файла
            print(f"Извлекаем данные из файла: {source_menu_path}")
            categories = self.extract_categorized_dishes(source_menu_path)
            
            # Отладочный вывод
            total_dishes = sum(len(dishes) for dishes in categories.values())
            print(f"Извлечено блюд по категориям:")
            for cat, dishes in categories.items():
                print(f"  {cat}: {len(dishes)} блюд")
            print(f"Всего блюд: {total_dishes}")
            
            if total_dishes == 0:
                return False, "Не удалось извлечь блюда из исходного файла"
            
            # Извлекаем дату из исходного файла
            menu_date = self.extract_date_from_menu(source_menu_path)
            
            # Открываем шаблон
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # Обновляем дату в шаблоне
            self.update_template_date(ws, menu_date)
            
            # Заполняем колонки данными
            total_filled = 0
            categories_filled = 0
            
            for category, dishes in categories.items():
                if not dishes:
                    continue
                
                # Находим соответствующую колонку в шаблоне
                template_header = self.categories_mapping.get(category)
                if not template_header:
                    print(f"Предупреждение: Не найдено соответствие для категории '{category}'")
                    continue
                
                col = self.find_column_by_header(ws, template_header)
                if not col:
                    print(f"Предупреждение: Не найдена колонка для '{template_header}'")
                    continue
                
                # Находим начальную строку для данных
                start_row = self.find_data_start_row(ws, col)
                
                # Заполняем колонку
                filled_count = self.fill_template_column(ws, col, dishes, start_row)
                total_filled += filled_count
                categories_filled += 1
                
                print(f"Категория '{category}' -> колонка {col}: добавлено {filled_count} блюд")
            
            # Сохраняем результат
            wb.save(output_path)
            
            date_str = menu_date.strftime("%d.%m.%Y") if menu_date else "текущая дата"
            message = f"Шаблон меню заполнен для даты {date_str}\n"
            message += f"Заполнено категорий: {categories_filled}\n"
            message += f"Всего добавлено блюд: {total_filled}"
            
            return True, message
            
        except Exception as e:
            return False, f"Ошибка при заполнении шаблона: {str(e)}"


def fill_menu_template_from_source(template_path: str, source_menu_path: str, output_path: str) -> Tuple[bool, str]:
    """Удобная функция для заполнения шаблона меню"""
    filler = MenuTemplateFiller()
    return filler.fill_menu_template(template_path, source_menu_path, output_path)
