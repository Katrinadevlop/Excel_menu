#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import pandas as pd

def debug_excel_structure():
    """Debug the Excel file structure to understand column layout."""
    menu_file = 'templates/Шаблон меню пример.xlsx'
    
    print("=== DEBUGGING EXCEL STRUCTURE ===\n")
    
    try:
        # Используем openpyxl для точного анализа
        wb = openpyxl.load_workbook(menu_file, data_only=True)
        ws = wb.active
        
        print(f"Worksheet name: {ws.title}")
        print(f"Max rows: {ws.max_row}, Max columns: {ws.max_column}\n")
        
        # Найдем строку с заголовками
        header_row = None
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_text = ''
            for col_idx in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    row_text += str(cell_val).strip() + ' '
            
            if 'ЗАВТРАКИ' in row_text.upper():
                header_row = row_idx
                break
        
        if header_row:
            print(f"Header row found at: {header_row}")
            print("Header row content:")
            for col_idx in range(1, min(ws.max_column + 1, 10)):
                cell_val = ws.cell(row=header_row, column=col_idx).value
                if cell_val:
                    print(f"  Column {col_idx} ({chr(64+col_idx)}): {cell_val}")
            print()
            
            # Показываем первые несколько строк после заголовка
            print("First 15 rows after header:")
            for row_idx in range(header_row + 1, min(header_row + 16, ws.max_row + 1)):
                print(f"Row {row_idx}:")
                for col_idx in range(1, min(ws.max_column + 1, 8)):
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    col_letter = chr(64 + col_idx)
                    if cell_val:
                        print(f"  {col_letter}: {str(cell_val)[:50]}")
                print("  ---")
        else:
            print("Header row not found!")
            
    except Exception as e:
        print(f"Error: {e}")

if __name__ == '__main__':
    debug_excel_structure()
